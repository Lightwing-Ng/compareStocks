"""
Shared web runtime and route handlers.

Code version: v0.73.2
- Fixed: Investment intraday requests refresh a completed requested trading
  day when its regular-session bar set is incomplete, instead of treating a
  partial but date-current store as fresh and carrying one close across the
  rest of the equity curve.
- Fixed: When Yahoo does not provide a complete completed trading day,
  Investment intraday requests use the configured Longbridge one-minute
  fallback instead of carrying the partial day's last close.
- Added: Shared local pagination ellipses now carry grouped hidden-page ranges
  for the accessible range picker rendered on every pagination surface.
- Fixed: All configured money-market funds use the same standard MMF token in
  Cash-equivalent Settings; quote currency remains a separate text badge.
- Added: Cash-equivalent Settings separates configurable listed securities
  from the configured money-market funds and exposes the latter with their
  standard token identities.
- Fixed: Investment intraday requests fall back to the configured Longbridge
  one-minute source when Yahoo cannot fill a requested active trading day.
- Fixed: The Investment transactions cache schema now invalidates payloads
  generated before the verified HSBC DRAM and EUV tax-lot conventions were
  added, preventing a backend restart from retaining the stale summary.
- Fixed: Daily investment close payloads reject non-finite and non-positive
  values and deterministically collapse duplicate ticker/date rows before the
  frontend builds its valuation index.
- Added: Longbridge paired-file imports preserve the exact uploaded Fund
  Details and History Orders bytes as immutable source evidence.
- Fixed: Investment intraday history responses exclude non-positive and malformed OHLC bars without rewriting local stores.
- Added: Settings -> Investment persists one shared buy/sell lot-matching preference and exposes it in every Investment payload.
- Added: Settings URL state uses canonical section paths, language tabs, and pagination with legacy query aliases readable during migration.
- Added: CSRF-protected Schwab security-transfer source-account confirmation persists metadata only and refreshes fail-closed aggregate reconciliation.
- Changed: Local Market Store pagination now mirrors the shared Investment page-boundary controls.
- Added: BOCHK Consolidated Statement PDF imports preserve HKD, CNY, and USD subaccounts across batches.
- Added: Investment payloads include date-aware USD FX history for HKD, CNY, and CNH cash conversion.
- Fixed: Investment FX payloads are bounded to the ledger date range instead of returning unused provider history.
"""

from __future__ import annotations
from collections.abc import Callable
from datetime import datetime
from http.client import RemoteDisconnected
import json
import logging
import re
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, cast
from urllib.parse import urlencode
import hashlib
import pandas as pd
from flask import g, jsonify, make_response, redirect, render_template, request, send_from_directory, session, url_for, send_file
from openpyxl import Workbook, load_workbook
from werkzeug.exceptions import RequestEntityTooLarge

from app.core.backtest_settings import load_backtest_execution_mode, save_backtest_execution_mode
from app.core.cash_equivalent_settings import (
    load_cash_equivalent_tickers,
    save_cash_equivalent_tickers,
)
from app.core.investment_settings import (
    load_investment_cost_basis_method,
    save_investment_cost_basis_method,
)
from app.core.debug_reporting import load_optional_debug_endpoint, post_debug_event
from app.core.date_display_settings import (
    load_date_display_settings,
    save_full_date_display_format,
    save_short_date_display_format,
)
from app.core.language_settings import (
    HTML_LANG_BY_LANGUAGE,
    LANGUAGE_LABELS,
    SUPPORTED_LANGUAGE_CODES,
    build_translation_map,
    load_language_settings,
    save_language_code,
    save_language_settings,
    translate_nested_text,
    translate_labels,
    translate_text,
)
from app.core.live_trading_security import (
    LIVE_TRADING_TOKEN_HEADER,
    authorize_live_trading_api_request,
    validate_live_trading_pin,
)
from app.web.request_security import validate_investment_browser_write_request
from app.infrastructure.broker_market_data import (
    classify_daily_store_status,
    classify_one_minute_store_status,
    has_recent_one_minute_store,
    is_one_minute_store_fresh,
    one_minute_lookback_start,
    test_broker_connection,
)
from app.infrastructure.longbridge_cli import (
    get_longbridge_cli_auth_status,
    start_longbridge_cli_browser_oauth,
    test_longbridge_cli_connection,
)
from app.core.broker_settings import (
    BrokerSettings,
    load_broker_settings,
    sanitize_broker_settings_for_view,
    save_broker_settings,
)
from app.services.comparisons import (
    align_intraday_datasets_for_compare,
    build_series_payload,
    calculate_ttm_dividend_yield,
    complete_market_local_trading_days,
    fill_intraday_market_session_gaps,
    filter_intraday_dataset_to_regular_session,
    market_trading_date_for_timestamp,
    prepare_intraday_dataset_for_compare,
    resolve_effective_period_for_datasets,
    shift_intraday_compare_axis_to_trading_date,
    slice_dataset_for_period,
    slice_datasets_for_compare_period,
    slice_intraday_datasets_for_compare_period,
)
from app.core.email_settings import (
    SmtpSettings,
    YAHOO_SMTP_HOST,
    YAHOO_SMTP_PORT,
    clear_oauth_settings,
    load_smtp_settings,
    sanitize_smtp_settings_for_view,
    save_smtp_settings,
    test_smtp_connection,
)
from strategies.backtest import run_single_ticker_backtest
from strategies.loader import instantiate_strategy, list_enabled_strategies, get_strategy_definition
from app.infrastructure.connectivity import (
    has_remote_market_access,
    network_transport_note,
    run_network_self_check,
    reset_connectivity_caches,
)
from app.core.config import (
    BASE_CURRENCY,
    BASE_TIMEZONE,
    CODE_VERSION,
    DEFAULT_INTERVAL,
    DEFAULT_PERIOD,
    DEFAULT_TICKERS,
    COMPARE_PERIODS_1D,
    MARKET_STORE_DIR,
    PERIOD_DAY_SPANS,
    PERIOD_LABELS,
    PERIOD_MONTH_SPANS,
    PERIOD_OFFSETS,
    SETTINGS_STORE_DIR,
    SUPPORTED_PERIODS_1D,
    SUPPORTED_PERIODS_1M,
)
from app.core.upload_limits import MAX_INVESTMENT_IMPORT_REQUEST_MIB
from app.services.date_constraints import (
    build_date_constraint_payload,
    build_date_constraint_availability,
    is_nyse_early_close,
    latest_completed_nyse_trading_day,
    nyse_market_session_state,
    nyse_recent_trading_days,
)
from app.services.dca import simulate_recurring_investment
from app.services.market_cap import (
    build_market_cap_series_payload,
    extract_stock_split_events,
    fetch_usd_exchange_rate_history,
)
from app.services.range_options import (
    COMPARE_INTRADAY_PERIODS,
    build_supported_compare_periods,
    build_supported_periods_from_dates,
    resolve_requested_period_from_supported,
)
from app.services.investment_import import (
    build_investment_internal_transfer_binding_index,
    merge_investment_payloads,
    normalize_investment_internal_transfer_bindings,
    normalize_investment_internal_transfer_ignored_source_keys,
    normalize_investment_payload_tickers,
    normalize_investment_security_transfer_attributions,
    parse_investment_payload,
    refresh_investment_security_transfer_reconciliation,
    validate_hsbc_pasted_text,
    validate_investment_internal_transfer_binding,
    validate_investment_security_transfer_attribution,
)
from app.services.investment_import_registry import commit_investment_import
from app.services.zircon_hk_import import (
    STANDARD_INVESTMENT_EXPORT_FILENAME,
    ZIRCON_HK_MAX_TRANSACTION_ROWS,
    ZIRCON_HK_TEMPLATE_FILENAME,
    build_standard_investment_xlsx,
    build_zircon_hk_template_xlsx,
)

from app.services.live_trading import (
    load_longbridge_account_balances,
    load_longbridge_account_label,
    load_longbridge_stock_positions,
    submit_longbridge_limit_order,
)
from app.services.logos import fetch_quote_profile, has_valid_ticker_format, normalize_ticker_input, refresh_quote_profile_cache, \
    resolve_stored_logo_url, search_tickers
from app.services.market_data import (
    canonical_compare_overnight_ticker,
    fetch_compare_one_day_extended_history,
    fetch_compare_one_day_overnight_history,
    fetch_history,
    fetch_longbridge_realtime_quotes,
    fetch_one_minute_history_for_trading_date,
    fetch_yfinance_realtime_quotes,
    has_compare_overnight_market_data_source,
    infer_ticker_market,
    list_available_market_intervals,
    refresh_history_store,
    refresh_one_minute_store,
    refresh_one_minute_store_with_longbridge,
    refresh_recent_one_minute_store_with_yfinance,
    resolve_compare_overnight_tickers,
    select_price_series,
    supports_compare_extended_hours,
    supports_compare_overnight,
)
from app.services.market_freshness import (
    ensure_latest_daily_caches,
    ensure_latest_investment_daily_caches,
    extract_open_investment_tickers,
)
from app.services.market_freshness import ensure_latest_backtest_caches
from app.models.schemas import DateConstraintPayload, QuoteProfile, SeriesPayload
from app.services.presentation import (
    build_series_colors,
    format_display_date,
    format_display_datetime,
    format_period_label,
    format_short_display_date,
    hex_to_rgba,
)
from app.core.settings import get_settings
from app.infrastructure.storage import (
    INVESTMENT_STORE_PATH,
    LOGOS_STORE_DIR,
    clear_non_historical_market_cache,
    clear_investment_store,
    delete_ticker_data,
    has_logo_asset,
    has_profile_record,
    history_store_path_for,
    investment_store_exists,
    investment_store_path_for,
    intraday_history_store_path_for,
    is_ticker_fallback_company_name,
    list_local_tickers,
    list_historical_tickers,
    load_investment_store_payload,
    load_profile_record,
    market_ticker_store_aliases,
    market_store_file_lock,
    materialize_investment_source_artifacts,
    investment_ticker_identity_store_aliases,
    investment_ticker_lineage_payload,
    investment_ticker_store_aliases,
    known_ticker_company_names_payload,
    normalize_ticker,
    propagate_investment_lineage_identity_profiles,
    resolve_known_ticker_company_name,
    record_ticker_usage,
    record_strategy_usage,
    save_investment_store_payload,
    top_used_strategies,
    update_investment_store_payload,
    verify_investment_source_artifacts,
    write_json_atomic,
)
from app.web.form_parsing import (
    build_default_weights,
    ensure_positive_portfolio_weights,
    normalize_portfolio_weights,
    parse_bool_flag_from_args,
    parse_float_value,
    parse_int_value,
    parse_portfolio_allocation_mode_from_args,
    parse_range_request_args_from_args,
    parse_requested_shares_from_args,
    parse_requested_tickers_from_args,
    parse_requested_weights_from_args,
    resolve_workspace_dividend_mode,
)
from app.web.navigation import (
    BACKTEST_VIEWS,
    MAX_TICKERS,
    MIN_TICKERS,
    VIEW_PATHS,
    build_settings_path,
    build_settings_state_url,
    build_settings_url,
    build_trade_path,
    build_trade_url,
    build_view_path,
    build_view_url,
    max_tickers_for_view,
    normalize_settings_page,
    normalize_settings_section,
    normalize_settings_tab,
    normalize_trade_section,
    normalize_view_name,
)
from app.web.market_history import (
    align_datasets_on_common_dates,
    build_supported_periods_for_history_store,
    extract_union_dates,
)
from app.web.strategy_forms import (
    build_strategy_form_fields as build_strategy_form_fields_for_strategy,
    build_strategy_option_groups as build_strategy_option_groups_for_recent,
    build_strategy_settings_rows as build_strategy_settings_rows_for_factory,
)
from app.web.style_token_rows import (
    build_color_token_rows,
    build_export_image_rows,
    build_font_token_rows,
    build_material_token_rows,
    build_style_token_rows,
)

LOGGER = logging.getLogger(__name__)
FETCH_ABORT_DEBUG_CONFIG = load_optional_debug_endpoint(
    "frontend-fetch-aborts.env",
    "frontend-fetch-aborts",
)
PROJECT_SOURCE_URL = "https://github.com/Lightwing-Ng/compareStocks"
PROJECT_DISPLAY_URL = PROJECT_SOURCE_URL.removeprefix("https://").removeprefix("http://")


def report_fetch_abort_debug_event(
    hypothesis_id: str,
    location: str,
    msg: str,
    data: dict[str, Any] | None = None,
    run_id: str = "post-fix",
) -> None:
    # #region debug-point E:backend-fetch-abort
    post_debug_event(
        FETCH_ABORT_DEBUG_CONFIG,
        hypothesis_id=hypothesis_id,
        location=location,
        msg=msg,
        data=data,
        run_id=run_id,
        timeout_seconds=0.5,
    )
    # #endregion

PORTFOLIO_BENCHMARK_TICKERS = ("SPY", "QQQ")
INVESTMENT_TRANSACTIONS_CACHE_SCHEMA_VERSION = "investment-transactions-v11"
INVESTMENT_TRANSACTIONS_CACHE_PATH = SETTINGS_STORE_DIR / "investment_cache" / "transactions_payload.json"
INVESTMENT_REALTIME_QUOTE_TTL_SECONDS = 60.0
INVESTMENT_REALTIME_QUOTE_TIMEOUT_SECONDS = 30
REALTIME_BATCH_SIZE = 8
PORTFOLIO_BENCHMARK_COLORS = {
    "SPY": "#8e8e93",
    "QQQ": "#c7c7cc",
}
LOCAL_STORE_PAGE_SIZE = 10
SETTINGS_LANGUAGE_PAGE_SIZE = 10
SETTINGS_FEEDBACK_COOKIE = "antigravity_settings_feedback"
@dataclass(frozen=True)
class WebRuntime:
    """Callable handlers and helpers shared across split route modules."""

    root: Any
    compare_page: Any
    market_cap_compare_page: Any
    legacy_compare_page: Any
    price_compare_page: Any
    portfolio_page: Any
    legacy_portfolio_page: Any
    dca_page: Any
    legacy_dca_page: Any
    backtest_page: Any
    grid_trading_page: Any
    legacy_backtest_page: Any
    legacy_trade_messages_page: Any
    trade_root: Any
    trade_page: Any
    legacy_trade_root: Any
    legacy_trade_page: Any
    live_trading_unlock: Any
    settings_root: Any
    settings_page: Any
    export_transactions_api: Any
    general_settings_action: Any
    language_settings_api: Any
    language_cycle_api: Any
    language_download_api: Any
    backtest_settings_action: Any
    investment_settings_action: Any
    cash_equivalents_action: Any
    email_smtp_action: Any
    broker_access_action: Any
    longbridge_oauth_status_api: Any
    local_market_store_action: Any
    settings_cache_action: Any
    market_store_logo: Any
    favicon_icon: Any
    symbol_search: Any
    date_constraints_api: Any
    compare_live_api: Any
    trade_strategy_fields_api: Any
    settings_network_status_api: Any
    local_market_store_page_data_api: Any
    market_store_presence_api: Any
    investment_page: Any
    investment_get_transactions: Any
    investment_add_transaction: Any
    investment_download_zircon_hk_template: Any
    investment_export_standard_xlsx: Any
    investment_validate_zircon_hk_workbook: Any
    investment_validate_hsbc_pasted_text: Any
    investment_get_latest_price: Any
    investment_get_parquet: Any
    investment_get_intraday_history: Any
    investment_get_market_session: Any
    investment_get_realtime_quotes: Any
    investment_update_internal_transfer_binding: Any
    investment_update_security_transfer_attribution: Any
    live_trading_get_positions: Any
    live_trading_submit_order: Any


def extract_first_non_null_value(raw_value: object) -> object | None:
    if raw_value is None:
        return None
    if isinstance(raw_value, pd.DataFrame):
        if raw_value.empty:
            return None
        for column in raw_value.columns:
            extracted = extract_first_non_null_value(raw_value[column])
            if extracted is not None:
                return extracted
        return None
    if isinstance(raw_value, pd.Series):
        values = raw_value.dropna()
        if values.empty:
            return None
        return extract_first_non_null_value(values.iloc[0])
    if isinstance(raw_value, pd.Index):
        values = raw_value.dropna()
        if len(values) == 0:
            return None
        return extract_first_non_null_value(values[0])
    if isinstance(raw_value, (list, tuple)):
        for value in raw_value:
            extracted = extract_first_non_null_value(value)
            if extracted is not None:
                return extracted
        return None
    if hasattr(raw_value, "ndim") and hasattr(raw_value, "tolist") and not pd.api.types.is_scalar(raw_value):
        values = raw_value.tolist()
        return extract_first_non_null_value(values)
    return raw_value


def format_store_range_date_value(raw_value: object) -> str:
    candidate = extract_first_non_null_value(raw_value)
    if candidate is None:
        return ""
    timestamp = pd.Timestamp(candidate)
    if pd.isna(timestamp):
        return ""
    return format_short_display_date(timestamp)


def build_web_runtime() -> WebRuntime:
    settings = get_settings()
    defaults = settings["defaults"]
    base_labels = settings["ui"]["labels"]
    theme_settings = settings["ui"]["theme"]
    theme_light = theme_settings["light"]
    theme_dark = theme_settings["dark"]
    theme = theme_light
    chart_config = settings["ui"]["chart"]
    logos = settings["ui"]["logos"]
    app_meta = settings["app"]
    live_trading_pin = settings.get("security", {}).get("live_trading_pin", "")
    investment_settings = settings.get("investment", {}) if isinstance(settings.get("investment"), dict) else {}
    money_market_settings = (
        investment_settings.get("money_market_funds", {})
        if isinstance(investment_settings.get("money_market_funds"), dict)
        else {}
    )

    def quote_profile_to_json(profile: QuoteProfile) -> dict[str, str | None]:
        return {
            "ticker": profile.ticker,
            "company_name": profile.company_name,
            "website": profile.website,
            "logo_url": profile.logo_url,
        }

    def date_constraint_payload_to_json(
            payload: DateConstraintPayload,
    ) -> dict[str, object]:
        return {
            "min_date": payload.min_date,
            "max_date": payload.max_date,
            "trading_dates": list(payload.trading_dates),
            "adjusted_start": payload.adjusted_start,
            "adjusted_end": payload.adjusted_end,
            "message": payload.message,
            "availability": payload.availability,
        }

    def annotate_date_constraint_availability(
            payload: DateConstraintPayload,
            tickers: list[str],
            datasets: list[pd.DataFrame],
    ) -> DateConstraintPayload:
        payload.availability = build_date_constraint_availability(payload, tickers, datasets)
        return payload

    def fetch_request_compare_one_day_overnight_history(
            ticker: str,
            *,
            trading_date: object | None = None,
    ) -> pd.DataFrame:
        cache = getattr(g, "compare_overnight_history_cache", None)
        if cache is None:
            cache = {}
            g.compare_overnight_history_cache = cache
        parsed_trading_date = pd.to_datetime(trading_date, errors="coerce") if trading_date is not None else None
        cache_date = "" if parsed_trading_date is None or pd.isna(parsed_trading_date) else parsed_trading_date.date().isoformat()
        cache_key = (normalize_ticker_input(ticker), cache_date)
        if cache_key not in cache:
            cache[cache_key] = fetch_compare_one_day_overnight_history(
                ticker,
                trading_date=trading_date,
            )
        return cache[cache_key].copy()

    def market_local_trading_dates_frame(dataset: pd.DataFrame, ticker: str) -> pd.DataFrame:
        if dataset.empty or "Date" not in dataset.columns:
            return pd.DataFrame({"Date": pd.Series(dtype="datetime64[ns]")})
        def market_local_date(value: object) -> object:
            return market_trading_date_for_timestamp(value, ticker)

        dates = dataset["Date"].map(market_local_date).dropna().drop_duplicates().sort_values()
        return pd.DataFrame({"Date": pd.to_datetime(dates)})

    def format_compare_intraday_market_local_display_range(
            datasets: list[pd.DataFrame],
            tickers: list[str],
    ) -> str:
        local_dates: list[pd.Timestamp] = []
        for index, dataset in enumerate(datasets):
            ticker = tickers[index] if index < len(tickers) else ""
            frame = market_local_trading_dates_frame(dataset, ticker)
            if frame.empty:
                continue
            local_dates.extend(pd.to_datetime(frame["Date"], errors="coerce").dropna().tolist())
        if not local_dates:
            return ""
        return f"{format_display_date(min(local_dates))} - {format_display_date(max(local_dates))}"

    def build_one_day_intraday_date_constraint_payload(
            tickers: list[str],
            requested_start: str | None = None,
            requested_end: str | None = None,
            include_overnight_flag: bool = False,
    ) -> DateConstraintPayload:
        date_frames: list[pd.DataFrame] = []
        refresh_failures: list[str] = []
        live_session_date = pd.Timestamp.now(tz="Asia/Shanghai").date()
        has_live_session_date = False
        requested_dates = {
            parsed.date()
            for value in (requested_start, requested_end)
            if value and not pd.isna(parsed := pd.to_datetime(value, errors="coerce"))
        }
        for ticker in tickers:
            use_overnight_source = include_overnight_flag and infer_ticker_market(ticker) == "US"
            if use_overnight_source:
                try:
                    intraday_dataset = fetch_request_compare_one_day_overnight_history(
                        ticker,
                        trading_date=requested_start or requested_end,
                    )
                except (ImportError, OSError, ValueError, KeyError, TypeError) as exc:
                    LOGGER.warning("Unable to load overnight date constraints for %s: %s", ticker, exc)
                    refresh_failures.append(ticker)
                    intraday_dataset = fetch_history(
                        ticker,
                        include_dividends=False,
                        interval="1m",
                        dividend_mode="price",
                    )
            else:
                intraday_dataset = fetch_history(
                    ticker,
                    include_dividends=False,
                    interval="1m",
                    dividend_mode="price",
                )
            prepared_dataset = prepare_intraday_dataset_for_compare(
                intraday_dataset,
                ticker,
                regular_session_only=not use_overnight_source,
            )
            date_frame = market_local_trading_dates_frame(prepared_dataset, ticker)
            available_dates = set(date_frame["Date"].dt.date) if not date_frame.empty else set()
            should_refresh_public_data = (
                not use_overnight_source
                and (not requested_dates or not requested_dates.issubset(available_dates))
            )
            if should_refresh_public_data:
                try:
                    refresh_recent_one_minute_store_with_yfinance(ticker)
                    intraday_dataset = fetch_history(
                        ticker,
                        include_dividends=False,
                        interval="1m",
                        dividend_mode="price",
                    )
                    prepared_dataset = prepare_intraday_dataset_for_compare(
                        intraday_dataset,
                        ticker,
                        regular_session_only=True,
                    )
                    date_frame = market_local_trading_dates_frame(prepared_dataset, ticker)
                except (ImportError, OSError, ValueError, KeyError, TypeError) as exc:
                    LOGGER.warning("Unable to refresh 1-minute date constraints for %s: %s", ticker, exc)
                    refresh_failures.append(ticker)
            available_dates = set(date_frame["Date"].dt.date) if not date_frame.empty else set()
            missing_requested_dates = requested_dates - available_dates
            if not use_overnight_source:
                for requested_date in sorted(missing_requested_dates):
                    try:
                        exact_dataset = fetch_one_minute_history_for_trading_date(
                            ticker,
                            requested_date,
                            include_dividends=False,
                            dividend_mode="price",
                        )
                        prepared_exact_dataset = prepare_intraday_dataset_for_compare(
                            exact_dataset,
                            ticker,
                            regular_session_only=True,
                        )
                        exact_date_frame = market_local_trading_dates_frame(
                            prepared_exact_dataset,
                            ticker,
                        )
                        date_frame = (
                            pd.concat([date_frame, exact_date_frame], ignore_index=True)
                            .drop_duplicates(subset=["Date"])
                            .sort_values("Date")
                            .reset_index(drop=True)
                        )
                    except (ImportError, OSError, ValueError, KeyError, TypeError) as exc:
                        LOGGER.warning(
                            "Unable to load exact-day date constraints for %s on %s: %s",
                            ticker,
                            requested_date,
                            exc,
                        )
                        if ticker not in refresh_failures:
                            refresh_failures.append(ticker)
            has_live_session_date = has_live_session_date or bool(
                not date_frame.empty and (date_frame["Date"].dt.date == live_session_date).any()
            )
            date_frames.append(date_frame)

        payload = build_date_constraint_payload(
            *date_frames,
            requested_start=requested_start,
            requested_end=requested_end,
        )
        if has_live_session_date:
            live_date_value = pd.Timestamp(live_session_date).strftime("%Y-%m-%d")
            trading_dates = sorted({*payload.trading_dates, live_date_value})
            adjusted_start = payload.adjusted_start
            adjusted_end = payload.adjusted_end
            requested_start_date = pd.to_datetime(requested_start, errors="coerce") if requested_start else None
            requested_end_date = pd.to_datetime(requested_end, errors="coerce") if requested_end else None
            if requested_start_date is not None and not pd.isna(requested_start_date) and requested_start_date.date() == live_session_date:
                adjusted_start = live_date_value
            if requested_end_date is not None and not pd.isna(requested_end_date) and requested_end_date.date() == live_session_date:
                adjusted_end = live_date_value
            payload = DateConstraintPayload(
                min_date=payload.min_date or live_date_value,
                max_date=max([payload.max_date, live_date_value] if payload.max_date else [live_date_value]),
                trading_dates=trading_dates,
                adjusted_start=adjusted_start,
                adjusted_end=adjusted_end,
                message=payload.message,
            )
        if refresh_failures:
            failed_preview = ", ".join(refresh_failures)
            refresh_notice = (
                f"Could not refresh 1-minute trading dates for {failed_preview}. "
                "Using currently cached intraday dates."
            )
            payload.message = f"{payload.message} {refresh_notice}".strip() if payload.message else refresh_notice
        return annotate_date_constraint_availability(payload, tickers, date_frames)

    def build_short_intraday_date_constraint_payload(
            tickers: list[str],
            requested_start: str | None = None,
            requested_end: str | None = None,
    ) -> DateConstraintPayload:
        date_frames: list[pd.DataFrame] = []
        for ticker in tickers:
            intraday_dataset = fetch_history(
                ticker,
                include_dividends=False,
                interval="1m",
                dividend_mode="price",
            )
            prepared_dataset = prepare_intraday_dataset_for_compare(
                intraday_dataset,
                ticker,
                regular_session_only=True,
            )
            date_frames.append(market_local_trading_dates_frame(prepared_dataset, ticker))

        payload = build_date_constraint_payload(
            *date_frames,
            requested_start=requested_start,
            requested_end=requested_end,
        )
        return annotate_date_constraint_availability(payload, tickers, date_frames)

    def resolve_compare_axis_trading_date(
            tickers: list[str],
            requested_trading_date: object,
    ) -> str:
        requested_date = pd.to_datetime(requested_trading_date, errors="coerce")
        if pd.isna(requested_date):
            raise ValueError(f"Invalid compare trading date: {requested_trading_date}.")
        requested_date_value = requested_date.date()
        live_session_date = pd.Timestamp.now(tz="Asia/Shanghai").date()

        date_frames: list[pd.DataFrame] = []
        requested_date_is_complete = requested_date_value != live_session_date
        for ticker in tickers:
            intraday_dataset = fetch_history(
                ticker,
                include_dividends=False,
                interval="1m",
                dividend_mode="price",
            )
            prepared_dataset = prepare_intraday_dataset_for_compare(
                intraday_dataset,
                ticker,
                regular_session_only=True,
            )
            date_frame = market_local_trading_dates_frame(prepared_dataset, ticker)
            requested_date_is_complete = requested_date_is_complete and requested_date_value in complete_market_local_trading_days(
                prepared_dataset,
                ticker,
            )
            if not date_frame.empty:
                date_frame = date_frame[date_frame["Date"].dt.date < requested_date_value].copy()
            date_frames.append(date_frame)

        if requested_date_is_complete:
            return pd.Timestamp(requested_date_value).strftime("%Y-%m-%d")

        payload = build_date_constraint_payload(*date_frames)
        if not payload.trading_dates:
            return pd.Timestamp(requested_date_value).strftime("%Y-%m-%d")
        return payload.max_date or payload.trading_dates[-1]

    def canonicalize_money_market_ticker(value: object) -> str:
        raw_ticker = str(value or "").strip().upper()
        aliases = investment_ticker_store_aliases(raw_ticker)
        return str(aliases[0] if aliases else raw_ticker).strip().upper()

    configured_money_market_tickers = {
        canonicalize_money_market_ticker(value)
        for value in money_market_settings.get("tickers", [])
        if str(value).strip()
    }
    configured_money_market_quote_currencies = {
        canonicalize_money_market_ticker(ticker): str(currency).strip().upper()
        for ticker, currency in money_market_settings.get("quote_currency_overrides", {}).items()
        if (
            canonicalize_money_market_ticker(ticker) in configured_money_market_tickers
            and len(str(currency).strip()) == 3
        )
    }
    money_market_name_from_description = bool(money_market_settings.get("name_from_description", False))
    money_market_description_keywords = [
        str(value).strip().upper()
        for value in money_market_settings.get("description_keywords", [])
        if str(value).strip()
    ]

    def exclude_configured_money_market_tickers(tickers: list[str]) -> list[str]:
        return [
            ticker
            for ticker in tickers
            if not is_configured_money_market_ticker(ticker)
        ]

    investment_daily_refresh_lock = threading.Lock()
    investment_realtime_quote_cache_lock = threading.Lock()
    investment_realtime_quote_cache: dict[tuple[str, ...], tuple[float, list[dict[str, object]]]] = {}

    def is_configured_money_market_ticker(ticker: str) -> bool:
        return canonicalize_money_market_ticker(ticker) in configured_money_market_tickers

    def get_cash_equivalent_tickers() -> set[str]:
        try:
            raw = load_cash_equivalent_tickers()
            return {
                str(value).strip().upper()
                for value in raw
                if str(value).strip()
            }
        except Exception:  # noqa: BLE001
            return {"BOXX", "SGOV"}

    def apply_no_store_headers(response):
        response.headers["Cache-Control"] = "no-store, no-cache, max-age=0, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        return response

    def live_trading_api_authorization_failure_response():
        access_granted, error_status, error_message = authorize_live_trading_api_request(
            bool(session.get("live_trading_unlocked")),
            request.headers.get(LIVE_TRADING_TOKEN_HEADER),
        )
        if access_granted:
            return None

        response = jsonify({"success": False, "error": error_message})
        response.status_code = error_status
        if error_status == 401:
            response.headers["WWW-Authenticate"] = 'Bearer realm="antigravity-live-trading"'
        return apply_no_store_headers(response)

    def ensure_investment_transactions_cache_dir() -> None:
        INVESTMENT_TRANSACTIONS_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)

    def invalidate_investment_transactions_cache() -> None:
        try:
            ensure_investment_transactions_cache_dir()
            with market_store_file_lock(INVESTMENT_TRANSACTIONS_CACHE_PATH):
                if INVESTMENT_TRANSACTIONS_CACHE_PATH.exists():
                    INVESTMENT_TRANSACTIONS_CACHE_PATH.unlink()
        except OSError:
            LOGGER.warning(
                "Unable to invalidate the derived investment transaction cache; "
                "continuing with the portable investment ledger.",
                exc_info=True,
            )

    def build_file_fingerprint(path: Path) -> dict[str, object]:
        if not path.exists():
            return {
                "exists": False,
                "path": str(path),
                "size": 0,
                "mtime_ns": 0,
            }
        stat_result = path.stat()
        return {
            "exists": True,
            "path": str(path),
            "size": stat_result.st_size,
            "mtime_ns": stat_result.st_mtime_ns,
        }

    def build_investment_price_store_fingerprints(
            transactions: list[dict[str, Any]],
            open_tickers: list[str] | set[str] | tuple[str, ...],
    ) -> list[dict[str, object]]:
        open_ticker_set = {
            normalize_ticker_input(str(ticker))
            for ticker in (open_tickers or [])
            if str(ticker or "").strip()
        }
        fingerprints: list[dict[str, object]] = []
        for ticker in collect_investment_display_tickers(transactions):
            if is_configured_money_market_ticker(ticker):
                continue
            path = resolve_investment_history_store_path(ticker)
            fingerprints.append({
                "ticker": ticker,
                "is_open": ticker in open_ticker_set,
                "store": build_file_fingerprint(path) if path is not None else {
                    "exists": False,
                    "path": "",
                    "size": 0,
                    "mtime_ns": 0,
                },
            })
        return fingerprints

    def read_investment_transactions_cache(
            investment_store_fingerprint: dict[str, object],
    ) -> dict[str, Any] | None:
        try:
            ensure_investment_transactions_cache_dir()
            with market_store_file_lock(INVESTMENT_TRANSACTIONS_CACHE_PATH):
                if not INVESTMENT_TRANSACTIONS_CACHE_PATH.exists():
                    return None
                with open(INVESTMENT_TRANSACTIONS_CACHE_PATH, "r", encoding="utf-8") as f:
                    cached = json.load(f)
        except (json.JSONDecodeError, OSError, TypeError):
            return None

        if cached.get("schema_version") != INVESTMENT_TRANSACTIONS_CACHE_SCHEMA_VERSION:
            return None
        if cached.get("investment_store") != investment_store_fingerprint:
            return None
        payload = cached.get("payload")
        if not isinstance(payload, dict):
            return None
        section_freshness = payload.get("section_freshness")
        if not isinstance(section_freshness, dict):
            return None
        target_trading_day = latest_completed_nyse_trading_day().strftime("%Y-%m-%d")
        if section_freshness.get("target_trading_day") != target_trading_day:
            return None
        transactions = payload.get("transactions", [])
        if not isinstance(transactions, list):
            return None
        price_store_fingerprints = build_investment_price_store_fingerprints(
            cast(list[dict[str, Any]], transactions),
            section_freshness.get("open_tickers") or [],
        )
        if cached.get("price_stores") != price_store_fingerprints:
            return None
        return cast(
            dict[str, Any],
            refresh_investment_security_transfer_reconciliation(
                normalize_investment_payload_tickers(payload)
            ),
        )

    def write_investment_transactions_cache(
            *,
            investment_store_fingerprint: dict[str, object],
            price_store_fingerprints: list[dict[str, object]],
            payload: dict[str, Any],
    ) -> None:
        cache_payload = {
            "schema_version": INVESTMENT_TRANSACTIONS_CACHE_SCHEMA_VERSION,
            "investment_store": investment_store_fingerprint,
            "price_stores": price_store_fingerprints,
            "payload": payload,
        }
        try:
            ensure_investment_transactions_cache_dir()
            with market_store_file_lock(INVESTMENT_TRANSACTIONS_CACHE_PATH):
                write_json_atomic(INVESTMENT_TRANSACTIONS_CACHE_PATH, cache_payload)
        except OSError:
            LOGGER.warning(
                "Unable to write the derived investment transaction cache; "
                "continuing with the portable investment ledger.",
                exc_info=True,
            )

    def load_normalized_investment_payload() -> dict[str, Any]:
        return refresh_investment_security_transfer_reconciliation(
            normalize_investment_payload_tickers(
                load_investment_store_payload(INVESTMENT_STORE_PATH)
            )
        )

    def write_investment_payload(payload: dict[str, Any]) -> None:
        normalized_payload = refresh_investment_security_transfer_reconciliation(
            normalize_investment_payload_tickers(payload)
        )
        save_investment_store_payload(cast(dict[str, Any], normalized_payload), INVESTMENT_STORE_PATH)
        invalidate_investment_transactions_cache()

    def merge_and_write_investment_payload(imported_payload: dict[str, Any]) -> dict[str, Any]:
        def is_test_account_identifier(value: object) -> bool:
            normalized = re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())
            return normalized.endswith("TEST") or normalized.endswith("E2E")

        imported_broker = str(imported_payload.get("broker") or "").strip().lower()
        imported_records = [
            record
            for record in imported_payload.get("transactions", [])
            if isinstance(record, dict)
        ]
        has_test_account = imported_broker == "ibkr" and (
            is_test_account_identifier(imported_payload.get("account"))
            or any(
                is_test_account_identifier(record.get("account"))
                or is_test_account_identifier(
                    record.get("source", {}).get("account")
                    if isinstance(record.get("source"), dict)
                    else ""
                )
                for record in imported_records
            )
        )
        if has_test_account:
            raise ValueError(
                "Refusing to persist an IBKR test-fixture account into the investment store."
            )

        def normalize_payload(payload: dict[str, Any]) -> dict[str, Any]:
            return cast(dict[str, Any], normalize_investment_payload_tickers(payload))

        def update_store(updater):
            return cast(
                dict[str, Any],
                update_investment_store_payload(updater, INVESTMENT_STORE_PATH),
            )

        return commit_investment_import(
            imported_payload,
            normalize_payload=normalize_payload,
            merge_payloads=merge_investment_payloads,
            update_store=update_store,
            load_store=load_normalized_investment_payload,
            invalidate_cache=invalidate_investment_transactions_cache,
            materialize_payload=lambda payload: materialize_investment_source_artifacts(
                payload,
                INVESTMENT_STORE_PATH,
            ),
            verify_persisted_payload=lambda payload: verify_investment_source_artifacts(
                payload,
                INVESTMENT_STORE_PATH,
            ),
        )

    def refresh_investment_import_price_caches(
        imported_payload: dict[str, Any],
    ) -> list[str]:
        try:
            return ensure_latest_investment_daily_caches(
                exclude_configured_money_market_tickers(
                    extract_open_investment_tickers(imported_payload)
                )
            )
        except Exception:  # noqa: BLE001
            LOGGER.exception("Investment import price-cache refresh failed")
            return [
                "Price cache refresh failed after import. Retry it from Local Market Store."
            ]

    def load_local_investment_dividend_actions(
        tickers: set[str],
    ) -> dict[str, list[dict[str, str]]]:
        actions: dict[str, list[dict[str, str]]] = {}
        for ticker in sorted(tickers):
            path = history_store_path_for(ticker)
            if not path.exists():
                continue
            try:
                dataset = pd.read_parquet(path, columns=["Date", "Dividends"])
            except Exception:
                continue
            dividend_values = pd.to_numeric(dataset["Dividends"], errors="coerce").fillna(0.0)
            ticker_actions = [
                {
                    "date": pd.Timestamp(row_date).date().isoformat(),
                    "dividend_per_share": str(dividend_value),
                }
                for row_date, dividend_value in zip(dataset["Date"], dividend_values)
                if float(dividend_value) > 0
            ]
            if ticker_actions:
                actions[ticker] = ticker_actions
        return actions

    def refresh_investment_open_tickers_in_background(tickers: list[str]) -> None:
        if not investment_daily_refresh_lock.acquire(blocking=False):
            return
        try:
            ensure_latest_investment_daily_caches(
                exclude_configured_money_market_tickers(tickers)
            )
        finally:
            investment_daily_refresh_lock.release()

    def build_investment_section_freshness(payload: dict[str, Any]) -> dict[str, Any]:
        return {
            "scope": "section",
            "target_trading_day": latest_completed_nyse_trading_day().strftime("%Y-%m-%d"),
            "open_tickers": sorted(
                exclude_configured_money_market_tickers(
                    extract_open_investment_tickers(payload)
                )
            ),
        }

    def build_investment_fx_rate_history_payload(
        transactions: list[dict[str, Any]],
    ) -> dict[str, dict[str, Any]]:
        """Build local-currency-per-USD daily rates for the investment frontend."""
        transaction_currencies = {
            str(transaction.get("currency") or "").strip().upper()
            for transaction in (transactions if isinstance(transactions, list) else [])
        }
        requested_currencies: list[str] = []
        if "HKD" in transaction_currencies:
            requested_currencies.append("HKD")
        if transaction_currencies.intersection({"CNY", "CNH", "RMB"}):
            requested_currencies.append("CNY")
        if not requested_currencies:
            return {}

        parsed_dates: list[pd.Timestamp] = []
        for transaction in (transactions if isinstance(transactions, list) else []):
            parsed = pd.to_datetime(transaction.get("date"), errors="coerce")
            if pd.isna(parsed):
                continue
            parsed_timestamp = pd.Timestamp(parsed)
            if parsed_timestamp.tzinfo is not None:
                parsed_timestamp = parsed_timestamp.tz_convert("America/New_York").tz_localize(None)
            parsed_dates.append(parsed_timestamp.normalize())

        end_date = pd.Timestamp.now(tz="America/New_York").tz_localize(None).normalize()
        start_date = min(parsed_dates) if parsed_dates else end_date
        if start_date > end_date:
            start_date = end_date

        payload: dict[str, dict[str, Any]] = {}
        for currency in requested_currencies:
            try:
                history = fetch_usd_exchange_rate_history(currency, start_date, end_date)
            except Exception as exc:  # noqa: BLE001
                LOGGER.warning("Unable to build investment %s FX history: %s", currency, exc)
                continue

            history = history.copy()
            history["Date"] = pd.to_datetime(history["Date"], errors="coerce").dt.normalize()
            history = history.loc[
                history["Date"].notna()
                & (history["Date"] >= start_date)
                & (history["Date"] <= end_date)
            ]
            values: dict[str, float] = {}
            for row in history.itertuples(index=False):
                row_date = pd.to_datetime(getattr(row, "Date", None), errors="coerce")
                usd_per_unit = pd.to_numeric(getattr(row, "UsdPerUnit", None), errors="coerce")
                if pd.isna(row_date) or pd.isna(usd_per_unit) or float(usd_per_unit) <= 0:
                    continue
                normalized_date = pd.Timestamp(row_date)
                if normalized_date.tzinfo is not None:
                    normalized_date = normalized_date.tz_convert("America/New_York").tz_localize(None)
                values[normalized_date.date().isoformat()] = 1.0 / float(usd_per_unit)
            if values:
                dates = sorted(values)
                payload[currency] = {
                    "dates": dates,
                    "values": {date: values[date] for date in dates},
                }

        # Yahoo has a CNY history but no separate CNH mapping in this project.
        # Use it as the RMB fallback; transaction- or statement-specific rates
        # are applied later by the frontend and therefore remain authoritative.
        if "CNY" in payload and "CNH" in transaction_currencies and "CNH" not in payload:
            payload["CNH"] = {
                "dates": list(payload["CNY"]["dates"]),
                "values": dict(payload["CNY"]["values"]),
            }
        return payload

    def collect_investment_display_tickers(transactions: list[dict[str, Any]]) -> list[str]:
        tickers: list[str] = []
        seen: set[str] = set()
        excluded_types = {"forex_trade", "forex_trade_component", "fx_translation_pnl"}
        for txn in transactions:
            ticker = str(txn.get("ticker") or "").strip().upper()
            if not ticker:
                continue
            normalized_type = str(txn.get("type") or "").replace(" ", "_").lower()
            if normalized_type in excluded_types or ticker in seen:
                continue
            seen.add(ticker)
            tickers.append(ticker)
        return tickers

    def resolve_money_market_company_name(
            ticker: str,
            transactions: list[dict[str, Any]],
    ) -> str | None:
        if not is_configured_money_market_ticker(ticker) or not money_market_name_from_description:
            return None

        preferred_transaction_types = {"buy", "sell"}
        fallback_candidate = None
        for txn in transactions:
            if canonicalize_money_market_ticker(txn.get("ticker")) != canonicalize_money_market_ticker(ticker):
                continue
            description = " ".join(str(txn.get("description") or "").split()).strip()
            if not description:
                continue
            description_upper = description.upper()
            if money_market_description_keywords and not any(
                    keyword in description_upper for keyword in money_market_description_keywords
            ):
                continue
            normalized_type = str(txn.get("type") or "").replace(" ", "_").lower()
            if normalized_type in preferred_transaction_types:
                return description
            if fallback_candidate is None:
                fallback_candidate = description
        return fallback_candidate

    def load_investment_realtime_quotes(open_tickers: list[str] | set[str] | tuple[str, ...]) -> list[dict[str, object]]:
        requested_tickers = list(dict.fromkeys(
            str(ticker).strip().upper()
            for ticker in (open_tickers or [])
            if str(ticker or "").strip()
        ))
        requested_tickers = exclude_configured_money_market_tickers(requested_tickers)
        if not requested_tickers:
            return []
        cache_key = tuple(sorted(requested_tickers))
        now_monotonic = time.monotonic()
        with investment_realtime_quote_cache_lock:
            cached_entry = investment_realtime_quote_cache.get(cache_key)
            if cached_entry is not None and now_monotonic - cached_entry[0] <= INVESTMENT_REALTIME_QUOTE_TTL_SECONDS:
                return [dict(item) for item in cached_entry[1]]
        # Prefer configured Longbridge live quotes during supported US sessions.
        # Batch yfinance requests recover only the unresolved tickers, preserving
        # Longbridge as the authoritative source whenever it returned a quote.
        quotes = fetch_longbridge_realtime_quotes(requested_tickers)
        resolved_tickers = {str(item.get("ticker") or "").strip().upper() for item in quotes}
        unresolved_tickers = [ticker for ticker in requested_tickers if ticker not in resolved_tickers]
        if unresolved_tickers:
            for quote in fetch_yfinance_realtime_quotes(unresolved_tickers):
                quote_ticker = str(quote.get("ticker") or "").strip().upper()
                if quote_ticker and quote_ticker not in resolved_tickers:
                    quotes.append(quote)
                    resolved_tickers.add(quote_ticker)
        resolved_tickers = {str(item.get("ticker") or "").strip().upper() for item in quotes}
        if resolved_tickers.issuperset(requested_tickers):
            with investment_realtime_quote_cache_lock:
                investment_realtime_quote_cache[cache_key] = (
                    time.monotonic(),
                    [dict(item) for item in quotes],
                )
        return quotes

    def build_investment_ticker_profiles(
            transactions: list[dict[str, Any]],
            open_tickers: list[str] | set[str] | tuple[str, ...],
    ) -> dict[str, dict[str, str]]:
        open_ticker_set = {
            normalize_ticker_input(str(ticker))
            for ticker in open_tickers
            if str(ticker or "").strip()
        }
        ticker_profiles: dict[str, dict[str, str]] = {}
        for raw_ticker in collect_investment_display_tickers(transactions):
            company_name, logo_url = resolve_ticker_identity_snapshot(
                raw_ticker,
                allow_remote_refresh=(
                    raw_ticker in open_ticker_set
                    and not is_configured_money_market_ticker(raw_ticker)
                ),
            )
            if company_name == raw_ticker:
                known_company_name = resolve_known_ticker_company_name(raw_ticker)
                if known_company_name:
                    company_name = known_company_name
            if company_name == raw_ticker:
                inferred_money_market_name = resolve_money_market_company_name(raw_ticker, transactions)
                if inferred_money_market_name:
                    company_name = inferred_money_market_name
            profile_entry = {
                "ticker": raw_ticker,
                "company_name": company_name,
                "logo_url": logo_url,
            }
            ticker_profiles[raw_ticker] = profile_entry
            if raw_ticker.endswith(".US"):
                bare_ticker = raw_ticker[:-3].strip()
                if bare_ticker and bare_ticker not in ticker_profiles:
                    ticker_profiles[bare_ticker] = {
                        **profile_entry,
                        "ticker": bare_ticker,
                    }
        propagate_investment_lineage_identity_profiles(ticker_profiles)
        return ticker_profiles

    def iter_investment_store_ticker_aliases(ticker: str) -> list[str]:
        return investment_ticker_store_aliases(ticker)

    def resolve_investment_history_store_path(
            ticker: str,
            *,
            interval: str = "1d",
            include_proxy: bool = True,
    ) -> Path | None:
        alias_candidates = (
            iter_investment_store_ticker_aliases(ticker)
            if include_proxy
            else investment_ticker_identity_store_aliases(ticker)
        )
        for candidate in alias_candidates:
            path = intraday_history_store_path_for(candidate, interval) if interval == "1m" else history_store_path_for(candidate)
            if path.exists() and path.stat().st_size > 0:
                return path
        return None

    def load_price_history_series(path: Path) -> list[dict[str, Any]]:
        dataset = pd.read_parquet(path, columns=["Date", "Close"])
        dataset = dataset.assign(
            _parsed_date=pd.to_datetime(dataset["Date"], errors="coerce"),
            _parsed_close=pd.to_numeric(dataset["Close"], errors="coerce"),
        )
        dataset = dataset.loc[
            dataset["_parsed_date"].notna()
            & dataset["_parsed_close"].notna()
            & (dataset["_parsed_close"] > 0)
        ].sort_values(["_parsed_date", "_parsed_close"], kind="mergesort")
        prices_by_date: dict[str, float] = {}
        for date_val, close_val in dataset[["_parsed_date", "_parsed_close"]].itertuples(
                index=False, name=None):
            if pd.isna(date_val) or pd.isna(close_val) or not float(close_val) > 0:
                continue
            date_str = pd.Timestamp(date_val).date().isoformat()
            # The lowest valid close is a deterministic tie-breaker for
            # duplicate provider prints, independent of parquet row order.
            prices_by_date.setdefault(date_str, float(close_val))
        return [
            {"date": date_str, "close": prices_by_date[date_str]}
            for date_str in sorted(prices_by_date)
        ]

    def load_investment_price_histories(
            transactions: list[dict[str, Any]],
            *,
            open_tickers: list[str] | set[str] | tuple[str, ...] | None = None,
    ) -> tuple[dict[str, list[dict[str, Any]]], list[dict[str, str]]]:
        price_history_by_ticker: dict[str, list[dict[str, Any]]] = {}
        failures: list[dict[str, str]] = []
        open_ticker_set = {
            normalize_ticker_input(str(ticker))
            for ticker in (open_tickers or [])
            if str(ticker or "").strip()
        }
        for ticker in collect_investment_display_tickers(transactions):
            if is_configured_money_market_ticker(ticker):
                continue
            try:
                path = resolve_investment_history_store_path(ticker, include_proxy=False)
                should_refresh_live_cache = ticker in open_ticker_set
                if path is None:
                    failures.append({
                        "ticker": ticker,
                        "reason": "missing_store",
                        "message": (
                            f"No local market history is available for {ticker}."
                            if should_refresh_live_cache
                            else f"No cached market history is available for the closed position {ticker}; ledger trade prices will be used instead."
                        ),
                    })
                    continue
                prices = load_price_history_series(path)
                if not prices:
                    failures.append({
                        "ticker": ticker,
                        "reason": "empty_store",
                        "message": f"No closing-price rows are available for {ticker}.",
                    })
                    continue
                price_history_by_ticker[ticker] = prices
            except Exception:  # noqa: BLE001
                LOGGER.exception("Could not read local market history for %s", ticker)
                failures.append({
                    "ticker": ticker,
                    "reason": "read_failed",
                    "message": f"Could not read local market history for {ticker}.",
                })
        return price_history_by_ticker, failures

    # Backtest result cache: skip redundant computation when config doesn't change
    _cached_backtest: dict[str, tuple] = {}

    def _get_backtest_cache_key() -> str:
        """Generate a cache key from all backtest configuration parameters."""
        requested_ticker = request.args.get("ticker", "").strip()
        if not requested_ticker:
            requested_ticker = str(defaults.get("backtest_ticker", DEFAULT_TICKERS[0]))
        normalized_ticker = normalize_ticker_input(requested_ticker)
        daily_path = history_store_path_for(normalized_ticker) if normalized_ticker else None
        intraday_path = intraday_history_store_path_for(normalized_ticker, "1m") if normalized_ticker else None
        params = [
            request.path,
            request.args.get("ticker", ""),
            request.args.get("strategy", ""),
            request.args.get("capital", ""),
            request.args.get("period", ""),
            request.args.get("range", ""),
            request.args.get("from", ""),
            request.args.get("to", ""),
            request.args.get("interval", ""),
            str(request.args.get("price_only", "")),
            str(request.args.get("dividends", "")),
            # Include all strategy parameters in cache key
            sorted([(k, request.args.get(k, "")) for k in request.args.keys() if k not in {
                "ticker", "strategy", "capital", "period", "range", "from", "to", "interval", "price_only", "dividends",
                "view", "section", "view", "tickers", "weight",
            }]),
            {
                "daily_mtime_ns": daily_path.stat().st_mtime_ns if daily_path and daily_path.exists() else None,
                "intraday_mtime_ns": intraday_path.stat().st_mtime_ns if intraday_path and intraday_path.exists() else None,
            },
        ]
        key_string = json.dumps(params, sort_keys=True)
        return hashlib.sha256(key_string.encode("utf-8")).hexdigest()[:16]

    def _read_settings_feedback() -> dict[str, str]:
        raw_feedback = request.cookies.get(SETTINGS_FEEDBACK_COOKIE, "").strip()
        if not raw_feedback:
            return {}
        try:
            payload = json.loads(raw_feedback)
        except json.JSONDecodeError:
            return {}
        if not isinstance(payload, dict):
            return {}
        return {
            key: str(value).strip()
            for key, value in payload.items()
            if key in {
                "notice",
                "error",
                "broker_test_status",
                "broker_test_message",
                "broker_test_checked_at",
                "longbridge_oauth_pending",
            }
               and str(value).strip()
        }

    def _redirect_with_settings_feedback(
            section_name: str,
            *,
            notice: str = "",
            error: str = "",
            broker_test_status: str = "",
            broker_test_message: str = "",
            broker_test_checked_at: str = "",
            longbridge_oauth_pending: str = "",
            query_params: dict[str, Any] | None = None,
    ):
        target_path = build_settings_path(section_name)
        if query_params:
            target_path = build_settings_state_url(
                section_name,
                tab=query_params.get("tab", query_params.get("settings_tab", "current")),
                page=query_params.get(
                    "page",
                    query_params.get("settings_page", query_params.get("local_page", 1)),
                ),
            )
        response = make_response(redirect(target_path, code=303))
        payload = {
            key: value.strip()
            for key, value in {
                "notice": notice,
                "error": error,
                "broker_test_status": broker_test_status,
                "broker_test_message": broker_test_message,
                "broker_test_checked_at": broker_test_checked_at,
                "longbridge_oauth_pending": longbridge_oauth_pending,
            }.items()
            if value and value.strip()
        }
        if payload:
            response.set_cookie(
                SETTINGS_FEEDBACK_COOKIE,
                json.dumps(payload, separators=(",", ":")),
                max_age=60,
                httponly=True,
                samesite="Lax",
                path="/settings",
            )
        else:
            response.delete_cookie(SETTINGS_FEEDBACK_COOKIE, path="/settings")
        return response

    def validate_ticker_or_raise(raw_ticker: str) -> str:
        normalized_ticker = normalize_ticker_input(raw_ticker)
        if not has_valid_ticker_format(normalized_ticker):
            raise ValueError(f"Invalid ticker format: {raw_ticker}.")
        return normalized_ticker

    def parse_requested_tickers(view_name: str | None = None) -> list[str]:
        return parse_requested_tickers_from_args(
            request.args,
            max_tickers=max_tickers_for_view(view_name),
            normalize=normalize_ticker_input,
            getlist=request.args.getlist,
        )

    def parse_requested_weights(slot_count: int) -> list[int]:
        return parse_requested_weights_from_args(
            request.args,
            slot_count,
            getlist=request.args.getlist,
        )

    def parse_portfolio_allocation_mode() -> str:
        return parse_portfolio_allocation_mode_from_args(request.args)

    def parse_requested_shares(slot_count: int) -> list[int]:
        return parse_requested_shares_from_args(
            request.args,
            slot_count,
            getlist=request.args.getlist,
        )

    def parse_bool_flag(*names: str, default: bool = False) -> bool:
        return parse_bool_flag_from_args(
            request.args,
            *names,
            default=default,
            getlist=request.args.getlist,
        )

    def parse_range_request_args() -> tuple[str, str, str, str]:
        return parse_range_request_args_from_args(
            request.args,
            default_range_mode=str(defaults.get("range_mode", "period")),
            default_period=str(defaults.get("period", DEFAULT_PERIOD)),
        )

    def build_exact_range_bounds(start_value: str, end_value: str) -> tuple[pd.Timestamp, pd.Timestamp]:
        start_bound = pd.to_datetime(start_value).normalize()
        end_bound = pd.to_datetime(end_value).replace(hour=23, minute=59, second=59)
        return start_bound, end_bound

    def slice_dataset_to_exact_range(
            dataset: pd.DataFrame,
            adjusted_start: str,
            adjusted_end: str,
    ) -> pd.DataFrame:
        start_bound, end_bound = build_exact_range_bounds(adjusted_start, adjusted_end)
        return dataset[(dataset["Date"] >= start_bound) & (dataset["Date"] <= end_bound)].copy()

    def slice_datasets_to_exact_range(
            datasets: list[pd.DataFrame],
            adjusted_start: str,
            adjusted_end: str,
    ) -> list[pd.DataFrame]:
        return [
            slice_dataset_to_exact_range(dataset, adjusted_start, adjusted_end)
            for dataset in datasets
        ]

    def exact_trading_dates_in_range(payload: DateConstraintPayload) -> list[str]:
        if not payload.trading_dates or not payload.adjusted_start or not payload.adjusted_end:
            return []
        start_date = pd.to_datetime(payload.adjusted_start).date()
        end_date = pd.to_datetime(payload.adjusted_end).date()
        return [
            trading_date
            for trading_date in payload.trading_dates
            if start_date <= pd.to_datetime(trading_date).date() <= end_date
        ]

    def slice_intraday_dataset_to_trading_date(dataset: pd.DataFrame, trading_date: object) -> pd.DataFrame:
        target_date = pd.to_datetime(trading_date).date()
        return dataset[dataset["Date"].dt.date == target_date].copy()

    def market_timezone_for_ticker(ticker: str) -> str:
        market = infer_ticker_market(ticker)
        if market == "HK":
            return "Asia/Hong_Kong"
        if market == "KR":
            return "Asia/Seoul"
        if market == "JP":
            return "Asia/Tokyo"
        if market == "CN":
            return "Asia/Shanghai"
        if market == "UK":
            return "Europe/London"
        if market == "SG":
            return "Asia/Singapore"
        if market == "AU":
            return "Australia/Sydney"
        if market == "CA":
            return "America/Toronto"
        if market == "EU":
            return "Europe/Paris"
        if market == "FI":
            return "Europe/Helsinki"
        if market == "IN":
            return "Asia/Kolkata"
        if market == "TW":
            return "Asia/Taipei"
        if market == "MY":
            return "Asia/Kuala_Lumpur"
        if market == "TH":
            return "Asia/Bangkok"
        if market == "ID":
            return "Asia/Jakarta"
        if market == "NZ":
            return "Pacific/Auckland"
        if market == "BR":
            return "America/Sao_Paulo"
        if market == "LATAM":
            return "America/Mexico_City"
        if market == "IL":
            return "Asia/Jerusalem"
        if market == "SA":
            return "Asia/Riyadh"
        if market == "ZA":
            return "Africa/Johannesburg"
        if market == "QA":
            return "Asia/Qatar"
        return "America/New_York"

    def market_close_minute_for_ticker(ticker: str) -> int | None:
        market = infer_ticker_market(ticker)
        if market in {"KR", "JP"}:
            return (15 * 60) + 30
        if market == "HK":
            return (15 * 60) + 59
        if market == "CN":
            return (14 * 60) + 59
        if market == "UK":
            return (16 * 60) + 29
        if market in {"AU", "CA", "ID"}:
            return (16 * 60) - 1
        if market == "SG":
            return (17 * 60) - 1
        if market in {"BR", "ZA"}:
            return (17 * 60) - 1
        if market in {"EU", "FI", "IL"}:
            return (17 * 60) + 30
        if market == "IN":
            return (15 * 60) + 30
        if market == "TW":
            return (13 * 60) + 30
        if market == "MY":
            return (17 * 60) - 1
        if market == "TH":
            return (16 * 60) + 30
        if market == "NZ":
            return (16 * 60) + 44
        if market == "LATAM":
            return (15 * 60) - 1
        if market == "SA":
            return (15 * 60) - 1
        if market == "QA":
            return (13 * 60) + 9
        return None

    def market_session_segments_for_ticker(ticker: str) -> list[tuple[int, int]]:
        market = infer_ticker_market(ticker)
        if market == "HK":
            return [((9 * 60) + 30, 12 * 60), (13 * 60, 16 * 60)]
        if market == "CN":
            return [((9 * 60) + 30, (11 * 60) + 30), (13 * 60, 15 * 60)]
        if market == "KR":
            return [(9 * 60, (15 * 60) + 30)]
        if market == "JP":
            return [(9 * 60, (11 * 60) + 30), ((12 * 60) + 30, (15 * 60) + 30)]
        if market == "UK":
            return [(8 * 60, (16 * 60) + 30)]
        if market == "SG":
            return [(9 * 60, 12 * 60), (13 * 60, 17 * 60)]
        if market in {"AU", "MY", "EU", "FI", "ID", "ZA"}:
            return [(9 * 60, market_close_minute_for_ticker(ticker) + 1)]
        if market == "CA":
            return [((9 * 60) + 30, 16 * 60)]
        if market == "IN":
            return [((9 * 60) + 15, (15 * 60) + 30)]
        if market == "TW":
            return [(9 * 60, (13 * 60) + 30)]
        if market == "TH":
            return [(10 * 60, (16 * 60) + 30)]
        if market == "NZ":
            return [(10 * 60, (16 * 60) + 45)]
        if market == "BR":
            return [(10 * 60, 17 * 60)]
        if market == "LATAM":
            return [((8 * 60) + 30, 15 * 60)]
        if market == "IL":
            return [((9 * 60) + 30, (17 * 60) + 30)]
        if market == "SA":
            return [(10 * 60, 15 * 60)]
        if market == "QA":
            return [((9 * 60) + 30, (13 * 60) + 10)]
        return [((9 * 60) + 30, 16 * 60)]

    def is_market_regular_session_active_for_ticker(
            ticker: str,
            reference_timestamp: object | None = None,
    ) -> bool:
        current_timestamp = (
            pd.Timestamp.now(tz="UTC")
            if reference_timestamp is None
            else pd.to_datetime(reference_timestamp, errors="coerce")
        )
        if pd.isna(current_timestamp):
            return False
        if current_timestamp.tzinfo is None:
            current_timestamp = current_timestamp.tz_localize("UTC")
        else:
            current_timestamp = current_timestamp.tz_convert("UTC")
        localized = current_timestamp.tz_convert(market_timezone_for_ticker(ticker))
        if int(localized.weekday()) >= 5:
            return False
        minute_of_day = (int(localized.hour) * 60) + int(localized.minute)
        return any(
            start_minute <= minute_of_day < end_minute
            for start_minute, end_minute in market_session_segments_for_ticker(ticker)
        )

    def apply_market_close_anchor(
            intraday_dataset: pd.DataFrame,
            daily_dataset: pd.DataFrame,
            ticker: str,
            trading_date: object,
    ) -> pd.DataFrame:
        close_minute = market_close_minute_for_ticker(ticker)
        if close_minute is None or intraday_dataset.empty or daily_dataset.empty or "Close" not in daily_dataset.columns:
            return intraday_dataset

        target_date = pd.to_datetime(trading_date).date()
        daily_dates = pd.to_datetime(daily_dataset["Date"], errors="coerce").dt.date
        daily_rows = daily_dataset.loc[daily_dates == target_date]
        if daily_rows.empty:
            return intraday_dataset

        close_value = pd.to_numeric(pd.Series([daily_rows.iloc[-1]["Close"]]), errors="coerce").iloc[0]
        if pd.isna(close_value):
            return intraday_dataset

        market_timezone = market_timezone_for_ticker(ticker)
        close_local = pd.Timestamp(
            year=int(target_date.year),
            month=int(target_date.month),
            day=int(target_date.day),
            hour=close_minute // 60,
            minute=close_minute % 60,
            tz=market_timezone,
        )
        close_timestamp = close_local.tz_convert("America/New_York").tz_localize(None)

        anchored_dataset = intraday_dataset.copy()
        intraday_close_values = pd.to_numeric(anchored_dataset.get("Close", pd.Series(dtype="float64")), errors="coerce")
        comparable_closes = intraday_close_values.loc[
            intraday_close_values.notna()
            & (anchored_dataset["Date"] <= close_timestamp)
        ]
        if not comparable_closes.empty:
            last_intraday_close = float(comparable_closes.iloc[-1])
            if last_intraday_close > 0:
                close_gap_ratio = abs(float(close_value) - last_intraday_close) / last_intraday_close
                if close_gap_ratio > 0.20:
                    LOGGER.warning(
                        "Skipped market close anchor for %s on %s because daily close %.4f is %.2f%% away from intraday close %.4f.",
                        ticker,
                        trading_date,
                        float(close_value),
                        close_gap_ratio * 100,
                        last_intraday_close,
                    )
                    return intraday_dataset

        matching_rows = anchored_dataset["Date"] == close_timestamp
        if not matching_rows.any():
            row: dict[str, object] = {column: pd.NA for column in anchored_dataset.columns}
            row["Date"] = close_timestamp
            anchored_dataset = pd.concat([anchored_dataset, pd.DataFrame([row])], ignore_index=True)
            matching_rows = anchored_dataset["Date"] == close_timestamp

        for column in ("Open", "High", "Low", "Close", "Adj Close"):
            if column in anchored_dataset.columns:
                anchored_dataset.loc[matching_rows, column] = float(close_value)
        if "Volume" in anchored_dataset.columns:
            anchored_dataset.loc[matching_rows, "Volume"] = 0.0
        if "Turnover" in anchored_dataset.columns:
            anchored_dataset.loc[matching_rows, "Turnover"] = 0.0

        return anchored_dataset.drop_duplicates(subset=["Date"], keep="last").sort_values("Date").reset_index(drop=True)

    def slice_intraday_dataset_to_market_trading_date(
            dataset: pd.DataFrame,
            ticker: str,
            trading_date: object,
    ) -> pd.DataFrame:
        target_date = pd.to_datetime(trading_date).date()
        def date_for_market(value: object) -> object:
            return market_trading_date_for_timestamp(value, ticker)

        market_dates = dataset["Date"].map(date_for_market)
        return dataset[market_dates == target_date].copy()

    def slice_intraday_dataset_to_trading_dates(
            dataset: pd.DataFrame,
            ticker: str,
            trading_dates: list[str],
    ) -> pd.DataFrame:
        selected_dates = {pd.to_datetime(trading_date).date() for trading_date in trading_dates}
        def date_for_market(value: object) -> object:
            return market_trading_date_for_timestamp(value, ticker)

        market_dates = dataset["Date"].map(date_for_market)
        return dataset[market_dates.isin(selected_dates)].copy()

    def load_local_compare_one_day_intraday_dataset(
            ticker: str,
            *,
            refresh_stale: bool = False,
    ) -> pd.DataFrame:
        path = next(
            (
                candidate_path
                for candidate in market_ticker_store_aliases(ticker)
                if (candidate_path := intraday_history_store_path_for(candidate, "1m")).exists()
                and candidate_path.stat().st_size > 0
            ),
            intraday_history_store_path_for(ticker, "1m"),
        )
        if not path.exists() or path.stat().st_size == 0:
            refresh_one_minute_store(ticker)
            path = next(
                (
                    candidate_path
                    for candidate in market_ticker_store_aliases(ticker)
                    if (candidate_path := intraday_history_store_path_for(candidate, "1m")).exists()
                    and candidate_path.stat().st_size > 0
                ),
                intraday_history_store_path_for(ticker, "1m"),
            )
        elif refresh_stale and not is_one_minute_store_fresh(ticker):
            try:
                refresh_one_minute_store(ticker)
            except (ImportError, OSError, ValueError, KeyError, TypeError) as exc:
                LOGGER.warning(
                    "Unable to refresh stale local 1-minute compare data for %s; using the existing cache: %s",
                    ticker,
                    exc,
                )
            path = next(
                (
                    candidate_path
                    for candidate in market_ticker_store_aliases(ticker)
                    if (candidate_path := intraday_history_store_path_for(candidate, "1m")).exists()
                    and candidate_path.stat().st_size > 0
                ),
                intraday_history_store_path_for(ticker, "1m"),
            )
        if not path.exists() or path.stat().st_size == 0:
            raise ValueError(f"Local 1-minute market data for {ticker} is unavailable.")
        with market_store_file_lock(path):
            dataset = pd.read_parquet(path)
        dataset = select_price_series(dataset, include_dividends=False, dividend_mode="price")
        if dataset.empty:
            raise ValueError(f"Local 1-minute market data for {ticker} is empty.")
        return dataset

    def load_compare_one_day_intraday_dataset(
            ticker: str,
            *,
            include_extended_hours_flag: bool,
            include_overnight_flag: bool = False,
            trading_date: object | None = None,
            refresh_stale_local: bool = False,
    ) -> pd.DataFrame:
        if infer_ticker_market(ticker) == "US" and include_overnight_flag:
            intraday_dataset = fetch_request_compare_one_day_overnight_history(
                ticker,
                trading_date=trading_date,
            )
            if trading_date is not None:
                intraday_dataset = slice_intraday_dataset_to_market_trading_date(
                    intraday_dataset,
                    ticker,
                    trading_date,
                )
            if intraday_dataset.empty:
                raise ValueError(
                    f"Overnight companion data for {ticker} does not include {trading_date}."
                )
            return intraday_dataset

        if trading_date is None:
            intraday_dataset = load_local_compare_one_day_intraday_dataset(
                ticker,
                refresh_stale=refresh_stale_local,
            )
            if infer_ticker_market(ticker) == "US" and not include_extended_hours_flag:
                intraday_dataset = filter_intraday_dataset_to_regular_session(intraday_dataset)
            if intraday_dataset.empty:
                raise ValueError(f"The latest local intraday store does not contain usable data for {ticker}.")
            return intraday_dataset

        if infer_ticker_market(ticker) != "US":
            intraday_dataset = load_local_compare_one_day_intraday_dataset(ticker)
            dated_dataset = slice_intraday_dataset_to_market_trading_date(
                intraday_dataset,
                ticker,
                trading_date,
            )
            if not dated_dataset.empty:
                return dated_dataset
            try:
                intraday_dataset = fetch_one_minute_history_for_trading_date(
                    ticker,
                    trading_date,
                    include_dividends=False,
                    dividend_mode="price",
                )
            except (ImportError, OSError, ValueError, KeyError, TypeError) as exc:
                LOGGER.warning(
                    "Unable to fetch exact-day 1-minute compare data for %s on %s: %s",
                    ticker,
                    trading_date,
                    exc,
                )
                intraday_dataset = fetch_history(
                    ticker,
                    include_dividends=False,
                    interval="1m",
                    dividend_mode="price",
                )
            intraday_dataset = slice_intraday_dataset_to_market_trading_date(intraday_dataset, ticker, trading_date)
            if intraday_dataset.empty:
                raise ValueError(f"The selected trading date does not contain shared intraday data for {ticker}.")
            return intraday_dataset

        try:
            intraday_dataset = fetch_compare_one_day_extended_history(
                ticker,
                trading_date=trading_date,
            )
            dated_dataset = slice_intraday_dataset_to_market_trading_date(intraday_dataset, ticker, trading_date)
            if dated_dataset.empty:
                raise ValueError(f"Extended-hours data for {ticker} does not include {trading_date}.")
            intraday_dataset = dated_dataset
            if not include_extended_hours_flag:
                intraday_dataset = filter_intraday_dataset_to_regular_session(intraday_dataset)
            return intraday_dataset
        except (ImportError, OSError, ValueError, KeyError, TypeError) as exc:
            LOGGER.warning(
                "Unable to fetch extended-hours compare 1d data for %s: %s",
                ticker,
                exc,
            )

        intraday_dataset = fetch_history(
            ticker,
            include_dividends=False,
            interval="1m",
            dividend_mode="price",
        )
        intraday_dataset = slice_intraday_dataset_to_market_trading_date(intraday_dataset, ticker, trading_date)
        if not include_extended_hours_flag:
            intraday_dataset = filter_intraday_dataset_to_regular_session(intraday_dataset)
        if intraday_dataset.empty:
            raise ValueError(f"The selected trading date does not contain shared intraday data for {ticker}.")
        return intraday_dataset

    def market_minute_key_for_compare_axis(value: object, ticker: str) -> str:
        timestamp = pd.Timestamp(value)
        if timestamp.tzinfo is None:
            timestamp = timestamp.tz_localize("America/New_York")
        else:
            timestamp = timestamp.tz_convert("America/New_York")
        localized = timestamp.tz_convert(market_timezone_for_ticker(ticker))
        return f"{int(localized.hour):02d}:{int(localized.minute):02d}"

    def load_live_compare_one_day_intraday_dataset(
            ticker: str,
            *,
            live_trading_date: object,
            include_extended_hours_flag: bool,
            force_refresh: bool,
            include_overnight_flag: bool = False,
    ) -> tuple[pd.DataFrame, str]:
        if infer_ticker_market(ticker) == "US" and include_overnight_flag:
            intraday_dataset = fetch_request_compare_one_day_overnight_history(
                ticker,
                trading_date=live_trading_date,
            )
            source = str(
                intraday_dataset.attrs.get("market_data_source")
                or "overnight_companion"
            )
            intraday_dataset = slice_intraday_dataset_to_market_trading_date(
                intraday_dataset,
                ticker,
                live_trading_date,
            )
            if intraday_dataset.empty:
                raise ValueError(
                    f"Overnight companion data for {ticker} does not include {live_trading_date}."
                )
            return intraday_dataset, source

        if infer_ticker_market(ticker) == "US" and include_extended_hours_flag:
            try:
                intraday_dataset = fetch_compare_one_day_extended_history(
                    ticker,
                    trading_date=live_trading_date,
                )
                source = str(
                    intraday_dataset.attrs.get("market_data_source")
                    or "extended_hours"
                )
                intraday_dataset = slice_intraday_dataset_to_market_trading_date(
                    intraday_dataset,
                    ticker,
                    live_trading_date,
                )
                if intraday_dataset.empty:
                    raise ValueError(
                        f"Extended-hours data for {ticker} does not include {live_trading_date}."
                    )
                return intraday_dataset, source
            except (ImportError, OSError, ValueError, KeyError, TypeError) as exc:
                LOGGER.warning(
                    "Unable to load default extended-hours data for %s on %s: %s",
                    ticker,
                    live_trading_date,
                    exc,
                )

        source = "local"
        if force_refresh:
            try:
                refresh_result = refresh_one_minute_store(ticker)
                source = refresh_result.source
            except Exception as exc:  # noqa: BLE001
                LOGGER.warning("Unable to refresh live 1-minute compare data for %s: %s", ticker, exc)

        intraday_dataset = load_local_compare_one_day_intraday_dataset(ticker)
        intraday_dataset = slice_intraday_dataset_to_market_trading_date(
            intraday_dataset,
            ticker,
            live_trading_date,
        )
        if intraday_dataset.empty and infer_ticker_market(ticker) != "US":
            try:
                intraday_dataset = fetch_one_minute_history_for_trading_date(
                    ticker,
                    live_trading_date,
                    include_dividends=False,
                    dividend_mode="price",
                )
                source = str(intraday_dataset.attrs.get("market_data_source") or "yfinance_exact")
            except (ImportError, OSError, ValueError, KeyError, TypeError) as exc:
                LOGGER.warning(
                    "Unable to fetch missing live 1-minute compare data for %s on %s: %s",
                    ticker,
                    live_trading_date,
                    exc,
                )
        if infer_ticker_market(ticker) == "US" and not include_extended_hours_flag:
            intraday_dataset = filter_intraday_dataset_to_regular_session(intraday_dataset)
        if intraday_dataset.empty:
            raise ValueError(f"Live 1-minute data for {ticker} does not include {live_trading_date}.")
        return intraday_dataset, source

    def load_target_compare_one_day_intraday_dataset(
            ticker: str,
            *,
            target_trading_date: object,
            include_extended_hours_flag: bool,
            include_overnight_flag: bool = False,
            live_session_date: object | None = None,
            force_refresh: bool = False,
    ) -> pd.DataFrame:
        parsed_target_date = pd.to_datetime(target_trading_date, errors="coerce")
        if pd.isna(parsed_target_date):
            raise ValueError(f"Invalid compare trading date: {target_trading_date}.")

        target_date_value = parsed_target_date.date()
        if live_session_date is None:
            current_live_session_date = pd.Timestamp.now(tz="Asia/Shanghai").date()
        else:
            parsed_live_session_date = pd.to_datetime(live_session_date, errors="coerce")
            if pd.isna(parsed_live_session_date):
                raise ValueError(f"Invalid live session date: {live_session_date}.")
            current_live_session_date = parsed_live_session_date.date()

        if target_date_value == current_live_session_date:
            return load_live_compare_one_day_intraday_dataset(
                ticker,
                live_trading_date=target_date_value,
                include_extended_hours_flag=include_extended_hours_flag,
                include_overnight_flag=include_overnight_flag,
                force_refresh=force_refresh,
            )[0]

        return load_compare_one_day_intraday_dataset(
            ticker,
            include_extended_hours_flag=include_extended_hours_flag,
            include_overnight_flag=include_overnight_flag,
            trading_date=target_date_value,
        )

    def append_live_compare_intraday_dataset(
            ticker: str,
            intraday_dataset: pd.DataFrame,
            *,
            live_trading_date: object,
            include_extended_hours_flag: bool,
            force_refresh: bool,
    ) -> tuple[pd.DataFrame, str | None]:
        try:
            live_dataset, source = load_live_compare_one_day_intraday_dataset(
                ticker,
                live_trading_date=live_trading_date,
                include_extended_hours_flag=include_extended_hours_flag,
                force_refresh=force_refresh,
            )
        except Exception as exc:  # noqa: BLE001
            LOGGER.warning("Unable to append live intraday compare data for %s: %s", ticker, exc)
            return intraday_dataset, None
        combined = (
            pd.concat([intraday_dataset, live_dataset], ignore_index=True)
            .drop_duplicates(subset=["Date"], keep="last")
            .sort_values("Date")
            .reset_index(drop=True)
        )
        return combined, source

    def map_live_intraday_dataset_to_reference_axis(
            reference_dataset: pd.DataFrame,
            live_dataset: pd.DataFrame,
            ticker: str,
    ) -> pd.DataFrame:
        reference_frame = build_empty_compare_axis_dataset(reference_dataset)

        live_prepared = fill_intraday_market_session_gaps(
            prepare_intraday_dataset_for_compare(live_dataset, ticker),
            ticker,
        )
        if live_prepared.empty:
            return reference_frame

        live_rows_by_minute: dict[str, pd.Series] = {}
        for _, live_row in live_prepared.sort_values("Date").iterrows():
            live_rows_by_minute[market_minute_key_for_compare_axis(live_row["Date"], ticker)] = live_row

        for row_index, reference_date in reference_frame["Date"].items():
            live_row = live_rows_by_minute.get(market_minute_key_for_compare_axis(reference_date, ticker))
            if live_row is None:
                continue
            for column in ("Open", "High", "Low", "Close", "Adj Close", "Volume"):
                if column in live_row.index:
                    reference_frame.at[row_index, column] = live_row[column]

        return reference_frame

    def build_empty_compare_axis_dataset(reference_dataset: pd.DataFrame) -> pd.DataFrame:
        reference_frame = reference_dataset[["Date"]].copy().sort_values("Date").reset_index(drop=True)
        for column in ("Open", "High", "Low", "Close", "Adj Close", "Volume"):
            reference_frame[column] = pd.NA
        return reference_frame

    def truncate_intraday_datasets_to_common_live_timestamp(datasets: list[pd.DataFrame]) -> list[pd.DataFrame]:
        if not datasets:
            return []

        valid_end_dates: list[pd.Timestamp] = []
        for dataset in datasets:
            if dataset.empty or "Date" not in dataset.columns or "Close" not in dataset.columns:
                return datasets
            close_values = pd.to_numeric(dataset["Close"], errors="coerce")
            valid_dates = pd.to_datetime(dataset.loc[close_values.notna(), "Date"], errors="coerce").dropna()
            if valid_dates.empty:
                return datasets
            valid_end_dates.append(pd.Timestamp(valid_dates.max()))

        common_live_end = min(valid_end_dates)
        price_columns = ("Open", "High", "Low", "Close", "Adj Close")
        muted_columns = (*price_columns, "Volume", "Turnover")
        truncated_datasets: list[pd.DataFrame] = []
        for dataset in datasets:
            truncated = dataset.copy()
            parsed_dates = pd.to_datetime(truncated["Date"], errors="coerce")
            trailing_mask = parsed_dates > common_live_end
            if trailing_mask.any():
                for column in muted_columns:
                    if column in truncated.columns:
                        truncated.loc[trailing_mask, column] = pd.NA
            truncated_datasets.append(truncated)
        return truncated_datasets

    def build_empty_compare_axis_series_payload(
            ticker: str,
            reference_dataset: pd.DataFrame,
            color: str | None = None,
    ) -> SeriesPayload:
        reference_dates = reference_dataset["Date"].map(lambda value: pd.Timestamp(value).strftime("%Y-%m-%d %H:%M")).tolist()
        display_dates = reference_dataset["Date"].map(lambda value: format_display_datetime(value)).tolist()
        # Keep the ticker on the shared axis even before that market has live bars.
        return SeriesPayload(
            ticker=ticker.upper(),
            dates=display_dates,
            raw_dates=reference_dates,
            normalized_returns=[None for _ in reference_dates],
            color=color,
            glow=False,
            candlestick_returns=[
                {"x": index, "o": None, "h": None, "l": None, "c": None, "v": None, "synthetic": True}
                for index, _value in enumerate(reference_dates)
            ],
            candlestick_prices=[
                {"x": index, "o": None, "h": None, "l": None, "c": None, "v": None, "synthetic": True}
                for index, _value in enumerate(reference_dates)
            ],
            prices=[None for _ in reference_dates],
        )

    def build_compare_series_payload(
            ticker: str,
            dataset: pd.DataFrame,
            color: str | None = None,
    ) -> SeriesPayload:
        try:
            return build_series_payload(ticker, dataset, color=color)
        except ValueError:
            if dataset.empty or "Date" not in dataset.columns:
                raise
            has_intraday_timestamps = dataset["Date"].map(
                lambda value: pd.Timestamp(value).hour != 0 or pd.Timestamp(value).minute != 0
            ).any()
            close_values = pd.to_numeric(dataset.get("Close", pd.Series(dtype="float64")), errors="coerce").dropna()
            if not has_intraday_timestamps or not close_values.empty:
                raise
            return build_empty_compare_axis_series_payload(ticker, dataset, color=color)

    def build_ttm_dividend_yield_map(
            tickers: list[str],
            end_date: object | None = None,
    ) -> dict[str, float | None]:
        yields: dict[str, float | None] = {}
        for ticker in tickers:
            normalized_ticker = normalize_ticker_input(ticker)
            try:
                price_dataset = fetch_history(normalized_ticker, False, dividend_mode="price")
                yields[normalized_ticker] = calculate_ttm_dividend_yield(price_dataset, end_date=end_date)
            except Exception as exc:  # noqa: BLE001
                LOGGER.info("Unable to calculate TTM dividend yield for %s: %s", normalized_ticker, exc)
                yields[normalized_ticker] = None
        return yields

    def best_numeric_metric(values: list[float | None]) -> float | None:
        numeric_values = [value for value in values if value is not None]
        return max(numeric_values) if numeric_values else None

    def format_store_range_date(raw_value: object) -> str:
        return format_store_range_date_value(raw_value)

    def build_portfolio_series_payload(datasets: list[pd.DataFrame], weights: list[int], color: str):
        first_dataset = datasets[0]
        cumulative_growth = pd.Series(0.0, index=first_dataset.index)
        if len(datasets) != len(weights):
            raise ValueError("Portfolio datasets and weights must have the same length.")
        for dataset, weight in zip(datasets, weights):
            first_close = float(dataset["Close"].iloc[0])
            cumulative_growth += (weight / 100.0) * (dataset["Close"] / first_close)
        portfolio_frame = pd.DataFrame(
            {
                "Date": first_dataset["Date"],
                "Close": cumulative_growth,
            }
        )
        return build_series_payload("Portfolio", portfolio_frame, color=color)

    def build_portfolio_series_payload_for_shares(datasets: list[pd.DataFrame], shares: list[int], color: str):
        first_dataset = datasets[0]
        portfolio_value = pd.Series(0.0, index=first_dataset.index)
        if len(datasets) != len(shares):
            raise ValueError("Portfolio datasets and shares must have the same length.")
        for dataset, share_count in zip(datasets, shares):
            portfolio_value += max(int(share_count), 0) * dataset["Close"]
        if float(portfolio_value.iloc[0]) <= 0:
            raise ValueError("Each selected ticker must have at least 1 share.")
        portfolio_frame = pd.DataFrame(
            {
                "Date": first_dataset["Date"],
                "Close": portfolio_value,
            }
        )
        return build_series_payload("Portfolio", portfolio_frame, color=color)

    def normalize_portfolio_share_weights(datasets: list[pd.DataFrame], shares: list[int]) -> list[int]:
        if not datasets:
            return []
        initial_values = [
            max(int(share_count), 0) * float(dataset["Close"].iloc[0])
            for dataset, share_count in zip(datasets, shares)
        ]
        total = sum(initial_values)
        if total <= 0:
            return [0 for _value in initial_values]
        scaled = [int((value * 100) / total) for value in initial_values]
        remainder = 100 - sum(scaled)
        order = sorted(range(len(initial_values)), key=lambda index: initial_values[index], reverse=True)
        for index in order:
            if remainder <= 0:
                break
            scaled[index] += 1
            remainder -= 1
        return scaled

    def build_portfolio_growth_multipliers(datasets: list[pd.DataFrame]) -> list[float]:
        return [
            float(dataset["Close"].iloc[-1]) / float(dataset["Close"].iloc[0])
            for dataset in datasets
        ]

    def build_benchmark_series_payloads(
            reference_dates: pd.Series,
            include_dividends: bool,
            price_only: bool,
    ) -> tuple[list[SeriesPayload], list[QuoteProfile]]:
        benchmark_series: list[SeriesPayload] = []
        benchmark_profiles: list[QuoteProfile] = []
        reference_date_frame = pd.DataFrame({"Date": reference_dates})
        dividend_mode = resolve_workspace_dividend_mode(price_only, include_dividends)
        for ticker in PORTFOLIO_BENCHMARK_TICKERS:
            try:
                dataset = fetch_history(ticker, include_dividends, dividend_mode=dividend_mode)
            except (ImportError, OSError, ValueError, KeyError, TypeError):
                continue
            aligned = pd.merge(
                reference_date_frame,
                dataset[["Date", "Close"]],
                on="Date",
                how="inner",
            ).sort_values("Date")
            if aligned.empty or len(aligned) != len(reference_date_frame):
                continue
            benchmark_series.append(
                build_series_payload(
                    ticker,
                    aligned,
                    color=PORTFOLIO_BENCHMARK_COLORS[ticker],
                    glow=False,
                )
            )
            benchmark_profiles.append(fetch_quote_profile(ticker, False))
        return benchmark_series, benchmark_profiles

    def resolve_view() -> str:
        return normalize_view_name(request.args.get("view", "tickers"))

    def build_legacy_workspace_redirect(view_name: str):
        query_string = request.query_string.decode().strip()
        target_path = build_view_path(view_name)
        return redirect(f"{target_path}?{query_string}" if query_string else target_path)

    def resolve_settings_section() -> str:
        for parameter_name in ("section", "settings_section"):
            requested_section = request.args.get(parameter_name, "").strip()
            if requested_section:
                return normalize_settings_section(requested_section)
        return "about"

    def resolve_settings_tab() -> str:
        for parameter_name in ("tab", "settings_tab", "language_tab"):
            requested_tab = request.args.get(parameter_name, "").strip()
            if requested_tab:
                return normalize_settings_tab(requested_tab)
        return "current"

    def settings_page_value() -> int:
        for parameter_name in ("page", "settings_page", "local_page", "language_page"):
            requested_page = request.args.get(parameter_name, "").strip()
            if requested_page:
                return normalize_settings_page(requested_page)
        return 1

    def should_use_modal_banner_message(message: str | None) -> bool:
        normalized = (message or "").strip()
        if not normalized:
            return False
        return (
                normalized.startswith("No market data returned for ")
                or normalized.startswith("Local market data for ")
                or normalized.startswith("Unknown or unsupported ticker: ")
                or normalized.startswith("has no local or remote market data")
                or normalized.startswith("Failed to perform, curl: (35) TLS connect error:")
        )

    def modal_banner_icon_class(message: str | None) -> str:
        normalized = (message or "").strip()
        if normalized.startswith("Backtest execution model updated:"):
            return "icon-modal-dialog-banner-backtest-execution"
        return "icon-modal-dialog-banner-default"

    def build_local_store_page_url(page_number: int) -> str:
        return build_settings_state_url("local-market-store", page=page_number)

    def _run_backtest_from_request():
        backtest_execution_mode = load_backtest_execution_mode()
        requested_tickers = parse_requested_tickers()
        if not requested_tickers:
            requested_tickers = [normalize_ticker_input(str(defaults.get("backtest_ticker", DEFAULT_TICKERS[0])))]
        if not requested_tickers:
            raise ValueError("No ticker selected for backtest.")
        trade_ticker = validate_ticker_or_raise(requested_tickers[0])
        backtest_cache_refresh = ensure_latest_backtest_caches(trade_ticker)
        price_only = request.args.get("return", "").strip().lower() == "price" or parse_bool_flag("price_only", "price_return_only")
        include_dividends = False if price_only else (
            request.args.get("return", "").strip().lower() == "dividends"
            or parse_bool_flag("dividends", "include_dividends")
        )
        range_mode, period, exact_start, exact_end = parse_range_request_args()
        supported_intervals = list_available_market_intervals(trade_ticker)
        requested_interval = request.args.get("interval", defaults.get("backtest_interval", DEFAULT_INTERVAL)).strip().lower()
        if not request.args.get("interval") and period == "1w" and "1m" in supported_intervals:
            requested_interval = "1m"
        if requested_interval not in supported_intervals:
            requested_interval = supported_intervals[0]
        trade_dataset = fetch_history(
            trade_ticker,
            False,
            interval=requested_interval,
            dividend_mode="price",
        )

        if requested_interval == "1m":
            six_months_ago = one_minute_lookback_start().tz_localize(None)
            trade_dataset = trade_dataset[trade_dataset["Date"] >= six_months_ago]

        date_constraints = build_date_constraint_payload(
            trade_dataset,
            requested_start=exact_start or None,
            requested_end=exact_end or None,
        )
        if range_mode == "exact":
            if not date_constraints.trading_dates:
                raise ValueError("The selected exact range does not contain trading dates.")
            trade_dataset = slice_dataset_to_exact_range(
                trade_dataset,
                date_constraints.adjusted_start,
                date_constraints.adjusted_end,
            )
            if trade_dataset.empty:
                raise ValueError("The selected exact range does not contain trading dates.")
        else:
            common_end_date = trade_dataset["Date"].max()
            trade_dataset = slice_dataset_for_period(trade_dataset, period, common_end_date)

        strategy_options = list_enabled_strategies()
        is_grid_workspace = (
            request.path == VIEW_PATHS["grid-trading"]
            or request.args.get("workspace", "").strip().lower() == "grid-trading"
        )
        default_strategy_id = (
            "grid-trading"
            if is_grid_workspace
            else defaults.get("backtest_strategy", strategy_options[0]["id"] if strategy_options else "")
        )
        selected_strategy_id = (
            "grid-trading"
            if is_grid_workspace
            else request.args.get("strategy", default_strategy_id).strip()
        )
        strategy_ids = {str(item["id"]) for item in strategy_options}
        if selected_strategy_id not in strategy_ids and strategy_options:
            selected_strategy_id = str(strategy_options[0]["id"])
        selected_strategy_params = collect_strategy_form_values(selected_strategy_id) if selected_strategy_id else {}
        backtest_initial_capital = max(
            parse_float_value(
                request.args.get("capital", request.args.get("initial_capital")),
                float(defaults.get("backtest_capital", 10000.0)),
            ),
            1.0,
        )

        strategy = instantiate_strategy(selected_strategy_id)
        signal_result = strategy.compute_signals(trade_dataset, selected_strategy_params)
        backtest_result = run_single_ticker_backtest(
            signal_result,
            backtest_initial_capital,
            execution_mode=backtest_execution_mode,
            interval=requested_interval,
            reinvest_cash_dividends=include_dividends,
            include_cash_dividends=not price_only,
        )
        return (
            backtest_result,
            trade_ticker,
            requested_interval,
            date_constraints,
            trade_dataset,
            selected_strategy_id,
            selected_strategy_params,
            backtest_cache_refresh,
        )

    def build_strategy_option_groups(strategy_options: list[dict[str, object]]) -> list[dict[str, object]]:
        return build_strategy_option_groups_for_recent(
            strategy_options,
            top_used_strategies(limit=3),
        )

    def collect_strategy_form_values(strategy_id: str) -> dict[str, Any]:
        strategy = instantiate_strategy(strategy_id)
        raw_values: dict[str, Any] = {}
        for definition in strategy.get_parameter_definitions():
            raw_value = request.args.get(definition.key)
            if raw_value is None or str(raw_value).strip() == "":
                raw_values[definition.key] = definition.default
            else:
                raw_values[definition.key] = raw_value
        return strategy.normalize_params(raw_values)

    def build_strategy_form_fields(strategy_id: str, values: dict[str, Any] | None = None) -> list[dict[str, object]]:
        return build_strategy_form_fields_for_strategy(
            strategy_id,
            values,
            strategy_factory=instantiate_strategy,
        )

    def build_strategy_settings_rows(strategy_options: list[dict[str, object]]) -> list[dict[str, object]]:
        return build_strategy_settings_rows_for_factory(
            strategy_options,
            strategy_factory=instantiate_strategy,
        )

    def build_local_store_pagination_ranges(
            first_page: int,
            last_page: int,
            chunk_size: int = 5,
    ) -> list[tuple[int, int]]:
        if first_page > last_page:
            return []
        ranges = [
            (range_start, min(range_start + chunk_size - 1, last_page))
            for range_start in range(first_page, last_page + 1, chunk_size)
        ]
        if len(ranges) > 1 and ranges[-1][1] - ranges[-1][0] + 1 < chunk_size:
            ranges[-2] = (ranges[-2][0], ranges[-1][1])
            ranges.pop()
        return ranges

    def build_local_store_pagination_items(
            current_page: int,
            total_pages: int,
    ) -> list[dict[str, Any]]:
        page_group_index = (current_page - 1) // 5
        page_start = (page_group_index * 5) + 1
        page_end = min(page_start + 4, total_pages)
        items: list[dict[str, Any]] = []

        if total_pages <= 1:
            return items
        if total_pages <= 5:
            return [
                {
                    "kind": "page",
                    "page": page_number,
                    "is_active": page_number == current_page,
                }
                for page_number in range(1, total_pages + 1)
            ]

        if page_start > 1:
            items.extend((
                {"kind": "previous", "page": page_start - 1},
                {"kind": "page", "page": 1, "is_active": current_page == 1},
                {
                    "kind": "ellipsis",
                    "position": "leading",
                    "ranges": build_local_store_pagination_ranges(1, page_start - 1),
                },
            ))

        items.extend(
            {
                "kind": "page",
                "page": page_number,
                "is_active": page_number == current_page,
            }
            for page_number in range(page_start, page_end + 1)
        )

        if page_end < total_pages:
            items.extend((
                {
                    "kind": "ellipsis",
                    "position": "trailing",
                    "ranges": build_local_store_pagination_ranges(page_end + 1, total_pages),
                },
                {"kind": "page", "page": total_pages, "is_active": current_page == total_pages},
                {"kind": "next", "page": page_end + 1},
            ))
        return items

    def has_local_profile_snapshot(ticker: str) -> bool:
        return has_profile_record(ticker)

    def has_local_logo_snapshot(ticker: str) -> bool:
        return has_logo_asset(ticker)

    def list_local_market_tickers() -> list[str]:
        # Project canonical form for US stocks is bare symbol (e.g. "BAC").
        # Support legacy files named "XXX.US.parquet" coming from Longbridge imports
        # without polluting the displayed symbol.
        def _has_usable_history(t: str) -> bool:
            p = history_store_path_for(t)
            if p.exists() and p.stat().st_size > 0:
                return True
            # legacy polluted name from Longbridge
            legacy = MARKET_STORE_DIR / "historical" / f"{normalize_ticker(t)}.US.parquet"
            return legacy.exists() and legacy.stat().st_size > 0
        return [
            ticker
            for ticker in list_local_tickers()
            if _has_usable_history(ticker)
               and has_local_profile_snapshot(ticker) and has_local_logo_snapshot(ticker)
        ]

    def load_local_profile_snapshot(ticker: str) -> tuple[str, str] | None:
        normalized_ticker = normalize_ticker_input(ticker)
        fallback_logo_url = ""
        for candidate in iter_investment_store_ticker_aliases(ticker):
            profile_record = load_profile_record(candidate)
            if profile_record is None:
                continue
            logo_url = resolve_stored_logo_url(candidate)
            if not logo_url:
                continue
            company_name = str(profile_record.get("company_name") or "").strip()
            if not is_ticker_fallback_company_name(company_name, normalized_ticker):
                return company_name, logo_url
            if company_name:
                fallback_logo_url = fallback_logo_url or logo_url
        if normalized_ticker and fallback_logo_url:
            return (
                resolve_known_ticker_company_name(normalized_ticker) or normalized_ticker,
                fallback_logo_url,
            )
        if normalized_ticker:
            return None
        return None

    def resolve_ticker_identity_snapshot(
            ticker: str,
            *,
            allow_remote_refresh: bool = True,
    ) -> tuple[str, str]:
        profile_snapshot = load_local_profile_snapshot(ticker)
        if profile_snapshot is not None:
            return profile_snapshot

        normalized_ticker = normalize_ticker_input(ticker)
        company_name = normalized_ticker
        logo_url = ""
        for candidate in iter_investment_store_ticker_aliases(ticker):
            profile_record = load_profile_record(candidate) or {}
            candidate_company_name = str(profile_record.get("company_name") or "").strip()
            candidate_logo_url = resolve_stored_logo_url(candidate)
            if not is_ticker_fallback_company_name(candidate_company_name, normalized_ticker):
                company_name = candidate_company_name
            elif not company_name:
                company_name = candidate_company_name or normalized_ticker
            if candidate_logo_url and not logo_url:
                logo_url = candidate_logo_url
            if logo_url and company_name and company_name.upper() != normalized_ticker:
                break

        if allow_remote_refresh and (not logo_url or company_name.upper() == normalized_ticker):
            for candidate in iter_investment_store_ticker_aliases(ticker):
                profile = fetch_quote_profile(candidate, force_refresh=True)
                profile_company_name = str(profile.company_name or "").strip()
                if profile_company_name and profile_company_name.upper() != str(candidate).upper():
                    company_name = profile_company_name
                logo_url = str(profile.logo_url or "").strip() or logo_url
                if logo_url and company_name and company_name.upper() != normalized_ticker:
                    break
        if company_name.upper() == normalized_ticker:
            known_company_name = resolve_known_ticker_company_name(ticker)
            if known_company_name:
                company_name = known_company_name
        return company_name, logo_url

    def build_local_market_rows_for_tickers(
            tickers: list[str],
            *,
            include_ranges: bool,
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for ticker in tickers:
            history_path = history_store_path_for(ticker)
            if not history_path.exists() or history_path.stat().st_size == 0:
                # fallback for Longbridge-polluted "BAC.US.parquet" etc.
                legacy_path = MARKET_STORE_DIR / "historical" / f"{normalize_ticker(ticker)}.US.parquet"
                if legacy_path.exists() and legacy_path.stat().st_size > 0:
                    history_path = legacy_path
                else:
                    continue
            profile_snapshot = load_local_profile_snapshot(ticker)
            if profile_snapshot is None:
                continue
            company_name, logo_url = profile_snapshot
            range_start = ""
            range_end = ""
            if include_ranges:
                try:
                    dataset = pd.read_parquet(history_path, columns=["Date"])
                    if dataset.empty:
                        continue
                    date_values = dataset["Date"]
                    if isinstance(date_values, pd.DataFrame):
                        date_values = date_values.iloc[:, 0]
                    range_start = format_store_range_date(date_values.min())
                    range_end = format_store_range_date(date_values.max())
                except (ImportError, OSError, ValueError, KeyError, TypeError):
                    pass
            daily_store_status = classify_daily_store_status(ticker)
            intraday_store_status = classify_one_minute_store_status(ticker)
            rows.append(
                {
                    "ticker": ticker,
                    "company_name": company_name,
                    "logo_url": logo_url,
                    "range_start": range_start,
                    "range_end": range_end,
                    "range": f"{range_start} - {range_end}" if range_start and range_end else "",
                    "has_1m": intraday_store_status == "fresh",
                    "has_1d": daily_store_status == "fresh",
                    "daily_store_status": daily_store_status,
                    "intraday_store_status": intraday_store_status,
                }
            )
        return rows

    def build_network_service_rows(
        *,
        pending: bool,
        service_labels: dict[str, str] | None = None,
        translate_fn: Callable[[str], str] | None = None,
    ) -> list[dict[str, Any]]:
        if service_labels is None or translate_fn is None:
            current_language_settings = load_language_settings()
            current_translations = build_translation_map(current_language_settings)
            service_labels = service_labels or translate_labels(
                base_labels,
                current_language_settings,
            )
            if translate_fn is None:
                def translate_fn(value: str) -> str:
                    return translate_text(
                        value,
                        current_language_settings.language,
                        current_translations,
                    )

        translate = translate_fn

        def service_logo_url(filename: str) -> str:
            return url_for("static", filename=f"images/{filename}")

        def format_checked_at(value: float | None) -> str:
            prefix = translate("Last checked:")
            if value is None:
                return f"{prefix} {translate('Not checked yet.')}"
            stamp = pd.Timestamp(value, unit="s")
            return f"{prefix} {format_display_datetime(stamp, include_seconds=True)}"

        service_definitions = (
            {
                "key": "market",
                "name": "yfinance",
                "logo": "Yahoo-Logo.svg",
                "pending_note": "Checking Yahoo Finance Chart through the verified HTTP(S) and yfinance transports from this application host.",
            },
            {
                "key": "sec",
                "name": "SEC EDGAR",
                "logo": "network.svg",
                "pending_note": "Checking SEC EDGAR submissions through the verified HTTP(S) transport from this application host.",
            },
            {
                "key": "longbridge",
                "name": "Longbridge OpenAPI",
                "logo": "network.svg",
                "pending_note": "Checking the Longbridge OpenAPI transport without sending credentials or trading requests.",
            },
            {
                "key": "logo",
                "name": service_labels["logo_network"],
                "logo": "apple.logo.svg",
                "pending_note": "Checking the primary ticker logo provider and its fallback providers from this application host.",
            },
            {
                "key": "google-hk",
                "name": translate("Google (Hong Kong)"),
                "logo": "Google__G__logo.svg",
                "pending_note": "Checking Google (Hong Kong) and its global fallback from this application host.",
            },
            {
                "key": "smtp",
                "name": "Yahoo Mail SMTP",
                "logo": "envelope.fill.svg",
                "pending_note": "Checking the configured SMTP host and STARTTLS transport without submitting mailbox credentials.",
            },
        )

        if pending:
            return [
                {
                    "key": definition["key"],
                    "name": definition["name"],
                    "status": translate("Checking..."),
                    "note": translate(definition["pending_note"]),
                    "pending_note": translate(definition["pending_note"]),
                    "checked_at_text": f"{translate('Last checked:')} {translate('Checking...')}",
                    "logo_url": service_logo_url(definition["logo"]),
                    "is_available": False,
                    "is_pending": True,
                }
                for definition in service_definitions
            ]

        raw_payload = run_network_self_check(
            smtp_settings=load_smtp_settings(),
            broker_settings=load_broker_settings(),
        )
        raw_rows = {
            str(item.get("key")): item
            for item in raw_payload.get("rows", [])
            if isinstance(item, dict)
        }
        status_labels = {
            "available": service_labels["service_ok"],
            "unavailable": service_labels["service_down"],
            "disabled": translate("Disabled by configuration"),
            "not_configured": translate("Not configured"),
            "not_installed": translate("Not installed"),
            "not_applicable": translate("Not applicable"),
        }
        rows: list[dict[str, Any]] = []
        for definition in service_definitions:
            raw_row = raw_rows.get(definition["key"], {})
            state = str(raw_row.get("state") or "unavailable")
            checked_at_value = raw_row.get("checked_at")
            try:
                checked_at = float(checked_at_value) if checked_at_value is not None else None
            except (TypeError, ValueError):
                checked_at = None
            rows.append(
                {
                    "key": definition["key"],
                    "name": definition["name"],
                    "status": status_labels.get(state, service_labels["service_down"]),
                    "note": str(raw_row.get("note") or translate("No diagnostic was returned.")),
                    "pending_note": translate(definition["pending_note"]),
                    "checked_at_text": format_checked_at(checked_at),
                    "logo_url": service_logo_url(definition["logo"]),
                    "is_available": bool(raw_row.get("is_available")),
                    "is_pending": False,
                    "state": state,
                    "latency_ms": raw_row.get("latency_ms"),
                }
            )
        return rows

    def maintain_local_market_store() -> dict[str, Any]:
        historical_tickers = list_historical_tickers()
        if not historical_tickers:
            return {
                "total_count": 0,
                "history_refreshed_count": 0,
                "metadata_refreshed_count": 0,
                "metadata_blocked_count": 0,
                "history_failed_tickers": [],
            }

        def refresh_local_entry(ticker: str) -> tuple[str, bool]:
            refresh_history_store(ticker)
            metadata_was_refreshed = refresh_quote_profile_cache(ticker, force_refresh=True)
            if not metadata_was_refreshed:
                fetch_quote_profile(ticker, force_refresh=False)
            return ticker, metadata_was_refreshed

        history_refreshed_count = 0
        metadata_refreshed_count = 0
        history_failed_tickers: list[str] = []
        worker_count = min(6, len(historical_tickers))
        with ThreadPoolExecutor(max_workers=worker_count) as executor:
            futures = {executor.submit(refresh_local_entry, ticker): ticker for ticker in historical_tickers}
            for future in as_completed(futures):
                ticker = futures[future]
                try:
                    _, entry_metadata_refreshed = future.result()
                    history_refreshed_count += 1
                    if entry_metadata_refreshed:
                        metadata_refreshed_count += 1
                except (ImportError, OSError, ValueError, KeyError, TypeError):
                    history_failed_tickers.append(ticker)
        return {
            "total_count": len(historical_tickers),
            "history_refreshed_count": history_refreshed_count,
            "metadata_refreshed_count": metadata_refreshed_count,
            "metadata_blocked_count": max(history_refreshed_count - metadata_refreshed_count, 0),
            "history_failed_tickers": history_failed_tickers,
        }

    def local_store_page_value() -> int:
        return settings_page_value()

    def build_modern_query_pairs(view_name: str | None = None) -> list[tuple[str, str]]:
        pairs: list[tuple[str, str]] = []
        view_max_tickers = max_tickers_for_view(view_name)

        for ticker in parse_requested_tickers(view_name):
            pairs.append(("ticker", ticker))

        has_weight_args = bool(request.args.getlist("weight")) or any(
            key.startswith("weight_") for key in request.args.keys()
        )
        allocation_mode = parse_portfolio_allocation_mode()
        has_share_args = bool(request.args.getlist("shares")) or any(
            key.startswith("shares_") for key in request.args.keys()
        )
        if allocation_mode == "shares" and has_share_args:
            pairs.append(("allocation", "shares"))
            for share_count in parse_requested_shares(view_max_tickers):
                pairs.append(("shares", str(share_count)))
        elif has_weight_args:
            for weight in parse_requested_weights(view_max_tickers):
                pairs.append(("weight", str(weight)))

        raw_range_value = request.args.get("range", request.args.get("range_mode", "")).strip().lower()
        period_value = request.args.get("period", "").strip().lower()
        if raw_range_value in {"exact", "custom"}:
            pairs.append(("range", "custom"))
            if period_value:
                pairs.append(("period", period_value))
        elif period_value and period_value != str(defaults.get("period", DEFAULT_PERIOD)).strip().lower():
            pairs.append(("range", period_value))
        elif raw_range_value and raw_range_value not in {"period"}:
            pairs.append(("range", raw_range_value))

        date_value = request.args.get("date", request.args.get("trading_date", request.args.get("exact_trading_date", ""))).strip()
        start_value = request.args.get("from", request.args.get("exact_start", "")).strip()
        end_value = request.args.get("to", request.args.get("exact_end", "")).strip()
        if date_value and (period_value == "1d" or raw_range_value in {"exact", "custom"}):
            pairs.append(("date", date_value))
        else:
            if start_value:
                pairs.append(("from", start_value))
            if end_value:
                pairs.append(("to", end_value))

        return_value = request.args.get("return", "").strip().lower()
        price_only_value = request.args.get("price_only", request.args.get("price_return_only", "")).strip()
        dividends_value = request.args.get("dividends", request.args.get("include_dividends", "")).strip()
        if return_value in {"price", "total"}:
            if return_value == "price":
                pairs.append(("return", "price"))
        elif price_only_value == "1":
            pairs.append(("return", "price"))
        elif dividends_value == "1":
            pairs.append(("dividends", "1"))

        overnight_value = request.args.get("overnight", request.args.get("include_overnight", "")).strip()
        if overnight_value == "1":
            pairs.append(("overnight", "1"))

        extended_hours_value = request.args.get("extended-hours", request.args.get("extended_hours", request.args.get("include_extended_hours", ""))).strip()
        if extended_hours_value == "1":
            pairs.append(("extended-hours", "1"))

        strategy_value = request.args.get("strategy", "").strip()
        if strategy_value:
            pairs.append(("strategy", strategy_value))

        capital_value = request.args.get("capital", request.args.get("initial_capital", "")).strip()
        if capital_value:
            pairs.append(("capital", capital_value))

        amount_value = request.args.get("amount", "").strip()
        if amount_value:
            pairs.append(("amount", amount_value))

        frequency_value = request.args.get("frequency", "").strip().lower()
        if frequency_value:
            pairs.append(("frequency", frequency_value))

        weekday_value = request.args.get("weekday", "").strip()
        if weekday_value:
            pairs.append(("weekday", weekday_value))

        month_day_value = request.args.get("month-day", request.args.get("month_day", "")).strip()
        if month_day_value:
            pairs.append(("month-day", month_day_value))

        tab_value = request.args.get("tab", request.args.get("trade_detail_tab", "")).strip().lower()
        if tab_value == "transactions":
            pairs.append(("tab", "transactions"))

        page_value = request.args.get("page", request.args.get("local_page", "")).strip()
        if page_value:
            pairs.append(("page", page_value))

        passthrough_keys = {
            "ticker",
            "tickers",
            "weight",
            "allocation",
            "shares",
            "period",
            "range",
            "range_mode",
            "date",
            "trading_date",
            "exact_trading_date",
            "from",
            "to",
            "exact_start",
            "exact_end",
            "dividends",
            "include_dividends",
            "return",
            "price_only",
            "price_return_only",
            "extended-hours",
            "extended_hours",
            "include_extended_hours",
            "overnight",
            "include_overnight",
            "strategy",
            "capital",
            "initial_capital",
            "amount",
            "frequency",
            "weekday",
            "month_day",
            "month-day",
            "tab",
            "trade_detail_tab",
            "page",
            "local_page",
            "view",
            "section",
            "ticker_a",
            "ticker_b",
        }
        passthrough_keys.update({f"ticker_{index}" for index in range(1, view_max_tickers + 1)})
        passthrough_keys.update({f"weight_{index}" for index in range(1, view_max_tickers + 1)})
        passthrough_keys.update({f"shares_{index}" for index in range(1, view_max_tickers + 1)})

        strategy_param_keys: set[str] = set()
        strategy_value = request.args.get("strategy", "").strip()
        if strategy_value:
            try:
                strategy = instantiate_strategy(strategy_value)
                strategy_param_keys = {definition.key for definition in strategy.get_parameter_definitions()}
            except (AttributeError, ImportError, TypeError, ValueError):
                strategy_param_keys = set()

        for key in request.args.keys():
            if key in {"view", "section"}:
                continue
            if key in passthrough_keys and key not in strategy_param_keys:
                continue
            for value in request.args.getlist(key):
                cleaned = str(value).strip()
                if cleaned:
                    pairs.append((key, cleaned))

        return pairs

    def resolve_effective_period_for_many(requested_period: str, datasets: list[pd.DataFrame]) -> tuple[str, str | None]:
        return resolve_effective_period_for_datasets(requested_period, datasets)

    def render_workspace_page(current_view: str, settings_section: str = "about", trade_section: str = "investment"):
        backtest_execution_mode = load_backtest_execution_mode()
        investment_cost_basis_method = load_investment_cost_basis_method()
        date_display_settings = load_date_display_settings()
        language_settings = load_language_settings()
        labels = translate_labels(base_labels, language_settings)
        language_translations = build_translation_map(language_settings)
        def translate_ui(value: str) -> str:
            return translate_text(value, language_settings.language, language_translations)
        language_history_rows = [
            {
                "timestamp": str(entry.get("timestamp", "")),
                "change": change,
            }
            for entry in reversed(language_settings.history)
            for change in entry.get("changes", [])
        ]
        is_dock_prefetch = request.headers.get("X-Requested-With") == "dock-prefetch"
        view_max_tickers = max_tickers_for_view(current_view)
        requested_tickers = parse_requested_tickers(current_view)
        range_mode, period, exact_start, exact_end = parse_range_request_args()
        price_only = request.args.get("return", "").strip().lower() == "price" or parse_bool_flag("price_only", "price_return_only", default=bool(defaults.get("price_only", False)))
        include_dividends = False if price_only else (
            request.args.get("return", "").strip().lower() == "dividends"
            or parse_bool_flag("dividends", "include_dividends")
        )
        if current_view in {"prices", "market-caps"}:
            price_only = True
            include_dividends = False
        include_extended_hours = current_view in {"tickers", "market-caps", "prices"} and period == "1d"
        include_overnight = (
            current_view in {"tickers", "market-caps", "prices"}
            and period == "1d"
            and parse_bool_flag("overnight", "include_overnight")
        )
        show_extended_hours_toggle = False
        show_overnight_toggle = False

        if current_view in {"tickers", "market-caps", "prices"} and not requested_tickers:
            requested_tickers = [
                normalize_ticker_input(defaults.get("ticker_a", DEFAULT_TICKERS[0])),
                normalize_ticker_input(defaults.get("ticker_b", DEFAULT_TICKERS[1])),
            ]
            include_dividends = False
        elif current_view == "portfolio" and not requested_tickers:
            requested_tickers = [
                                    normalize_ticker_input(value)
                                    for value in defaults.get("portfolio_tickers", ["NVDA", "AAPL", "QQQ"])
                                    if normalize_ticker_input(value)
                                ][:view_max_tickers]
            include_dividends = False
        elif current_view in BACKTEST_VIEWS and not requested_tickers:
            default_trade_ticker = normalize_ticker_input(
                defaults.get("backtest_ticker", defaults.get("ticker_a", DEFAULT_TICKERS[0]))
            )
            requested_tickers = [default_trade_ticker] if default_trade_ticker else [DEFAULT_TICKERS[0]]
            include_dividends = False if price_only else bool(defaults.get("backtest_include_dividends", False))
        elif current_view == "dca" and not requested_tickers:
            default_trade_ticker = normalize_ticker_input(
                defaults.get("dca_ticker", defaults.get("ticker_a", DEFAULT_TICKERS[0]))
            )
            requested_tickers = [default_trade_ticker] if default_trade_ticker else [DEFAULT_TICKERS[0]]
            include_dividends = False if price_only else bool(defaults.get("dca_include_dividends", False))

        settings_feedback = _read_settings_feedback() if current_view == "settings" else {}
        error = (
            settings_feedback.get("error")
            if current_view == "settings"
            else (request.args.get("error", "").strip() or None)
        )
        notice = (
            settings_feedback.get("notice")
            if current_view == "settings"
            else (request.args.get("notice", "").strip() or None)
        )
        broker_test_status = (
            settings_feedback.get("broker_test_status", "").lower() or None
            if current_view == "settings"
            else (request.args.get("broker_test_status", "").strip().lower() or None)
        )
        broker_test_message = (
            settings_feedback.get("broker_test_message")
            if current_view == "settings"
            else (request.args.get("broker_test_message", "").strip() or None)
        )
        broker_test_checked_at = (
            settings_feedback.get("broker_test_checked_at")
            if current_view == "settings"
            else (request.args.get("broker_test_checked_at", "").strip() or None)
        )
        longbridge_oauth_pending = (
            settings_feedback.get("longbridge_oauth_pending") == "1"
            if current_view == "settings"
            else request.args.get("longbridge_oauth_pending", "").strip() == "1"
        )
        floating_banner_icon_class = "icon-modal-dialog-banner-default"
        if notice and "Successfully connected" in notice:
            floating_banner_icon_class = "icon-settings-broker"
        elif error:
            floating_banner_icon_class = "icon-modal-dialog-banner-default"  # Or some error icon
        exact_start_value = exact_start
        exact_end_value = exact_end
        chart_trading_date_value = ""
        display_range = ""
        profiles: list[QuoteProfile] = []
        series: list[SeriesPayload] = []
        performance_items = []
        portfolio_items = []
        portfolio_weights = []
        portfolio_shares = []
        portfolio_allocation_mode = parse_portfolio_allocation_mode()
        portfolio_total_return = None
        validated_tickers: list[str] = []
        control_tickers: list[str] = []
        strategy_options = list_enabled_strategies()
        strategy_option_groups = build_strategy_option_groups(strategy_options)
        selected_strategy_id = (
            "grid-trading"
            if current_view == "grid-trading"
            else request.args.get(
                "strategy",
                defaults.get("backtest_strategy", strategy_options[0]["id"] if strategy_options else "")
                if current_view == "backtest"
                else (strategy_options[0]["id"] if strategy_options else ""),
            ).strip()
        )
        strategy_ids = {str(item["id"]) for item in strategy_options}
        if selected_strategy_id not in strategy_ids and strategy_options:
            selected_strategy_id = str(strategy_options[0]["id"])
        selected_strategy_params = collect_strategy_form_values(selected_strategy_id) if selected_strategy_id else {}
        strategy_form_fields = build_strategy_form_fields(selected_strategy_id, selected_strategy_params) if selected_strategy_id else []
        backtest_initial_capital = max(
            parse_float_value(
                request.args.get("capital", request.args.get("initial_capital")),
                float(defaults.get("backtest_capital", 10000.0)) if current_view in BACKTEST_VIEWS else 10000.0,
            ),
            1.0,
        )
        dca_amount = max(
            parse_float_value(
                request.args.get("amount"),
                1000.0,
            ),
            1.0,
        )
        dca_frequency = (
            "weekly"
            if request.args.get("frequency", str(defaults.get("dca_frequency", "monthly"))).strip().lower() == "weekly"
            else "monthly"
        )
        dca_weekday = min(max(parse_int_value(request.args.get("weekday"), parse_int_value(defaults.get("dca_weekday"), 0)), 0), 4)
        dca_month_day = min(max(parse_int_value(request.args.get("month-day", request.args.get("month_day")), parse_int_value(defaults.get("dca_month_day"), 15)), 1), 28)
        requested_interval = request.args.get("interval", defaults.get("backtest_interval", DEFAULT_INTERVAL)).strip().lower()
        supported_intervals = ["1d"]
        if current_view in BACKTEST_VIEWS and requested_tickers:
            try:
                trade_ticker = validate_ticker_or_raise(requested_tickers[0])
                supported_intervals = list_available_market_intervals(trade_ticker)
            except ValueError:
                pass

        # Smart default for 1w period if interval is not specified
        if not request.args.get("interval") and period == "1w" and "1m" in supported_intervals:
            requested_interval = "1m"

        if requested_interval not in supported_intervals:
            requested_interval = supported_intervals[0]
        if current_view == "dca":
            requested_interval = "1d"
            supported_intervals = ["1d"]

        backtest_result = None
        dca_result = None
        backtest_market_refresh: dict[str, str | bool | None] | None = None
        ticker_slots = requested_tickers.copy() if requested_tickers else ["", ""]
        requested_weights = parse_requested_weights(max(len(ticker_slots), MIN_TICKERS)) if current_view == "portfolio" else []
        requested_shares = parse_requested_shares(max(len(ticker_slots), MIN_TICKERS)) if current_view == "portfolio" else []
        has_weight_query = bool(request.args.getlist("weight")) or any(
            key.startswith("weight_") for key in request.args.keys()
        )
        has_share_query = bool(request.args.getlist("shares")) or any(
            key.startswith("shares_") for key in request.args.keys()
        )
        if current_view == "portfolio" and not has_weight_query:
            requested_weights = [
                                    min(max(parse_int_value(value, 0), 0), 100)
                                    for value in defaults.get("portfolio_weights", [25, 25, 50])
                                ][:max(len(requested_tickers), MIN_TICKERS)]
        if current_view == "portfolio" and portfolio_allocation_mode != "shares":
            requested_shares = [0] * max(len(ticker_slots), MIN_TICKERS)
        if current_view == "portfolio" and portfolio_allocation_mode == "shares" and not has_share_query:
            requested_shares = [0] * max(len(ticker_slots), MIN_TICKERS)
        period_label = format_period_label(period)
        page_title = labels["hero_title"]
        report_heading = labels["performance_summary"]
        chart_heading = labels["chart_summary"]
        settings_title = labels["about"]
        settings_service_rows: list[dict[str, Any]] = []
        strategy_settings_rows: list[dict[str, object]] = []
        style_token_rows: list[dict[str, object]] = []
        color_token_rows: list[dict[str, object]] = []
        export_image_rows: list[dict[str, object]] = []
        material_token_rows: list[dict[str, object]] = []
        cash_equivalent_rows: list[dict[str, object]] = []
        cash_equivalent_fund_rows: list[dict[str, object]] = []
        font_token_rows: list[dict[str, object]] = []
        smtp_settings = sanitize_smtp_settings_for_view(load_smtp_settings())
        broker_settings = sanitize_broker_settings_for_view(load_broker_settings())
        local_market_rows: list[dict[str, Any]] = []
        local_store_total_pages = 1
        local_store_current_page = 1
        local_store_pagination_items: list[dict[str, Any]] = []
        settings_tab = "current"
        settings_page_number = 1
        backtest_periods_by_interval: dict[str, list[str]] = {
            "1d": list(SUPPORTED_PERIODS_1D),
            "1m": list(SUPPORTED_PERIODS_1M),
        }

        settings_section = normalize_settings_section(settings_section)
        trade_section = normalize_trade_section(trade_section)

        if current_view == "settings":
            settings_tab = resolve_settings_tab() if settings_section == "general" else "current"
            if settings_section == "general":
                settings_page_number = settings_page_value()
                language_row_count = (
                    len(language_settings.translations)
                    if settings_tab == "current"
                    else max(len(language_history_rows), 1)
                )
                language_total_pages = max(
                    (language_row_count - 1) // SETTINGS_LANGUAGE_PAGE_SIZE + 1,
                    1,
                )
                settings_page_number = min(settings_page_number, language_total_pages)
            elif settings_section == "local-market-store":
                settings_page_number = settings_page_value()

        if current_view == "settings" and settings_section == "about":
            error = None
            notice = None

        if current_view == "prices":
            page_title = labels.get("dock_prices", "Price performance")
            report_heading = labels.get("dock_prices", "Price performance")
            chart_heading = "Price history"
        elif current_view == "market-caps":
            page_title = labels.get("dock_market_caps", "Market cap comparison")
            report_heading = labels.get("dock_market_caps", "Market cap comparison")
            chart_heading = "Market cap history"
        elif current_view == "portfolio":
            page_title = labels["portfolio_title"]
            report_heading = labels["portfolio_summary"]
            chart_heading = labels["portfolio_chart"]
        elif current_view == "dca":
            page_title = labels["dca_title"]
            report_heading = labels["dca_metrics"]
            chart_heading = labels["dca_chart"]
        elif current_view in BACKTEST_VIEWS:
            page_title = labels.get("grid_trading_title", "Grid trading") if current_view == "grid-trading" else labels["backtest_title"]
        elif current_view == "settings":
            page_title = labels["settings_title"]
            if settings_section == "network":
                settings_title = labels["network_self_check"]
            elif settings_section == "general":
                settings_title = translate_ui("General")
            elif settings_section == "backtest":
                settings_title = translate_ui("Backtest")
            elif settings_section == "investment":
                settings_title = translate_ui("Investment")
            elif settings_section == "font-tokens":
                settings_title = translate_ui("Font tokens")
            elif settings_section == "color-tokens":
                settings_title = translate_ui("Color tokens")
            elif settings_section == "material-tokens":
                settings_title = translate_ui("Material tokens")
            elif settings_section == "strategies":
                settings_title = labels["strategy_settings"]
            elif settings_section == "email-smtp":
                settings_title = labels["email_smtp"]
            elif settings_section == "broker-access":
                settings_title = translate_ui("Broker access")
            elif settings_section == "local-market-store":
                settings_title = labels["local_market_store"]
            elif settings_section == "clear-caches":
                settings_title = translate_ui("Clear caches")
            elif settings_section == "style-tokens":
                settings_title = translate_ui("Style tokens")
            elif settings_section == "export-image":
                settings_title = translate_ui("Export images")
            elif settings_section == "cash-equivalents":
                settings_title = translate_ui("Cash equivalents")
        elif current_view == "trade":
            page_title = labels["trade_title"]
            settings_title = labels["trade_title"]
            if trade_section == "investment":
                page_title = "Investment"
                settings_title = "Investment"
            elif trade_section == "live-trading":
                page_title = "Live trading"
                settings_title = "Live trading"

        supported_periods = (
            list(COMPARE_PERIODS_1D)
            if current_view in {"tickers", "market-caps", "prices", "portfolio"}
            else list(SUPPORTED_PERIODS_1M) if requested_interval == "1m" and "1m" in supported_intervals
            else list(SUPPORTED_PERIODS_1D)
        )

        if period not in supported_periods:
            period = supported_periods[0] if supported_periods else DEFAULT_PERIOD

        def handle_fetch_history_failure(
                ticker: str,
                include_dividends_flag: bool,
                dividend_mode: str | None = None,
        ) -> pd.DataFrame:
            """
            If fetching fails because remote data cannot be retrieved, try to load whatever local data exists.
            If any local parquet exists, even if incomplete, return it instead of raising immediately.
            """
            path = history_store_path_for(ticker)
            if path.exists():
                try:
                    return select_price_series(
                        pd.read_parquet(path),
                        include_dividends_flag,
                        dividend_mode=dividend_mode,
                    )
                except (ImportError, OSError, ValueError, KeyError, TypeError):
                    pass
            raise ValueError(f"No market data returned for {ticker}.")

        try:
            if current_view in BACKTEST_VIEWS:
                if requested_tickers:
                    backtest_refresh_ticker = validate_ticker_or_raise(requested_tickers[0])
                    backtest_market_refresh = ensure_latest_backtest_caches(backtest_refresh_ticker)
                # Check cache: skip re-computation if config unchanged
                cache_key = _get_backtest_cache_key()
                if cache_key in _cached_backtest:
                    # Cache hit - use cached result directly
                    (
                        backtest_result,
                        trade_ticker,
                        requested_interval,
                        date_constraints,
                        trade_dataset,
                        selected_strategy_id,
                        selected_strategy_params,
                    ) = _cached_backtest[cache_key]
                else:
                    # Cache miss - need to recompute and cache
                    (
                        backtest_result,
                        trade_ticker,
                        requested_interval,
                        date_constraints,
                        trade_dataset,
                        selected_strategy_id,
                        selected_strategy_params,
                        _,
                    ) = _run_backtest_from_request()
                    _cached_backtest[cache_key] = (
                        backtest_result, trade_ticker, requested_interval, date_constraints,
                        trade_dataset, selected_strategy_id, selected_strategy_params
                    )
                    # Limit cache size to prevent memory growth (keep last 8 cached results)
                    if len(_cached_backtest) > 8:
                        # Remove oldest entry
                        oldest_key = next(iter(_cached_backtest.keys()))
                        del _cached_backtest[oldest_key]
                record_strategy_usage(selected_strategy_id)
                ticker_slots = [trade_ticker]
                profiles = [fetch_quote_profile(trade_ticker, False)]
                backtest_periods_by_interval = {
                    "1d": build_supported_periods_for_history_store(trade_ticker, "1d"),
                    "1m": build_supported_periods_for_history_store(trade_ticker, "1m"),
                }
                if not trade_dataset.empty:
                    backtest_periods_by_interval[requested_interval] = build_supported_periods_from_dates(
                        trade_dataset["Date"],
                        interval=requested_interval,
                    )
                if backtest_market_refresh:
                    refresh_notices: list[str] = []
                    if backtest_market_refresh.get("daily_error"):
                        refresh_notices.append(
                            "Could not refresh the latest 1d cache automatically, so the backtest reused the newest local daily data."
                        )
                    if backtest_market_refresh.get("intraday_error"):
                        refresh_notices.append(
                            "Could not refresh the latest 1m cache automatically, so the backtest reused the newest local intraday data when available."
                        )
                    if refresh_notices:
                        refresh_notice = " ".join(refresh_notices)
                        if notice is None:
                            notice = refresh_notice
                        else:
                            notice += " " + refresh_notice
                supported_periods = backtest_periods_by_interval.get(requested_interval, list(SUPPORTED_PERIODS_1D))
                period, period_notice = resolve_requested_period_from_supported(
                    period,
                    supported_periods,
                    earliest_available=trade_dataset["Date"].min() if not trade_dataset.empty else None,
                )
                if period_notice and notice is None:
                    notice = period_notice
                elif period_notice:
                    notice += " " + period_notice
                exact_start_value = trade_dataset["Date"].min().strftime("%Y-%m-%d")
                exact_end_value = trade_dataset["Date"].max().strftime("%Y-%m-%d")
                if range_mode == "exact":
                    period_label = "Exact range"
                else:
                    period_label = format_period_label(period)
                display_range = f"{format_display_date(trade_dataset['Date'].min())} - {format_display_date(trade_dataset['Date'].max())}"
            elif current_view == "dca":
                if requested_tickers:
                    dca_ticker = validate_ticker_or_raise(requested_tickers[0])
                    dca_refresh_failures = ensure_latest_daily_caches([dca_ticker])
                else:
                    raise ValueError("No ticker selected for recurring investment.")

                try:
                    dca_dataset = fetch_history(dca_ticker, False, dividend_mode="price")
                except ValueError:
                    dca_dataset = handle_fetch_history_failure(dca_ticker, False, dividend_mode="price")

                date_constraints = build_date_constraint_payload(
                    dca_dataset,
                    requested_start=exact_start or None,
                    requested_end=exact_end or None,
                )
                if range_mode == "exact":
                    if not date_constraints.trading_dates:
                        raise ValueError("The selected exact range does not contain trading dates.")
                    dca_dataset = slice_dataset_to_exact_range(
                        dca_dataset,
                        date_constraints.adjusted_start,
                        date_constraints.adjusted_end,
                    )
                else:
                    supported_periods = build_supported_periods_from_dates(dca_dataset["Date"], interval="1d")
                    period, period_notice = resolve_requested_period_from_supported(
                        period,
                        supported_periods,
                        earliest_available=dca_dataset["Date"].min() if not dca_dataset.empty else None,
                    )
                    if period_notice and notice is None:
                        notice = period_notice
                    elif period_notice:
                        notice += " " + period_notice
                    dca_dataset = slice_dataset_for_period(dca_dataset, period, dca_dataset["Date"].max())

                if dca_dataset.empty:
                    raise ValueError(f"No market data available for {dca_ticker} in the selected range.")

                range_start = pd.Timestamp(dca_dataset["Date"].min()).strftime("%Y-%m-%d")
                range_end = pd.Timestamp(dca_dataset["Date"].max()).strftime("%Y-%m-%d")
                dca_result = simulate_recurring_investment(
                    dca_ticker,
                    dca_dataset,
                    amount_per_period=dca_amount,
                    frequency=dca_frequency,
                    weekday=dca_weekday,
                    month_day=dca_month_day,
                    reinvest_cash_dividends=include_dividends,
                    include_cash_dividends=not price_only,
                )
                profiles = [fetch_quote_profile(dca_ticker, False)]
                ticker_slots = [dca_ticker]
                exact_start_value = range_start
                exact_end_value = range_end
                period_label = "Exact range" if range_mode == "exact" else format_period_label(period)
                display_range = f"{format_display_date(dca_dataset['Date'].min())} - {format_display_date(dca_dataset['Date'].max())}"
                if dca_refresh_failures:
                    failed_preview = ", ".join(dca_refresh_failures)
                    refresh_notice = (
                        f"Could not refresh the latest trading-day cache for {failed_preview}. "
                        "Using the newest local daily data currently available."
                    )
                    if notice is None:
                        notice = refresh_notice
                    else:
                        notice += " " + refresh_notice
            elif current_view in {"tickers", "market-caps", "prices", "portfolio"}:
                if requested_tickers and len(requested_tickers) >= MIN_TICKERS:
                    if is_dock_prefetch:
                        validated_tickers = [normalize_ticker_input(t) or t for t in requested_tickers]
                        profiles = [
                            QuoteProfile(ticker=t, company_name=t, logo_url="")
                            for t in validated_tickers
                        ]
                        if current_view == "portfolio":
                            portfolio_weights = requested_weights or [0] * len(validated_tickers)
                            portfolio_shares = requested_shares or [0] * len(validated_tickers)
                            portfolio_items = [
                                {
                                    "ticker": t,
                                    "company_name": t,
                                    "logo_url": "",
                                    "weight": w,
                                    "shares": s,
                                    "growth_multiple": 1.0,
                                    "color": "transparent",
                                }
                                for t, w, s in zip(validated_tickers, portfolio_weights, portfolio_shares)
                            ]
                            portfolio_total_return = 0.0
                        else:
                            series = [
                                SeriesPayload(
                                    ticker=t,
                                    dates=[""],
                                    raw_dates=[""],
                                    normalized_returns=[0.0],
                                    color="transparent",
                                    glow=False,
                                )
                                for t in validated_tickers
                            ]
                            performance_items = [
                                {
                                    "ticker": t,
                                    "company_name": t,
                                    "logo_url": "",
                                    "ending_return": 0.0,
                                    "ttm_dividend_yield": None,
                                    "color": "transparent",
                                    "shadow_color": "transparent",
                                    "is_winner": False,
                                    "is_dividend_yield_winner": False,
                                }
                                for t in validated_tickers]
                        display_range = "Loading range..."
                        ticker_slots = (control_tickers or validated_tickers).copy()
                        continue_process_tickers = False
                    else:
                        validated_tickers = [validate_ticker_or_raise(ticker) for ticker in requested_tickers]
                        continue_process_tickers = True
                    if continue_process_tickers:
                        if len(set(validated_tickers)) != len(validated_tickers):
                            raise ValueError("Ticker symbols must be unique.")

                        control_tickers = validated_tickers.copy()
                        overnight_tickers = resolve_compare_overnight_tickers(control_tickers)
                        show_overnight_toggle = (
                            current_view in {"tickers", "market-caps", "prices"}
                            and len(overnight_tickers) <= view_max_tickers
                            and supports_compare_overnight(control_tickers, period)
                            and has_compare_overnight_market_data_source()
                        )
                        if not show_overnight_toggle:
                            include_overnight = False
                        elif include_overnight:
                            control_tickers = [
                                canonical_compare_overnight_ticker(ticker)
                                for ticker in control_tickers
                            ]
                            if len(set(control_tickers)) != len(control_tickers):
                                raise ValueError("SKHYV and SKHY identify the same security.")
                            validated_tickers = overnight_tickers

                        freshness_refresh_failures: list[str] = []
                        is_local_intraday_price_request = (
                            current_view in {"market-caps", "prices"}
                            and range_mode != "exact"
                            and period in {"1d", "3d", "1w"}
                        )
                        if (
                                current_view in {"tickers", "market-caps", "prices", "portfolio"}
                                and not is_local_intraday_price_request
                        ):
                            freshness_refresh_failures = ensure_latest_daily_caches(validated_tickers)

                        # Try to fetch datasets, handle missing remote data by falling back to any available local data
                        datasets: list[pd.DataFrame] = []
                        failed_fetches: list[str] = []
                        completely_missing: list[str] = []
                        dividend_mode = resolve_workspace_dividend_mode(price_only, include_dividends)
                        for ticker in validated_tickers:
                            try:
                                datasets.append(fetch_history(ticker, include_dividends, dividend_mode=dividend_mode))
                            except ValueError as fetch_exc:
                                if "No market data returned" in str(fetch_exc) or "Local market data for" in str(fetch_exc):
                                    try:
                                        dataset = handle_fetch_history_failure(ticker, include_dividends, dividend_mode=dividend_mode)
                                        datasets.append(dataset)
                                        failed_fetches.append(ticker)
                                    except ValueError:
                                        completely_missing.append(ticker)
                                else:
                                    raise

                        # If any ticker is completely missing (no local + no remote), replace it with the first available local ticker from usage history
                        if completely_missing:
                            local_tickers = [t for t in list_local_market_tickers() if t not in completely_missing]
                            if not local_tickers:
                                # If no local tickers available at all, use the default tickers to guarantee something renders
                                local_tickers = [normalize_ticker_input(t) for t in DEFAULT_TICKERS if normalize_ticker_input(t) not in completely_missing]

                            for missing_ticker in completely_missing:
                                # Pick the first available local ticker that has data
                                replacement = local_tickers[0] if len(local_tickers) > 0 else DEFAULT_TICKERS[0]
                                replacement = normalize_ticker_input(replacement)
                                # Replace in validated_tickers
                                idx = validated_tickers.index(missing_ticker)
                                validated_tickers[idx] = replacement
                                # Fetch dataset for replacement
                                try:
                                    dataset = fetch_history(replacement, include_dividends, dividend_mode=dividend_mode)
                                    # Remove the placeholder None we appended when skipping
                                    if len(datasets) > idx:
                                        datasets.pop(idx)
                                    datasets.insert(idx, dataset)
                                except (ImportError, OSError, ValueError, KeyError, TypeError):
                                    # This should not happen since we filtered local_tickers to only include available ones
                                    pass
                                # Add to notice
                                if notice is None:
                                    notice = f"{missing_ticker} has no local or remote market data, automatically replaced with {replacement}."
                                else:
                                    notice += f" {missing_ticker} has no local or remote market data, automatically replaced with {replacement}."

                        profiles = [fetch_quote_profile(ticker, False) for ticker in validated_tickers]
                        market_cap_split_events = {
                            ticker: extract_stock_split_events(dataset)
                            for ticker, dataset in zip(validated_tickers, datasets)
                        } if current_view == "market-caps" else {}
                        market_cap_split_actions_authoritative = {
                            ticker: bool(dataset.attrs.get("stock_split_actions_authoritative"))
                            for ticker, dataset in zip(validated_tickers, datasets)
                        } if current_view == "market-caps" else {}
                        include_extended_hours = (
                            current_view in {"tickers", "market-caps", "prices"}
                            and supports_compare_extended_hours(validated_tickers, period)
                        )
                        show_extended_hours_toggle = False
                        intraday_period_candidates = set(COMPARE_INTRADAY_PERIODS)
                        intraday_period_sets = [
                            {
                                candidate
                                for candidate in build_supported_periods_for_history_store(ticker, "1m")
                                if candidate in intraday_period_candidates
                            }
                            for ticker in validated_tickers
                        ]
                        supported_periods = build_supported_compare_periods(
                            extract_union_dates(datasets),
                            intraday_period_sets,
                        )
                        intraday_supported_periods = [
                            candidate
                            for candidate in supported_periods
                            if candidate in intraday_period_candidates
                        ]
                        if range_mode != "exact" and period not in intraday_period_candidates:
                            period, period_notice = resolve_requested_period_from_supported(
                                period,
                                supported_periods,
                                earliest_available=min(dataset["Date"].min() for dataset in datasets),
                            )
                            if period_notice and notice is None:
                                notice = period_notice
                            elif period_notice:
                                notice = f"{notice} {period_notice}"
                        date_constraints = build_date_constraint_payload(
                            *datasets,
                            requested_start=exact_start or None,
                            requested_end=exact_end or None,
                        )

                        # Auto-switch to Exact mode if we're in Relative mode and any ticker couldn't fetch full recent data
                        if range_mode != "exact" and len(failed_fetches) > 0:
                            # Get the minimal max date across all datasets (latest available data is bounded by the ticker with incomplete data)
                            common_max_end = min(dataset["Date"].max() for dataset in datasets)
                            # The requested period offset stays the same but end is now at the latest available common date
                            requested_start = (
                                max(dataset["Date"].min() for dataset in datasets).normalize()
                                if period == "max"
                                else (common_max_end - PERIOD_OFFSETS[period]).normalize()
                            )
                            adjusted_start = requested_start.strftime("%Y-%m-%d")
                            adjusted_end = common_max_end.strftime("%Y-%m-%d")
                            # Rebuild date constraints with the new exact range
                            date_constraints = build_date_constraint_payload(
                                *datasets,
                                requested_start=adjusted_start,
                                requested_end=adjusted_end,
                            )
                            range_mode = "exact"
                            period_label = "Exact range"
                            ticker_list = ", ".join(failed_fetches)
                            auto_notice = (
                                f"Could not retrieve the latest market data for {ticker_list}. "
                                f"Automatically switched to exact range from {format_display_date(pd.to_datetime(adjusted_start))} "
                                f"to {format_display_date(common_max_end)} based on available local data."
                            )
                            if notice is None:
                                notice = auto_notice
                            else:
                                notice += " " + auto_notice

                        if freshness_refresh_failures:
                            failed_preview = ", ".join(freshness_refresh_failures)
                            freshness_notice = (
                                f"Could not refresh the latest trading-day cache for {failed_preview}. "
                                "Using the newest local daily data currently available."
                            )
                            if notice is None:
                                notice = freshness_notice
                            else:
                                notice += " " + freshness_notice

                        is_exact_one_day_compare = current_view in {"tickers", "market-caps", "prices"} and range_mode == "exact" and period == "1d"
                        is_intraday_compare_period = range_mode != "exact" and period in {"1d", "3d", "1w"}
                        exact_range_trading_dates = exact_trading_dates_in_range(date_constraints)
                        is_exact_short_intraday_compare = (
                            current_view in {"tickers", "market-caps", "prices"}
                            and range_mode == "exact"
                            and period in {"3d", "1w"}
                            and not is_exact_one_day_compare
                            and 2 <= len(exact_range_trading_dates) <= 5
                        )
                        if is_exact_one_day_compare:
                            date_constraints = build_one_day_intraday_date_constraint_payload(
                                validated_tickers,
                                requested_start=exact_start or None,
                                requested_end=exact_end or None,
                                include_overnight_flag=include_overnight,
                            )
                        elif current_view in {"tickers", "market-caps", "prices"} and range_mode == "exact" and period in {"3d", "1w"}:
                            intraday_date_constraints = build_short_intraday_date_constraint_payload(
                                validated_tickers,
                                requested_start=exact_start or None,
                                requested_end=exact_end or None,
                            )
                            if intraday_date_constraints.trading_dates:
                                date_constraints = intraday_date_constraints
                                exact_range_trading_dates = exact_trading_dates_in_range(date_constraints)
                                is_exact_short_intraday_compare = 2 <= len(exact_range_trading_dates) <= 5

                        if is_exact_one_day_compare:
                            if not date_constraints.trading_dates:
                                raise ValueError("The selected tickers do not share any common trading dates.")
                            target_trading_date = date_constraints.adjusted_start or date_constraints.adjusted_end or date_constraints.max_date
                            if not target_trading_date:
                                raise ValueError("Select a shared trading date.")
                            live_session_date = pd.Timestamp.now(tz="Asia/Shanghai").date()
                            selected_markets = {
                                infer_ticker_market(ticker)
                                for ticker in validated_tickers
                            }
                            axis_trading_date = resolve_compare_axis_trading_date(
                                validated_tickers,
                                target_trading_date,
                            ) if target_trading_date == date_constraints.max_date else target_trading_date
                            reference_intraday_datasets = [
                                load_compare_one_day_intraday_dataset(
                                    ticker,
                                    include_extended_hours_flag=include_extended_hours,
                                    include_overnight_flag=include_overnight,
                                    trading_date=axis_trading_date,
                                )
                                for ticker in validated_tickers
                            ]
                            reference_common_end_date = min(dataset["Date"].max() for dataset in reference_intraday_datasets)
                            reference_aligned_datasets = slice_intraday_datasets_for_compare_period(
                                reference_intraday_datasets,
                                "1d",
                                reference_common_end_date,
                                validated_tickers,
                            )
                            if axis_trading_date != target_trading_date:
                                display_reference_aligned_datasets = [
                                    shift_intraday_compare_axis_to_trading_date(
                                        dataset,
                                        axis_trading_date,
                                        target_trading_date,
                                    )
                                    for dataset in reference_aligned_datasets
                                ]
                                aligned_datasets = []
                                for reference_dataset, ticker in zip(display_reference_aligned_datasets, validated_tickers):
                                    try:
                                        target_dataset = load_target_compare_one_day_intraday_dataset(
                                            ticker,
                                            target_trading_date=target_trading_date,
                                            include_extended_hours_flag=include_extended_hours,
                                            include_overnight_flag=include_overnight,
                                            force_refresh=False,
                                        )
                                        aligned_datasets.append(
                                            map_live_intraday_dataset_to_reference_axis(reference_dataset, target_dataset, ticker)
                                        )
                                    except Exception as exc:  # noqa: BLE001
                                        LOGGER.info("No live compare bars for %s on %s yet: %s", ticker, target_trading_date, exc)
                                        aligned_datasets.append(build_empty_compare_axis_dataset(reference_dataset))
                            else:
                                aligned_datasets = [
                                    apply_market_close_anchor(
                                        dataset,
                                        daily_dataset,
                                        ticker,
                                        target_trading_date,
                                    )
                                    for dataset, daily_dataset, ticker in zip(reference_aligned_datasets, datasets, validated_tickers)
                                ]
                            if (
                                    pd.to_datetime(target_trading_date, errors="coerce").date() == live_session_date
                                    and len(selected_markets) == 1
                            ):
                                aligned_datasets = truncate_intraday_datasets_to_common_live_timestamp(aligned_datasets)
                            chart_trading_date_value = pd.to_datetime(axis_trading_date).strftime("%Y-%m-%d")
                            exact_start_value = pd.to_datetime(target_trading_date).strftime("%Y-%m-%d")
                            exact_end_value = exact_start_value
                            period_label = "Trading date"
                        elif range_mode == "exact":
                            if not date_constraints.trading_dates:
                                raise ValueError("The selected tickers do not share any common trading dates.")
                            if is_exact_short_intraday_compare:
                                intraday_datasets = []
                                live_session_date = pd.Timestamp.now(
                                    tz=market_timezone_for_ticker(validated_tickers[0])
                                ).date()
                                should_append_exact_live = bool(exact_range_trading_dates) and pd.to_datetime(
                                    exact_range_trading_dates[-1],
                                    errors="coerce",
                                ).date() == live_session_date and any(
                                    is_market_regular_session_active_for_ticker(ticker)
                                    for ticker in validated_tickers
                                )
                                for ticker in validated_tickers:
                                    raw_intraday_dataset = fetch_history(
                                        ticker,
                                        include_dividends=False,
                                        interval="1m",
                                        dividend_mode="price",
                                    )
                                    if should_append_exact_live:
                                        raw_intraday_dataset = append_live_compare_intraday_dataset(
                                            ticker,
                                            raw_intraday_dataset,
                                            live_trading_date=live_session_date,
                                            include_extended_hours_flag=include_extended_hours,
                                            force_refresh=True,
                                        )[0]
                                    intraday_dataset = prepare_intraday_dataset_for_compare(
                                        raw_intraday_dataset,
                                        ticker,
                                        regular_session_only=True,
                                    )
                                    intraday_datasets.append(
                                        slice_intraday_dataset_to_trading_dates(
                                            intraday_dataset,
                                            ticker,
                                            exact_range_trading_dates,
                                        )
                                    )
                                aligned_datasets = align_intraday_datasets_for_compare(
                                    intraday_datasets,
                                    validated_tickers,
                                )
                                if should_append_exact_live and len({
                                    infer_ticker_market(ticker)
                                    for ticker in validated_tickers
                                }) == 1:
                                    aligned_datasets = truncate_intraday_datasets_to_common_live_timestamp(aligned_datasets)
                            else:
                                aligned_datasets = align_datasets_on_common_dates(datasets)
                                aligned_datasets = slice_datasets_to_exact_range(
                                    aligned_datasets,
                                    date_constraints.adjusted_start,
                                    date_constraints.adjusted_end,
                                )
                            if any(dataset.empty for dataset in aligned_datasets):
                                raise ValueError("The selected exact range does not contain shared trading dates.")
                            exact_start_value = date_constraints.adjusted_start or date_constraints.min_date or ""
                            exact_end_value = date_constraints.adjusted_end or date_constraints.max_date or ""
                            period_label = "Exact range"
                        elif is_intraday_compare_period:
                            if period not in intraday_supported_periods:
                                earliest_intraday = min(dataset["Date"].min() for dataset in datasets) if datasets else None
                                period, intraday_notice = resolve_requested_period_from_supported(
                                    period,
                                    intraday_supported_periods or list(COMPARE_PERIODS_1D),
                                    earliest_intraday,
                                )
                                if intraday_notice and notice is None:
                                    notice = intraday_notice
                                elif intraday_notice:
                                    notice = f"{notice} {intraday_notice}"
                            intraday_datasets: list[pd.DataFrame] = []
                            live_session_date = pd.Timestamp.now(
                                tz=market_timezone_for_ticker(validated_tickers[0])
                            ).date()
                            should_append_relative_live = any(
                                is_market_regular_session_active_for_ticker(ticker)
                                for ticker in validated_tickers
                            )
                            if period == "1d":
                                all_selected_tickers_are_us = all(
                                    infer_ticker_market(ticker) == "US"
                                    for ticker in validated_tickers
                                )
                                refresh_stale_local = len({
                                    infer_ticker_market(ticker)
                                    for ticker in validated_tickers
                                }) > 1
                                loaded_intraday_datasets: list[pd.DataFrame | None] = []
                                first_intraday_error: Exception | None = None
                                for ticker in validated_tickers:
                                    try:
                                        loaded_intraday_datasets.append(
                                            load_compare_one_day_intraday_dataset(
                                                ticker,
                                                include_extended_hours_flag=include_extended_hours,
                                                include_overnight_flag=include_overnight,
                                                refresh_stale_local=refresh_stale_local,
                                            )
                                        )
                                    except Exception as exc:  # noqa: BLE001
                                        if not all_selected_tickers_are_us:
                                            raise
                                        first_intraday_error = first_intraday_error or exc
                                        loaded_intraday_datasets.append(None)
                                        LOGGER.info(
                                            "Keeping %s pending on the one-day axis until its first regular-session quote: %s",
                                            ticker,
                                            exc,
                                        )
                                available_intraday_datasets = [
                                    dataset
                                    for dataset in loaded_intraday_datasets
                                    if dataset is not None and not dataset.empty
                                ]
                                if not available_intraday_datasets:
                                    if first_intraday_error is not None:
                                        raise first_intraday_error
                                    raise ValueError("The selected tickers do not have one-day intraday data yet.")
                                latest_available_timestamp = max(
                                    pd.Timestamp(dataset["Date"].max())
                                    for dataset in available_intraday_datasets
                                )
                                latest_available_day = latest_available_timestamp.date()
                                has_pending_ticker = any(dataset is None for dataset in loaded_intraday_datasets)
                                if (
                                        has_pending_ticker
                                        and not should_append_relative_live
                                        and latest_available_day != live_session_date
                                ):
                                    if first_intraday_error is not None:
                                        raise first_intraday_error
                                    raise ValueError("A selected ticker does not have one-day intraday data yet.")
                                reference_dataset = max(
                                    available_intraday_datasets,
                                    key=lambda dataset: pd.Timestamp(dataset["Date"].max()),
                                )
                                reference_latest_day = pd.Timestamp(reference_dataset["Date"].max()).date()
                                reference_latest_session = reference_dataset.loc[
                                    pd.to_datetime(reference_dataset["Date"], errors="coerce").dt.date
                                    == reference_latest_day
                                ].copy()
                                intraday_datasets = [
                                    dataset
                                    if dataset is not None
                                    else build_empty_compare_axis_dataset(reference_latest_session)
                                    for dataset in loaded_intraday_datasets
                                ]
                                common_end_date = min(dataset["Date"].max() for dataset in intraday_datasets)
                                try:
                                    aligned_datasets = slice_intraday_datasets_for_compare_period(
                                        intraday_datasets,
                                        period,
                                        common_end_date,
                                        validated_tickers,
                                    )
                                except ValueError as alignment_error:
                                    fallback_trading_days: list[object] = []
                                    preferred_fallback_date = pd.to_datetime(
                                        date_constraints.max_date,
                                        errors="coerce",
                                    )
                                    if not pd.isna(preferred_fallback_date):
                                        fallback_trading_days.append(preferred_fallback_date.date())

                                    latest_local_market_dates: list[object] = []
                                    for ticker, dataset in zip(validated_tickers, intraday_datasets):
                                        available_dates = {
                                            market_trading_date_for_timestamp(value, ticker)
                                            for value in dataset["Date"]
                                        }
                                        if available_dates:
                                            latest_local_market_dates.append(max(available_dates))
                                    if latest_local_market_dates:
                                        stale_market_fallback_date = min(latest_local_market_dates)
                                        if stale_market_fallback_date not in fallback_trading_days:
                                            fallback_trading_days.append(stale_market_fallback_date)

                                    recovered_alignment = False
                                    for fallback_trading_day in fallback_trading_days:
                                        restored_tickers: list[str] = []
                                        restored_intraday_datasets = intraday_datasets.copy()
                                        restore_failed = False
                                        for index, (ticker, dataset) in enumerate(
                                                zip(validated_tickers, intraday_datasets)
                                        ):
                                            available_trading_days = {
                                                market_trading_date_for_timestamp(value, ticker)
                                                for value in dataset["Date"]
                                            }
                                            if fallback_trading_day in available_trading_days:
                                                continue
                                            try:
                                                restored_intraday_datasets[index] = load_compare_one_day_intraday_dataset(
                                                    ticker,
                                                    include_extended_hours_flag=include_extended_hours,
                                                    include_overnight_flag=include_overnight,
                                                    trading_date=fallback_trading_day,
                                                )
                                                restored_tickers.append(ticker)
                                            except (ImportError, OSError, ValueError, KeyError, TypeError) as exc:
                                                LOGGER.info(
                                                    "Unable to restore the shared one-day comparison date for %s: %s",
                                                    ticker,
                                                    exc,
                                                )
                                                restore_failed = True
                                                break
                                        if restore_failed or not restored_tickers:
                                            continue

                                        try:
                                            common_end_date = min(
                                                dataset["Date"].max()
                                                for dataset in restored_intraday_datasets
                                            )
                                            aligned_datasets = slice_intraday_datasets_for_compare_period(
                                                restored_intraday_datasets,
                                                period,
                                                common_end_date,
                                                validated_tickers,
                                            )
                                        except ValueError:
                                            continue

                                        restored_preview = ", ".join(restored_tickers)
                                        restore_notice = (
                                            f"Loaded on-demand 1-minute data for {restored_preview} to restore "
                                            "the shared one-day market date."
                                        )
                                        if notice is None:
                                            notice = restore_notice
                                        else:
                                            notice += f" {restore_notice}"
                                        recovered_alignment = True
                                        break
                                    if not recovered_alignment:
                                        raise alignment_error
                                reference_timestamp = pd.Timestamp(aligned_datasets[0]["Date"].max())
                                if reference_timestamp.tzinfo is None:
                                    reference_timestamp = reference_timestamp.tz_localize("America/New_York")
                                else:
                                    reference_timestamp = reference_timestamp.tz_convert("America/New_York")
                                chart_trading_date_value = reference_timestamp.tz_convert(
                                    market_timezone_for_ticker(validated_tickers[0])
                                ).strftime("%Y-%m-%d")
                                exact_start_value = chart_trading_date_value
                                exact_end_value = chart_trading_date_value
                            for ticker in ([] if period == "1d" else validated_tickers):
                                intraday_dataset = fetch_history(
                                    ticker,
                                    include_dividends=False,
                                    interval="1m",
                                    dividend_mode="price",
                                )
                                if (
                                        period in {"3d", "1w"}
                                        and is_market_regular_session_active_for_ticker(ticker)
                                ):
                                    intraday_dataset = append_live_compare_intraday_dataset(
                                        ticker,
                                        intraday_dataset,
                                        live_trading_date=pd.Timestamp.now(
                                            tz=market_timezone_for_ticker(ticker)
                                        ).date(),
                                        include_extended_hours_flag=include_extended_hours,
                                        force_refresh=True,
                                    )[0]
                                intraday_datasets.append(intraday_dataset)
                            if period != "1d" and intraday_datasets:
                                common_end_date = min(dataset["Date"].max() for dataset in intraday_datasets)
                                aligned_datasets = slice_intraday_datasets_for_compare_period(
                                    intraday_datasets,
                                    period,
                                    common_end_date,
                                    validated_tickers,
                                )
                                exact_start_value = aligned_datasets[0]["Date"].min().strftime("%Y-%m-%d")
                                exact_end_value = aligned_datasets[0]["Date"].max().strftime("%Y-%m-%d")
                            period_label = format_period_label(period)
                        else:
                            period, notice_resolve = resolve_effective_period_for_many(period, datasets)
                            if notice_resolve and notice is None:
                                notice = notice_resolve
                            elif notice_resolve:
                                notice = (notice or "") + " " + notice_resolve
                            common_end_date = min(dataset["Date"].max() for dataset in datasets)
                            aligned_datasets = slice_datasets_for_compare_period(
                                datasets,
                                period,
                                common_end_date,
                            )
                            exact_start_value = aligned_datasets[0]["Date"].min().strftime("%Y-%m-%d")
                            exact_end_value = aligned_datasets[0]["Date"].max().strftime("%Y-%m-%d")
                            period_label = format_period_label(period)
                        supported_periods = supported_periods or list(COMPARE_PERIODS_1D)

                        colors = build_series_colors(len(validated_tickers), theme["accent_primary"], theme["accent_secondary"])
                        if current_view == "portfolio":
                            portfolio_shares = requested_shares[:len(validated_tickers)]
                            if len(portfolio_shares) < len(validated_tickers):
                                portfolio_shares.extend([0] * (len(validated_tickers) - len(portfolio_shares)))
                            growth_multipliers = build_portfolio_growth_multipliers(aligned_datasets)
                            if portfolio_allocation_mode == "shares":
                                if any(share_count <= 0 for share_count in portfolio_shares):
                                    raise ValueError("Each selected ticker must have at least 1 share.")
                                portfolio_weights = normalize_portfolio_share_weights(aligned_datasets, portfolio_shares)
                                portfolio_series = build_portfolio_series_payload_for_shares(
                                    aligned_datasets,
                                    portfolio_shares,
                                    theme["accent_primary"],
                                )
                            else:
                                ensure_positive_portfolio_weights(requested_weights, len(validated_tickers))
                                portfolio_weights = normalize_portfolio_weights(requested_weights, len(validated_tickers))
                                portfolio_series = build_portfolio_series_payload(
                                    aligned_datasets,
                                    portfolio_weights,
                                    theme["accent_primary"],
                                )
                            benchmark_series, benchmark_profiles = build_benchmark_series_payloads(
                                aligned_datasets[0]["Date"],
                                include_dividends,
                                price_only,
                            )
                            series = [portfolio_series, *benchmark_series]
                            profiles = [*profiles, *benchmark_profiles]
                            portfolio_total_return = portfolio_series.normalized_returns[-1]
                            portfolio_items = [
                                {
                                    "ticker": ticker,
                                    "company_name": profile.company_name,
                                    "logo_url": profile.logo_url,
                                    "weight": weight,
                                    "shares": share_count,
                                    "growth_multiple": growth_multiple,
                                    "color": color,
                                }
                                for ticker, profile, weight, share_count, growth_multiple, color in zip(
                                    validated_tickers,
                                    profiles[: len(validated_tickers)],
                                    portfolio_weights,
                                    portfolio_shares,
                                    growth_multipliers,
                                    colors,
                                )
                            ]
                        else:
                            if current_view == "market-caps":
                                series = [
                                    build_market_cap_series_payload(
                                        ticker,
                                        dataset,
                                        color=color,
                                        split_events=market_cap_split_events.get(ticker),
                                        resolve_missing_split_events=True,
                                        split_events_are_authoritative=(
                                            market_cap_split_actions_authoritative.get(ticker, False)
                                        ),
                                    )
                                    for ticker, dataset, color in zip(validated_tickers, aligned_datasets, colors)
                                ]
                            else:
                                series = [
                                    build_compare_series_payload(ticker, dataset, color=color)
                                    for ticker, dataset, color in zip(validated_tickers, aligned_datasets, colors)
                                ]
                        def last_valid_return(item: SeriesPayload) -> float | None:
                            valid_returns = [value for value in item.normalized_returns if value is not None]
                            return valid_returns[-1] if valid_returns else None

                        valid_performance_returns = [
                            value for value in (last_valid_return(item) for item in series) if value is not None
                        ]
                        best_return = max(valid_performance_returns) if valid_performance_returns else None
                        common_start = aligned_datasets[0]["Date"].min()
                        common_end = aligned_datasets[0]["Date"].max()
                        if current_view in {"tickers", "market-caps", "prices"} and period == "1d" and (range_mode == "exact" or is_intraday_compare_period):
                            display_range = format_display_date(pd.to_datetime(exact_start_value or common_start))
                        elif current_view in {"tickers", "market-caps", "prices"} and (
                            (is_intraday_compare_period and period in {"3d", "1w"})
                            or is_exact_short_intraday_compare
                        ):
                            display_range = format_compare_intraday_market_local_display_range(
                                aligned_datasets,
                                validated_tickers,
                            ) or f"{format_display_date(common_start)} - {format_display_date(common_end)}"
                        else:
                            display_range = f"{format_display_date(common_start)} - {format_display_date(common_end)}"
                        if current_view == "tickers":
                            dividend_yield_map = build_ttm_dividend_yield_map(validated_tickers, common_end)
                            best_dividend_yield = best_numeric_metric([
                                dividend_yield_map.get(ticker)
                                for ticker in validated_tickers
                            ])
                            performance_items = [
                                {
                                    "ticker": item.ticker,
                                    "company_name": profile.company_name,
                                    "logo_url": profile.logo_url,
                                    "ending_return": last_valid_return(item),
                                    "ttm_dividend_yield": dividend_yield_map.get(item.ticker),
                                    "color": item.color,
                                    "shadow_color": hex_to_rgba(item.color or theme["accent_primary"], 0.22),
                                    "is_winner": best_return is not None and last_valid_return(item) == best_return,
                                    "is_dividend_yield_winner": (
                                        best_dividend_yield is not None
                                        and dividend_yield_map.get(item.ticker) == best_dividend_yield
                                    ),
                                }
                                for item, profile in zip(series, profiles)
                            ]
                        ticker_slots = (control_tickers or validated_tickers).copy()
                        record_ticker_usage(validated_tickers)
        except Exception:  # noqa: BLE001
            LOGGER.exception("Unable to render %s workspace", current_view)
            error = "Unable to load this workspace. Check your local data and try again."
            if should_use_modal_banner_message(error):
                floating_banner_icon_class = modal_banner_icon_class(error)

        remote_market_access = True

        if current_view != "settings" and not error and not notice:
            requires_remote_probe = any(
                not history_store_path_for(ticker).exists()
                for ticker in validated_tickers
            )
            if requires_remote_probe:
                remote_market_access = has_remote_market_access()
                if not remote_market_access:
                    notice = "Using bundled local market_store data because remote market access is unavailable."

        top_tickers = []
        timing_selected_ticker = ""
        timing_metrics = []
        timing_summary = []
        timing_market = {}
        timing_error = ""
        live_trading_account_label = "Integrated A/C (Unavailable)"

        if current_view == "settings":
            if settings_section in {"general", "backtest", "email-smtp", "broker-access", "local-market-store", "clear-caches"} and (notice or error):
                floating_banner_icon_class = modal_banner_icon_class(error or notice)
            settings_service_rows = build_network_service_rows(
                pending=settings_section != "network",
                service_labels=labels,
                translate_fn=translate_ui,
            )
            strategy_settings_rows = translate_nested_text(
                build_strategy_settings_rows(strategy_options),
                language_settings.language,
                language_translations,
            )
            font_token_rows = translate_nested_text(
                build_font_token_rows(labels),
                language_settings.language,
                language_translations,
            )
            color_token_rows = translate_nested_text(
                build_color_token_rows(theme_light, theme_dark),
                language_settings.language,
                language_translations,
            )
            style_token_rows = translate_nested_text(
                build_style_token_rows(labels),
                language_settings.language,
                language_translations,
            )
            export_image_rows = translate_nested_text(
                build_export_image_rows(PROJECT_DISPLAY_URL),
                language_settings.language,
                language_translations,
            )
            material_token_rows = translate_nested_text(
                build_material_token_rows(),
                language_settings.language,
                language_translations,
            )
            cash_equivalent_tickers = load_cash_equivalent_tickers()
            cash_equivalent_rows = []
            for t in cash_equivalent_tickers:
                company = resolve_known_ticker_company_name(t) or t
                logo = resolve_stored_logo_url(t) or ""
                cash_equivalent_rows.append({
                    "ticker": t,
                    "company_name": company,
                    "logo_url": logo,
                })
            cash_equivalent_fund_rows = []
            for raw_ticker in money_market_settings.get("tickers", []):
                ticker = canonicalize_money_market_ticker(raw_ticker)
                if not ticker:
                    continue
                quote_currency = configured_money_market_quote_currencies.get(ticker, "")
                cash_equivalent_fund_rows.append({
                    "ticker": ticker,
                    "company_name": resolve_known_ticker_company_name(ticker) or ticker,
                    "quote_currency": quote_currency,
                    "token_logo_class": "investment-cash-equivalent-token-logo",
                })
            if settings_section == "local-market-store":
                all_local_market_tickers = list_local_market_tickers()
                local_store_current_page = local_store_page_value()
                local_store_total_pages = max((len(all_local_market_tickers) - 1) // LOCAL_STORE_PAGE_SIZE + 1, 1)
                local_store_current_page = min(local_store_current_page, local_store_total_pages)
                settings_page_number = local_store_current_page
                local_store_pagination_items = build_local_store_pagination_items(
                    local_store_current_page,
                    local_store_total_pages,
                )
                start_index = (local_store_current_page - 1) * LOCAL_STORE_PAGE_SIZE
                end_index = start_index + LOCAL_STORE_PAGE_SIZE
                local_market_rows = build_local_market_rows_for_tickers(
                    all_local_market_tickers[start_index:end_index],
                    include_ranges=True,
                )
        elif current_view == "trade":
            if trade_section == "live-trading":
                live_trading_account_label = load_longbridge_account_label(load_broker_settings())

        if current_view in {*BACKTEST_VIEWS, "dca"}:
            ticker_slots = ticker_slots[:1] if ticker_slots else [""]
        else:
            while len(ticker_slots) < MIN_TICKERS:
                ticker_slots.append("")
        if current_view == "portfolio":
            if not portfolio_weights and any(ticker_slots):
                portfolio_weights = build_default_weights(len([ticker for ticker in ticker_slots if ticker]))
            while len(portfolio_weights) < len(ticker_slots):
                portfolio_weights.append(0)
            if not portfolio_shares:
                portfolio_shares = requested_shares[:len(ticker_slots)] if requested_shares else []
            while len(portfolio_shares) < len(ticker_slots):
                portfolio_shares.append(0)

        template_name = {
            "tickers": "compare.html",
            "market-caps": "market_cap_compare.html",
            "prices": "price_compare.html",
            "portfolio": "portfolio.html",
            "dca": "dca.html",
            "backtest": "backtest.html",
            "grid-trading": "grid_trading.html",
            "trade": (
                "investment.html"
                if trade_section == "investment"
                else "live_trading.html"
                if trade_section == "live-trading"
                else "investment.html"
            ),
            "settings": "settings.html",
        }[current_view]

        response = make_response(render_template(
            template_name,
            error=error,
            notice=notice,
            floating_banner_icon_class=floating_banner_icon_class,
            period=period,
            period_label=period_label,
            display_range=display_range,
            periods=supported_periods,
            period_labels={item: format_period_label(item) for item in supported_periods},
            period_metadata={
                "labels": PERIOD_LABELS,
                "daySpans": PERIOD_DAY_SPANS,
                "monthSpans": PERIOD_MONTH_SPANS,
            },
            series=series,
            profiles_json=[quote_profile_to_json(profile) for profile in profiles],
            performance_items=performance_items,
            portfolio_items=portfolio_items,
            portfolio_weights=portfolio_weights,
            portfolio_shares=portfolio_shares,
            portfolio_allocation_mode=portfolio_allocation_mode,
            portfolio_total_return=portfolio_total_return,
            dca_result=dca_result,
            dca_amount=dca_amount,
            dca_frequency=dca_frequency,
            dca_weekday=dca_weekday,
            dca_month_day=dca_month_day,
            ticker_slots=ticker_slots,
            max_tickers=view_max_tickers,
            min_tickers=MIN_TICKERS,
            base_currency=BASE_CURRENCY,
            base_timezone=BASE_TIMEZONE,
            include_dividends=include_dividends,
            price_only=price_only,
            include_extended_hours=include_extended_hours,
            show_extended_hours_toggle=show_extended_hours_toggle,
            include_overnight=include_overnight,
            show_overnight_toggle=show_overnight_toggle,
            range_mode=range_mode,
            exact_start=exact_start_value,
            exact_end=exact_end_value,
            format_display_date=format_display_date,
            chart_trading_date=chart_trading_date_value or exact_start_value,
            version=app_meta.get("version", CODE_VERSION),
            updated_on=app_meta.get("updated_on", ""),
            current_view=current_view,
            settings_section=settings_section,
            trade_section=trade_section,
            top_tickers=top_tickers,
            timing_selected_ticker=timing_selected_ticker,
            timing_metrics=timing_metrics,
            timing_summary=timing_summary,
            timing_market=timing_market,
            timing_error=timing_error,
            remote_market_access=remote_market_access,
            settings_title=settings_title,
            settings_service_rows=settings_service_rows,
            strategy_settings_rows=strategy_settings_rows,
            font_token_rows=font_token_rows,
            color_token_rows=color_token_rows,
            style_token_rows=style_token_rows,
            export_image_rows=export_image_rows,
            material_token_rows=material_token_rows,
            cash_equivalent_rows=cash_equivalent_rows,
            cash_equivalent_fund_rows=cash_equivalent_fund_rows,
            backtest_execution_mode=backtest_execution_mode,
            investment_cost_basis_method=investment_cost_basis_method,
            date_display_full_format=date_display_settings.full_date_format,
            date_display_short_format=date_display_settings.short_date_format,
            language_code=language_settings.language,
            language_labels=LANGUAGE_LABELS,
            language_options=SUPPORTED_LANGUAGE_CODES,
            language_translations=list(language_settings.translations),
            language_history_rows=language_history_rows,
            language_html_lang=HTML_LANG_BY_LANGUAGE[language_settings.language],
            translate_ui=translate_ui,
            broker_settings=broker_settings,
            broker_test_status=broker_test_status,
            broker_test_message=broker_test_message,
            broker_test_checked_at=broker_test_checked_at,
            longbridge_oauth_pending=longbridge_oauth_pending,
            live_trading_account_label=live_trading_account_label,
            local_market_rows=local_market_rows,
            local_store_current_page=local_store_current_page,
            local_store_page_size=LOCAL_STORE_PAGE_SIZE,
            local_store_total_pages=local_store_total_pages,
            local_store_pagination_items=local_store_pagination_items,
            settings_tab=settings_tab,
            settings_page_number=settings_page_number,
            settings_language_page_size=SETTINGS_LANGUAGE_PAGE_SIZE,
            page_title=page_title,
            sidebar_title=labels["trade_title"] if current_view == "trade" else page_title,
            report_heading=report_heading,
            chart_heading=chart_heading,
            dock_urls={view_name: build_view_url(view_name) for view_name in ("tickers", "market-caps", "prices", "portfolio", "dca", "backtest", "grid-trading", "trade", "settings")},
            settings_urls={section_name: build_settings_url(section_name) for section_name in
                           ("about", "general", "investment", "backtest", "font-tokens", "color-tokens", "material-tokens", "network", "strategies", "email-smtp", "broker-access", "local-market-store", "clear-caches",
                            "style-tokens", "export-image", "cash-equivalents")},
            trade_urls={section_name: build_trade_url(section_name) for section_name in ("investment", "live-trading")},
            local_store_page_urls={page_number: build_local_store_page_url(page_number) for page_number in range(1, local_store_total_pages + 1)},
            labels=labels,
            theme=theme,
            theme_light=theme_light,
            theme_dark=theme_dark,
            project_source_url=PROJECT_SOURCE_URL,
            project_display_url=PROJECT_DISPLAY_URL,
            chart_config=chart_config,
            logos=logos,
            defaults=defaults,
            smtp_settings=smtp_settings,
            strategy_options=strategy_options,
            strategy_option_groups=strategy_option_groups,
            selected_strategy_id=selected_strategy_id,
            strategy_form_fields=strategy_form_fields,
            selected_strategy_params=selected_strategy_params,
            backtest_initial_capital=backtest_initial_capital,
            backtest_result=backtest_result,
            backtest_periods_by_interval=backtest_periods_by_interval,
            supported_intervals=supported_intervals,
            requested_interval=requested_interval,
            current_view_name=current_view,
            current_path=request.path,
            fetch_abort_debug_config=FETCH_ABORT_DEBUG_CONFIG,
            endpoints={
                "symbolSearch": "/api/symbol-search",
                "dateConstraints": "/api/date-constraints",
                "compareLive": "/api/compare/live",
                "strategyFields": "/api/trade-strategy-fields",
                "settingsNetworkStatus": "/api/settings/network-status",
                "localStorePageData": "/api/settings/local-market-store/page-data",
                "marketStorePresence": "/api/market-store/presence",
                "investmentIntraday": "/api/investment/intraday",
                "investmentMarketSession": "/api/market-session/us-equity",
                "investmentRealtimeQuotes": "/api/investment/realtime-quotes",
                "liveTradingPositions": "/api/live-trading/positions",
                "liveTradingOrder": "/api/live-trading/orders",
            },
        ))
        if current_view == "settings":
            response.delete_cookie(SETTINGS_FEEDBACK_COOKIE, path="/settings")
        if current_view == "trade" and trade_section == "investment":
            apply_no_store_headers(response)
        return response

    def export_transactions_api():
        try:
            # Re-run backtest to get the full transaction list
            (
                backtest_result,
                trade_ticker,
                requested_interval,
                _date_constraints,
                trade_dataset,
                strategy_id,
                strategy_params,
                _,
            ) = _run_backtest_from_request()

            raw_summary = backtest_result.get("summary", {})
            summary: dict[str, object] = (
                cast(dict[str, object], raw_summary)
                if isinstance(raw_summary, dict)
                else {}
            )
            raw_trades = backtest_result.get("trades", [])
            trades: list[dict[str, object]] = (
                [cast(dict[str, object], trade) for trade in raw_trades if isinstance(trade, dict)]
                if isinstance(raw_trades, list)
                else []
            )
            if not trades:
                return "No transactions to export.", 404

            strategy_definition = get_strategy_definition(strategy_id)
            # 0. Context for Filename
            start_str = trade_dataset["Date"].min().strftime("%Y%m%d")
            end_str = trade_dataset["Date"].max().strftime("%Y%m%d")
            strategy_name = strategy_definition.get('name', strategy_id)
            report_filename = f"{trade_ticker} Backtest Report {start_str} - {end_str} ({strategy_name}).md"
            period_start = pd.to_datetime(trade_dataset["Date"].min())
            period_end = pd.to_datetime(trade_dataset["Date"].max())
            period_label = f"{format_display_date(period_start)} - {format_display_date(period_end)}"
            dataset_export_date_format = "%Y-%m-%d %H:%M" if requested_interval == "1m" else "%Y-%m-%d"
            market_data_csv = trade_dataset.to_csv(index=False, date_format=dataset_export_date_format).rstrip()

            # 1. Performance Summary
            benchmark_alpha = float(summary.get("benchmark_alpha", 0) or 0)
            long_gain = float(summary.get("long_gain", 0) or 0)
            short_gain = float(summary.get("short_gain", 0) or 0)
            long_loss = float(summary.get("long_loss", 0) or 0)
            beat_bh_pct = float(summary.get("beat_bh_pct", 0) or 0)
            win_rate_pct = summary.get("win_rate_pct")
            win_rate_display = "N/A" if win_rate_pct is None else f"{parse_float_value(win_rate_pct, 0.0):,.2f}%"

            md_lines = [
                f"## Backtest Report: {trade_ticker}",
                f"**Generated on**: {format_display_datetime(pd.Timestamp.now(), include_seconds=True, timezone_suffix='HKT')}",
                f"**Algorithm**: {strategy_name}",
                f"**Period**: {period_label}",
                "",
                "### Performance Summary",
                f"- **Initial capital**: ${summary.get('initial_capital', 0):,.2f}",
                f"- **Final equity**: ${summary.get('final_equity', 0):,.2f}",
                f"- **Net return**: {summary.get('net_return_pct', 0):,.2f}%",
                f"- **Total trades**: {summary.get('total_trades', 0)}",
                f"- **Win rate**: {win_rate_display}",
                f"- **Beat B&H**: {beat_bh_pct:,.2f}%",
                f"- **Alpha vs B&H**: {'+' if benchmark_alpha >= 0 else '-'}${abs(benchmark_alpha):,.2f}",
                f"- **Realized long P&L**: {'+' if long_gain >= 0 else '-'}${abs(long_gain):,.2f}",
                f"- **Realized short P&L**: {'+' if short_gain >= 0 else '-'}${abs(short_gain):,.2f}",
                f"- **Realized long loss**: {'-' if long_loss > 0 else '+'}${abs(long_loss):,.2f}",
                "",
            ]

            # 2. Transaction Details
            md_lines.extend([
                "### Transaction History",
                "",
                "| No. | Date | Side | Price | Shares | P&L | Cash | Equity |",
                "| ---: | :--- | :--- | ---: | ---: | ---: | ---: | ---: |"
            ])
            for i, trade in enumerate(trades):
                if trade.get("_virtual_close"):
                    continue  # Skip virtual closing trade, same as table display
                trade_date = format_display_datetime(pd.to_datetime(trade.get("date")), use_short_date=True) if trade.get("date") else "N/A"
                md_lines.append(
                    f"| {i + 1} | {trade_date} | {trade.get('side')} | "
                    f"{trade.get('price', 0):,.2f} | {trade.get('shares', 0):,.0f} | "
                    f"{trade.get('pnl', 0):,.2f} | {trade.get('cash', 0):,.2f} | {trade.get('equity', 0):,.2f} |"
                )

            # 3. Strategy Context
            md_lines.extend([
                "",
                "### Strategy Context",
                "",
                "#### Parameters",
                "",
                "| Parameter | Value |",
                "| :--- | :--- |"
            ])
            for key, val in strategy_params.items():
                md_lines.append(f"| {key} | {val} |")

            # Read Strategy Code
            try:
                module_name = strategy_definition.get("module", "")
                if module_name:
                    file_name = module_name.split(".")[-1] + ".py"
                    algo_path = Path(__file__).resolve().parent.parent / "strategies" / "algorithms" / file_name
                    if algo_path.exists():
                        with open(algo_path, "r", encoding="utf-8") as f:
                            strategy_code = f.read()
                        md_lines.extend([
                            "",
                            "#### Strategy Implementation",
                            "```python",
                            strategy_code,
                            "```",
                            ""
                        ])
            except Exception:  # noqa: BLE001
                LOGGER.exception("Unable to load strategy source for exported backtest report")
                md_lines.append("\n*(Strategy source was unavailable for this export.)*")

            # 4. LLM Strategy Developer Prompt
            md_lines.extend([
                "",
                "### LLM Strategy Developer Prompt",
                "",
                "*Copy and paste the prompt below into any SOTA LLMs to recreate or iterate on this strategy.*",
                "",
                "````",
                "You are an elite quantitative trading developer and Python engineer. Your task is to write a trading strategy plugin for the `antigravity` trading system.",
                "",
                "The user will provide a trading logic or indicator concept. You must output a fully functional, production-ready Python file named `strategy_{strategy_id}.py` that acts as a drop-in component for the `strategies/algorithms/` directory.",
                "",
                "### Core Architecture & Constraints",
                "",
                "1. **Imports & Inheritance**:",
                "   - Must include `from __future__ import annotations` at the very top.",
                "   - Must import `import pandas as pd` and `import numpy as np`.",
                "   - Must import: `from ..base import BaseStrategy, StrategyParameterDefinition, StrategySignalResult, StrategySupportMatrix`.",
                "   - The strategy class must inherit from `BaseStrategy`.",
                "",
                "2. **Class Metadata (Class Attributes)**:",
                "   Every strategy must define the following class-level attributes exactly:",
                "   - `strategy_id` (str): Unique snake_case identifier (e.g., \"macd\", \"rsi_reversion\").",
                "   - `strategy_name` (str): Human-readable name for the UI.",
                "   - `strategy_description` (str): Clear, concise description of the logic.",
                "   - `strategy_category` (str): Must be one of: \"trend\", \"momentum\", \"mean_reversion\", \"volatility\", \"volume\", or \"machine_learning\".",
                "   - `strategy_display_order` (int): An integer (10-90) indicating UI sorting priority.",
                "   - `strategy_supports` (StrategySupportMatrix): Usually `StrategySupportMatrix(single_ticker=True, multi_ticker=False, long_only=True, short=False)`.",
                "",
                "3. **Parameter Definitions (`get_parameter_definitions`)**:",
                "   Override `def get_parameter_definitions(self) -> tuple[StrategyParameterDefinition, ...]:` to declare all user-configurable parameters.",
                "   Supported kwargs for `StrategyParameterDefinition`: `key`, `label`, `kind` (\"integer\", \"number\", \"boolean\", \"choice\", \"string\"), `default`, `minimum`, `maximum`, `step`, `options`, `help_text`, `unit_hint`.",
                "",
                "4. **Signal Computation (`compute_signals`)**:",
                "   Override `def compute_signals(self, dataset: pd.DataFrame, params: dict | None = None) -> StrategySignalResult:`.",
                "   - **CRITICAL**: Never modify `dataset` in-place. Always start with `frame = dataset.copy()`.",
                "   - **CRITICAL**: Always normalize parameters first via `normalized_params = self.normalize_params(params)`.",
                "   - Extract your parameters with explicit type casting.",
                "   - Compute your indicators using vectorized pandas operations. Avoid loops for performance unless mathematically required.",
                "   - Create exactly two boolean signal columns: `\"buy_signal\"` and `\"sell_signal\"`.",
                "   - **CRITICAL**: Use `.fillna(False)` on signal columns.",
                "   - Return: `return StrategySignalResult(frame=frame, buy_signal_column=\"buy_signal\", sell_signal_column=\"sell_signal\")`.",
                "",
                "5. **Output Format**:",
                "   - Provide ONLY the Python code block. No surrounding markdown explanations.",
                "   - The code must be pristine, strictly typed (Python 3.10+), and adhere to standard Black formatting.",
                "",
                "### Gold Standard Reference (MACD Strategy)",
                "```python",
                "from __future__ import annotations",
                "import pandas as pd",
                "from ..base import BaseStrategy, StrategyParameterDefinition, StrategySignalResult, StrategySupportMatrix",
                "",
                "class MacdStrategy(BaseStrategy):",
                "    strategy_id = \"macd\"",
                "    strategy_name = \"MACD\"",
                "    strategy_description = \"MACD crossover strategy using default settings.\"",
                "    strategy_category = \"momentum\"",
                "    strategy_display_order = 20",
                "    strategy_supports = StrategySupportMatrix(single_ticker=True, multi_ticker=False, long_only=True, short=False)",
                "",
                "    def get_parameter_definitions(self) -> tuple[StrategyParameterDefinition, ...]:",
                "        return (",
                "            StrategyParameterDefinition(key=\"fast_span\", label=\"Fast EMA\", kind=\"integer\", default=12, minimum=1),",
                "            StrategyParameterDefinition(key=\"slow_span\", label=\"Slow EMA\", kind=\"integer\", default=26, minimum=2),",
                "            StrategyParameterDefinition(key=\"signal_span\", label=\"Signal EMA\", kind=\"integer\", default=9, minimum=1),",
                "        )",
                "",
                "    def compute_signals(self, dataset: pd.DataFrame, params: dict | None = None) -> StrategySignalResult:",
                "        frame = dataset.copy()",
                "        normalized_params = self.normalize_params(params)",
                "        fast_span, slow_span, signal_span = int(normalized_params[\"fast_span\"]), int(normalized_params[\"slow_span\"]), int(normalized_params[\"signal_span\"])",
                "        ema_fast = frame[\"Close\"].ewm(span=fast_span, adjust=False).mean()",
                "        ema_slow = frame[\"Close\"].ewm(span=slow_span, adjust=False).mean()",
                "        frame[\"macd_line\"] = ema_fast - ema_slow",
                "        frame[\"signal_line\"] = frame[\"macd_line\"].ewm(span=signal_span, adjust=False).mean()",
                "        frame[\"buy_signal\"] = ((frame[\"macd_line\"] > frame[\"signal_line\"]) & (frame[\"macd_line\"].shift(1) <= frame[\"signal_line\"].shift(1))).fillna(False)",
                "        frame[\"sell_signal\"] = ((frame[\"macd_line\"] < frame[\"signal_line\"]) & (frame[\"macd_line\"].shift(1) >= frame[\"signal_line\"].shift(1))).fillna(False)",
                "        return StrategySignalResult(frame=frame, buy_signal_column=\"buy_signal\", sell_signal_column=\"sell_signal\")",
                "```",
                "````",
            ])

            # 5. Source market data
            md_lines.extend([
                "### Source market data",
                "",
                "```text",
                market_data_csv,
                "```",
                ""
            ])

            md_content = "\n".join(md_lines)

            return send_file(
                BytesIO(md_content.encode("utf-8")),
                mimetype='text/markdown',
                as_attachment=True,
                download_name=report_filename
            )
        except Exception:  # noqa: BLE001
            LOGGER.exception("Unable to export backtest transactions")
            return "Unable to export backtest transactions. Try again later.", 500

    def root():
        legacy_view = request.args.get("view")
        if request.args:
            target_view = resolve_view() if legacy_view else "tickers"
            if target_view == "settings":
                return redirect(
                    build_settings_state_url(
                        resolve_settings_section(),
                        tab=resolve_settings_tab(),
                        page=settings_page_value(),
                    )
                )
            target_path = build_view_path(target_view)
            query_string = urlencode(build_modern_query_pairs(target_view), doseq=True)
            return redirect(f"{target_path}?{query_string}" if query_string else target_path)
        return redirect(build_view_path("tickers"))

    def compare_page():
        return render_workspace_page("tickers")

    def market_cap_compare_page():
        return render_workspace_page("market-caps")

    def legacy_compare_page():
        return build_legacy_workspace_redirect("tickers")

    def price_compare_page():
        return render_workspace_page("prices")

    def portfolio_page():
        return render_workspace_page("portfolio")

    def legacy_portfolio_page():
        return build_legacy_workspace_redirect("portfolio")

    def dca_page():
        return render_workspace_page("dca")

    def legacy_dca_page():
        return build_legacy_workspace_redirect("dca")

    def backtest_page():
        return render_workspace_page("backtest")

    def grid_trading_page():
        return render_workspace_page("grid-trading")

    def legacy_backtest_page():
        return build_legacy_workspace_redirect("backtest")

    def legacy_trade_messages_page():
        return build_legacy_workspace_redirect("backtest")

    def trade_root():
        return redirect(build_trade_path("investment"))

    def trade_page(section_name: str):
        normalized_section = normalize_trade_section(section_name)
        if normalized_section != (section_name or "").strip().lower():
            return redirect(build_trade_path(normalized_section))
        if normalized_section == "live-trading" and not session.get("live_trading_unlocked"):
            response = make_response(render_template(
                "live_trading_unlock.html",
                error_message="",
                theme_dark=theme_dark,
                theme_light=theme_light,
                version=app_meta.get("version", CODE_VERSION),
            ))
            return apply_no_store_headers(response)
        return render_workspace_page("trade", trade_section=normalized_section)

    def live_trading_unlock():
        access_granted, error_status, error_message = validate_live_trading_pin(
            request.form.get("pin"),
            live_trading_pin,
        )
        if not access_granted:
            response = make_response(render_template(
                "live_trading_unlock.html",
                error_message=error_message,
                theme_dark=theme_dark,
                theme_light=theme_light,
                version=app_meta.get("version", CODE_VERSION),
            ), error_status)
            return apply_no_store_headers(response)

        session.clear()
        session["live_trading_unlocked"] = True
        return redirect(build_trade_path("live-trading"), code=303)

    def legacy_trade_root():
        return redirect(build_trade_path("investment"))

    def legacy_trade_page(section_name: str):
        return redirect(build_trade_path(normalize_trade_section(section_name)))

    def settings_root():
        return redirect(
            build_settings_state_url(
                resolve_settings_section(),
                tab=resolve_settings_tab(),
                page=settings_page_value(),
            )
        )

    def settings_page(section_name: str):
        normalized_section = normalize_settings_section(section_name)
        canonical_url = build_settings_state_url(
            normalized_section,
            tab=resolve_settings_tab(),
            page=settings_page_value(),
        )
        current_url = request.path
        if request.query_string:
            current_url = f"{current_url}?{request.query_string.decode()}"
        if current_url != canonical_url:
            return redirect(canonical_url)
        report_fetch_abort_debug_event(
            "E",
            "runtime.py:settings_page",
            "settings page request received",
            {
                "section_name": normalized_section,
                "path": request.path,
                "query_string": request.query_string.decode(),
            },
        )
        return render_workspace_page("settings", normalized_section)

    def _language_rows_from_request_form() -> list[dict[str, str]]:
        return [
            {
                "en": english,
                "zh_hant_hk": zh_hant_hk,
                "zh_hans_cn": zh_hans_cn,
            }
            for english, zh_hant_hk, zh_hans_cn in zip(
                request.form.getlist("translation_en"),
                request.form.getlist("translation_zh_hant_hk"),
                request.form.getlist("translation_zh_hans_cn"),
            )
        ]

    def _language_rows_from_xlsx_bytes(payload: bytes) -> list[dict[str, str]]:
        workbook = load_workbook(filename=BytesIO(payload), data_only=True)
        worksheet = workbook.active
        headers = [
            str(worksheet.cell(row=1, column=column_index).value or "").strip()
            for column_index in range(1, 5)
        ]
        header_to_column = {header: index + 1 for index, header in enumerate(headers)}
        english_column = header_to_column.get("English", 2)
        traditional_column = header_to_column.get("繁體中文（香港）", 3)
        simplified_column = header_to_column.get("简体中文（中国大陆）", 4)
        rows: list[dict[str, str]] = []
        for row_index in range(2, worksheet.max_row + 1):
            english = str(worksheet.cell(row=row_index, column=english_column).value or "").strip()
            if not english:
                continue
            rows.append(
                {
                    "en": english,
                    "zh_hant_hk": str(worksheet.cell(row=row_index, column=traditional_column).value or "").strip(),
                    "zh_hans_cn": str(worksheet.cell(row=row_index, column=simplified_column).value or "").strip(),
                }
            )
        if not rows:
            raise ValueError("The uploaded spreadsheet does not contain any language mapping rows.")
        return rows

    def general_settings_action():
        notices: list[str] = []
        wants_async_language_response = request.headers.get("X-Settings-Async") == "1"
        selected_language_settings = load_language_settings()
        language_action = str(request.form.get("language_action", "save")).strip()
        language_file = request.files.get("language_mapping_xlsx")
        settings_state_from_form = {
            "tab": request.form.get("settings_tab", ""),
            "page": request.form.get("settings_page", ""),
        }
        if language_action == "upload" and language_file and language_file.filename:
            try:
                selected_language_settings = save_language_settings(
                    language=request.form.get("language_code", load_language_settings().language),
                    translations=_language_rows_from_xlsx_bytes(language_file.read()),
                    history_label="Spreadsheet upload",
                )
                notices.append(f"Language translations imported from {language_file.filename}.")
                if (
                    selected_language_settings.language in {"zh_hant_hk", "zh_hans_cn"}
                    and load_date_display_settings().full_date_format == "d_mmm_yyyy"
                ):
                    save_full_date_display_format("yyyy_mm_dd_cjk")
                elif (
                    selected_language_settings.language == "en"
                    and load_date_display_settings().full_date_format == "yyyy_mm_dd_cjk"
                ):
                    save_full_date_display_format("d_mmm_yyyy")
            except ValueError as exc:
                return _redirect_with_settings_feedback(
                    "general",
                    error=(
                        "Language spreadsheet import failed: "
                        f"{str(exc).strip() or 'check the file and try again.'}"
                    ),
                    query_params=settings_state_from_form,
                )
            except Exception:  # noqa: BLE001
                LOGGER.exception("Language spreadsheet import failed")
                return _redirect_with_settings_feedback(
                    "general",
                    error="Language spreadsheet import failed. Check the file and try again.",
                    query_params=settings_state_from_form,
                )
        elif "language_code" in request.form or "translation_en" in request.form:
            current_language_settings = load_language_settings()
            translation_rows = _language_rows_from_request_form()
            selected_language_settings = save_language_settings(
                language=request.form.get("language_code", current_language_settings.language),
                translations=translation_rows if translation_rows else None,
                history_label="Manual edit",
            )
            if (
                selected_language_settings.language in {"zh_hant_hk", "zh_hans_cn"}
                and load_date_display_settings().full_date_format == "d_mmm_yyyy"
            ):
                save_full_date_display_format("yyyy_mm_dd_cjk")
            elif (
                selected_language_settings.language == "en"
                and load_date_display_settings().full_date_format == "yyyy_mm_dd_cjk"
            ):
                save_full_date_display_format("d_mmm_yyyy")
            if selected_language_settings.language != current_language_settings.language:
                notices.append(f"Language updated: {LANGUAGE_LABELS[selected_language_settings.language]}.")
            elif translation_rows:
                notices.append("Language translations updated.")
        if wants_async_language_response:
            response = jsonify({
                "success": True,
                "notice": " ".join(notices),
                "language": selected_language_settings.language,
                "label": LANGUAGE_LABELS[selected_language_settings.language],
            })
            return apply_no_store_headers(response)
        if "full_date_format" in request.form:
            current_full = load_date_display_settings().full_date_format
            selected_full = save_full_date_display_format(request.form.get("full_date_format", current_full))
            if selected_full != current_full:
                full_labels = {
                    "d_mmm_yyyy": "D Mmm yyyy",
                    "dd_mmm_yyyy": "DD Mmm yyyy",
                    "yyyy_mmm_d": "yyyy Mmm D",
                    "yyyy_mmm_dd": "yyyy Mmm DD",
                    "yyyy_mm_dd_cjk": "yyyy年mm月dd日",
                }
                notices.append(f"Full date format updated: {full_labels[selected_full]}.")
        if "short_date_format" in request.form:
            current_short = load_date_display_settings().short_date_format
            selected_short = save_short_date_display_format(request.form.get("short_date_format", current_short))
            if selected_short != current_short:
                short_labels = {
                    "yyyy_mm_dd": "yyyy/mm/dd",
                    "dd_mm_yyyy": "dd/mm/yyyy",
                }
                notices.append(f"Compact date format updated: {short_labels[selected_short]}.")
        notice = " ".join(notices)
        return _redirect_with_settings_feedback(
            "general",
            notice=notice,
            query_params=settings_state_from_form,
        )

    def language_download_api():
        settings = load_language_settings()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "i18n mapping"
        worksheet.append(["No.", "English", "繁體中文（香港）", "简体中文（中国大陆）"])
        for index, row in enumerate(settings.translations, start=1):
            worksheet.append([index, row["en"], row["zh_hant_hk"], row["zh_hans_cn"]])
        worksheet.freeze_panes = "A2"
        widths = {"A": 8, "B": 52, "C": 52, "D": 52}
        for column_letter, width in widths.items():
            worksheet.column_dimensions[column_letter].width = width
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return send_file(
            buffer,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="antigravity-i18n-mapping.xlsx",
        )

    def language_settings_api():
        payload = request.get_json(silent=True) or {}
        language = str(payload.get("language", "")).strip()
        current_language = load_language_settings().language
        selected_language = save_language_code(language or current_language)
        current_full_date_format = load_date_display_settings().full_date_format
        if selected_language in {"zh_hant_hk", "zh_hans_cn"} and current_full_date_format == "d_mmm_yyyy":
            save_full_date_display_format("yyyy_mm_dd_cjk")
        elif selected_language == "en" and current_full_date_format == "yyyy_mm_dd_cjk":
            save_full_date_display_format("d_mmm_yyyy")
        response = jsonify({
            "success": True,
            "language": selected_language,
            "htmlLang": HTML_LANG_BY_LANGUAGE[selected_language],
            "label": LANGUAGE_LABELS[selected_language],
            "dateDisplay": {
                "full": load_date_display_settings().full_date_format,
                "short": load_date_display_settings().short_date_format,
            },
        })
        return apply_no_store_headers(response)

    def language_cycle_api():
        current_language = load_language_settings().language
        current_index = SUPPORTED_LANGUAGE_CODES.index(current_language)
        selected_language = SUPPORTED_LANGUAGE_CODES[(current_index + 1) % len(SUPPORTED_LANGUAGE_CODES)]
        save_language_code(selected_language)
        current_full_date_format = load_date_display_settings().full_date_format
        if selected_language in {"zh_hant_hk", "zh_hans_cn"} and current_full_date_format == "d_mmm_yyyy":
            save_full_date_display_format("yyyy_mm_dd_cjk")
        elif selected_language == "en" and current_full_date_format == "yyyy_mm_dd_cjk":
            save_full_date_display_format("d_mmm_yyyy")
        response = jsonify({
            "success": True,
            "language": selected_language,
            "htmlLang": HTML_LANG_BY_LANGUAGE[selected_language],
            "label": LANGUAGE_LABELS[selected_language],
            "dateDisplay": {
                "full": load_date_display_settings().full_date_format,
                "short": load_date_display_settings().short_date_format,
            },
        })
        return apply_no_store_headers(response)

    def backtest_settings_action():
        notice = ""
        if "backtest_execution_mode" in request.form:
            current_mode = load_backtest_execution_mode()
            selected_mode = save_backtest_execution_mode(request.form.get("backtest_execution_mode", "next_open"))
            if selected_mode != current_mode:
                selected_label = "Signal bar close" if selected_mode == "signal_close" else "Next bar open"
                notice = f"Backtest execution model updated: {selected_label}."
        return _redirect_with_settings_feedback("backtest", notice=notice)

    def investment_settings_action():
        current_method = load_investment_cost_basis_method()
        selected_method = save_investment_cost_basis_method(
            request.form.get("investment_cost_basis_method", current_method),
        )
        if selected_method == current_method:
            return _redirect_with_settings_feedback("investment")
        selected_labels = {
            "lowest_cost_first": "Lowest-cost lots first",
            "fifo": "First in, first out (FIFO)",
            "lifo": "Last in, first out (LIFO)",
            "moving_average": "Moving average cost",
        }
        return _redirect_with_settings_feedback(
            "investment",
            notice=(
                "Investment cost basis method updated: "
                f"{selected_labels[selected_method]}."
            ),
        )

    def cash_equivalents_action():
        action = str(request.form.get("action", "save")).strip().lower()
        current = load_cash_equivalent_tickers()
        if action == "add":
            raw = request.form.get("ticker", "") or request.form.get("tickers", "")
            new_ticker = str(raw).strip().upper()
            if new_ticker:
                updated = list(dict.fromkeys(current + [new_ticker]))  # preserve order, dedup
                save_cash_equivalent_tickers(updated)
            return _redirect_with_settings_feedback("cash-equivalents", notice="Cash equivalent added.")
        if action == "remove":
            target = str(request.form.get("ticker", "")).strip().upper()
            if target:
                updated = [t for t in current if t != target]
                save_cash_equivalent_tickers(updated)
            return _redirect_with_settings_feedback("cash-equivalents", notice="Cash equivalent removed.")
        if action == "set":
            # accept repeated tickers or comma
            raw_list = request.form.getlist("ticker") or []
            if not raw_list:
                csv = request.form.get("tickers", "")
                raw_list = [x for x in csv.split(",") if x.strip()]
            updated = _normalize_ticker_list_for_cash(raw_list)
            save_cash_equivalent_tickers(updated)
            return _redirect_with_settings_feedback("cash-equivalents", notice="Cash equivalents updated.")
        # default: redirect
        return _redirect_with_settings_feedback("cash-equivalents")

    def _normalize_ticker_list_for_cash(raw_values: list) -> list[str]:
        result: list[str] = []
        seen: set[str] = set()
        for v in raw_values or []:
            t = str(v or "").strip().upper()
            if t and t not in seen:
                seen.add(t)
                result.append(t)
        return result

    def email_smtp_action():
        action = request.form.get("action", "save").strip().lower()
        current_settings = load_smtp_settings()
        mailbox = request.form.get("from_email", current_settings.from_email or current_settings.username).strip()
        updated_settings = SmtpSettings(
            host=YAHOO_SMTP_HOST,
            port=YAHOO_SMTP_PORT,
            username=mailbox,
            password=request.form.get("password", ""),
            from_email=mailbox,
            use_starttls=request.form.getlist("use_starttls")[-1] == "1" if request.form.getlist("use_starttls") else False,
        )
        if not updated_settings.password:
            updated_settings.password = current_settings.password
        clear_oauth_settings(updated_settings)
        save_smtp_settings(updated_settings)
        if action == "test":
            success, message, updated_settings = test_smtp_connection(updated_settings)
            save_smtp_settings(updated_settings)
        else:
            success, message = True, "Yahoo SMTP settings saved."
        return _redirect_with_settings_feedback(
            "email-smtp",
            notice=message if success else "",
            error="" if success else message,
        )

    def broker_access_action():
        current_settings = load_broker_settings()
        selected_broker = str(
            request.form.get("selected_broker", current_settings.selected_broker)
        ).strip().lower() or "longbridge"
        longbridge_auth_mode = current_settings.longbridge_auth_mode
        action = request.form.get("action", "save")
        if action == "authorize" and selected_broker == "longbridge":
            longbridge_auth_mode = "cli_oauth"

        updated_settings = BrokerSettings(
            selected_broker=selected_broker,
            longbridge_auth_mode=longbridge_auth_mode,
            longbridge_cli_path=str(request.form.get("longbridge_cli_path", "")).strip() or current_settings.longbridge_cli_path,
            longbridge_cli_home=str(request.form.get("longbridge_cli_home", "")).strip() or current_settings.longbridge_cli_home,
            # Legacy Longbridge API credentials remain read-only for backward compatibility.
            # This endpoint intentionally never accepts or stores new Longbridge secrets.
            longbridge_app_key=current_settings.longbridge_app_key,
            longbridge_app_secret=current_settings.longbridge_app_secret,
            longbridge_access_token=current_settings.longbridge_access_token,
            ibkr_account_id=str(request.form.get("ibkr_account_id", "")).strip() or current_settings.ibkr_account_id,
        )
        save_broker_settings(updated_settings)
        if action == "authorize":
            if selected_broker != "longbridge":
                return _redirect_with_settings_feedback(
                    "broker-access",
                    error="Select Longbridge before starting browser authorization.",
                )
            success, message = start_longbridge_cli_browser_oauth(updated_settings)
            return _redirect_with_settings_feedback(
                "broker-access",
                notice=message if success else "",
                error="" if success else message,
                longbridge_oauth_pending="1" if success else "",
            )
        if action == "test":
            success, message = test_broker_connection(updated_settings)
            checked_at = datetime.now().astimezone()
            checked_at_label = format_display_datetime(
                checked_at,
                include_seconds=True,
                timezone_suffix=checked_at.strftime("%Z"),
            )
            return _redirect_with_settings_feedback(
                "broker-access",
                broker_test_status="success" if success else "error",
                broker_test_message=message,
                broker_test_checked_at=checked_at_label,
            )
        else:
            notice = (
                "Broker settings were saved only on this device. "
                "This project is open source, and the developer cannot retrieve your local secrets."
            )
            return _redirect_with_settings_feedback("broker-access", notice=notice)

    def longbridge_oauth_status_api():
        settings = load_broker_settings()
        if settings.selected_broker != "longbridge":
            response = jsonify({
                "status": "error",
                "message": "Select Longbridge before checking browser authorization.",
            })
            response.status_code = 400
            return apply_no_store_headers(response)

        try:
            auth_status = get_longbridge_cli_auth_status(settings)
        except Exception:  # noqa: BLE001
            LOGGER.exception("Longbridge authorization status check failed")
            response = jsonify({
                "status": "error",
                "message": "Longbridge authorization status is temporarily unavailable. Try again later.",
            })
            response.status_code = 503
            return apply_no_store_headers(response)

        token_status = str(((auth_status.get("token") or {}).get("status") or "")).strip().lower()
        if token_status == "refresh_pending":
            return apply_no_store_headers(jsonify({
                "status": "pending",
                "message": "Waiting for Longbridge browser authorization to finish.",
                "token_status": token_status,
            }))

        if token_status != "valid":
            token_failure_messages = {
                "expired": "Longbridge browser authorization expired. Start authorization again.",
                "error": "Longbridge browser authorization failed. Start authorization again.",
                "missing": "Longbridge authorization is unavailable. Start authorization again.",
            }
            return apply_no_store_headers(jsonify({
                "status": "error",
                "message": token_failure_messages.get(
                    token_status,
                    "Longbridge authorization did not report a usable token. Start authorization again.",
                ),
                "token_status": token_status or "unknown",
            }))

        success, message = test_longbridge_cli_connection(settings)
        response = jsonify({
            "status": "success" if success else "error",
            "message": message,
            "token_status": token_status,
        })
        response.delete_cookie(SETTINGS_FEEDBACK_COOKIE, path="/settings")
        return apply_no_store_headers(response)

    def local_market_store_action():
        ticker = normalize_ticker_input(request.form.get("ticker", ""))
        action = request.form.get("action", "").strip().lower()
        page = normalize_settings_page(request.form.get("page", request.form.get("local_page")))

        def build_local_store_redirect(**extra_params: str) -> str:
            return build_settings_state_url(
                "local-market-store",
                page=extra_params.get("page", page),
            )

        redirect_url = build_local_store_redirect()

        try:
            if action == "maintain":
                maintenance = maintain_local_market_store()
                total_count = int(maintenance["total_count"])
                history_refreshed_count = int(maintenance["history_refreshed_count"])
                metadata_refreshed_count = int(maintenance["metadata_refreshed_count"])
                metadata_blocked_count = int(maintenance["metadata_blocked_count"])
                history_failed_tickers = list(maintenance["history_failed_tickers"])
                if history_failed_tickers and history_refreshed_count == 0:
                    failed_preview = ", ".join(history_failed_tickers[:3])
                    return _redirect_with_settings_feedback(
                        "local-market-store",
                        error=f"Unable to refresh historical market data for {failed_preview}.",
                        query_params={"page": page},
                    )

                notice_parts: list[str] = []
                if total_count == 0:
                    notice = "Local Market Store is already up to date."
                    return _redirect_with_settings_feedback(
                        "local-market-store",
                        notice=notice,
                        query_params={"page": page},
                    )

                if history_refreshed_count > 0:
                    notice_parts.append(
                        f"Updated {history_refreshed_count:,} historical parquet dataset"
                        f"{'' if history_refreshed_count == 1 else 's'}."
                    )
                if metadata_refreshed_count > 0:
                    notice_parts.append(
                        f"Refreshed {metadata_refreshed_count:,} logo and company profile entr"
                        f"{'y' if metadata_refreshed_count == 1 else 'ies'}."
                    )
                if metadata_blocked_count > 0:
                    notice_parts.append(
                        f"Yahoo blocked {metadata_blocked_count:,} metadata refresh request"
                        f"{'' if metadata_blocked_count == 1 else 's'}, so cached logos and profiles were kept."
                    )
                if history_failed_tickers:
                    failed_count = len(history_failed_tickers)
                    preview = ", ".join(history_failed_tickers[:3])
                    notice_parts.append(
                        f"{failed_count:,} historical dataset"
                        f"{'' if failed_count == 1 else 's'} could not be refreshed yet"
                        f"{': ' + preview if preview else '.'}"
                    )
                notice = " ".join(part.rstrip(".") + "." for part in notice_parts if part)
                return _redirect_with_settings_feedback(
                    "local-market-store",
                    notice=notice,
                    query_params={"page": page},
                )
            if not ticker:
                return redirect(redirect_url, code=303)
            if action == "refresh":
                refresh_history_store(ticker)
                try:
                    fetch_quote_profile(ticker, force_refresh=True)
                except (AttributeError, ImportError, OSError, ValueError, KeyError, TypeError, RemoteDisconnected):
                    try:
                        fetch_quote_profile(ticker, force_refresh=False)
                    except (AttributeError, ImportError, OSError, ValueError, KeyError, TypeError, RemoteDisconnected):
                        pass
                notice = f"Saved the latest daily market data for {ticker} to local cache."
                return _redirect_with_settings_feedback(
                    "local-market-store",
                    notice=notice,
                    query_params={"page": page},
                )
            elif action == "refresh-1m":
                refresh_result = refresh_one_minute_store(ticker)
                if refresh_result.source == "longbridge_fallback":
                    notice = (
                        f"Saved the latest 6 months of 1-minute market data for {ticker} "
                        "to local cache (via optional Longbridge fallback after yfinance failed)."
                    )
                elif refresh_result.source == "yfinance_30d":
                    notice = (
                        "Saved the latest "
                        f"{refresh_result.fetched_days} days of 1-minute market data to local cache "
                        f"for {ticker} (via the default yfinance window stitching)."
                    )
                else:
                    notice = (
                        "Saved the latest "
                        f"{refresh_result.fetched_days} days of 1-minute market data to local cache "
                        f"for {ticker} (via the default yfinance source)."
                    )
                return _redirect_with_settings_feedback(
                    "local-market-store",
                    notice=notice,
                    query_params={"page": page},
                )
            elif action == "delete":
                delete_ticker_data(ticker)
                notice = f"Removed all cached data for {ticker} from local storage."
                return _redirect_with_settings_feedback(
                    "local-market-store",
                    notice=notice,
                    query_params={"page": page},
                )
        except Exception:  # noqa: BLE001
            LOGGER.exception("Unable to update local market cache for %s", ticker)
            return _redirect_with_settings_feedback(
                "local-market-store",
                error=f"Unable to update local cache for {ticker}. Try again later.",
                query_params={"page": page},
            )

        return redirect(redirect_url, code=303)

    def settings_cache_action():
        section_name = normalize_settings_section(request.form.get("section", "clear-caches"))
        action = str(request.form.get("action", "market-data")).strip().lower() or "market-data"
        try:
            if action == "investment-transactions":
                if clear_investment_store(INVESTMENT_STORE_PATH):
                    notice = "Cleared the local broker transaction record stored in settings_store/investment.parquet."
                else:
                    notice = "No local broker transaction record was found in settings_store/investment.parquet."
                invalidate_investment_transactions_cache()
            else:
                cache_summary = clear_non_historical_market_cache()
                reset_connectivity_caches()
                notice = (
                    f"Cleared {cache_summary['removed_search_queries']:,} market search cache entr"
                    f"{'y' if cache_summary['removed_search_queries'] == 1 else 'ies'}, "
                    f"{cache_summary['removed_profiles']:,} non-local market profile entr"
                    f"{'y' if cache_summary['removed_profiles'] == 1 else 'ies'}, "
                    f"{cache_summary['removed_logos']:,} non-local market logo image"
                    f"{'' if cache_summary['removed_logos'] == 1 else 's'}. "
                    f"Protected {cache_summary['protected_tickers']:,} Local Market Store ticker entr"
                    f"{'y' if cache_summary['protected_tickers'] == 1 else 'ies'}, "
                    f"kept {cache_summary['protected_search_queries']:,} matching market search cache entr"
                    f"{'y' if cache_summary['protected_search_queries'] == 1 else 'ies'}, "
                    "and left ticker usage records untouched."
                )
            return _redirect_with_settings_feedback(section_name, notice=notice)
        except Exception:  # noqa: BLE001
            LOGGER.exception("Unable to clear cached settings data")
            return _redirect_with_settings_feedback(
                section_name,
                error="Unable to clear cached settings data. Try again later.",
            )

    def market_store_logo(filename: str):
        candidate = LOGOS_STORE_DIR / filename
        if candidate.exists():
            return send_from_directory(LOGOS_STORE_DIR, filename)
        return "Not Found", 404

    def favicon_icon():
        candidate = LOGOS_STORE_DIR / "favicon.svg"
        if candidate.exists():
            return send_from_directory(LOGOS_STORE_DIR, "favicon.svg")
        return "Not Found", 404

    def symbol_search():
        query = normalize_ticker_input(request.args.get("q", ""))
        limit = min(max(parse_int_value(request.args.get("limit"), 5), 1), 5)
        report_fetch_abort_debug_event(
            "E",
            "runtime.py:symbol_search",
            "symbol search request received",
            {
                "query": query,
                "limit": limit,
                "path": request.path,
            },
        )
        if not query:
            return jsonify(search_tickers("", limit=limit))
        return jsonify([] if not has_valid_ticker_format(query) else search_tickers(query, limit=limit))

    def date_constraints_api():
        requested_view = normalize_view_name(
            request.args.get("view", request.args.get("mode", "tickers"))
        )
        requested_tickers = parse_requested_tickers(requested_view)
        minimum_required = 1 if requested_view in {"backtest", "dca"} else MIN_TICKERS
        if len(requested_tickers) < minimum_required:
            return jsonify(date_constraint_payload_to_json(build_date_constraint_payload()))
        validated_tickers = [validate_ticker_or_raise(ticker) for ticker in requested_tickers]
        if len(set(validated_tickers)) != len(validated_tickers):
            return jsonify(date_constraint_payload_to_json(build_date_constraint_payload()))
        price_only_flag = request.args.get("return", "").strip().lower() == "price" or request.args.get("price_only", request.args.get("price_return_only", "0")) == "1"
        include_dividends_flag = False if price_only_flag else (
            request.args.get("return", "").strip().lower() == "dividends"
            or request.args.get("dividends", request.args.get("include_dividends", "0")) == "1"
        )
        dividend_mode = resolve_workspace_dividend_mode(price_only_flag, include_dividends_flag)
        requested_start = request.args.get("from", request.args.get("exact_start", "")).strip() or None
        requested_end = request.args.get("to", request.args.get("exact_end", "")).strip() or None
        requested_range = request.args.get("range", request.args.get("range_mode", "")).strip().lower()
        if requested_range == "custom":
            requested_range = "exact"
        requested_period = request.args.get("period", "").strip().lower()
        freshness_refresh_failures: list[str] = []
        if requested_view in {"tickers", "market-caps", "prices"} and requested_range == "exact" and requested_period == "1d":
            payload = build_one_day_intraday_date_constraint_payload(
                validated_tickers,
                requested_start=requested_start,
                requested_end=requested_end,
            )
            return jsonify(date_constraint_payload_to_json(payload))
        if requested_view in {"tickers", "market-caps"} and requested_range == "exact" and requested_period in {"3d", "1w"}:
            payload = build_short_intraday_date_constraint_payload(
                validated_tickers,
                requested_start=requested_start,
                requested_end=requested_end,
            )
            if payload.trading_dates:
                return jsonify(date_constraint_payload_to_json(payload))
        if requested_view in {"tickers", "market-caps", "portfolio", "dca"}:
            freshness_refresh_failures = ensure_latest_daily_caches(validated_tickers)
        datasets = [
            fetch_history(ticker, include_dividends_flag, dividend_mode=dividend_mode)
            for ticker in validated_tickers
        ]
        payload = build_date_constraint_payload(*datasets, requested_start=requested_start, requested_end=requested_end)
        annotate_date_constraint_availability(payload, validated_tickers, datasets)
        if freshness_refresh_failures:
            failed_preview = ", ".join(freshness_refresh_failures)
            freshness_notice = (
                f"Could not refresh the latest trading-day cache for {failed_preview}. "
                "Using the newest local daily data currently available."
            )
            payload.message = f"{payload.message} {freshness_notice}".strip() if payload.message else freshness_notice
        return jsonify(date_constraint_payload_to_json(payload))

    def compare_live_api():
        raw_tickers = [value.strip() for value in request.args.getlist("ticker") if value.strip()]
        if not raw_tickers:
            repeated = str(request.args.get("tickers", "")).strip()
            raw_tickers = [value.strip() for value in repeated.split(",") if value.strip()]
        if len(raw_tickers) < MIN_TICKERS:
            response = jsonify({"success": False, "error": "At least two tickers are required for live comparison."})
            response.status_code = 400
            return apply_no_store_headers(response)

        try:
            validated_tickers = list(dict.fromkeys(validate_ticker_or_raise(ticker) for ticker in raw_tickers))
            if len(validated_tickers) < MIN_TICKERS:
                raise ValueError("At least two distinct tickers are required for live comparison.")

            requested_period = request.args.get("period", "1d").strip().lower() or "1d"
            requested_overnight = request.args.get(
                "overnight",
                request.args.get("include_overnight", "0"),
            ) == "1"
            include_overnight_flag = (
                requested_overnight
                and supports_compare_overnight(validated_tickers, requested_period)
                and has_compare_overnight_market_data_source()
            )
            if include_overnight_flag:
                validated_tickers = resolve_compare_overnight_tickers(validated_tickers)
                if len(validated_tickers) > MAX_TICKERS:
                    raise ValueError(f"Overnight comparison supports at most {MAX_TICKERS} tickers.")

            live_date_value = request.args.get("live_date", "").strip()
            live_trading_date = (
                pd.to_datetime(live_date_value, errors="coerce")
                if live_date_value
                else pd.Timestamp.now(tz="Asia/Shanghai")
            )
            if pd.isna(live_trading_date):
                raise ValueError(f"Invalid live trading date: {live_date_value}.")
            live_trading_date = live_trading_date.date()
            current_live_session_date = pd.Timestamp.now(tz="Asia/Shanghai").date()
            selected_markets = {
                infer_ticker_market(ticker)
                for ticker in validated_tickers
            }
            live_session_active = (
                live_trading_date == current_live_session_date
                and any(is_market_regular_session_active_for_ticker(ticker) for ticker in validated_tickers)
            )

            include_extended_hours_flag = supports_compare_extended_hours(
                validated_tickers,
                requested_period,
            )
            force_refresh = request.args.get("refresh", "1").strip() != "0"

            if requested_period in {"3d", "1w"}:
                live_sources: list[str] = []
                intraday_datasets: list[pd.DataFrame] = []
                for ticker in validated_tickers:
                    intraday_dataset = fetch_history(
                        ticker,
                        include_dividends=False,
                        interval="1m",
                        dividend_mode="price",
                    )
                    intraday_dataset, source = append_live_compare_intraday_dataset(
                        ticker,
                        intraday_dataset,
                        live_trading_date=live_trading_date,
                        include_extended_hours_flag=include_extended_hours_flag,
                        force_refresh=force_refresh,
                    )
                    intraday_datasets.append(intraday_dataset)
                    live_sources.append(source or "local")

                common_end_date = min(dataset["Date"].max() for dataset in intraday_datasets)
                aligned_datasets = slice_intraday_datasets_for_compare_period(
                    intraday_datasets,
                    requested_period,
                    common_end_date,
                    validated_tickers,
                )
                if len(selected_markets) == 1:
                    aligned_datasets = truncate_intraday_datasets_to_common_live_timestamp(aligned_datasets)
                colors = build_series_colors(len(validated_tickers), theme["accent_primary"], theme["accent_secondary"])
                series = [
                    build_compare_series_payload(ticker, dataset, color=color)
                    for ticker, dataset, color in zip(validated_tickers, aligned_datasets, colors)
                ]

                def last_valid_return(item: SeriesPayload) -> float | None:
                    valid_returns = [value for value in item.normalized_returns if value is not None]
                    return valid_returns[-1] if valid_returns else None

                valid_performance_returns = [
                    value for value in (last_valid_return(item) for item in series) if value is not None
                ]
                best_return = max(valid_performance_returns) if valid_performance_returns else None
                dividend_yield_map = build_ttm_dividend_yield_map(validated_tickers, common_end_date)
                best_dividend_yield = best_numeric_metric([
                    dividend_yield_map.get(ticker)
                    for ticker in validated_tickers
                ])
                performance_items = [
                    {
                        "ticker": item.ticker,
                        "ending_return": last_valid_return(item),
                        "ttm_dividend_yield": dividend_yield_map.get(item.ticker),
                        "color": item.color,
                        "is_winner": best_return is not None and last_valid_return(item) == best_return,
                        "is_dividend_yield_winner": (
                            best_dividend_yield is not None
                            and dividend_yield_map.get(item.ticker) == best_dividend_yield
                        ),
                    }
                    for item in series
                ]
                response = jsonify({
                    "success": True,
                    "series": [asdict(item) for item in series],
                    "performanceItems": performance_items,
                    "period": requested_period,
                    "liveDate": pd.Timestamp(live_trading_date).strftime("%Y-%m-%d"),
                    "liveSessionActive": live_session_active,
                    "displayRange": format_compare_intraday_market_local_display_range(
                        aligned_datasets,
                        validated_tickers,
                    ),
                    "sources": {
                        ticker: source
                        for ticker, source in zip(validated_tickers, live_sources)
                    },
                    "fetchedAt": pd.Timestamp.now(tz="UTC").isoformat(),
                })
                return apply_no_store_headers(response)

            axis_date_value = request.args.get("axis_date", request.args.get("trading_date", "")).strip()
            if not axis_date_value and requested_period == "1d":
                axis_date_value = resolve_compare_axis_trading_date(
                    validated_tickers,
                    live_trading_date,
                )
            if not axis_date_value:
                raise ValueError("A reference axis trading date is required.")
            axis_trading_date = pd.to_datetime(axis_date_value, errors="coerce")
            if pd.isna(axis_trading_date):
                raise ValueError(f"Invalid reference axis trading date: {axis_date_value}.")

            reference_datasets = [
                load_compare_one_day_intraday_dataset(
                    ticker,
                    include_extended_hours_flag=include_extended_hours_flag,
                    include_overnight_flag=include_overnight_flag,
                    trading_date=axis_trading_date,
                )
                for ticker in validated_tickers
            ]
            reference_common_end = min(dataset["Date"].max() for dataset in reference_datasets)
            reference_aligned_datasets = slice_intraday_datasets_for_compare_period(
                reference_datasets,
                "1d",
                reference_common_end,
                validated_tickers,
            )
            if pd.Timestamp(axis_trading_date).date() != pd.Timestamp(live_trading_date).date():
                reference_aligned_datasets = [
                    shift_intraday_compare_axis_to_trading_date(
                        dataset,
                        axis_trading_date,
                        live_trading_date,
                    )
                    for dataset in reference_aligned_datasets
                ]

            live_sources: list[str] = []
            colors = build_series_colors(len(validated_tickers), theme["accent_primary"], theme["accent_secondary"])
            mapped_live_datasets: list[pd.DataFrame] = []
            for reference_dataset, ticker in zip(reference_aligned_datasets, validated_tickers):
                try:
                    live_dataset, source = load_live_compare_one_day_intraday_dataset(
                        ticker,
                        live_trading_date=live_trading_date,
                        include_extended_hours_flag=include_extended_hours_flag,
                        force_refresh=force_refresh,
                        include_overnight_flag=include_overnight_flag,
                    )
                    mapped_live_datasets.append(
                        map_live_intraday_dataset_to_reference_axis(reference_dataset, live_dataset, ticker)
                    )
                    live_sources.append(source)
                except Exception as exc:  # noqa: BLE001
                    LOGGER.info("No live compare bars for %s on %s yet: %s", ticker, live_trading_date, exc)
                    mapped_live_datasets.append(build_empty_compare_axis_dataset(reference_dataset))
                    live_sources.append("pending")
            if len(selected_markets) == 1:
                mapped_live_datasets = truncate_intraday_datasets_to_common_live_timestamp(mapped_live_datasets)
            series = [
                build_compare_series_payload(ticker, dataset, color=color)
                for ticker, dataset, color in zip(validated_tickers, mapped_live_datasets, colors)
            ]

            def last_valid_return(item: SeriesPayload) -> float | None:
                valid_returns = [value for value in item.normalized_returns if value is not None]
                return valid_returns[-1] if valid_returns else None

            valid_performance_returns = [
                value for value in (last_valid_return(item) for item in series) if value is not None
            ]
            best_return = max(valid_performance_returns) if valid_performance_returns else None
            dividend_yield_map = build_ttm_dividend_yield_map(validated_tickers, live_trading_date)
            best_dividend_yield = best_numeric_metric([
                dividend_yield_map.get(ticker)
                for ticker in validated_tickers
            ])
            performance_items = [
                {
                    "ticker": item.ticker,
                    "ending_return": last_valid_return(item),
                    "ttm_dividend_yield": dividend_yield_map.get(item.ticker),
                    "color": item.color,
                    "is_winner": best_return is not None and last_valid_return(item) == best_return,
                    "is_dividend_yield_winner": (
                        best_dividend_yield is not None
                        and dividend_yield_map.get(item.ticker) == best_dividend_yield
                    ),
                }
                for item in series
            ]

            response = jsonify({
                "success": True,
                "series": [asdict(item) for item in series],
                "performanceItems": performance_items,
                "axisDate": axis_trading_date.strftime("%Y-%m-%d"),
                "liveDate": pd.Timestamp(live_trading_date).strftime("%Y-%m-%d"),
                "liveSessionActive": live_session_active,
                "displayRange": format_display_date(pd.Timestamp(live_trading_date)),
                "sources": {
                    ticker: source
                    for ticker, source in zip(validated_tickers, live_sources)
                },
                "fetchedAt": pd.Timestamp.now(tz="UTC").isoformat(),
            })
            return apply_no_store_headers(response)
        except Exception:  # noqa: BLE001
            LOGGER.exception("Live comparison request failed")
            response = jsonify({
                "success": False,
                "error": "Live comparison is temporarily unavailable. Try again later.",
            })
            response.status_code = 500
            return apply_no_store_headers(response)

    def trade_strategy_fields_api():
        strategy_id = request.args.get("strategy", "").strip()
        if not strategy_id:
            return jsonify({"is_tunable": False, "html": ""})

        strategy_ids = {str(item["id"]) for item in list_enabled_strategies()}
        if strategy_id not in strategy_ids:
            return jsonify({"is_tunable": False, "html": ""})

        strategy_form_fields = build_strategy_form_fields(strategy_id)
        html = render_template(
            "_trade_strategy_params_panel.html",
            strategy_form_fields=strategy_form_fields,
        )
        return jsonify(
            {
                "is_tunable": bool(strategy_form_fields),
                "html": html,
            }
        )

    def settings_network_status_api():
        if request.args.get("refresh", "").strip() == "1":
            reset_connectivity_caches()
        return jsonify({
            "rows": build_network_service_rows(pending=False),
            "transport_note": network_transport_note(),
        })

    def local_market_store_page_data_api():
        current_page = normalize_settings_page(request.args.get("page"))
        all_local_market_tickers = list_local_market_tickers()
        total_pages = max((len(all_local_market_tickers) - 1) // LOCAL_STORE_PAGE_SIZE + 1, 1)
        current_page = min(current_page, total_pages)
        start_index = (current_page - 1) * LOCAL_STORE_PAGE_SIZE
        end_index = start_index + LOCAL_STORE_PAGE_SIZE
        rows = build_local_market_rows_for_tickers(
            all_local_market_tickers[start_index:end_index],
            include_ranges=True,
        )
        return jsonify(
            {
                "page": current_page,
                "total_pages": total_pages,
                "rows": rows,
            }
        )

    def market_store_presence_api():
        raw_tickers = [value.strip() for value in request.args.getlist("ticker") if value.strip()]
        normalized_tickers: list[str] = []
        for raw_ticker in raw_tickers:
            try:
                normalized_tickers.append(validate_ticker_or_raise(raw_ticker))
            except ValueError:
                continue
        unique_tickers = list(dict.fromkeys(normalized_tickers))
        missing_history = [
            ticker
            for ticker in unique_tickers
            if not history_store_path_for(ticker).exists()
        ]
        has_1m_mapping = {
            ticker: has_recent_one_minute_store(ticker)
            for ticker in unique_tickers
        }
        period_options_mapping = {
            ticker: {
                "1d": build_supported_periods_for_history_store(ticker, "1d"),
                "1m": build_supported_periods_for_history_store(ticker, "1m"),
            }
            for ticker in unique_tickers
        }
        return jsonify(
            {
                "tickers": unique_tickers,
                "missingHistory": missing_history,
                "hasMissingHistory": bool(missing_history),
                "has1m": has_1m_mapping,
                "periodOptions": period_options_mapping,
            }
        )

    def investment_page():
        query_string = request.query_string.decode().strip()
        target_path = build_trade_path("investment")
        return redirect(f"{target_path}?{query_string}" if query_string else target_path)

    def investment_get_transactions():
        """Get all saved investment transactions from local storage."""
        report_fetch_abort_debug_event(
            "E",
            "runtime.py:investment_get_transactions",
            "investment transactions request received",
            {
                "path": request.path,
                "store_exists": investment_store_exists(INVESTMENT_STORE_PATH),
            },
        )
        if not investment_store_exists(INVESTMENT_STORE_PATH):
            invalidate_investment_transactions_cache()
            response = jsonify({
                "transactions": [],
                "ticker_profiles": {},
                "price_history_by_ticker": {},
                "fx_rate_history_by_currency": {},
                "price_history_failures": [],
                "money_market_tickers": sorted(configured_money_market_tickers),
                "money_market_quote_currencies": configured_money_market_quote_currencies,
                "cash_equivalent_tickers": sorted(get_cash_equivalent_tickers()),
                "ticker_lineage": investment_ticker_lineage_payload(),
                "known_ticker_company_names": known_ticker_company_names_payload(),
                "investment_cost_basis_method": load_investment_cost_basis_method(),
                "realtime_quotes": [],
                "section_freshness": build_investment_section_freshness({}),
                "success": True,
            })
            report_fetch_abort_debug_event(
                "E",
                "runtime.py:investment_get_transactions",
                "investment transactions returned empty store payload",
                {
                    "status": 200,
                    "transaction_count": 0,
                },
            )
            return apply_no_store_headers(response)
        try:
            investment_store_fingerprint = build_file_fingerprint(investment_store_path_for(INVESTMENT_STORE_PATH))
            cached_data = read_investment_transactions_cache(investment_store_fingerprint)
            if cached_data is not None:
                section_freshness = cached_data.get("section_freshness", {})
                cached_data["realtime_quotes"] = load_investment_realtime_quotes(
                    section_freshness.get("open_tickers", [])
                    if isinstance(section_freshness, dict)
                    else []
                )
                cached_data["money_market_tickers"] = sorted(configured_money_market_tickers)
                cached_data["money_market_quote_currencies"] = configured_money_market_quote_currencies
                cached_data["cash_equivalent_tickers"] = sorted(get_cash_equivalent_tickers())
                cached_data["ticker_lineage"] = investment_ticker_lineage_payload()
                cached_data["known_ticker_company_names"] = known_ticker_company_names_payload()
                cached_data["investment_cost_basis_method"] = load_investment_cost_basis_method()
                cached_data["success"] = True
                cached_data["investment_store_version"] = str(
                    investment_store_fingerprint.get("mtime_ns", 0)
                )
                cached_data["investment_cache"] = {
                    "status": "hit",
                    "schema_version": INVESTMENT_TRANSACTIONS_CACHE_SCHEMA_VERSION,
                }
                response = jsonify(cached_data)
                report_fetch_abort_debug_event(
                    "E",
                    "runtime.py:investment_get_transactions",
                    "investment transactions returned cached payload",
                    {
                        "status": 200,
                        "transaction_count": len(cached_data.get("transactions", [])),
                    },
                )
                return apply_no_store_headers(response)

            data = load_normalized_investment_payload()
            section_freshness = build_investment_section_freshness(data)
            freshness_refresh_failures = ensure_latest_investment_daily_caches(
                section_freshness["open_tickers"]
            )
            investment_store_fingerprint = build_file_fingerprint(investment_store_path_for(INVESTMENT_STORE_PATH))
            transactions = data.get("transactions", [])
            data["fx_rate_history_by_currency"] = build_investment_fx_rate_history_payload(transactions)
            price_history_by_ticker, price_history_failures = load_investment_price_histories(
                transactions,
                open_tickers=section_freshness["open_tickers"],
            )
            data["ticker_profiles"] = build_investment_ticker_profiles(
                transactions,
                section_freshness["open_tickers"],
            )
            data["price_history_by_ticker"] = price_history_by_ticker
            data["price_history_failures"] = price_history_failures
            data["money_market_tickers"] = sorted(configured_money_market_tickers)
            data["money_market_quote_currencies"] = configured_money_market_quote_currencies
            data["cash_equivalent_tickers"] = sorted(get_cash_equivalent_tickers())
            data["ticker_lineage"] = investment_ticker_lineage_payload()
            data["known_ticker_company_names"] = known_ticker_company_names_payload()
            data["investment_cost_basis_method"] = load_investment_cost_basis_method()
            data["realtime_quotes"] = load_investment_realtime_quotes(section_freshness["open_tickers"])
            data["freshness_refresh_failures"] = freshness_refresh_failures
            data["section_freshness"] = section_freshness
            data["success"] = True
            data["investment_store_version"] = str(
                investment_store_fingerprint.get("mtime_ns", 0)
            )
            data["investment_cache"] = {
                "status": "miss",
                "schema_version": INVESTMENT_TRANSACTIONS_CACHE_SCHEMA_VERSION,
            }
            price_store_fingerprints = build_investment_price_store_fingerprints(
                transactions,
                section_freshness["open_tickers"],
            )
            if not freshness_refresh_failures:
                cacheable_data = dict(data)
                cacheable_data["realtime_quotes"] = []
                cacheable_data["investment_cache"] = {
                    "status": "stored",
                    "schema_version": INVESTMENT_TRANSACTIONS_CACHE_SCHEMA_VERSION,
                }
                write_investment_transactions_cache(
                    investment_store_fingerprint=investment_store_fingerprint,
                    price_store_fingerprints=price_store_fingerprints,
                    payload=cacheable_data,
                )
            response = jsonify(data)
            report_fetch_abort_debug_event(
                "E",
                "runtime.py:investment_get_transactions",
                "investment transactions returned success payload",
                {
                    "status": 200,
                    "transaction_count": len(transactions),
                    "freshness_refresh_failure_count": len(freshness_refresh_failures),
                    "price_history_failure_count": len(price_history_failures),
                },
            )
            return apply_no_store_headers(response)
        except Exception:  # noqa: BLE001
            LOGGER.exception("Unable to load local investment transactions")
            response = jsonify({
                "success": False,
                "error": "Unable to load local investment transactions. Try again later.",
            })
            response.status_code = 500
            report_fetch_abort_debug_event(
                "E",
                "runtime.py:investment_get_transactions",
                "investment transactions failed",
                {
                    "status": 500,
                },
            )
            return apply_no_store_headers(response)

    def investment_download_zircon_hk_template():
        """Download the typed generic fallback investment workbook."""
        response = send_file(
            BytesIO(build_zircon_hk_template_xlsx()),
            as_attachment=True,
            download_name=ZIRCON_HK_TEMPLATE_FILENAME,
            mimetype=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            max_age=0,
        )
        return apply_no_store_headers(response)

    def investment_export_standard_xlsx():
        """Export selected rendered ledger rows in the round-trip XLSX contract."""
        try:
            security_error = validate_investment_browser_write_request(request)
            if security_error:
                response = jsonify({"success": False, "error": security_error})
                response.status_code = 403
                return apply_no_store_headers(response)
            payload = request.get_json(silent=True)
            if not isinstance(payload, dict):
                raise ValueError("The standard XLSX export request is invalid.")
            selected_transactions = payload.get("transactions")
            if not isinstance(selected_transactions, list) or not selected_transactions:
                raise ValueError("Select at least one transaction for standard XLSX export.")
            if len(selected_transactions) > ZIRCON_HK_MAX_TRANSACTION_ROWS:
                raise ValueError(
                    "A standard XLSX export supports at most "
                    f"{ZIRCON_HK_MAX_TRANSACTION_ROWS:,} transactions."
                )
            if not all(isinstance(transaction, dict) for transaction in selected_transactions):
                raise ValueError("The standard XLSX export contains an invalid transaction.")
            tickers = {
                str(transaction.get("ticker") or "").strip().upper()
                for transaction in selected_transactions
                if str(transaction.get("ticker") or "").strip()
            }
            filename = STANDARD_INVESTMENT_EXPORT_FILENAME
            if len(tickers) == 1:
                ticker = re.sub(r"[^A-Z0-9._-]+", "-", next(iter(tickers))).strip("-")
                if ticker:
                    filename = f"{ticker}_standard_investment_export.xlsx"
            response = send_file(
                BytesIO(build_standard_investment_xlsx(selected_transactions)),
                as_attachment=True,
                download_name=filename,
                mimetype=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
                max_age=0,
            )
            return apply_no_store_headers(response)
        except ValueError as exc:
            response = jsonify({"success": False, "error": str(exc)})
            response.status_code = 400
            return apply_no_store_headers(response)
        except Exception:  # noqa: BLE001
            LOGGER.exception("Unable to export the standard investment workbook")
            response = jsonify({
                "success": False,
                "error": "The standard investment workbook could not be exported.",
            })
            response.status_code = 500
            return apply_no_store_headers(response)

    def investment_validate_zircon_hk_workbook():
        """Validate a generic fallback workbook without writing investment data."""
        try:
            security_error = validate_investment_browser_write_request(request)
            if security_error:
                response = jsonify({"success": False, "error": security_error})
                response.status_code = 403
                return apply_no_store_headers(response)
            workbook_file = request.files.get("zircon_hk_transactions_xlsx")
            if workbook_file is None:
                response = jsonify({
                    "success": False,
                    "error": "Please upload the completed manual investment XLSX workbook.",
                })
                response.status_code = 400
                return apply_no_store_headers(response)
            workbook_bytes = workbook_file.read()
            if not workbook_bytes:
                response = jsonify({
                    "success": False,
                    "error": "The manual investment XLSX workbook is empty.",
                })
                response.status_code = 400
                return apply_no_store_headers(response)
            imported_payload = parse_investment_payload(
                "zircon_hk",
                "manual_xlsx",
                xlsx_bytes=workbook_bytes,
                filename=str(
                    getattr(workbook_file, "filename", "") or ""
                ).strip(),
            )
            transaction_count = len(imported_payload.get("transactions", []))
            response = jsonify({
                "success": True,
                "message": (
                    f"Validated {transaction_count:,} manual investment "
                    f"{'transaction' if transaction_count == 1 else 'transactions'}."
                ),
                "transaction_count": transaction_count,
                "summary": imported_payload.get("summary", {}),
            })
            return apply_no_store_headers(response)
        except RequestEntityTooLarge:
            response = jsonify({
                "success": False,
                "error": (
                    f"The workbook exceeds the {MAX_INVESTMENT_IMPORT_REQUEST_MIB} MiB "
                    "investment upload limit."
                ),
            })
            response.status_code = 413
            return apply_no_store_headers(response)
        except ValueError as exc:
            response = jsonify({"success": False, "error": str(exc)})
            response.status_code = 400
            return apply_no_store_headers(response)
        except Exception:  # noqa: BLE001
            LOGGER.exception("Unable to validate manual investment workbook")
            response = jsonify({
                "success": False,
                "error": (
                    "The manual investment workbook could not be validated. "
                    "Download a fresh template and try again."
                ),
            })
            response.status_code = 500
            return apply_no_store_headers(response)

    def investment_validate_hsbc_pasted_text():
        """Validate HSBC paste content without writing ledger or evidence data."""
        try:
            security_error = validate_investment_browser_write_request(request)
            if security_error:
                response = jsonify({"success": False, "error": security_error})
                response.status_code = 403
                return apply_no_store_headers(response)
            request_payload = request.get_json(silent=True)
            if not isinstance(request_payload, dict):
                response = jsonify({
                    "success": False,
                    "error": "HSBC pasted text validation requires a JSON request body.",
                })
                response.status_code = 400
                return apply_no_store_headers(response)
            validation = validate_hsbc_pasted_text(
                portfolio_text=str(request_payload.get("portfolio_text", "") or ""),
                order_status_text=str(request_payload.get("order_status_text", "") or ""),
                cash_account_text=str(request_payload.get("cash_account_text", "") or ""),
                dividend_action_loader=load_local_investment_dividend_actions,
            )
            return apply_no_store_headers(jsonify({"success": True, **validation}))
        except RequestEntityTooLarge:
            response = jsonify({
                "success": False,
                "error": (
                    f"The pasted HSBC text exceeds the {MAX_INVESTMENT_IMPORT_REQUEST_MIB} MiB "
                    "investment import limit."
                ),
            })
            response.status_code = 413
            return apply_no_store_headers(response)
        except ValueError as exc:
            response = jsonify({"success": False, "error": str(exc)})
            response.status_code = 400
            return apply_no_store_headers(response)
        except Exception:  # noqa: BLE001
            LOGGER.exception("Unable to validate HSBC pasted text")
            response = jsonify({
                "success": False,
                "error": "The HSBC pasted text could not be validated. Paste the full HSBC page again.",
            })
            response.status_code = 500
            return apply_no_store_headers(response)

    def investment_add_transactions():
        """Import or sync broker activity into the local investment store."""
        transactions_file = None
        positions_file = None
        try:
            security_error = validate_investment_browser_write_request(request)
            if security_error:
                response = jsonify({"success": False, "error": security_error})
                response.status_code = 403
                return apply_no_store_headers(response)
            broker = str(request.form.get("broker", "ibkr")).strip().lower() or "ibkr"
            report_fetch_abort_debug_event(
                "E",
                "runtime.py:investment_add_transactions",
                "investment import request received",
                {
                    "broker": broker,
                    "path": request.path,
                },
            )
            transactions_file = request.files.get("transactions_csv")
            positions_file = request.files.get("positions_csv")
            dry_run = False
            if broker == "ibkr":
                ibkr_import_mode = str(request.form.get("ibkr_import_mode", "csv")).strip().lower()
                if ibkr_import_mode == "gainskeeper":
                    gainskeeper_files = request.files.getlist("gainskeeper_files")
                    gainskeeper_payloads: list[tuple[bytes, str]] = []
                    for gainskeeper_file in gainskeeper_files:
                        if gainskeeper_file is None:
                            continue
                        file_payload = gainskeeper_file.read()
                        if not file_payload:
                            continue
                        gainskeeper_payloads.append((
                            file_payload,
                            str(getattr(gainskeeper_file, "filename", "") or "").strip(),
                        ))
                    if not gainskeeper_payloads:
                        return jsonify({
                            "success": False,
                            "error": "Please upload at least one IBKR GainsKeeper .gkx file.",
                        }), 400
                    imported_payload = parse_investment_payload(
                        "ibkr",
                        "gainskeeper",
                        files=gainskeeper_payloads,
                    )
                    success_message = (
                        "IBKR GainsKeeper import complete. OFX/GKX records were merged idempotently, "
                        "matching CSV records were upgraded with intraday trade timestamps where available, "
                        "and exact uploaded source files were retained locally as SHA-256-verified immutable evidence."
                    )
                elif ibkr_import_mode == "web_paste":
                    trade_notifications_text = str(
                        request.form.get("ibkr_trade_notifications_text", "")
                    ).strip()
                    if not trade_notifications_text:
                        return jsonify({
                            "success": False,
                            "error": "Please paste the IBKR Trade Notifications page text.",
                        }), 400
                    imported_payload = parse_investment_payload(
                        "ibkr",
                        "web_pasted_text",
                        trade_notifications_text=trade_notifications_text,
                    )
                    success_message = (
                        "IBKR web trade notification sync complete. Filled trades were merged "
                        "idempotently as a provisional current-moment capture; later matching "
                        "Transaction History CSV or GainsKeeper records replace their rounded "
                        "web values with authoritative file precision. Exact pasted text is retained "
                        "locally as SHA-256-verified immutable evidence."
                    )
                else:
                    if transactions_file is None or positions_file is None:
                        return jsonify({
                            "success": False,
                            "error": "Please upload both the Transaction History CSV and the Realized Summary CSV.",
                        }), 400

                    transactions_payload = transactions_file.read()
                    positions_payload = positions_file.read()
                    if not transactions_payload or not positions_payload:
                        return jsonify({
                            "success": False,
                            "error": "Both CSV files must be non-empty.",
                        }), 400

                    imported_payload = parse_investment_payload(
                        "ibkr",
                        "csv",
                        transaction_csv_bytes=transactions_payload,
                        positions_csv_bytes=positions_payload,
                        transaction_filename=str(
                            getattr(transactions_file, "filename", "") or ""
                        ).strip(),
                        positions_filename=str(
                            getattr(positions_file, "filename", "") or ""
                        ).strip(),
                    )
                    success_message = (
                        "IBKR import complete. Matching records were merged incrementally into the local investment store "
                        "without clearing older data first. Exact uploaded CSV source files are retained locally as "
                        "SHA-256-verified immutable evidence."
                    )
            elif broker == "longbridge_hk":
                hk_fund_details_file = request.files.get("longbridge_hk_fund_details_txt")
                hk_history_orders_file = request.files.get("longbridge_hk_history_orders_xlsx")
                if hk_fund_details_file is None or hk_history_orders_file is None:
                    return jsonify({
                        "success": False,
                        "error": "Please upload both the Fund Details text file and the History Orders spreadsheet.",
                    }), 400

                hk_fund_details_bytes = hk_fund_details_file.read()
                hk_fund_details_text = hk_fund_details_bytes.decode("utf-8", errors="replace")
                hk_history_orders_bytes = hk_history_orders_file.read()
                if not hk_fund_details_text.strip() or not hk_history_orders_bytes:
                    return jsonify({
                        "success": False,
                        "error": "Both Longbridge (HK) import files must be non-empty.",
                    }), 400

                imported_payload = parse_investment_payload(
                    "longbridge_hk",
                    "paired_files",
                    fund_details_text=hk_fund_details_text,
                    fund_details_bytes=hk_fund_details_bytes,
                    history_orders_xlsx_bytes=hk_history_orders_bytes,
                    fund_details_filename=str(getattr(hk_fund_details_file, "filename", "") or "").strip(),
                    history_orders_filename=str(getattr(hk_history_orders_file, "filename", "") or "").strip(),
                )
                success_message = (
                    "Longbridge (HK) import complete. Fund Details and History Orders files were parsed in memory and "
                    "merged incrementally into the local investment store without clearing older data first. Exact "
                    "uploaded files were retained locally as SHA-256-verified immutable evidence."
                )
            elif broker == "longbridge_sg":
                fund_details_file = request.files.get("longbridge_sg_fund_details_txt")
                history_orders_file = request.files.get("longbridge_sg_history_orders_xlsx")
                if fund_details_file is None or history_orders_file is None:
                    return jsonify({
                        "success": False,
                        "error": "Please upload both the Fund Details text file and the History Orders spreadsheet.",
                    }), 400

                fund_details_bytes = fund_details_file.read()
                fund_details_text = fund_details_bytes.decode("utf-8", errors="replace")
                history_orders_bytes = history_orders_file.read()
                if not fund_details_text.strip() or not history_orders_bytes:
                    return jsonify({
                        "success": False,
                        "error": "Both Longbridge (SG) import files must be non-empty.",
                    }), 400

                imported_payload = parse_investment_payload(
                    "longbridge_sg",
                    "paired_files",
                    fund_details_text=fund_details_text,
                    fund_details_bytes=fund_details_bytes,
                    history_orders_xlsx_bytes=history_orders_bytes,
                    fund_details_filename=str(getattr(fund_details_file, "filename", "") or "").strip(),
                    history_orders_filename=str(getattr(history_orders_file, "filename", "") or "").strip(),
                )
                success_message = (
                    "Longbridge (SG) import complete. Fund Details and History Orders files were parsed in memory and "
                    "merged incrementally into the local investment store without clearing older data first. Exact "
                    "uploaded files were retained locally as SHA-256-verified immutable evidence."
                )
            elif broker == "futuhk":
                statement_pdf_files = request.files.getlist("futuhk_statement_pdfs")
                statement_pdf_payloads: list[tuple[bytes, str]] = []
                for statement_pdf_file in statement_pdf_files:
                    if statement_pdf_file is None:
                        continue
                    pdf_bytes = statement_pdf_file.read()
                    if not pdf_bytes:
                        continue
                    statement_pdf_payloads.append(
                        (
                            pdf_bytes,
                            str(getattr(statement_pdf_file, "filename", "") or "").strip(),
                        )
                    )
                imported_payload = parse_investment_payload(
                    "futuhk",
                    "statement_pdfs",
                    statement_pdf_payloads=statement_pdf_payloads,
                )
                success_message = (
                    "Futu (HK) import complete. Monthly statement PDFs were parsed in memory and merged "
                    "incrementally into the local investment store without clearing older data first."
                )
            elif broker == "hsbc":
                hsbc_import_mode = str(request.form.get("hsbc_import_mode", "paste")).strip().lower()
                if hsbc_import_mode == "statement_pdf":
                    statement_pdf_payloads: list[tuple[bytes, str]] = []
                    for statement_pdf_file in request.files.getlist("hsbc_statement_pdfs"):
                        if statement_pdf_file is None:
                            continue
                        pdf_bytes = statement_pdf_file.read()
                        if not pdf_bytes:
                            continue
                        statement_pdf_payloads.append((
                            pdf_bytes,
                            str(getattr(statement_pdf_file, "filename", "") or "").strip(),
                        ))
                    imported_payload = parse_investment_payload(
                        "hsbc",
                        "statement_bundle",
                        statement_pdf_payloads=statement_pdf_payloads,
                    )
                    success_message = (
                        "HSBC statement import complete. Full monthly cash-account statements, or compatible "
                        "composite/investment statement pairs, were reconciled by period and account before the "
                        "committed store was read back."
                    )
                else:
                    imported_payload = parse_investment_payload(
                        "hsbc",
                        "pasted_text",
                        portfolio_text=str(
                            request.form.get("hsbc_portfolio_text", "")
                        ).strip(),
                        order_status_text=str(
                            request.form.get("hsbc_order_status_text", "")
                        ).strip(),
                        cash_account_text=str(
                            request.form.get("hsbc_cash_account_text", "")
                        ).strip(),
                        dividend_action_loader=load_local_investment_dividend_actions,
                    )
                    import_summary = (
                        imported_payload.get("summary")
                        if isinstance(imported_payload.get("summary"), dict)
                        else {}
                    )
                    if import_summary.get("hsbc_paste_import_scope") == "cash_only_non_usd":
                        success_message = (
                            "HSBC cash-only sync complete. The pasted HKD/CNH cash-account text was normalized and "
                            "merged incrementally without replacing the existing USD Portfolio or cash snapshot."
                        )
                    else:
                        success_message = (
                            "HSBC sync complete. The pasted cash-account, Portfolio, and Order Status text were normalized and "
                            "merged incrementally into the local investment store without "
                            "clearing older data first."
                        )
            elif broker == "schwab":
                schwab_transactions_file = request.files.get("transactions_csv")
                if schwab_transactions_file is None:
                    schwab_transactions_file = request.files.get("schwab_transactions_csv")
                schwab_positions_file = request.files.get("positions_csv")
                if schwab_positions_file is None:
                    schwab_positions_file = request.files.get("schwab_positions_csv")
                if schwab_transactions_file is None or schwab_positions_file is None:
                    return jsonify({
                        "success": False,
                        "error": "Please upload both the Schwab Transactions CSV and Positions CSV.",
                    }), 400
                schwab_transactions_payload = schwab_transactions_file.read()
                schwab_positions_payload = schwab_positions_file.read()
                if not schwab_transactions_payload or not schwab_positions_payload:
                    return jsonify({
                        "success": False,
                        "error": "The Schwab Transactions and Positions CSV files cannot be empty.",
                    }), 400
                imported_payload = parse_investment_payload(
                    "schwab",
                    "csv",
                    transaction_csv_bytes=schwab_transactions_payload,
                    positions_csv_bytes=schwab_positions_payload,
                    transaction_filename=str(
                        getattr(schwab_transactions_file, "filename", "") or ""
                    ).strip(),
                    positions_filename=str(
                        getattr(schwab_positions_file, "filename", "") or ""
                    ).strip(),
                )
                success_message = (
                    "Charles Schwab import complete. Transactions and the authoritative Positions snapshot were "
                    "merged incrementally into the local investment store without clearing older data first."
                )
            elif broker == "boc_hk" and request.files.getlist("boc_hk_statement_pdfs"):
                statement_pdf_payloads: list[tuple[bytes, str]] = []
                for statement_pdf_file in request.files.getlist("boc_hk_statement_pdfs"):
                    if statement_pdf_file is None:
                        continue
                    source_filename = str(
                        getattr(statement_pdf_file, "filename", "") or ""
                    ).strip()
                    if not source_filename:
                        return jsonify({
                            "success": False,
                            "error": "Every uploaded BOCHK statement must have a non-empty filename.",
                        }), 400
                    if not source_filename.lower().endswith(".pdf"):
                        return jsonify({
                            "success": False,
                            "error": (
                                f"The uploaded BOCHK statement '{source_filename}' must use a .pdf filename."
                            ),
                        }), 400
                    pdf_bytes = statement_pdf_file.read()
                    if not pdf_bytes:
                        return jsonify({
                            "success": False,
                            "error": (
                                f"The uploaded BOCHK statement PDF '{source_filename}' is empty."
                            ),
                        }), 400
                    statement_pdf_payloads.append((
                        pdf_bytes,
                        source_filename,
                    ))
                if not statement_pdf_payloads:
                    return jsonify({
                        "success": False,
                        "error": "Please upload at least one BOCHK Consolidated Statement PDF.",
                    }), 400
                imported_payload = parse_investment_payload(
                    "boc_hk",
                    "statement_pdfs",
                    statement_pdf_payloads=statement_pdf_payloads,
                )
                success_message = (
                    "BOCHK import complete. Consolidated Statement PDFs were parsed in memory, with each cash subaccount "
                    "and source currency preserved, then merged incrementally without clearing older data first."
                )
            elif broker == "boc_hk" and request.files.get("zircon_hk_transactions_xlsx") is None:
                return jsonify({
                    "success": False,
                    "error": "Please upload at least one BOCHK Consolidated Statement PDF.",
                }), 400
            elif broker in {
                "zircon_hk",
                "standard_xlsx",
                "cmb_cn",
                "boc_cn",
                "boc_hk",
                "icbc_cn",
                "icbc_hk",
                "ccb_cn",
                "ccb_hk",
            }:
                workbook_file = request.files.get("zircon_hk_transactions_xlsx")
                if workbook_file is None:
                    return jsonify({
                        "success": False,
                        "error": "Please upload the completed manual investment XLSX workbook.",
                    }), 400
                workbook_bytes = workbook_file.read()
                if not workbook_bytes:
                    return jsonify({
                        "success": False,
                        "error": "The manual investment XLSX workbook is empty.",
                    }), 400
                imported_payload = parse_investment_payload(
                    "zircon_hk",
                    "manual_xlsx",
                    xlsx_bytes=workbook_bytes,
                    filename=str(
                        getattr(workbook_file, "filename", "") or ""
                    ).strip(),
                )
                success_message = (
                    "Manual investment workbook import complete. Validated records were "
                    "merged incrementally into the local investment store, and the exact "
                    "uploaded XLSX was retained as SHA-256-verified immutable evidence."
                )
            elif broker in {"tigertrade", "usmart_hk"}:
                field_name = f"{broker}_statement_pdfs"
                statement_pdf_payloads: list[tuple[bytes, str]] = []
                for statement_pdf_file in request.files.getlist(field_name):
                    if statement_pdf_file is None:
                        continue
                    pdf_bytes = statement_pdf_file.read()
                    if not pdf_bytes:
                        continue
                    statement_pdf_payloads.append((
                        pdf_bytes,
                        str(getattr(statement_pdf_file, "filename", "") or "").strip(),
                    ))
                if broker == "tigertrade":
                    imported_payload = parse_investment_payload(
                        "tigertrade",
                        "statement_pdfs",
                        statement_pdf_payloads=statement_pdf_payloads,
                    )
                    broker_label = "Tiger Trade"
                else:
                    imported_payload = parse_investment_payload(
                        "usmart_hk",
                        "statement_pdfs",
                        statement_pdf_payloads=statement_pdf_payloads,
                    )
                    broker_label = "uSMART (HK)"
                success_message = (
                    f"{broker_label} import complete. Statement PDFs were parsed in memory and merged "
                    "incrementally into the local investment store without clearing older data first."
                )
            else:
                return jsonify({
                    "success": False,
                    "error": f"{broker.upper()} investment import is not implemented yet.",
                }), 400

            if dry_run:
                investment_payload = imported_payload
            else:
                investment_payload = merge_and_write_investment_payload(imported_payload)
            
            # Run the price cache refresh in the background (skip for dry-run)
            if not dry_run:
                def background_refresh(payload: dict[str, Any]) -> None:
                    try:
                        refresh_investment_import_price_caches(payload)
                    except Exception:  # noqa: BLE001
                        LOGGER.exception("Investment import background refresh failed")
                
                threading.Thread(target=background_refresh, args=(imported_payload,), daemon=True).start()
            freshness_refresh_failures: list[str] = []

            return jsonify({
                "success": True,
                "message": success_message,
                "summary": investment_payload.get("summary", {}),
                "freshness_refresh_failures": freshness_refresh_failures,
                "investment_store_version": str(build_file_fingerprint(
                    investment_store_path_for(INVESTMENT_STORE_PATH)
                ).get("mtime_ns", 0)),
                "transaction_count": len(investment_payload.get("transactions", [])),
            })
        except RequestEntityTooLarge:
            response = jsonify({
                "success": False,
                "error": (
                    f"The investment import exceeds the {MAX_INVESTMENT_IMPORT_REQUEST_MIB} MiB total upload limit. "
                    "Upload fewer or smaller files."
                ),
            })
            response.status_code = 413
            report_fetch_abort_debug_event(
                "E",
                "runtime.py:investment_add_transactions",
                "investment import rejected because the request exceeded its size limit",
                {
                    "status": 413,
                },
            )
            return apply_no_store_headers(response)
        except ValueError as exc:
            report_fetch_abort_debug_event(
                "E",
                "runtime.py:investment_add_transactions",
                "investment import rejected with value error",
                {
                    "status": 400,
                    "error": str(exc),
                },
            )
            return jsonify({"success": False, "error": str(exc)}), 400
        except Exception:  # noqa: BLE001
            LOGGER.exception("Investment import failed")
            report_fetch_abort_debug_event(
                "E",
                "runtime.py:investment_add_transactions",
                "investment import failed",
                {
                    "status": 500,
                },
            )
            return jsonify({
                "success": False,
                "error": "The investment import could not be completed. Try again later.",
            }), 500

    def investment_update_internal_transfer_binding():
        """Persist a manual internal-transfer binding into the local investment store."""
        try:
            security_error = validate_investment_browser_write_request(request)
            if security_error:
                response = jsonify({"success": False, "error": security_error})
                response.status_code = 403
                return apply_no_store_headers(response)
            payload = request.get_json(silent=True) or {}
            requested_source_key = str(payload.get("source_key", "")).strip()
            requested_target_key = str(payload.get("target_key", "")).strip()
            requested_action = str(payload.get("action", "")).strip().lower()
            if requested_action not in {"", "bind", "ignore", "restore"}:
                return jsonify({
                    "success": False,
                    "error": "The internal-transfer action is invalid.",
                }), 400
            if not requested_source_key:
                return jsonify({
                    "success": False,
                    "error": "A source transfer key is required.",
                }), 400
            if not investment_store_exists(INVESTMENT_STORE_PATH):
                return jsonify({
                    "success": False,
                    "error": "No local investment store exists yet.",
                }), 400

            def update_bindings(
                current_payload: dict[str, object],
            ) -> tuple[dict[str, object], dict[str, object]]:
                investment_payload = normalize_investment_payload_tickers(current_payload)
                transactions = investment_payload.get("transactions")
                ignored_source_keys = normalize_investment_internal_transfer_ignored_source_keys(
                    investment_payload.get("manual_internal_transfer_ignored_source_keys"),
                    transactions=transactions,
                )
                if requested_action in {"ignore", "restore"}:
                    normalized_requested_source_keys = normalize_investment_internal_transfer_ignored_source_keys(
                        [requested_source_key],
                        transactions=transactions,
                    )
                    source_key = (
                        normalized_requested_source_keys[0]
                        if len(normalized_requested_source_keys) == 1
                        else requested_source_key
                    )
                    binding_index = build_investment_internal_transfer_binding_index(
                        transactions
                    )
                    if len(binding_index.get(source_key, [])) != 1:
                        raise ValueError("The source transfer key is missing or ambiguous.")
                    next_bindings = normalize_investment_internal_transfer_bindings(
                        investment_payload.get("manual_internal_transfer_bindings"),
                        transactions=transactions,
                    )
                    if requested_action == "ignore":
                        next_bindings.pop(source_key, None)
                        if source_key not in ignored_source_keys:
                            ignored_source_keys.append(source_key)
                    else:
                        ignored_source_keys = [
                            key for key in ignored_source_keys if key != source_key
                        ]
                    investment_payload["manual_internal_transfer_bindings"] = next_bindings
                    investment_payload["manual_internal_transfer_ignored_source_keys"] = (
                        ignored_source_keys
                    )
                    updated_payload = refresh_investment_security_transfer_reconciliation(
                        investment_payload
                    )
                    return cast(dict[str, object], updated_payload), {
                        "manual_internal_transfer_bindings": next_bindings,
                        "manual_internal_transfer_ignored_source_keys": (
                            updated_payload.get(
                                "manual_internal_transfer_ignored_source_keys", []
                            )
                        ),
                        "summary": updated_payload.get("summary", {}),
                    }
                requested_pair = normalize_investment_internal_transfer_bindings(
                    {
                        requested_source_key: requested_target_key or requested_source_key,
                    },
                    transactions=transactions,
                )
                source_key = next(iter(requested_pair), requested_source_key)
                target_key = requested_pair.get(source_key, "") if requested_target_key else ""
                if requested_target_key:
                    validate_investment_internal_transfer_binding(
                        transactions,
                        source_key,
                        target_key,
                    )
                next_bindings = normalize_investment_internal_transfer_bindings(
                    investment_payload.get("manual_internal_transfer_bindings"),
                    transactions=transactions,
                )
                ignored_source_keys = [
                    key for key in ignored_source_keys if key != source_key
                ]
                if target_key:
                    for existing_source_key, existing_target_key in next_bindings.items():
                        if existing_source_key != source_key and existing_target_key == target_key:
                            raise ValueError(
                                "The selected internal-transfer counterpart is already bound to another source record. Remove that binding first."
                            )
                    next_bindings[source_key] = target_key
                else:
                    next_bindings.pop(source_key, None)
                investment_payload["manual_internal_transfer_bindings"] = next_bindings
                investment_payload["manual_internal_transfer_ignored_source_keys"] = (
                    ignored_source_keys
                )
                updated_payload = refresh_investment_security_transfer_reconciliation(
                    investment_payload
                )
                return cast(dict[str, object], updated_payload), {
                    "manual_internal_transfer_bindings": next_bindings,
                    "manual_internal_transfer_ignored_source_keys": (
                        updated_payload.get(
                            "manual_internal_transfer_ignored_source_keys", []
                        )
                    ),
                    "summary": updated_payload.get("summary", {}),
                }

            update_result = cast(
                dict[str, object],
                update_investment_store_payload(update_bindings, INVESTMENT_STORE_PATH),
            )
            invalidate_investment_transactions_cache()
            return jsonify({
                "success": True,
                "manual_internal_transfer_bindings": update_result.get(
                    "manual_internal_transfer_bindings", {}
                ),
                "manual_internal_transfer_ignored_source_keys": update_result.get(
                    "manual_internal_transfer_ignored_source_keys", []
                ),
                "summary": update_result.get("summary", {}),
            })
        except ValueError as exc:
            return jsonify({
                "success": False,
                "error": str(exc),
            }), 400
        except Exception:  # noqa: BLE001
            LOGGER.exception("Unable to update internal transfer binding")
            return jsonify({
                "success": False,
                "error": "Unable to update the internal transfer binding. Try again later.",
            }), 500

    def investment_update_security_transfer_attribution():
        """Persist a user-attested source account for one Schwab receipt only."""
        try:
            security_error = validate_investment_browser_write_request(request)
            if security_error:
                response = jsonify({"success": False, "error": security_error})
                response.status_code = 403
                return apply_no_store_headers(response)
            request_payload = request.get_json(silent=True) or {}
            requested_receipt_key = str(request_payload.get("receipt_key", "")).strip()
            raw_source_broker = str(request_payload.get("source_broker", "")).strip()
            raw_source_account = str(request_payload.get("source_account", "")).strip()
            if not requested_receipt_key:
                return jsonify({
                    "success": False,
                    "error": "A Schwab transfer receipt key is required.",
                }), 400
            if bool(raw_source_broker) != bool(raw_source_account):
                return jsonify({
                    "success": False,
                    "error": "Select both a source broker and a source account, or clear the attribution.",
                }), 400
            if not investment_store_exists(INVESTMENT_STORE_PATH):
                return jsonify({
                    "success": False,
                    "error": "No local investment store exists yet.",
                }), 400

            def update_attributions(
                current_payload: dict[str, object],
            ) -> tuple[dict[str, object], dict[str, object]]:
                investment_payload = normalize_investment_payload_tickers(current_payload)
                transactions = investment_payload.get("transactions")
                next_attributions = normalize_investment_security_transfer_attributions(
                    investment_payload.get("manual_security_transfer_attributions"),
                    transactions=transactions,
                )
                if raw_source_broker:
                    validate_investment_security_transfer_attribution(
                        transactions,
                        requested_receipt_key,
                        raw_source_broker,
                        raw_source_account,
                        existing_attributions=next_attributions,
                    )
                    next_attributions[requested_receipt_key] = {
                        "schema_version": "1",
                        "source_broker": raw_source_broker,
                        "source_account": raw_source_account,
                        "attested_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    }
                else:
                    next_attributions.pop(requested_receipt_key, None)
                investment_payload["manual_security_transfer_attributions"] = (
                    next_attributions
                )
                updated_payload = refresh_investment_security_transfer_reconciliation(
                    investment_payload
                )
                return cast(dict[str, object], updated_payload), {
                    "manual_security_transfer_attributions": updated_payload.get(
                        "manual_security_transfer_attributions", {}
                    ),
                    "summary": updated_payload.get("summary", {}),
                }

            update_result = cast(
                dict[str, object],
                update_investment_store_payload(
                    update_attributions,
                    INVESTMENT_STORE_PATH,
                ),
            )
            invalidate_investment_transactions_cache()
            return jsonify({
                "success": True,
                "manual_security_transfer_attributions": update_result.get(
                    "manual_security_transfer_attributions", {}
                ),
                "summary": update_result.get("summary", {}),
            })
        except ValueError as exc:
            return jsonify({
                "success": False,
                "error": str(exc),
            }), 400
        except Exception:  # noqa: BLE001
            LOGGER.exception("Unable to update Schwab transfer attribution")
            return jsonify({
                "success": False,
                "error": "Unable to update the Schwab transfer attribution. Try again later.",
            }), 500

    def investment_get_latest_price():
        """Get the latest closing price for a ticker from local market store."""
        ticker = request.args.get("ticker", "").strip().upper()
        if not ticker:
            response = jsonify({"success": False, "error": "No ticker provided"})
            response.status_code = 400
            return apply_no_store_headers(response)

        try:
            path = resolve_investment_history_store_path(ticker)
            if path is None:
                response = jsonify({"success": False, "error": f"No local data for {ticker}"})
                response.status_code = 404
                return apply_no_store_headers(response)

            df = pd.read_parquet(path)
            if df.empty or "Close" not in df.columns:
                response = jsonify({"success": False, "error": f"No price data for {ticker}"})
                response.status_code = 404
                return apply_no_store_headers(response)

            # Get the latest close price (last row)
            latest_row = df.sort_values("Date").iloc[-1]
            latest_close = float(latest_row["Close"])
            latest_date = str(latest_row["Date"].date()) if hasattr(latest_row["Date"], "date") else str(latest_row["Date"])

            response = jsonify({
                "success": True,
                "ticker": ticker,
                "latest_close": latest_close,
                "latest_date": latest_date
            })
            return apply_no_store_headers(response)
        except Exception:  # noqa: BLE001
            LOGGER.exception("Unable to load the latest local price for %s", ticker)
            response = jsonify({
                "success": False,
                "error": f"Unable to load the latest local price for {ticker}. Try again later.",
            })
            response.status_code = 500
            return apply_no_store_headers(response)

    def investment_get_parquet():
        """Get all date -> close price mappings from the parquet file for a ticker."""
        ticker = request.args.get("ticker", "").strip().upper()
        if not ticker:
            response = jsonify({"success": False, "error": "No ticker provided"})
            response.status_code = 400
            return apply_no_store_headers(response)

        try:
            path = resolve_investment_history_store_path(ticker)
            freshness_scope = request.args.get("freshness_scope", "").strip().lower()
            section_freshness = build_investment_section_freshness(
                load_normalized_investment_payload()
            )
            should_refresh_ticker = (
                ticker in set(section_freshness["open_tickers"])
                and not is_configured_money_market_ticker(ticker)
            )
            if path is None:
                if not should_refresh_ticker:
                    response = jsonify({"success": False, "error": f"No local data for {ticker}"})
                    response.status_code = 404
                    return apply_no_store_headers(response)
                fetch_history(ticker, include_dividends=False)
                path = resolve_investment_history_store_path(ticker)
            else:
                if should_refresh_ticker:
                    ensure_latest_investment_daily_caches([ticker])
                    path = resolve_investment_history_store_path(ticker) or path

            if path is None:
                response = jsonify({"success": False, "error": f"No local data for {ticker}"})
                response.status_code = 404
                return apply_no_store_headers(response)
            prices = load_price_history_series(path)
            if not prices:
                response = jsonify({"success": False, "error": f"No price/date data for {ticker}"})
                response.status_code = 404
                return apply_no_store_headers(response)

            response = jsonify({
                "success": True,
                "ticker": ticker,
                "prices": prices,
                "count": len(prices),
                "freshness_scope": freshness_scope or "ticker",
                "target_trading_day": section_freshness["target_trading_day"] if freshness_scope == "section" else "",
            })
            return apply_no_store_headers(response)
        except Exception:  # noqa: BLE001
            LOGGER.exception("Unable to load local price history for %s", ticker)
            response = jsonify({
                "success": False,
                "error": f"Unable to load local price history for {ticker}. Try again later.",
            })
            response.status_code = 500
            return apply_no_store_headers(response)

    def normalize_investment_intraday_ohlc(dataset: pd.DataFrame) -> pd.DataFrame:
        """Keep only positive, structurally valid OHLC bars at the API boundary."""
        price_columns = ["Open", "High", "Low", "Close"]
        if not all(column in dataset.columns for column in ["Date", *price_columns]):
            return dataset.iloc[0:0].copy()
        normalized = dataset.copy()
        normalized["Date"] = pd.to_datetime(normalized["Date"], errors="coerce")
        normalized[price_columns] = normalized[price_columns].apply(pd.to_numeric, errors="coerce")
        normalized = normalized.dropna(subset=["Date", *price_columns])
        positive_prices = (normalized[price_columns] > 0).all(axis=1)
        valid_structure = (
            (normalized["High"] >= normalized["Open"])
            & (normalized["High"] >= normalized["Close"])
            & (normalized["Low"] <= normalized["Open"])
            & (normalized["Low"] <= normalized["Close"])
            & (normalized["High"] >= normalized["Low"])
        )
        return normalized.loc[positive_prices & valid_structure].sort_values("Date")

    def investment_get_intraday_history():
        """Get local 1-minute OHLC history for Investment charts."""
        ticker = request.args.get("ticker", "").strip().upper()
        requested_range = request.args.get("range", "").strip().lower() or "1w"
        ensure_store = request.args.get("ensure_store", "").strip() == "1"
        requested_days = [
            day
            for day in (part.strip() for part in request.args.get("days", "").split(","))
            if re.match(r"^\d{4}-\d{2}-\d{2}$", day)
        ]
        if not ticker:
            response = jsonify({"success": False, "error": "No ticker provided"})
            response.status_code = 400
            return apply_no_store_headers(response)

        try:
            normalized_ticker = validate_ticker_or_raise(ticker)
            if is_configured_money_market_ticker(normalized_ticker):
                response = jsonify({
                    "success": True,
                    "ticker": normalized_ticker,
                    "interval": "1m",
                    "range": requested_range,
                    "days": requested_days,
                    "rows": [],
                    "count": 0,
                    "refreshed": False,
                    "source": "money_market_anchor",
                })
                return apply_no_store_headers(response)
            refresh_result = None
            intraday_path = resolve_investment_history_store_path(normalized_ticker, interval="1m")
            if ensure_store and not is_one_minute_store_fresh(normalized_ticker):
                refresh_result = refresh_one_minute_store(normalized_ticker)
                intraday_path = resolve_investment_history_store_path(normalized_ticker, interval="1m")

            if intraday_path is not None:
                dataset = pd.read_parquet(intraday_path)
            else:
                dataset = fetch_history(normalized_ticker, include_dividends=False, interval="1m")
            if dataset.empty:
                response = jsonify({"success": False, "error": f"No 1-minute market data for {normalized_ticker}"})
                response.status_code = 404
                return apply_no_store_headers(response)

            intraday = normalize_investment_intraday_ohlc(dataset)
            if intraday.empty:
                response = jsonify({"success": False, "error": f"No 1-minute OHLC data for {normalized_ticker}"})
                response.status_code = 404
                return apply_no_store_headers(response)
            if ensure_store and requested_days:
                session_state = nyse_market_session_state(include_overnight=True)
                recent_day_count = 23 if requested_range == "1m" else 5
                refreshable_days = set(nyse_recent_trading_days(
                    session_state.get("as_of"),
                    day_count=recent_day_count,
                ))
                eligible_requested_days = [
                    day for day in requested_days if day in refreshable_days
                ]
                intraday_day_keys = intraday["Date"].dt.strftime("%Y-%m-%d")
                available_days = set(intraday_day_keys.drop_duplicates().tolist())
                available_bar_counts = intraday_day_keys.value_counts()
                active_session_day = (
                    str(session_state.get("session_date") or "")
                    if session_state.get("is_realtime_allowed")
                    and session_state.get("session") == "intraday"
                    else ""
                )

                def expected_regular_bar_count(day_key: str) -> int:
                    regular_close_minute = (13 * 60) if is_nyse_early_close(day_key) else (16 * 60)
                    return regular_close_minute - ((9 * 60) + 30)

                missing_days = [
                    day for day in eligible_requested_days if day not in available_days
                ]
                incomplete_days = [
                    day
                    for day in eligible_requested_days
                    if day != active_session_day
                    and int(available_bar_counts.get(day, 0)) < expected_regular_bar_count(day)
                ]
                refresh_days = set(missing_days).union(incomplete_days)
                if refresh_days:
                    try:
                        refresh_result = refresh_recent_one_minute_store_with_yfinance(
                            normalized_ticker,
                            days=30,
                        )
                        intraday_path = resolve_investment_history_store_path(normalized_ticker, interval="1m")
                        dataset = pd.read_parquet(intraday_path) if intraday_path is not None else dataset
                        intraday = normalize_investment_intraday_ohlc(dataset)
                    except Exception as exc:  # noqa: BLE001
                        LOGGER.debug(
                            "Unable to fill requested Investment intraday days for %s with Yahoo: %s",
                            normalized_ticker,
                            exc,
                        )
                    available_days = set(
                        intraday["Date"].dt.strftime("%Y-%m-%d").drop_duplicates().tolist()
                    )
                    available_bar_counts = intraday["Date"].dt.strftime("%Y-%m-%d").value_counts()
                    fallback_days = [
                        day
                        for day in refresh_days
                        if (
                            day != active_session_day
                            and int(available_bar_counts.get(day, 0)) < expected_regular_bar_count(day)
                        )
                        or (day == active_session_day and day not in available_days)
                    ]
                    if fallback_days:
                        try:
                            longbridge_refresh_result = refresh_one_minute_store_with_longbridge(
                                normalized_ticker
                            )
                            intraday_path = resolve_investment_history_store_path(
                                normalized_ticker,
                                interval="1m",
                            )
                            dataset = pd.read_parquet(intraday_path) if intraday_path is not None else dataset
                            intraday = normalize_investment_intraday_ohlc(dataset)
                            available_days = set(
                                intraday["Date"].dt.strftime("%Y-%m-%d").drop_duplicates().tolist()
                            )
                            available_bar_counts = intraday["Date"].dt.strftime("%Y-%m-%d").value_counts()
                            if all(
                                day in available_days
                                and (
                                    day == active_session_day
                                    or int(available_bar_counts.get(day, 0)) >= expected_regular_bar_count(day)
                                )
                                for day in fallback_days
                            ):
                                refresh_result = longbridge_refresh_result
                            else:
                                LOGGER.debug(
                                    "Longbridge Investment intraday refresh for %s did not fill %s",
                                    normalized_ticker,
                                    ", ".join(fallback_days),
                                )
                        except Exception as exc:  # noqa: BLE001
                            LOGGER.debug(
                                "Unable to fill requested Investment intraday days for %s with fallback data: %s",
                                normalized_ticker,
                                exc,
                            )

            latest_timestamp = intraday["Date"].max()
            if requested_days:
                requested_day_set = set(requested_days)
                intraday = intraday.loc[intraday["Date"].dt.strftime("%Y-%m-%d").isin(requested_day_set)].copy()
            elif requested_range == "current-day":
                latest_day = latest_timestamp.strftime("%Y-%m-%d")
                intraday = intraday.loc[intraday["Date"].dt.strftime("%Y-%m-%d") == latest_day].copy()
            elif requested_range == "3d":
                trading_days = intraday["Date"].dt.strftime("%Y-%m-%d").drop_duplicates().tolist()
                selected_days = set(trading_days[-3:])
                if selected_days:
                    intraday = intraday.loc[intraday["Date"].dt.strftime("%Y-%m-%d").isin(selected_days)].copy()
            elif requested_range == "1w":
                trading_days = nyse_recent_trading_days(latest_timestamp, day_count=5)
                selected_days = set(trading_days)
                if selected_days:
                    intraday = intraday.loc[intraday["Date"].dt.strftime("%Y-%m-%d").isin(selected_days)].copy()
            elif requested_range == "1m":
                trading_days = nyse_recent_trading_days(latest_timestamp, day_count=23)
                selected_days = set(trading_days)
                if selected_days:
                    intraday = intraday.loc[intraday["Date"].dt.strftime("%Y-%m-%d").isin(selected_days)].copy()

            rows = [
                {
                    "date": timestamp.strftime("%Y-%m-%d %H:%M"),
                    "open": float(open_value),
                    "high": float(high_value),
                    "low": float(low_value),
                    "close": float(close_value),
                }
                for timestamp, open_value, high_value, low_value, close_value in zip(
                    intraday["Date"],
                    intraday["Open"],
                    intraday["High"],
                    intraday["Low"],
                    intraday["Close"],
                )
            ]
            response = jsonify({
                "success": True,
                "ticker": normalized_ticker,
                "interval": "1m",
                "range": requested_range,
                "days": requested_days,
                "rows": rows,
                "count": len(rows),
                "refreshed": refresh_result is not None,
                "source": refresh_result.source if refresh_result is not None else "local",
            })
            return apply_no_store_headers(response)
        except Exception:  # noqa: BLE001
            LOGGER.exception("Unable to load 1-minute market data for %s", ticker)
            response = jsonify({
                "success": False,
                "error": f"Unable to load 1-minute market data for {ticker}. Try again later.",
            })
            response.status_code = 500
            return apply_no_store_headers(response)

    def investment_get_realtime_quotes():
        """Get Longbridge-first realtime quotes with yfinance fallback."""
        failures: list[dict[str, str]] = []

        def collect_valid_tickers(raw_values: list[str]) -> list[str]:
            valid_tickers: list[str] = []
            for raw_value in raw_values:
                raw_ticker = str(raw_value or "").strip()
                if not raw_ticker:
                    continue
                try:
                    valid_tickers.append(validate_ticker_or_raise(raw_ticker))
                except ValueError as exc:
                    failures.append({
                        "ticker": raw_ticker,
                        "error": str(exc),
                    })
            return valid_tickers

        requested_tickers = collect_valid_tickers(request.args.getlist("ticker"))
        if not requested_tickers and not failures:
            repeated = str(request.args.get("tickers", "")).strip()
            requested_tickers = collect_valid_tickers(repeated.split(","))
        requested_tickers = list(dict.fromkeys(requested_tickers))
        if not requested_tickers:
            response = jsonify({
                "success": False,
                "error": "No valid tickers provided" if failures else "No tickers provided",
                "quotes": [],
                "failures": failures,
                "count": 0,
                "source": "yfinance",
            })
            response.status_code = 400
            return apply_no_store_headers(response)

        quotes: list[dict[str, object]] = []
        fetched_at = pd.Timestamp.now(tz="UTC")
        try:
            quotes = load_investment_realtime_quotes(requested_tickers)
        except Exception:  # noqa: BLE001
            LOGGER.exception("Unable to load realtime investment quotes")
            failures = [
                {
                    "ticker": ticker,
                    "error": "Realtime quote data is temporarily unavailable.",
                }
                for ticker in requested_tickers
            ]

        # Always return 200 with whatever we got (partial is common for large ticker sets
        # or slow symbols). The 502 was causing visible errors and making import feel stuck.
        # Failures are reported to the caller for diagnostics.
        response = jsonify({
            "success": bool(quotes) or (len(requested_tickers) == 0),
            "quotes": quotes,
            "failures": failures,
            "count": len(quotes),
            "source": (
                "mixed" if len({str(item.get("source") or "") for item in quotes}) > 1
                else str(quotes[0].get("source") or "yfinance") if quotes
                else "yfinance"
            ),
            "fetched_at": fetched_at.strftime("%Y-%m-%d %H:%M:%S%z"),
        })
        response.status_code = 200
        return apply_no_store_headers(response)

    def investment_get_market_session():
        """Get US-equity market session state for frontend-safe gating of realtime refresh."""
        try:
            reference = request.args.get("as_of")
            try:
                requested_day_count = int(request.args.get("day_count", "5"))
            except (TypeError, ValueError):
                requested_day_count = 5
            requested_day_count = max(1, min(365, requested_day_count))
            session_state = nyse_market_session_state(
                reference if reference else None,
                include_overnight=True,
            )
            trading_days = nyse_recent_trading_days(
                reference if reference else None,
                day_count=requested_day_count,
            )
            response = jsonify({"success": True, "trading_days": trading_days, **session_state})
            response.status_code = 200
            return apply_no_store_headers(response)
        except Exception:  # noqa: BLE001
            LOGGER.exception("Unable to load US-equity market session state")
            response = jsonify({
                "success": False,
                "error": "US-equity market session information is temporarily unavailable.",
                "market": "us_equity",
                "is_trading_day": False,
                "is_early_close": False,
                "session": "off",
                "session_date": "",
                "as_of": pd.Timestamp.now(tz="America/New_York").isoformat(),
                "timezone": "America/New_York",
                "is_realtime_allowed": False,
                "overnight_open": "20:00",
                "overnight_close": "04:00",
                "premarket_open": "04:00",
                "regular_open": "09:30",
                "regular_close": "16:00",
                "postmarket_close": "20:00",
                "next_session_open": "",
                "next_session_close": "",
                "trading_days": [],
            })
            response.status_code = 500
            return apply_no_store_headers(response)

    def live_trading_get_positions():
        """Load current Longbridge stock positions for the Live trading workspace."""
        authorization_failure = live_trading_api_authorization_failure_response()
        if authorization_failure is not None:
            return authorization_failure
        try:
            settings = load_broker_settings()
            account_balances = load_longbridge_account_balances(settings)
            positions = load_longbridge_stock_positions(settings)
            response = jsonify({
                "success": True,
                "account_balances": [
                    {
                        "total_cash": item.total_cash,
                        "max_finance_amount": item.max_finance_amount,
                        "remaining_finance_amount": item.remaining_finance_amount,
                        "risk_level": item.risk_level,
                        "margin_call": item.margin_call,
                        "currency": item.currency,
                        "market": item.market,
                        "net_assets": item.net_assets,
                        "init_margin": item.init_margin,
                        "maintenance_margin": item.maintenance_margin,
                        "buy_power": item.buy_power,
                        "cash_infos": [
                            {
                                "withdraw_cash": cash_item.withdraw_cash,
                                "available_cash": cash_item.available_cash,
                                "frozen_cash": cash_item.frozen_cash,
                                "settling_cash": cash_item.settling_cash,
                                "currency": cash_item.currency,
                            }
                            for cash_item in item.cash_infos
                        ],
                        "frozen_transaction_fees": [
                            {
                                "currency": fee_item.currency,
                                "frozen_transaction_fee": fee_item.frozen_transaction_fee,
                            }
                            for fee_item in item.frozen_transaction_fees
                        ],
                    }
                    for item in account_balances
                ],
                "positions": [
                    {
                        "symbol": item.symbol,
                        "symbol_name": item.symbol_name,
                        "quantity": item.quantity,
                        "available_quantity": item.available_quantity,
                        "cost_price": item.cost_price,
                        "currency": item.currency,
                        "market": item.market,
                        "account_channel": item.account_channel,
                    }
                    for item in positions
                ],
            })
            return apply_no_store_headers(response)
        except ValueError as exc:
            response = jsonify({"success": False, "error": str(exc)})
            response.status_code = 400
            return apply_no_store_headers(response)
        except Exception:  # noqa: BLE001
            LOGGER.exception("Unable to load live trading account data")
            response = jsonify({
                "success": False,
                "error": "Live trading account data is temporarily unavailable. Try again later.",
            })
            response.status_code = 500
            return apply_no_store_headers(response)

    def live_trading_submit_order():
        """Submit a Longbridge live limit order from the Live trading workspace."""
        authorization_failure = live_trading_api_authorization_failure_response()
        if authorization_failure is not None:
            return authorization_failure
        payload = request.get_json(silent=True) or {}
        try:
            order = submit_longbridge_limit_order(
                load_broker_settings(),
                ticker=str(payload.get("ticker", "")).strip(),
                side=str(payload.get("side", "")).strip(),
                price=str(payload.get("price", "")).strip(),
                quantity=str(payload.get("quantity", "")).strip(),
                remark=str(payload.get("remark", "")).strip(),
            )
            response = jsonify({
                "success": True,
                "message": f"{order.side.title()} order submitted for {order.symbol}.",
                "order": {
                    "order_id": order.order_id,
                    "symbol": order.symbol,
                    "side": order.side,
                    "price": order.price,
                    "quantity": order.quantity,
                    "order_type": order.order_type,
                    "time_in_force": order.time_in_force,
                    "status": order.status,
                    "remark": order.remark,
                },
            })
            return apply_no_store_headers(response)
        except ValueError as exc:
            response = jsonify({"success": False, "error": str(exc)})
            response.status_code = 400
            return apply_no_store_headers(response)
        except Exception:  # noqa: BLE001
            LOGGER.exception("Unable to submit live trading order")
            response = jsonify({
                "success": False,
                "error": "The live order could not be submitted. Try again later.",
            })
            response.status_code = 500
            return apply_no_store_headers(response)

    return WebRuntime(
        root=root,
        compare_page=compare_page,
        market_cap_compare_page=market_cap_compare_page,
        legacy_compare_page=legacy_compare_page,
        price_compare_page=price_compare_page,
        portfolio_page=portfolio_page,
        legacy_portfolio_page=legacy_portfolio_page,
        dca_page=dca_page,
        legacy_dca_page=legacy_dca_page,
        backtest_page=backtest_page,
        grid_trading_page=grid_trading_page,
        legacy_backtest_page=legacy_backtest_page,
        legacy_trade_messages_page=legacy_trade_messages_page,
        trade_root=trade_root,
        trade_page=trade_page,
        legacy_trade_root=legacy_trade_root,
        legacy_trade_page=legacy_trade_page,
        live_trading_unlock=live_trading_unlock,
        settings_root=settings_root,
        settings_page=settings_page,
        export_transactions_api=export_transactions_api,
        general_settings_action=general_settings_action,
        language_settings_api=language_settings_api,
        language_cycle_api=language_cycle_api,
        language_download_api=language_download_api,
        backtest_settings_action=backtest_settings_action,
        investment_settings_action=investment_settings_action,
        cash_equivalents_action=cash_equivalents_action,
        email_smtp_action=email_smtp_action,
        broker_access_action=broker_access_action,
        longbridge_oauth_status_api=longbridge_oauth_status_api,
        local_market_store_action=local_market_store_action,
        settings_cache_action=settings_cache_action,
        market_store_logo=market_store_logo,
        favicon_icon=favicon_icon,
        symbol_search=symbol_search,
        date_constraints_api=date_constraints_api,
        compare_live_api=compare_live_api,
        trade_strategy_fields_api=trade_strategy_fields_api,
        settings_network_status_api=settings_network_status_api,
        local_market_store_page_data_api=local_market_store_page_data_api,
        market_store_presence_api=market_store_presence_api,
        investment_page=investment_page,
        investment_get_transactions=investment_get_transactions,
        investment_add_transaction=investment_add_transactions,
        investment_download_zircon_hk_template=investment_download_zircon_hk_template,
        investment_export_standard_xlsx=investment_export_standard_xlsx,
        investment_validate_zircon_hk_workbook=investment_validate_zircon_hk_workbook,
        investment_validate_hsbc_pasted_text=investment_validate_hsbc_pasted_text,
        investment_get_latest_price=investment_get_latest_price,
        investment_get_parquet=investment_get_parquet,
        investment_get_intraday_history=investment_get_intraday_history,
        investment_get_realtime_quotes=investment_get_realtime_quotes,
        investment_get_market_session=investment_get_market_session,
        investment_update_internal_transfer_binding=investment_update_internal_transfer_binding,
        investment_update_security_transfer_attribution=(
            investment_update_security_transfer_attribution
        ),
        live_trading_get_positions=live_trading_get_positions,
        live_trading_submit_order=live_trading_submit_order,
    )
