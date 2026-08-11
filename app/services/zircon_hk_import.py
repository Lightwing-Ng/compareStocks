"""
Generic manual investment XLSX template and import parser.

Code version: v0.12.0
"""

from __future__ import annotations

import base64
import hashlib
import json
import math
import re
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any
from zipfile import BadZipFile, ZipFile
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

from app.core.broker_catalog import BROKER_CATALOG, sorted_broker_entries
from app.infrastructure.storage import normalize_ticker
from app.services.investment_record_basics import (
    build_normalized_transaction_view,
    decimal_to_str,
    normalize_import_text,
    normalize_import_whitespace,
)


ZIRCON_HK_IMPORTER_VERSION = "0.9.0"
ZIRCON_HK_BROKER_CODE = "zircon_hk"
ZIRCON_HK_BROKER_LABEL = "Zircon (HK)"
ZIRCON_HK_TEMPLATE_FILENAME = "Manual_investment_import.xlsx"
STANDARD_INVESTMENT_EXPORT_FILENAME = "Standard_investment_export.xlsx"
ZIRCON_HK_TRANSACTION_SHEET = "Transactions"
ZIRCON_HK_LISTS_SHEET = "Lists"
ZIRCON_HK_MAX_TRANSACTION_ROWS = 10_000
ZIRCON_HK_TEMPLATE_INPUT_ROWS = 2_000
ZIRCON_HK_MAX_REPORTED_ERRORS = 100
ZIRCON_HK_MAX_XLSX_BYTES = 16 * 1024 * 1024
ZIRCON_HK_MAX_ARCHIVE_ENTRIES = 256
ZIRCON_HK_MAX_UNCOMPRESSED_BYTES = 32 * 1024 * 1024
ZIRCON_HK_TIMEZONE = ZoneInfo("Asia/Hong_Kong")
INVESTMENT_LEDGER_TIMEZONE = ZoneInfo("America/New_York")

ZIRCON_HK_HEADERS = (
    "Broker",
    "Account",
    "Transaction Date / Time (Hong Kong)",
    "Type",
    "Currency",
    "Ticker",
    "Quantity",
    "Trade Price",
    "Amount",
    "Commission",
    "Description",
    "Reference ID",
)

ZIRCON_HK_TYPE_LABELS: dict[str, str] = {
    "Buy": "buy",
    "Sell": "sell",
    "Deposit": "deposit",
    "Withdrawal": "withdrawal",
    "Virtual deposit": "virtual_deposit",
    "Virtual withdrawal": "virtual_withdrawal",
    "Virtual balance reset": "virtual_balance_reset",
    "Dividend": "dividend",
    "Dividend reinvestment": "dividend_reinvestment",
    "Fee": "fee",
    "Credit interest": "credit_interest",
    "Debit interest": "debit_interest",
    "Foreign tax withholding": "foreign_tax_withholding",
    "Payment in lieu": "payment_in_lieu",
    "Adjustment": "adjustment",
    "Grant": "grant",
    "KOL reward": "kol_reward",
    "Forex trade component": "forex_trade_component",
    "FX translation P&L": "fx_translation_pnl",
    "Transfer In": "transfer_in",
    "Transfer Out": "transfer_out",
}
ZIRCON_HK_CURRENCIES = ("HKD", "USD", "CNH", "CNY", "SGD")
ZIRCON_HK_BROKER_ENTRIES = tuple(
    sorted_broker_entries(
        tuple(code for code in BROKER_CATALOG if code != "standard_xlsx")
    )
)

_TYPE_LOOKUP = {
    **{label.casefold(): value for label, value in ZIRCON_HK_TYPE_LABELS.items()},
    **{value.casefold(): value for value in ZIRCON_HK_TYPE_LABELS.values()},
}
_BROKER_LOOKUP = {
    alias.casefold(): entry.code
    for entry in ZIRCON_HK_BROKER_ENTRIES
    for alias in (entry.code, entry.label)
}
_BROKER_LOOKUP["zircon hk".casefold()] = ZIRCON_HK_BROKER_CODE
_SECURITY_TYPES = {
    "buy",
    "sell",
    "dividend",
    "dividend_reinvestment",
    "foreign_tax_withholding",
    "grant",
    "payment_in_lieu",
    "transfer_in",
    "transfer_out",
}
_QUANTITY_TYPES = {
    "buy",
    "sell",
    "dividend_reinvestment",
    "grant",
    "transfer_in",
    "transfer_out",
}
_PRICE_TYPES = {"buy", "sell", "dividend_reinvestment"}
_CASHLESS_SECURITY_TRANSFER_TYPES = {"transfer_in", "transfer_out"}
_IBKR_STATEMENT_BASE_CURRENCY = "USD"
_IBKR_BASE_CURRENCY_CASH_TYPES = {
    "deposit",
    "withdrawal",
    "virtual_deposit",
    "virtual_withdrawal",
}
_AMOUNTLESS_SECURITY_TYPES = {
    "buy",
    "sell",
    "dividend_reinvestment",
    "grant",
    *_CASHLESS_SECURITY_TRANSFER_TYPES,
}
_POSITIVE_CASH_TYPES = {
    "deposit",
    "virtual_deposit",
    "dividend",
    "credit_interest",
    "payment_in_lieu",
    "kol_reward",
}
_NEGATIVE_CASH_TYPES = {
    "withdrawal",
    "virtual_withdrawal",
    "virtual_balance_reset",
    "fee",
    "debit_interest",
    "foreign_tax_withholding",
}
_REQUIRED_CASH_TYPES = (
    _POSITIVE_CASH_TYPES
    | _NEGATIVE_CASH_TYPES
    | {"forex_trade_component", "fx_translation_pnl"}
)
_FORMULA_PREFIXES = ("=",)

_HEADER_FILL = "E7E6E6"
_HEADER_TEXT = "000000"
_REQUIRED_XLSX_MEMBERS = {
    "[Content_Types].xml",
    "xl/workbook.xml",
}
_FORBIDDEN_XLSX_MEMBER_PREFIXES = (
    "xl/embeddings/",
    "xl/externalLinks/",
    "xl/oleObjects/",
)
_FORBIDDEN_XLSX_MEMBER_NAMES = {
    "xl/connections.xml",
    "xl/vbaProject.bin",
}

_STANDARD_REFERENCE_SOURCE_FIELDS = (
    "file_kind",
    "source_filename",
    "source_file_sha256",
    "source_sheet",
    "source_row",
    "row_number",
    "file_index",
    "fitid",
    "secid",
    "fiid",
    "statement_order_id",
    "history_order_id",
    "history_order_row_number",
    "fund_details_entry_number",
    "cash_flow_contract_row_number",
)

_STANDARD_HEADER_COMMENTS = {
    "Broker": (
        "Choose the institution that produced this activity. The selected broker "
        "is authoritative for this row."
    ),
    "Account": (
        "Enter the broker account label or number exactly as it appears in the "
        "source statement."
    ),
    "Transaction Date / Time (Hong Kong)": (
        "Enter a native Excel date or date-time in Asia/Hong_Kong time. A date "
        "without a time defaults to 23:00 Hong Kong time."
    ),
    "Type": (
        "For Buy, Sell, and Dividend reinvestment, leave Amount blank: it is "
        "derived from Quantity, Trade Price, and Commission."
    ),
    "Currency": "Choose the currency of the row's Amount or trade price.",
    "Ticker": (
        "Use the broker or provider symbol. Security movements normally require "
        "a valid Ticker."
    ),
    "Quantity": (
        "Enter a positive native Excel number when the selected Type requires a "
        "position quantity."
    ),
    "Trade Price": (
        "Enter a positive native Excel number for Buy, Sell, or Dividend "
        "reinvestment."
    ),
    "Amount": (
        "Use a signed native Excel number for non-trade activity: cash received "
        "is positive and cash paid is negative. Adjustment may be zero when the "
        "Description documents a non-cash event."
    ),
    "Commission": (
        "Enter the non-negative fee magnitude. The imported ledger stores it as "
        "a negative commission."
    ),
    "Description": (
        "Keep the broker wording or explain a manual/non-cash event. This field "
        "is retained in the ledger and is limited to 500 characters."
    ),
    "Reference ID": (
        "Use a unique source identifier per broker/account. FX requires exactly "
        "two rows sharing one Reference ID."
    ),
}


def _validate_xlsx_archive(xlsx_bytes: bytes) -> None:
    if len(xlsx_bytes) > ZIRCON_HK_MAX_XLSX_BYTES:
        raise ValueError(
            "The manual investment workbook exceeds the 16 MiB XLSX file limit."
        )
    try:
        with ZipFile(BytesIO(xlsx_bytes)) as archive:
            members = archive.infolist()
            names = {member.filename for member in members}
            if len(members) > ZIRCON_HK_MAX_ARCHIVE_ENTRIES:
                raise ValueError(
                    "The manual investment workbook contains too many archive entries."
                )
            if not _REQUIRED_XLSX_MEMBERS.issubset(names):
                raise ValueError(
                    "The manual investment workbook is missing required XLSX components."
                )
            uncompressed_total = 0
            for member in members:
                normalized_name = member.filename.replace("\\", "/")
                name_parts = normalized_name.split("/")
                if (
                    normalized_name.startswith("/")
                    or ".." in name_parts
                    or member.flag_bits & 0x1
                ):
                    raise ValueError(
                        "The manual investment workbook contains an unsafe archive entry."
                    )
                if (
                    normalized_name in _FORBIDDEN_XLSX_MEMBER_NAMES
                    or normalized_name.startswith(
                        _FORBIDDEN_XLSX_MEMBER_PREFIXES
                    )
                ):
                    raise ValueError(
                        "The manual investment workbook must not contain macros, embedded "
                        "objects, data connections, or external workbook links."
                    )
                uncompressed_total += member.file_size
                if uncompressed_total > ZIRCON_HK_MAX_UNCOMPRESSED_BYTES:
                    raise ValueError(
                        "The manual investment workbook expands beyond the 32 MiB "
                        "safe parsing limit."
                    )
                if (
                    member.file_size > 1024 * 1024
                    and member.file_size
                    / max(member.compress_size, 1)
                    > 200
                ):
                    raise ValueError(
                        "The manual investment workbook contains a suspiciously compressed "
                        "archive entry."
                    )
    except BadZipFile as exc:
        raise ValueError(
            "The manual investment workbook is not a readable XLSX file."
        ) from exc


def _style_template_workbook(workbook: Workbook) -> None:
    workbook.properties.title = "Manual investment import"
    workbook.properties.subject = "Validated fallback broker transaction import"
    workbook.properties.creator = "antigravity"
    workbook.properties.description = (
        "Complete the Transactions sheet with real broker activity, then upload it "
        "through Trade > Investment. Date-only entries default to 23:00 Hong Kong "
        "time. Currency conversions use two Forex trade component rows with one "
        "shared Reference ID."
    )


def _add_defined_name(workbook: Workbook, name: str, reference: str) -> None:
    workbook.defined_names.add(DefinedName(name, attr_text=reference))


def _configure_lists_sheet(sheet: Any) -> None:
    sheet.sheet_state = "hidden"
    sheet["A1"] = "Transaction Types"
    for row, label in enumerate(ZIRCON_HK_TYPE_LABELS, start=2):
        sheet.cell(row=row, column=1, value=label)
    sheet["B1"] = "Currencies"
    for row, currency in enumerate(ZIRCON_HK_CURRENCIES, start=2):
        sheet.cell(row=row, column=2, value=currency)
    sheet["C1"] = "Brokers"
    for row, entry in enumerate(ZIRCON_HK_BROKER_ENTRIES, start=2):
        sheet.cell(row=row, column=3, value=entry.label)


def _add_list_validation(
    sheet: Any,
    cell_range: str,
    *,
    formula: str,
    title: str,
    message: str,
) -> None:
    validation = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=False,
        errorStyle="stop",
        errorTitle=title,
        error=message,
        showErrorMessage=True,
        showInputMessage=True,
        promptTitle=title,
        prompt=message,
    )
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def _add_decimal_validation(
    sheet: Any,
    cell_range: str,
    *,
    minimum: str,
    maximum: str,
    title: str,
    message: str,
) -> None:
    validation = DataValidation(
        type="decimal",
        operator="between",
        formula1=minimum,
        formula2=maximum,
        allow_blank=True,
        errorStyle="stop",
        errorTitle=title,
        error=message,
        showErrorMessage=True,
        showInputMessage=True,
        promptTitle=title,
        prompt=message,
    )
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def _configure_transactions_sheet(
    sheet: Any,
    *,
    input_rows: int = ZIRCON_HK_TEMPLATE_INPUT_ROWS,
) -> None:
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    for column, header in enumerate(ZIRCON_HK_HEADERS, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
        cell.font = Font(color=_HEADER_TEXT, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        comment_text = _STANDARD_HEADER_COMMENTS.get(header)
        if comment_text:
            cell.comment = Comment(comment_text, "antigravity")
    sheet.row_dimensions[1].height = 30

    widths = (20, 18, 29, 26, 12, 18, 14, 14, 16, 14, 40, 22)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for row in range(2, input_rows + 2):
        sheet.cell(row=row, column=3).number_format = "yyyy-mm-dd hh:mm"
        for column in range(7, 11):
            sheet.cell(row=row, column=column).number_format = (
                '#,##0.########;[Red]-#,##0.########;-'
            )

    _add_list_validation(
        sheet,
        f"A2:A{input_rows + 1}",
        formula="ZirconBrokers",
        title="Choose broker",
        message="Select the broker that produced this transaction.",
    )
    _add_list_validation(
        sheet,
        f"D2:D{input_rows + 1}",
        formula="ZirconTypes",
        title="Choose transaction type",
        message=(
            "Select one supported transaction type. A currency conversion uses "
            "Forex trade component on both signed currency legs."
        ),
    )
    _add_list_validation(
        sheet,
        f"E2:E{input_rows + 1}",
        formula="ZirconCurrencies",
        title="Choose currency",
        message="Select one supported currency.",
    )

    date_validation = DataValidation(
        type="date",
        operator="between",
        formula1="DATE(2000,1,1)",
        formula2="DATE(2100,12,31)",
        allow_blank=False,
        errorStyle="stop",
        errorTitle="Enter a real date-time",
        error="Use an Excel date or date-time in Hong Kong time, not text.",
        showErrorMessage=True,
        showInputMessage=True,
        promptTitle="Hong Kong transaction time",
        prompt=(
            "Enter an Excel date or date-time. A date without a time defaults "
            "to 23:00 Hong Kong time."
        ),
    )
    sheet.add_data_validation(date_validation)
    date_validation.add(f"C2:C{input_rows + 1}")

    _add_decimal_validation(
        sheet,
        f"G2:G{input_rows + 1}",
        minimum="0",
        maximum="1000000000000",
        title="Enter a numeric quantity",
        message="Quantity must be a non-negative numeric cell.",
    )
    _add_decimal_validation(
        sheet,
        f"H2:H{input_rows + 1}",
        minimum="0",
        maximum="1000000000000",
        title="Enter a numeric trade price",
        message="Trade Price must be a non-negative numeric cell.",
    )
    _add_decimal_validation(
        sheet,
        f"I2:I{input_rows + 1}",
        minimum="-1000000000000",
        maximum="1000000000000",
        title="Enter a signed amount",
        message=(
            "Amount is only for non-trade activity; cash received is positive "
            "and cash paid is negative. For FX, enter the sold leg as negative "
            "and the acquired leg as positive."
        ),
    )
    _add_decimal_validation(
        sheet,
        f"J2:J{input_rows + 1}",
        minimum="0",
        maximum="1000000000000",
        title="Enter a numeric commission",
        message="Commission must be a non-negative numeric fee magnitude.",
    )


def build_zircon_hk_template_xlsx() -> bytes:
    """Return a plain fallback XLSX template with typed validation rules."""
    return build_standard_investment_xlsx()


def _export_decimal(value: Any) -> int | float | None:
    if value is None or value == "":
        return None
    try:
        parsed = Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError) as exc:
        raise ValueError(f"Investment export contains an invalid numeric value: {value!r}.") from exc
    if not parsed.is_finite():
        raise ValueError("Investment export contains a non-finite numeric value.")
    integral = parsed.to_integral_value()
    if parsed == integral:
        return int(integral)
    return float(parsed)


def _export_transaction_datetime(transaction: dict[str, Any]) -> datetime:
    raw_datetime = normalize_import_text(transaction.get("datetime"))
    if not raw_datetime:
        raw_datetime = normalize_import_text(transaction.get("date"))
    try:
        parsed = datetime.fromisoformat(raw_datetime)
    except ValueError as exc:
        ledger_no = transaction.get("ledger_no")
        raise ValueError(
            f"Investment transaction {ledger_no or 'without a ledger number'} "
            "has no exportable date-time."
        ) from exc
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=INVESTMENT_LEDGER_TIMEZONE)
    return parsed.astimezone(ZIRCON_HK_TIMEZONE).replace(tzinfo=None, microsecond=0)


def _resolve_standard_export_currency(
    transaction: dict[str, Any],
    *,
    broker_code: str,
    transaction_type: str,
) -> str:
    currency = normalize_import_text(transaction.get("currency")).upper()
    if transaction_type == "forex_trade_component":
        description = normalize_import_whitespace(transaction.get("description")).upper()
        match = re.search(r"FX FROM\s+([A-Z]{3})\s+TO\s+([A-Z]{3})", description)
        if match:
            normalized = transaction.get("normalized") or {}
            amount = _export_decimal_from_fields(
                transaction,
                normalized,
                "net_amount_raw",
                "net_amount",
                "amount",
                "gross_amount_raw",
                "gross_amount",
            )
            if amount is not None:
                return (match.group(1) if amount < 0 else match.group(2))
    if currency in ZIRCON_HK_CURRENCIES:
        return currency

    source = transaction.get("source") or {}
    source_broker = normalize_import_text(source.get("broker")).casefold()
    source_kind = normalize_import_text(source.get("file_kind")).casefold()
    ticker = normalize_import_text(transaction.get("ticker"))
    if (
        not currency
        and broker_code == "ibkr"
        and source_broker == "ibkr"
        and source_kind == "transactions"
        and transaction_type in _IBKR_BASE_CURRENCY_CASH_TYPES
        and not ticker
    ):
        return _IBKR_STATEMENT_BASE_CURRENCY

    ledger_no = transaction.get("ledger_no")
    raise ValueError(
        f"Investment transaction {ledger_no or 'without a ledger number'} "
        f"uses unsupported standard XLSX currency {currency!r}."
    )


def _export_decimal_from_fields(
    transaction: dict[str, Any],
    normalized: dict[str, Any],
    *field_names: str,
) -> int | float | None:
    for container in (transaction, normalized):
        for field_name in field_names:
            value = _export_decimal(container.get(field_name))
            if value is not None:
                return value
    return None


def _standard_export_ticker(
    transaction: dict[str, Any],
    *,
    transaction_type: str,
) -> str:
    source = transaction.get("source") or {}
    for value in (
        transaction.get("ticker"),
        source.get("source_symbol_raw"),
        source.get("symbol"),
    ):
        ticker = normalize_import_text(value).upper()
        if ticker:
            return ticker

    if transaction_type in {"dividend", "foreign_tax_withholding", "payment_in_lieu"}:
        description = normalize_import_whitespace(transaction.get("description")).upper()
        match = re.match(
            r"^([A-Z][A-Z0-9._-]{0,15})\s+\d[\d,.]*\s+SHARES\s+"
            r"(?:DIVIDENDS?|WITHHOLDING|PAYMENT)",
            description,
        )
        if match:
            inferred_ticker = normalize_ticker(match.group(1))
            if inferred_ticker:
                return inferred_ticker
    return ""


def _stable_standard_reference_id(
    transaction: dict[str, Any],
    *,
    broker_code: str,
    transaction_type: str,
    currency: str,
    quantity: int | float | None,
    price: int | float | None,
    economic_amount: int | float | None,
    commission: int | float | None,
    ticker: str,
) -> str:
    source = transaction.get("source") or {}
    source_identity = {
        field_name: source.get(field_name)
        for field_name in _STANDARD_REFERENCE_SOURCE_FIELDS
        if source.get(field_name) not in {None, ""}
    }
    execution_key = normalize_import_text(source.get("execution_key"))
    if transaction_type == "forex_trade_component" and execution_key:
        for currency_code in ZIRCON_HK_CURRENCIES:
            currency_suffix = f":{currency_code.casefold()}"
            if execution_key.casefold().endswith(currency_suffix):
                execution_key = execution_key[: -len(currency_suffix)]
                break
        source_identity["execution_key"] = execution_key
        source_identity.pop("source_row", None)
        source_identity.pop("row_number", None)
    fingerprint = {
        "broker": broker_code,
        "account": normalize_import_text(
            transaction.get("account")
            or source.get("account")
            or source.get("account_number")
        ),
        "datetime": normalize_import_text(
            transaction.get("datetime") or transaction.get("date")
        ),
        "type": transaction_type,
        "currency": "" if execution_key and transaction_type == "forex_trade_component" else currency,
        "ticker": "" if execution_key and transaction_type == "forex_trade_component" else ticker,
        "quantity": None if execution_key and transaction_type == "forex_trade_component" else quantity,
        "price": None if execution_key and transaction_type == "forex_trade_component" else price,
        "amount": None if execution_key and transaction_type == "forex_trade_component" else economic_amount,
        "commission": None if execution_key and transaction_type == "forex_trade_component" else commission,
        "description": "" if execution_key and transaction_type == "forex_trade_component" else normalize_import_whitespace(transaction.get("description")),
        "source": source_identity,
    }
    encoded = json.dumps(
        fingerprint,
        ensure_ascii=True,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    ).encode("utf-8")
    digest = hashlib.sha256(encoded).hexdigest()[:40]
    return f"antigravity-{digest}"


def _standard_export_fallback(
    transaction: dict[str, Any],
    *,
    transaction_type: str,
    type_label: str,
    ticker: str,
    quantity: int | float | None,
    price: int | float | None,
    economic_amount: int | float | None,
) -> tuple[str, int | float | None, list[str]]:
    source = transaction.get("source") or {}
    reasons: list[str] = []
    if transaction_type in _SECURITY_TYPES and not ticker:
        reasons.append("Ticker is unavailable")
    if transaction_type in _QUANTITY_TYPES and quantity is None:
        reasons.append("Quantity is unavailable")
    if transaction_type in _PRICE_TYPES and price is None:
        reasons.append("Trade Price is unavailable")

    if transaction_type in _POSITIVE_CASH_TYPES and (
        economic_amount is None or economic_amount <= 0
    ):
        reasons.append("the signed cash amount is a reversal or non-standard positive flow")
    if transaction_type in _NEGATIVE_CASH_TYPES and (
        economic_amount is None or economic_amount >= 0
    ):
        reasons.append("the signed cash amount is a reversal or non-standard negative flow")
    if (
        transaction_type == "forex_trade_component"
        and not normalize_import_text(source.get("reference_id"))
        and not normalize_import_text(source.get("execution_key"))
    ):
        reasons.append("the source does not provide a stable two-leg FX identity")

    if not reasons:
        amount = (
            None
            if transaction_type in _AMOUNTLESS_SECURITY_TYPES
            else economic_amount
        )
        return transaction_type, amount, reasons

    if transaction_type in _QUANTITY_TYPES and (
        economic_amount is None or economic_amount == 0
    ):
        ledger_no = transaction.get("ledger_no")
        raise ValueError(
            f"Investment transaction {ledger_no or 'without a ledger number'} "
            f"({type_label}) cannot be exported safely: {', '.join(reasons)} "
            "and no signed cash Amount is available for a reversible Adjustment row."
        )
    if economic_amount is None:
        ledger_no = transaction.get("ledger_no")
        raise ValueError(
            f"Investment transaction {ledger_no or 'without a ledger number'} "
            f"({type_label}) cannot be exported safely: {', '.join(reasons)} "
            "and no signed cash Amount is available."
        )
    return "adjustment", economic_amount, reasons


def _standard_export_row(
    transaction: dict[str, Any],
    *,
    reference_id_override: str | None = None,
) -> tuple[Any, ...]:
    transaction_type = normalize_import_text(transaction.get("type")).casefold()
    type_label_by_value = {
        value: label for label, value in ZIRCON_HK_TYPE_LABELS.items()
    }
    type_label = type_label_by_value.get(transaction_type)
    ledger_no = transaction.get("ledger_no")
    if type_label is None:
        raise ValueError(
            f"Investment transaction {ledger_no or 'without a ledger number'} "
            f"uses unsupported standard XLSX type {transaction_type!r}."
        )

    broker_code = normalize_import_text(
        transaction.get("broker")
        or (transaction.get("source") or {}).get("broker")
    ).casefold()
    broker_entry = BROKER_CATALOG.get(broker_code)
    if broker_entry is None:
        raise ValueError(
            f"Investment transaction {ledger_no or 'without a ledger number'} "
            f"uses unsupported broker {broker_code!r}."
        )

    currency = _resolve_standard_export_currency(
        transaction,
        broker_code=broker_code,
        transaction_type=transaction_type,
    )

    normalized = transaction.get("normalized") or {}
    quantity = _export_decimal_from_fields(
        transaction,
        normalized,
        "quantity_raw",
        "position_quantity",
        "quantity_abs",
        "quantity",
    )
    if quantity is not None:
        quantity = abs(quantity)

    price = _export_decimal_from_fields(
        transaction,
        normalized,
        "price_raw",
        "unit_price",
        "price",
    )
    if price is not None:
        price = abs(price)

    economic_amount = _export_decimal_from_fields(
        transaction,
        normalized,
        "net_amount_raw",
        "net_amount",
        "amount",
        "gross_amount_raw",
        "gross_amount",
    )

    commission = _export_decimal_from_fields(
        transaction,
        normalized,
        "commission_raw",
        "commission",
    )
    if commission is not None:
        commission = abs(commission)

    ticker = _standard_export_ticker(
        transaction,
        transaction_type=transaction_type,
    )
    export_type, amount, fallback_reasons = _standard_export_fallback(
        transaction,
        transaction_type=transaction_type,
        type_label=type_label,
        ticker=ticker,
        quantity=quantity,
        price=price,
        economic_amount=economic_amount,
    )
    export_type_label = type_label_by_value[export_type]

    source = transaction.get("source") or {}
    reference_id = normalize_import_text(
        reference_id_override or source.get("reference_id")
    )
    if not reference_id:
        reference_id = _stable_standard_reference_id(
            transaction,
            broker_code=broker_code,
            transaction_type=transaction_type,
            currency=currency,
            quantity=quantity,
            price=price,
            economic_amount=economic_amount,
            commission=commission,
            ticker=ticker,
        )

    description = normalize_import_whitespace(transaction.get("description"))
    if fallback_reasons:
        description = normalize_import_whitespace(
            f"[Standard XLSX fallback: original Type {type_label!r} exported as "
            f"{export_type_label!r}; {'; '.join(fallback_reasons)}.] {description}"
        )

    return (
        broker_entry.label,
        normalize_import_text(
            transaction.get("account")
            or source.get("account")
            or source.get("account_number")
        ),
        _export_transaction_datetime(transaction),
        export_type_label,
        currency,
        ticker,
        quantity,
        price,
        amount,
        commission,
        description[:500],
        reference_id[:100],
    )


def build_standard_investment_xlsx(
    transactions: list[dict[str, Any]] | tuple[dict[str, Any], ...] = (),
) -> bytes:
    """Build the neutral standard workbook, optionally populated for round-trip import."""
    workbook = Workbook()
    transaction_sheet = workbook.active
    transaction_sheet.title = ZIRCON_HK_TRANSACTION_SHEET
    lists = workbook.create_sheet(ZIRCON_HK_LISTS_SHEET)
    _style_template_workbook(workbook)
    _configure_lists_sheet(lists)
    _add_defined_name(
        workbook,
        "ZirconTypes",
        f"'{ZIRCON_HK_LISTS_SHEET}'!$A$2:$A${len(ZIRCON_HK_TYPE_LABELS) + 1}",
    )
    _add_defined_name(
        workbook,
        "ZirconCurrencies",
        f"'{ZIRCON_HK_LISTS_SHEET}'!$B$2:$B${len(ZIRCON_HK_CURRENCIES) + 1}",
    )
    _add_defined_name(
        workbook,
        "ZirconBrokers",
        (
            f"'{ZIRCON_HK_LISTS_SHEET}'!$C$2:"
            f"$C${len(ZIRCON_HK_BROKER_ENTRIES) + 1}"
        ),
    )
    _configure_transactions_sheet(
        transaction_sheet,
        input_rows=max(ZIRCON_HK_TEMPLATE_INPUT_ROWS, len(transactions)),
    )
    reference_occurrences: dict[tuple[str, str, str], int] = {}
    reference_types: dict[tuple[str, str, str], set[str]] = {}
    for row_number, transaction in enumerate(transactions, start=2):
        exported_row = _standard_export_row(transaction)
        reference_key = (exported_row[0], exported_row[1], exported_row[11])
        occurrence = reference_occurrences.get(reference_key, 0) + 1
        reference_occurrences[reference_key] = occurrence
        existing_types = reference_types.setdefault(reference_key, set())
        if occurrence > 1 and (
            exported_row[3] != "Forex trade component"
            or existing_types != {"Forex trade component"}
        ):
            collision_payload = json.dumps(
                {
                    "row": exported_row[:11],
                    "occurrence": occurrence,
                },
                default=str,
                sort_keys=True,
                separators=(",", ":"),
            ).encode("utf-8")
            collision_suffix = hashlib.sha256(collision_payload).hexdigest()[:16]
            exported_row = _standard_export_row(
                transaction,
                reference_id_override=(
                    f"{exported_row[11]}::antigravity-{collision_suffix}"
                ),
            )
        existing_types.add(exported_row[3])
        for column_number, value in enumerate(exported_row, start=1):
            transaction_sheet.cell(
                row=row_number,
                column=column_number,
                value=value,
            )
    workbook.active = 0
    buffer = BytesIO()
    workbook.save(buffer)
    workbook_bytes = buffer.getvalue()
    if transactions:
        build_investment_payload_from_zircon_hk_manual_xlsx(
            xlsx_bytes=workbook_bytes,
            filename=STANDARD_INVESTMENT_EXPORT_FILENAME,
        )
    return workbook_bytes


def _cell_error(cell: Any, message: str) -> str:
    return f"{ZIRCON_HK_TRANSACTION_SHEET}!{cell.coordinate} ({cell.column_letter}): {message}"


def _plain_text(cell: Any, *, maximum_length: int, field_name: str) -> str:
    value = cell.value
    if value is None:
        return ""
    if cell.data_type == "f" or (
        isinstance(value, str) and value.lstrip().startswith(_FORMULA_PREFIXES)
    ):
        raise ValueError(_cell_error(cell, f"{field_name} must not contain a formula."))
    if not isinstance(value, str):
        value = str(value)
    normalized = normalize_import_whitespace(value)
    if len(normalized) > maximum_length:
        raise ValueError(
            _cell_error(
                cell,
                f"{field_name} exceeds the {maximum_length}-character limit.",
            )
        )
    return normalized


def _decimal_cell(
    cell: Any,
    *,
    field_name: str,
    required: bool = False,
    positive: bool = False,
    non_negative: bool = False,
) -> Decimal | None:
    value = cell.value
    if value is None or value == "":
        if required:
            raise ValueError(_cell_error(cell, f"{field_name} is required."))
        return None
    if cell.data_type == "f" or isinstance(value, bool) or not isinstance(value, (int, float, Decimal)):
        raise ValueError(
            _cell_error(
                cell,
                f"{field_name} must be a numeric Excel cell, not text or a formula.",
            )
        )
    if isinstance(value, float) and not math.isfinite(value):
        raise ValueError(_cell_error(cell, f"{field_name} must be finite."))
    try:
        parsed = Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError) as exc:
        raise ValueError(_cell_error(cell, f"{field_name} is not a valid number.")) from exc
    if positive and parsed <= 0:
        raise ValueError(_cell_error(cell, f"{field_name} must be greater than zero."))
    if non_negative and parsed < 0:
        raise ValueError(_cell_error(cell, f"{field_name} cannot be negative."))
    if abs(parsed) > Decimal("1000000000000"):
        raise ValueError(_cell_error(cell, f"{field_name} exceeds the supported magnitude."))
    return parsed


def _datetime_cell(cell: Any) -> tuple[datetime, bool]:
    value = cell.value
    parsed: datetime
    if cell.data_type == "f":
        raise ValueError(_cell_error(cell, "Transaction Date / Time must not contain a formula."))
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, time())
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            parsed_value = from_excel(value)
        except (TypeError, ValueError, OverflowError) as exc:
            raise ValueError(
                _cell_error(cell, "Transaction Date / Time is not a valid Excel date-time.")
            ) from exc
        parsed = (
            parsed_value
            if isinstance(parsed_value, datetime)
            else datetime.combine(parsed_value, time())
        )
    else:
        raise ValueError(
            _cell_error(
                cell,
                "Transaction Date / Time must be a typed Excel date or date-time, not text.",
            )
        )
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(ZIRCON_HK_TIMEZONE).replace(tzinfo=None)
    defaulted_time = (
        parsed.hour == 0
        and parsed.minute == 0
        and parsed.second == 0
        and parsed.microsecond == 0
    )
    if defaulted_time:
        parsed = parsed.replace(hour=23)
    if not datetime(2000, 1, 1) <= parsed <= datetime(2100, 12, 31, 23, 59, 59):
        raise ValueError(
            _cell_error(
                cell,
                "Transaction Date / Time must fall between 1 Jan 2000 and 31 Dec 2100.",
            )
        )
    return parsed.replace(microsecond=0), defaulted_time


def _normalized_broker(cell: Any) -> str:
    raw_broker = _plain_text(cell, maximum_length=64, field_name="Broker")
    normalized = _BROKER_LOOKUP.get(raw_broker.casefold())
    if normalized is None:
        allowed = ", ".join(entry.label for entry in ZIRCON_HK_BROKER_ENTRIES)
        raise ValueError(
            _cell_error(
                cell,
                f"unsupported Broker {raw_broker!r}. Choose one of: {allowed}.",
            )
        )
    return normalized


def _normalized_type(cell: Any) -> str:
    raw_type = _plain_text(cell, maximum_length=64, field_name="Type")
    normalized = _TYPE_LOOKUP.get(raw_type.casefold())
    if normalized is None:
        allowed = ", ".join(ZIRCON_HK_TYPE_LABELS)
        raise ValueError(
            _cell_error(
                cell,
                f"unsupported Type {raw_type!r}. Choose one of: {allowed}.",
            )
        )
    return normalized


def _normalized_currency(cell: Any) -> str:
    currency = _plain_text(cell, maximum_length=3, field_name="Currency").upper()
    if currency not in ZIRCON_HK_CURRENCIES:
        allowed = ", ".join(ZIRCON_HK_CURRENCIES)
        raise ValueError(
            _cell_error(cell, f"unsupported Currency {currency!r}. Choose {allowed}.")
        )
    return currency


def _transaction_from_row(
    cells: tuple[Any, ...],
    *,
    source_filename: str,
    source_sha256: str,
) -> dict[str, Any]:
    broker_cell, account_cell, datetime_cell, type_cell, currency_cell = cells[:5]
    ticker_cell, quantity_cell, price_cell, amount_cell, commission_cell = cells[5:10]
    description_cell, reference_cell = cells[10:12]
    broker = _normalized_broker(broker_cell)
    account = _plain_text(account_cell, maximum_length=64, field_name="Account")
    source_datetime, source_time_defaulted = _datetime_cell(datetime_cell)
    mapped_type = _normalized_type(type_cell)
    currency = _normalized_currency(currency_cell)
    raw_ticker = _plain_text(ticker_cell, maximum_length=32, field_name="Ticker").upper()
    ticker = normalize_ticker(raw_ticker) if raw_ticker else ""
    if mapped_type in _SECURITY_TYPES and not ticker:
        raise ValueError(_cell_error(ticker_cell, "Ticker is required for this Type."))
    if raw_ticker and not ticker:
        raise ValueError(_cell_error(ticker_cell, f"Ticker {raw_ticker!r} is invalid."))
    quantity = _decimal_cell(
        quantity_cell,
        field_name="Quantity",
        required=mapped_type in _QUANTITY_TYPES,
        positive=mapped_type in _QUANTITY_TYPES,
    )
    price = _decimal_cell(
        price_cell,
        field_name="Trade Price",
        required=mapped_type in _PRICE_TYPES,
        positive=mapped_type in _PRICE_TYPES,
        non_negative=mapped_type == "grant",
    )
    amount = _decimal_cell(
        amount_cell,
        field_name="Amount",
        required=mapped_type in _REQUIRED_CASH_TYPES,
    )
    commission_magnitude = _decimal_cell(
        commission_cell,
        field_name="Commission",
        non_negative=True,
    ) or Decimal("0")
    description = _plain_text(
        description_cell,
        maximum_length=500,
        field_name="Description",
    )
    reference_id = _plain_text(
        reference_cell,
        maximum_length=100,
        field_name="Reference ID",
    )
    if mapped_type == "adjustment" and amount is None:
        amount = Decimal("0")
    if mapped_type == "adjustment" and amount == 0 and not (description or reference_id):
        raise ValueError(
            _cell_error(
                amount_cell,
                "A zero Adjustment must include a Description or Reference ID "
                "that documents the non-cash event.",
            )
        )
    if mapped_type == "forex_trade_component" and not reference_id:
        raise ValueError(
            _cell_error(
                reference_cell,
                "Reference ID is required for both legs of a currency conversion.",
            )
        )

    if mapped_type in _POSITIVE_CASH_TYPES and (amount is None or amount <= 0):
        raise ValueError(
            _cell_error(amount_cell, "Amount must be positive for this Type.")
        )
    if mapped_type in _NEGATIVE_CASH_TYPES and (amount is None or amount >= 0):
        raise ValueError(
            _cell_error(amount_cell, "Amount must be negative for this Type.")
        )
    if mapped_type in {"forex_trade_component", "fx_translation_pnl"}:
        if amount is None or amount == 0:
            raise ValueError(
                _cell_error(amount_cell, "Amount must be non-zero for this Type.")
            )

    commission = -abs(commission_magnitude)
    position_quantity = quantity
    if mapped_type == "sell" and quantity is not None:
        position_quantity = -abs(quantity)
    elif quantity is not None:
        position_quantity = abs(quantity)

    if mapped_type == "grant":
        price = price or Decimal("0")
        if amount not in {None, Decimal("0")}:
            raise ValueError(
                _cell_error(amount_cell, "Amount must be blank or zero for Grant.")
            )
        gross_amount = Decimal("0")
        net_amount = Decimal("0")
    elif mapped_type in _CASHLESS_SECURITY_TRANSFER_TYPES:
        if amount not in {None, Decimal("0")}:
            raise ValueError(
                _cell_error(
                    amount_cell,
                    "Amount must be blank or zero for an in-kind security transfer.",
                )
            )
        gross_amount = Decimal("0")
        net_amount = Decimal("0")
    elif mapped_type in {"buy", "sell", "dividend_reinvestment"}:
        assert quantity is not None
        assert price is not None
        calculated_gross = abs(quantity * price)
        if mapped_type in {"buy", "dividend_reinvestment"}:
            calculated_gross = -calculated_gross
        if amount not in {None, Decimal("0")}:
            raise ValueError(
                _cell_error(
                    amount_cell,
                    "Amount must be blank for trades; it is calculated from "
                    "Quantity, Trade Price, and Commission.",
                )
            )
        gross_amount = calculated_gross
        net_amount = gross_amount + commission
    else:
        assert amount is not None
        net_amount = amount
        gross_amount = net_amount - commission if commission else net_amount

    if mapped_type == "fee" and commission_magnitude == 0:
        commission = net_amount

    localized_datetime = source_datetime.replace(tzinfo=ZIRCON_HK_TIMEZONE)
    ledger_datetime = localized_datetime.astimezone(INVESTMENT_LEDGER_TIMEZONE)
    normalized = build_normalized_transaction_view(
        mapped_type,
        position_quantity,
        price,
        gross_amount,
        commission if commission != 0 else None,
        net_amount,
        is_cash_flow_override=(
            False
            if mapped_type in {
                "buy",
                "sell",
                "grant",
                "fx_translation_pnl",
                *_CASHLESS_SECURITY_TRANSFER_TYPES,
            }
            else True
        ),
        side_override=(
            "buy"
            if mapped_type in {
                "buy",
                "grant",
                "dividend_reinvestment",
                "transfer_in",
            }
            else (
                "sell"
                if mapped_type in {"sell", "transfer_out"}
                else None
            )
        ),
    )
    source = {
        "file_kind": "manual_investment_xlsx",
        "source_filename": source_filename,
        "source_file_sha256": source_sha256,
        "source_sheet": ZIRCON_HK_TRANSACTION_SHEET,
        "source_row": broker_cell.row,
        "source_datetime_raw": source_datetime.strftime("%Y-%m-%d %H:%M:%S"),
        "source_timezone": "Asia/Hong_Kong",
        "source_time_defaulted_to_2300": source_time_defaulted,
        "broker": broker,
        "account": account,
        "reference_id": reference_id,
        "importer_version": ZIRCON_HK_IMPORTER_VERSION,
    }
    record: dict[str, Any] = {
        "date": ledger_datetime.strftime("%Y-%m-%d"),
        "datetime": ledger_datetime.strftime("%Y-%m-%d %H:%M:%S"),
        "type": mapped_type,
        "currency": currency,
        "description": description,
        "source": source,
        "gross_amount_raw": decimal_to_str(gross_amount),
        "net_amount_raw": decimal_to_str(net_amount),
        "broker": broker,
        "account": account or None,
        "normalized": normalized,
    }
    if ticker:
        record["ticker"] = ticker
    if position_quantity is not None:
        record["quantity_raw"] = decimal_to_str(position_quantity)
        record["quantity_abs"] = decimal_to_str(abs(position_quantity))
    if price is not None:
        record["price_raw"] = decimal_to_str(price)
    if commission != 0:
        record["commission_raw"] = decimal_to_str(commission)
        record["commission_abs"] = decimal_to_str(abs(commission))
    return record


def _manual_reference_group_key(
    transaction: dict[str, Any],
) -> tuple[str, str, str] | None:
    source = transaction.get("source") or {}
    reference_id = normalize_import_text(source.get("reference_id"))
    if not reference_id:
        return None
    return (
        normalize_import_text(transaction.get("broker")).casefold(),
        normalize_import_text(transaction.get("account")).casefold(),
        reference_id.casefold(),
    )


def _reference_cell_error(transaction: dict[str, Any], message: str) -> str:
    source = transaction.get("source") or {}
    row = int(source.get("source_row") or 0)
    return f"{ZIRCON_HK_TRANSACTION_SHEET}!L{row} (L): {message}"


def _validate_and_enrich_reference_groups(
    transactions: list[dict[str, Any]],
) -> list[str]:
    groups: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
    for transaction in transactions:
        key = _manual_reference_group_key(transaction)
        if key is not None:
            groups.setdefault(key, []).append(transaction)

    errors: list[str] = []
    for records in groups.values():
        forex_records = [
            record
            for record in records
            if record.get("type") == "forex_trade_component"
        ]
        if not forex_records:
            if len(records) > 1:
                errors.append(
                    _reference_cell_error(
                        records[1],
                        "Reference ID must be unique within each account.",
                    )
                )
            continue
        if len(forex_records) != len(records):
            errors.append(
                _reference_cell_error(
                    records[-1],
                    "A currency-conversion Reference ID must not be shared with "
                    "another transaction Type.",
                )
            )
            continue
        if len(forex_records) != 2:
            errors.append(
                _reference_cell_error(
                    forex_records[-1],
                    "A currency conversion requires exactly two Forex trade "
                    "component rows with the same Reference ID.",
                )
            )
            continue
        currencies = {
            normalize_import_text(record.get("currency")).upper()
            for record in forex_records
        }
        if len(currencies) != 2:
            errors.append(
                _reference_cell_error(
                    forex_records[1],
                    "The two currency-conversion rows must use different currencies.",
                )
            )
            continue
        datetimes = {
            normalize_import_text(record.get("datetime"))
            for record in forex_records
        }
        if len(datetimes) != 1:
            errors.append(
                _reference_cell_error(
                    forex_records[1],
                    "The two currency-conversion rows must use the same transaction "
                    "date and time.",
                )
            )
            continue
        negative_records = [
            record
            for record in forex_records
            if Decimal(record["net_amount_raw"]) < 0
        ]
        positive_records = [
            record
            for record in forex_records
            if Decimal(record["net_amount_raw"]) > 0
        ]
        if len(negative_records) != 1 or len(positive_records) != 1:
            errors.append(
                _reference_cell_error(
                    forex_records[1],
                    "A currency conversion requires one negative sold-currency "
                    "Amount and one positive acquired-currency Amount.",
                )
            )
            continue

        sold = negative_records[0]
        acquired = positive_records[0]
        sold_currency = normalize_import_text(sold.get("currency")).upper()
        acquired_currency = normalize_import_text(acquired.get("currency")).upper()
        sold_amount = abs(Decimal(sold["net_amount_raw"]))
        acquired_amount = Decimal(acquired["net_amount_raw"])
        exchange_rate = sold_amount / acquired_amount
        pair_ticker = f"{acquired_currency}.{sold_currency}"
        description = f"FX from {sold_currency} to {acquired_currency}"
        reference_id = normalize_import_text(
            (sold.get("source") or {}).get("reference_id")
        )
        for record, leg in ((sold, "sold"), (acquired, "acquired")):
            if not normalize_import_text(record.get("description")):
                record["description"] = description
            source = record["source"]
            source["execution_key"] = (
                f"manual_fx:{reference_id.casefold()}:{record['currency']}"
            )
            source["forex_pair_reference_id"] = reference_id
            source["forex_leg"] = leg
            source["forex_pair"] = pair_ticker
            source["exchange_rate_raw"] = decimal_to_str(exchange_rate)
            source["exchange_rate_convention"] = (
                f"{sold_currency} per {acquired_currency}"
            )
    return errors


def build_investment_payload_from_zircon_hk_manual_xlsx(
    *,
    xlsx_bytes: bytes,
    filename: str,
) -> dict[str, Any]:
    """Validate and parse a manually completed fallback investment workbook."""
    if not xlsx_bytes:
        raise ValueError("The manual investment XLSX file is empty.")
    normalized_filename = normalize_import_text(filename) or ZIRCON_HK_TEMPLATE_FILENAME
    if not normalized_filename.lower().endswith(".xlsx"):
        raise ValueError("Please upload the manual investment workbook as an .xlsx file.")
    _validate_xlsx_archive(xlsx_bytes)
    try:
        workbook = load_workbook(
            BytesIO(xlsx_bytes),
            read_only=False,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:  # noqa: BLE001
        raise ValueError(
            "The manual investment workbook is not a readable XLSX file."
        ) from exc
    if ZIRCON_HK_TRANSACTION_SHEET not in workbook.sheetnames:
        raise ValueError(
            f'The manual investment workbook is missing the required "{ZIRCON_HK_TRANSACTION_SHEET}" sheet.'
        )
    unexpected_sheets = set(workbook.sheetnames).difference({
        ZIRCON_HK_TRANSACTION_SHEET,
        ZIRCON_HK_LISTS_SHEET,
    })
    if unexpected_sheets:
        raise ValueError(
            "The manual investment workbook contains unexpected worksheets: "
            + ", ".join(sorted(unexpected_sheets))
            + ". Download a fresh template."
        )
    sheet = workbook[ZIRCON_HK_TRANSACTION_SHEET]
    if sheet.max_column > 50 or sheet.max_row > ZIRCON_HK_MAX_TRANSACTION_ROWS + 1:
        raise ValueError(
            "The manual investment workbook exceeds the supported "
            f"{ZIRCON_HK_MAX_TRANSACTION_ROWS:,} transaction rows or 50 columns."
        )
    headers = tuple(
        normalize_import_whitespace(sheet.cell(row=1, column=column).value)
        for column in range(1, len(ZIRCON_HK_HEADERS) + 1)
    )
    if headers != ZIRCON_HK_HEADERS:
        for column, (actual, expected) in enumerate(
            zip(headers, ZIRCON_HK_HEADERS, strict=True),
            start=1,
        ):
            if actual != expected:
                cell = sheet.cell(row=1, column=column)
                raise ValueError(
                    _cell_error(
                        cell,
                        f"expected header {expected!r}, found {actual!r}. "
                        "Download a fresh manual investment template.",
                    )
                )
    for column in range(len(ZIRCON_HK_HEADERS) + 1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=column)
        if normalize_import_text(cell.value):
            raise ValueError(
                _cell_error(
                    cell,
                    "unexpected extra column. Download a fresh manual investment template.",
                )
            )

    source_sha256 = hashlib.sha256(xlsx_bytes).hexdigest()
    transactions: list[dict[str, Any]] = []
    errors: list[str] = []
    omitted_error_count = 0
    accounts: set[str] = set()
    brokers: set[str] = set()
    for row_number in range(2, sheet.max_row + 1):
        cells = tuple(
            sheet.cell(row=row_number, column=column)
            for column in range(1, len(ZIRCON_HK_HEADERS) + 1)
        )
        if all(cell.value in {None, ""} for cell in cells[2:]):
            continue
        try:
            transaction = _transaction_from_row(
                cells,
                source_filename=normalized_filename,
                source_sha256=source_sha256,
            )
        except ValueError as exc:
            if len(errors) < ZIRCON_HK_MAX_REPORTED_ERRORS:
                errors.append(str(exc))
            else:
                omitted_error_count += 1
            continue
        transactions.append(transaction)
        brokers.add(normalize_import_text(transaction.get("broker")))
        account = normalize_import_text(transaction.get("account"))
        if account:
            accounts.add(account)
    if not errors:
        errors.extend(_validate_and_enrich_reference_groups(transactions))
    if errors:
        reported_error_count = len(errors)
        if omitted_error_count:
            errors.append(
                f"{omitted_error_count} additional validation error(s) were omitted; "
                "fix the listed cells first and upload again."
            )
        suffix = (
            " Fix the listed cells and upload the workbook again."
            if reported_error_count == 1
            else f" Fix these {reported_error_count} cells and upload the workbook again."
        )
        raise ValueError(
            "Manual investment workbook validation failed: "
            + " | ".join(errors)
            + suffix
        )
    if not transactions:
        raise ValueError(
            'The manual investment workbook has no completed transaction rows in the "Transactions" sheet.'
        )
    transactions.sort(
        key=lambda record: (
            normalize_import_text(record.get("datetime")),
            normalize_import_text((record.get("source") or {}).get("reference_id")),
            normalize_import_text(record.get("type")),
            normalize_import_text(record.get("ticker")),
        )
    )
    account = next(iter(accounts)) if len(accounts) == 1 else ("multiple" if accounts else None)
    payload_broker = next(iter(brokers)) if len(brokers) == 1 else "multiple"
    transaction_dates = [record["date"] for record in transactions]
    artifact = {
        "evidence_schema_version": "1.0",
        "sha256": source_sha256,
        "byte_count": len(xlsx_bytes),
        "filename": normalized_filename,
        "filenames": [normalized_filename],
        "broker": payload_broker,
        "account": account or "",
        "source_kind": "manual_investment_xlsx",
        "bundle_id": source_sha256,
        "bundle_ids": [source_sha256],
        "bundle_role": "manual_transactions",
        "related_sha256": "",
        "statement_title": "Manual investment transaction workbook",
        "statement_period": "",
        "statement_period_start": min(transaction_dates),
        "statement_period_end": max(transaction_dates),
        "statement_generated_at": "",
        "content_encoding": "base64",
        "content_base64": base64.b64encode(xlsx_bytes).decode("ascii"),
    }
    return {
        "schema_version": "3.0.0",
        "generator": {
            "name": "manual_xlsx_to_investment_json",
            "version": ZIRCON_HK_IMPORTER_VERSION,
            "generated_at": datetime.now(tz=ZoneInfo("UTC")).isoformat(),
        },
        "broker": payload_broker,
        "account": account,
        "datetime_policy": {
            "date_field_meaning": (
                "America/New_York calendar date converted from the entered Hong Kong time."
            ),
            "datetime_field_meaning": (
                "America/New_York ledger time converted from a typed Asia/Hong_Kong "
                "Excel date-time; date-only entries default to 23:00."
            ),
            "timezone": "America/New_York",
            "source_timezone": "Asia/Hong_Kong",
            "source_has_intraday_timestamp": True,
        },
        "summary": {
            "transaction_count": len(transactions),
            "warning_count": 0,
            "warnings": [],
            "source_row_count": len(transactions),
            "source_file_sha256": source_sha256,
        },
        "starting_cash": None,
        "ending_cash": None,
        "position_snapshot": {},
        "performance_snapshot": {},
        "transactions": transactions,
        "source_artifacts": [artifact],
    }
