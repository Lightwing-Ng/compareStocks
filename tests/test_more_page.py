"""
Tests for route stability across refactored web runtime branches.

Code version: v0.24.0
"""

from __future__ import annotations

import io
import json
from datetime import datetime
from decimal import Decimal
from html.parser import HTMLParser
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch

import pandas as pd
from flask import request, session
from openpyxl import load_workbook

from app import (
    INVESTMENT_IMPORT_MULTIPART_ALLOWANCE_BYTES,
    MAX_INVESTMENT_IMPORT_REQUEST_BYTES,
    create_app,
)
from app.infrastructure.storage import (
    INVESTMENT_STORE_PATH as REAL_INVESTMENT_STORE_PATH,
    MAX_INVESTMENT_SOURCE_ARTIFACT_BYTES,
    MAX_INVESTMENT_SOURCE_EVIDENCE_BYTES,
    investment_evidence_dir_for,
    load_investment_store_payload,
    save_investment_store_payload,
)
import app.services.investment_import as investment_import_service
from app.web.request_security import (
    INVESTMENT_CSRF_HEADER,
    INVESTMENT_CSRF_SESSION_KEY,
    validate_investment_browser_write_request,
)
from app.web.runtime import INVESTMENT_TRANSACTIONS_CACHE_SCHEMA_VERSION
from app.services.zircon_hk_import import (
    ZIRCON_HK_HEADERS,
    ZIRCON_HK_MAX_TRANSACTION_ROWS,
    ZIRCON_HK_TEMPLATE_INPUT_ROWS,
    build_investment_payload_from_zircon_hk_manual_xlsx,
    build_zircon_hk_template_xlsx,
)
from tests.factories.market import (
    FakeStrategy,
    backtest_result,
    fetch_history_stub,
    quote_profile_stub,
)


class _LocalStorePaginationStructureParser(HTMLParser):
    _VOID_ELEMENTS = frozenset({"area", "base", "br", "col", "embed", "hr", "img", "input", "link", "meta", "source", "track", "wbr"})

    def __init__(self) -> None:
        super().__init__()
        self._stack: list[tuple[str, dict[str, str | None]]] = []
        self._active_pagination_index: int | None = None
        self._active_control: tuple[str, int, int] | None = None
        self.pagination_parent_ids: list[str | None] = []
        self.pagination_parent_classes: list[str] = []
        self.pagination_classes: list[str] = []
        self.pagination_attributes: list[dict[str, str | None]] = []
        self.pagination_controls: list[list[dict[str, str | None]]] = []
        self.pagination_ellipsis_positions: list[list[str | None]] = []
        self.pagination_indicator_counts: list[int] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attributes = dict(attrs)
        if tag == "nav" and "data-local-store-pagination" in attributes:
            parent_attributes = self._stack[-1][1] if self._stack else {}
            self.pagination_parent_ids.append(parent_attributes.get("id"))
            self.pagination_parent_classes.append(parent_attributes.get("class") or "")
            self.pagination_classes.append(attributes.get("class") or "")
            self.pagination_attributes.append(attributes)
            self.pagination_controls.append([])
            self.pagination_ellipsis_positions.append([])
            self.pagination_indicator_counts.append(0)
            self._active_pagination_index = len(self.pagination_attributes) - 1
        elif self._active_pagination_index is not None:
            class_names = set((attributes.get("class") or "").split())
            if tag in {"a", "button"} and "local-store-page-button" in class_names:
                controls = self.pagination_controls[self._active_pagination_index]
                controls.append({
                    "tag": tag,
                    "class": attributes.get("class"),
                    "href": attributes.get("href"),
                    "target": attributes.get("data-pagination-target"),
                    "current": attributes.get("data-pagination-current"),
                    "aria_label": attributes.get("aria-label"),
                    "aria_current": attributes.get("aria-current"),
                    "text": "",
                })
                self._active_control = (
                    tag,
                    self._active_pagination_index,
                    len(controls) - 1,
                )
            if "local-store-pagination-indicator" in class_names:
                self.pagination_indicator_counts[self._active_pagination_index] += 1
            if "data-pagination-ellipsis" in attributes:
                self.pagination_ellipsis_positions[self._active_pagination_index].append(
                    attributes.get("data-pagination-ellipsis")
                )
        if tag not in self._VOID_ELEMENTS:
            self._stack.append((tag, attributes))

    def handle_data(self, data: str) -> None:
        if self._active_control is None:
            return
        _, pagination_index, control_index = self._active_control
        control = self.pagination_controls[pagination_index][control_index]
        control["text"] = f"{control['text'] or ''}{data}".strip()

    def handle_endtag(self, tag: str) -> None:
        if self._active_control is not None and self._active_control[0] == tag:
            self._active_control = None
        if tag == "nav" and self._active_pagination_index is not None:
            self._active_pagination_index = None
        for index in range(len(self._stack) - 1, -1, -1):
            if self._stack[index][0] == tag:
                del self._stack[index:]
                break


class MorePageTests(unittest.TestCase):
    def setUp(self) -> None:
        self._investment_temp_dir = TemporaryDirectory()
        temp_root = Path(self._investment_temp_dir.name)
        self.investment_store_path = temp_root / "investment.parquet"
        self.investment_cache_path = temp_root / "investment_cache" / "transactions_payload.json"
        self._investment_store_patch = patch(
            "app.web.runtime.INVESTMENT_STORE_PATH",
            self.investment_store_path,
        )
        self._investment_cache_patch = patch(
            "app.web.runtime.INVESTMENT_TRANSACTIONS_CACHE_PATH",
            self.investment_cache_path,
        )
        self._investment_store_patch.start()
        self._investment_cache_patch.start()
        self._private_evidence_patch = patch.object(
            investment_import_service,
            "_load_local_private_investment_evidence",
            return_value={
                "hsbc_performance_calibrations": {
                    "000-999999-999": {"RAM": "3.21"},
                },
                "verified_tax_lot_history": {
                    "hsbc": {
                        "000-999999-999": {
                            "DRAM": {
                                "expected_shares": "5",
                                "buy_count": 3,
                                "sell_count": 1,
                            },
                            "EUV": {
                                "expected_shares": "2",
                                "buy_count": 2,
                                "sell_count": 1,
                            },
                            "GOOGL": {
                                "buy_count": 2,
                                "sell_count": 1,
                                "calculation_method": "trade_price_and_commission",
                            },
                        },
                    },
                },
            },
        )
        self._private_evidence_patch.start()
        self.addCleanup(self._private_evidence_patch.stop)
        self.addCleanup(self._investment_cache_patch.stop)
        self.addCleanup(self._investment_store_patch.stop)
        self.addCleanup(self._investment_temp_dir.cleanup)

    def _post_investment_import(
        self,
        client,
        *args,
        origin: str = "http://localhost",
        **kwargs,
    ):
        client.get("/trade/investment", base_url=origin)
        with client.session_transaction() as browser_session:
            csrf_token = browser_session[INVESTMENT_CSRF_SESSION_KEY]
        headers = {
            "Origin": origin,
            "Sec-Fetch-Site": "same-origin",
            INVESTMENT_CSRF_HEADER: csrf_token,
            **kwargs.pop("headers", {}),
        }
        return client.post(
            *args,
            base_url=origin,
            headers=headers,
            **kwargs,
        )

    def _build_sample_transactions_csv(self) -> str:
        return "\n".join([
            "Statement,Header,Field Name,Field Value",
            "Statement,Data,Title,Transaction History",
            "Summary,Header,Field Name,Field Value",
            "Summary,Data,Starting Cash,1000.00",
            "Summary,Data,Ending Cash,899.00",
            "Transaction History,Header,Date,Account,Description,Transaction Type,Symbol,Quantity,Price,Price Currency,Gross Amount ,Commission,Net Amount",
            "Transaction History,Data,2026-03-01,U***00001,Example Buy,Buy,QQQ,1,100,USD,-100,-1,-101",
        ]) + "\n"

    def _build_sample_positions_csv(self) -> str:
        return "\n".join([
            "Statement,Header,Field Name,Field Value",
            "Statement,Data,Title,Realized Summary",
            "Realized & Unrealized Performance Summary,Header,Asset Category,Symbol,Cost Adj.,Realized S/T Profit,Realized S/T Loss,Realized L/T Profit,Realized L/T Loss,Realized Total,Unrealized S/T Profit,Unrealized S/T Loss,Unrealized L/T Profit,Unrealized L/T Loss,Unrealized Total,Total,Code",
            "Realized & Unrealized Performance Summary,Data,Stocks,QQQ,0,0,0,0,0,0,5,0,0,0,5,5,",
            "Open Positions,Header,DataDiscriminator,Asset Category,Currency,Symbol,Open,Quantity,Mult,Cost Price,Cost Basis,Close Price,Value,Unrealized P/L,Code",
            "Open Positions,Data,Summary,Stocks,USD,QQQ,-,1,1,100,100,105,105,5,",
            "Open Positions,Total,,Stocks,USD,,,,,,100,,105,5,",
        ]) + "\n"

    def _build_ibkr_web_trade_notifications_text(self) -> str:
        return """Orders & Trades
Trade Notifications
Trades Account Action Quantity Status Price Amount
DRAM
Bot 2 @ 45.50 on OVERNIGHT
U00000001 Bought 2
Filled
7/29/2026, 10:00 AM
45.50
91
Fees: 0.35
"""

    def _build_ibkr_current_web_trade_notifications_text(self) -> str:
        return """Orders & Trades
Trade Notifications
Trades Account Action Quantity Status Price Amount
ALFA
Bot 5 @ 14.00 on ARCA
U00000001 Bought 5
Filled
8/3/2025, 8:18 PM
14.00
70
Fees: 0.12
BETA
Bot 5 @ 15.35 on NASDAQ
U00000001 Bought 5
Filled
8/3/2025, 8:13 PM
15.35
76.75
Fees: 0.34
ALFA
Sold 5 @ 15.65 on OVERNIGHT
U00000001 Sold 5
Filled
8/3/2025, 9:56 AM
15.65
78.25
Fees: 0.35
ALFA
Sold 5 @ 15.65 on OVERNIGHT
U00000001 Sold 5
Filled
8/3/2025, 9:56 AM
15.65
78.25
Fees: 0.0
ALFA
Bot 3 @ 10.00 on OVERNIGHT
U00000001 Bought 3
Filled
8/1/2025, 12:57 PM
10.00
30
Fees: 0.12
"""

    def _build_hsbc_non_usd_cash_paste(self) -> str:
        return "\n".join([
            "HKD Current",
            "Account number:",
            "000-999999-999",
            "Ledger balance:",
            "1,000.00 HKD",
            "Available balance:",
            "1,000.00 HKD",
            "Post date Description Amount in Amount out Balance Additional options",
            "15 Jul 2026",
            "LONGBRIDGE KOL REWARD",
            "1,000.00",
            "1,000.00",
            "Download",
            "CNY Savings",
            "Account number:",
            "000-999999-999",
            "Ledger balance:",
            "12.00 CNY",
            "Available balance:",
            "12.00 CNY",
            "Post date Description Amount in Amount out Balance Additional options",
            "15 Jul 2026",
            "CNY INTEREST",
            "12.00",
            "12.00",
            "Download",
        ])

    def _build_hsbc_usd_cash_paste(self) -> str:
        return "\n".join([
            "USD Savings",
            "Account number:",
            "000-999999-999",
            "Ledger balance:",
            "60.99 USD",
            "Available balance:",
            "60.99 USD",
            "Post date Description Amount in Amount out Balance Additional options",
            "15 Jul 2026",
            "USD INTEREST",
            "60.99",
            "60.99",
            "Download",
        ])

    def _build_zircon_hk_workbook(self, *, invalid_date: bool = False) -> bytes:
        workbook = load_workbook(io.BytesIO(build_zircon_hk_template_xlsx()))
        sheet = workbook["Transactions"]
        values = (
            "Zircon (HK)",
            "Manual account",
            "not a typed date" if invalid_date else datetime(2026, 7, 30, 21, 15),
            "Buy",
            "HKD",
            "700.HK",
            100,
            20,
            None,
            5,
            "Manual activity entry",
            "zircon-reference-1",
        )
        for column, value in enumerate(values, start=1):
            sheet.cell(row=2, column=column, value=value)
        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def test_removed_timing_page_redirects_to_trade_investment(self) -> None:
        client = create_app().test_client()

        response = client.get("/trade/timing")

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], "/trade/investment")

    def test_more_investment_page_renders_from_more_section(self) -> None:
        client = create_app().test_client()

        response = client.get("/trade/investment")

        self.assertEqual(response.status_code, 200)
        body = response.get_data(as_text=True)

        self.assertIn("Investment", body)
        self.assertIn('<article class="report-card workspace-content-card trade-performance-card investment-report-card">', body)
        self.assertIn('class="investment-surface-stack investment-view-surface" id="investment_view_surface"', body)
        self.assertIn('class="surface-resizer surface-resizer--block surface-resizer--reveal investment-section-resizer"', body)
        self.assertIn('aria-orientation="horizontal"', body)
        self.assertIn('<article class="chart-surface investment-history-surface" id="investment_history_surface"', body)

    def test_missing_store_ignores_an_unusable_device_local_investment_cache(self) -> None:
        blocked_cache_parent = self.investment_cache_path.parent
        blocked_cache_parent.write_text("not a directory", encoding="utf-8")
        client = create_app().test_client()

        response = client.get("/api/investment/transactions")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["transactions"], [])
        self.assertTrue(response.get_json()["success"])

    def test_stale_investment_cache_schema_is_rebuilt(self) -> None:
        save_investment_store_payload(
            {
                "schema_version": "3.0.0",
                "broker": "multiple",
                "account": "multiple",
                "summary": {},
                "transactions": [],
            },
            self.investment_store_path,
        )
        client = create_app().test_client()

        first_response = client.get("/api/investment/transactions")

        self.assertEqual(first_response.status_code, 200)
        self.assertEqual(first_response.get_json()["investment_cache"]["status"], "miss")
        cached = json.loads(self.investment_cache_path.read_text(encoding="utf-8"))
        cached["schema_version"] = "investment-transactions-v2"
        cached["payload"]["stale_cache_marker"] = True
        self.investment_cache_path.write_text(
            json.dumps(cached),
            encoding="utf-8",
        )

        second_response = client.get("/api/investment/transactions")

        self.assertEqual(second_response.status_code, 200)
        second_payload = second_response.get_json()
        self.assertEqual(second_payload["investment_cache"]["status"], "miss")
        self.assertNotIn("stale_cache_marker", second_payload)
        rebuilt = json.loads(self.investment_cache_path.read_text(encoding="utf-8"))
        self.assertEqual(
            rebuilt["schema_version"],
            INVESTMENT_TRANSACTIONS_CACHE_SCHEMA_VERSION,
        )

    def test_current_investment_cache_reapplies_account_performance_calibrations(self) -> None:
        save_investment_store_payload(
            {
                "schema_version": "3.0.0",
                "broker": "multiple",
                "account": "multiple",
                "summary": {},
                "broker_summaries": {
                    "hsbc": {
                        "account": "000-999999-999",
                    },
                    "longbridge_hk": {
                        "account": "H99999999",
                        "performance_snapshot_authoritative": True,
                        "performance_snapshot": {
                            "SQQQ": {
                                "currency": "USD",
                                "realized_total": "2.22",
                                "calibration_source": "broker_reported_pnl",
                            },
                        },
                    },
                },
                "transactions": [],
            },
            self.investment_store_path,
        )
        client = create_app().test_client()

        first_response = client.get("/api/investment/transactions")

        self.assertEqual(first_response.status_code, 200)
        self.assertEqual(first_response.get_json()["investment_cache"]["status"], "miss")
        cached = json.loads(self.investment_cache_path.read_text(encoding="utf-8"))
        cached["payload"]["broker_summaries"]["longbridge_hk"]["performance_snapshot"][
            "SQQQ"
        ]["realized_total"] = "2.22"
        self.investment_cache_path.write_text(json.dumps(cached), encoding="utf-8")

        second_response = client.get("/api/investment/transactions")

        self.assertEqual(second_response.status_code, 200)
        second_payload = second_response.get_json()
        self.assertEqual(second_payload["investment_cache"]["status"], "hit")
        self.assertEqual(
            second_payload["broker_summaries"]["longbridge_hk"]["performance_snapshot"][
                "SQQQ"
            ]["realized_total"],
            "2.22",
        )
        googl_verification = second_payload["broker_summaries"]["hsbc"][
            "tax_lot_history_verifications"
        ]["GOOGL"]
        self.assertEqual(googl_verification["buy_count"], 2)
        self.assertEqual(googl_verification["sell_count"], 1)
        self.assertEqual(
            googl_verification["calculation_method"],
            "trade_price_and_commission",
        )
        hsbc_verifications = second_payload["broker_summaries"]["hsbc"][
            "tax_lot_history_verifications"
        ]
        self.assertEqual(hsbc_verifications["DRAM"]["expected_shares"], "5")
        self.assertEqual(hsbc_verifications["DRAM"]["buy_count"], 3)
        self.assertEqual(hsbc_verifications["DRAM"]["sell_count"], 1)
        self.assertEqual(hsbc_verifications["EUV"]["expected_shares"], "2")
        self.assertEqual(hsbc_verifications["EUV"]["buy_count"], 2)
        self.assertEqual(hsbc_verifications["EUV"]["sell_count"], 1)

    def test_more_investment_page_exposes_dual_csv_import_form(self) -> None:
        client = create_app().test_client()

        response = client.get("/trade/investment")
        body = response.get_data(as_text=True)

        self.assertIn('id="transactions_csv"', body)
        self.assertIn('id="positions_csv"', body)
        self.assertIn('value="web_paste"', body)
        self.assertIn('id="ibkr_trade_notifications_text"', body)
        self.assertIn('id="ibkr_trade_notifications_paste_button"', body)
        self.assertIn('id="futuhk_statement_pdfs"', body)
        self.assertIn('id="hsbc_statement_pdfs"', body)
        self.assertIn('id="hsbc_statement_pdfs_status"', body)
        self.assertIn('id="boc_hk_statement_pdfs"', body)
        self.assertIn('id="boc_hk_statement_pdfs_status"', body)
        self.assertRegex(
            body,
            r'id="boc_hk_statement_pdfs"[\s\S]*?type="file"[\s\S]*?accept="\.pdf,application/pdf"[\s\S]*?multiple',
        )
        self.assertIn('aria-label="Upload one or more BOCHK Consolidated Statement PDFs"', body)
        self.assertIn('id="boc_hk_statement_pdfs_hint"', body)
        self.assertIn(
            'class="investment-import-field-group investment-import-boc-hk-field-group" id="investment_import_boc_hk_fields"',
            body,
        )
        self.assertIn('data-import-field="boc_hk-statements"', body)
        self.assertIn('investment-import-label-trigger--single', body)
        self.assertNotIn('id="hsbc_composite_statement_pdfs"', body)
        self.assertNotIn('id="hsbc_investment_statement_pdfs"', body)
        self.assertIn('value="futuhk"', body)
        self.assertIn('Futu (HK)', body)
        self.assertIn('enctype="multipart/form-data"', body)
        self.assertIn('settings_store/investment.parquet', body)
        self.assertIn('id="investment_import_feedback_message"', body)
        self.assertIn('notice-floating-banner', body)
        self.assertIn('id="investment_import_submit_button"', body)
        self.assertIn('value="zircon_hk"', body)
        self.assertIn('id="investment_import_zircon_hk_fields"', body)
        self.assertIn('id="zircon_hk_template_download"', body)
        self.assertIn('id="zircon_hk_transactions_xlsx"', body)
        self.assertIn('id="zircon_hk_transactions_xlsx_status"', body)
        self.assertIn('value="standard_xlsx"', body)
        self.assertIn("No specified broker", body)
        self.assertIn('value="cmb_cn"', body)
        self.assertIn("China Merchants Bank", body)
        self.assertIn('value="boc_cn"', body)
        self.assertIn("Bank of China", body)
        self.assertIn('value="boc_hk"', body)
        self.assertIn("Bank of China (Hong Kong)", body)
        self.assertIn('value="icbc_cn"', body)
        self.assertIn("Industrial and Commercial Bank of China", body)
        self.assertIn('value="icbc_hk"', body)
        self.assertIn("Industrial and Commercial Bank of China (Asia)", body)
        self.assertIn('value="ccb_cn"', body)
        self.assertIn("China Construction Bank", body)
        self.assertIn('value="ccb_hk"', body)
        self.assertIn("China Construction Bank (Asia)", body)
        self.assertIn('id="export_standard_xlsx_button"', body)
        self.assertIn('"investmentCsrfToken":', body)
        self.assertIn(
            '<div class="field live-trading-broker-strip backtest-shared-select-field investment-import-broker-field investment-broker-summary-selector-shell">',
            body,
        )
        self.assertNotIn('class="investment-metrics-copy"', body)
        self.assertNotIn('investment-metrics-kicker', body)
        self.assertNotIn('investment-metrics-description', body)

    def test_zircon_hk_template_download_has_typed_workbook_contract(self) -> None:
        client = create_app().test_client()

        response = client.get("/api/investment/imports/zircon-hk/template.xlsx")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.headers["Cache-Control"],
            "no-store, no-cache, max-age=0, must-revalidate",
        )
        self.assertIn(
            "Manual_investment_import.xlsx",
            response.headers["Content-Disposition"],
        )
        workbook = load_workbook(io.BytesIO(response.data))
        self.assertEqual(
            tuple(cell.value for cell in workbook["Transactions"][1]),
            ZIRCON_HK_HEADERS,
        )

    def test_standard_xlsx_export_download_round_trips_through_import_parser(self) -> None:
        client = create_app().test_client()

        response = self._post_investment_import(
            client,
            "/api/investment/exports/standard.xlsx",
            json={
                "transactions": [{
                    "ledger_no": 7,
                    "broker": "ibkr",
                    "account": "U00000002",
                    "datetime": "2026-07-30 09:15:00",
                    "type": "buy",
                    "currency": "USD",
                    "ticker": "AAPL",
                    "quantity_raw": "2",
                    "price_raw": "210.25",
                    "commission_raw": "-1.25",
                    "description": "AAPL purchase",
                    "source": {},
                }],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "AAPL_standard_investment_export.xlsx",
            response.headers["Content-Disposition"],
        )
        payload = build_investment_payload_from_zircon_hk_manual_xlsx(
            xlsx_bytes=response.data,
            filename="AAPL_standard_investment_export.xlsx",
        )
        self.assertEqual(payload["summary"]["transaction_count"], 1)
        self.assertEqual(payload["transactions"][0]["broker"], "ibkr")
        self.assertEqual(payload["transactions"][0]["ticker"], "AAPL")

    def test_standard_xlsx_export_accepts_rendered_transaction_fields(self) -> None:
        client = create_app().test_client()

        response = self._post_investment_import(
            client,
            "/api/investment/exports/standard.xlsx",
            json={
                "transactions": [{
                    "ledger_no": 8,
                    "broker": "hsbc",
                    "source": {"account_number": "HSBC-001"},
                    "date": "2026-07-30",
                    "type": "buy",
                    "currency": "USD",
                    "ticker": "MSFT",
                    "quantity": 2,
                    "price": 210.25,
                    "amount": -420.5,
                    "commission": -1.25,
                    "description": "Rendered transaction",
                }],
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = build_investment_payload_from_zircon_hk_manual_xlsx(
            xlsx_bytes=response.data,
            filename="MSFT_standard_investment_export.xlsx",
        )
        transaction = payload["transactions"][0]
        self.assertEqual(transaction["broker"], "hsbc")
        self.assertEqual(transaction["account"], "HSBC-001")
        self.assertEqual(transaction["ticker"], "MSFT")
        self.assertEqual(transaction["quantity_raw"], "2")
        self.assertEqual(transaction["price_raw"], "210.25")
        self.assertEqual(transaction["commission_raw"], "-1.25")

    def test_standard_xlsx_export_accepts_full_history_scopes_above_2000_rows(self) -> None:
        client = create_app().test_client()
        transactions = [
            {
                "ledger_no": index,
                "broker": "ibkr",
                "account": "U-FULL-HISTORY",
                "datetime": "2026-07-30 09:15:00",
                "type": "buy",
                "currency": "USD",
                "ticker": "AAPL",
                "quantity_raw": "1",
                "price_raw": "100",
                "commission_raw": "0",
                "description": "Full history export regression row",
                "source": {"file_kind": "test_fixture", "row_number": index},
            }
            for index in range(1, 2_002)
        ]

        response = self._post_investment_import(
            client,
            "/api/investment/exports/standard.xlsx",
            json={"transactions": transactions},
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.data))
        sheet = workbook["Transactions"]
        expected_end_row = len(transactions) + 1
        self.assertGreater(len(transactions), ZIRCON_HK_TEMPLATE_INPUT_ROWS)
        self.assertEqual(sheet.max_row, expected_end_row)
        self.assertEqual(
            {
                (
                    validation.type,
                    validation.formula1,
                    validation.formula2,
                    str(validation.sqref),
                )
                for validation in sheet.data_validations.dataValidation
            },
            {
                ("list", "ZirconBrokers", None, f"A2:A{expected_end_row}"),
                ("list", "ZirconTypes", None, f"D2:D{expected_end_row}"),
                ("list", "ZirconCurrencies", None, f"E2:E{expected_end_row}"),
                (
                    "date",
                    "DATE(2000,1,1)",
                    "DATE(2100,12,31)",
                    f"C2:C{expected_end_row}",
                ),
                ("decimal", "0", "1000000000000", f"G2:G{expected_end_row}"),
                ("decimal", "0", "1000000000000", f"H2:H{expected_end_row}"),
                (
                    "decimal",
                    "-1000000000000",
                    "1000000000000",
                    f"I2:I{expected_end_row}",
                ),
                ("decimal", "0", "1000000000000", f"J2:J{expected_end_row}"),
            },
        )
        self.assertEqual(
            sheet.cell(row=expected_end_row, column=3).number_format,
            "yyyy-mm-dd hh:mm",
        )
        self.assertEqual(
            sheet.cell(row=expected_end_row, column=10).number_format,
            "#,##0.########;[Red]-#,##0.########;-",
        )
        payload = build_investment_payload_from_zircon_hk_manual_xlsx(
            xlsx_bytes=response.data,
            filename="Standard_investment_export.xlsx",
        )
        self.assertEqual(payload["summary"]["transaction_count"], 2_001)

    def test_standard_xlsx_export_rejects_more_than_10000_transactions(self) -> None:
        client = create_app().test_client()
        response = self._post_investment_import(
            client,
            "/api/investment/exports/standard.xlsx",
            json={"transactions": [{} for _ in range(ZIRCON_HK_MAX_TRANSACTION_ROWS + 1)]},
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("10,000", response.get_json()["error"])

    def test_zircon_hk_prevalidation_is_precise_and_never_writes_ledger(
        self,
    ) -> None:
        client = create_app().test_client()
        real_store_before = (
            REAL_INVESTMENT_STORE_PATH.read_bytes()
            if REAL_INVESTMENT_STORE_PATH.exists()
            else None
        )

        valid_response = self._post_investment_import(
            client,
            "/api/investment/imports/zircon-hk/validate",
            data={
                "zircon_hk_transactions_xlsx": (
                    io.BytesIO(self._build_zircon_hk_workbook()),
                    "zircon-valid.xlsx",
                ),
            },
            content_type="multipart/form-data",
        )
        invalid_response = self._post_investment_import(
            client,
            "/api/investment/imports/zircon-hk/validate",
            data={
                "zircon_hk_transactions_xlsx": (
                    io.BytesIO(
                        self._build_zircon_hk_workbook(invalid_date=True)
                    ),
                    "zircon-invalid.xlsx",
                ),
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(valid_response.status_code, 200)
        self.assertEqual(valid_response.get_json()["transaction_count"], 1)
        self.assertEqual(invalid_response.status_code, 400)
        self.assertIn("Transactions!C2", invalid_response.get_json()["error"])
        self.assertFalse(self.investment_store_path.exists())
        self.assertFalse(
            investment_evidence_dir_for(self.investment_store_path).exists()
        )
        self.assertEqual(
            REAL_INVESTMENT_STORE_PATH.read_bytes()
            if REAL_INVESTMENT_STORE_PATH.exists()
            else None,
            real_store_before,
        )

    def test_zircon_hk_import_commits_only_to_isolated_store_with_evidence(
        self,
    ) -> None:
        client = create_app().test_client()
        real_store_before = (
            REAL_INVESTMENT_STORE_PATH.read_bytes()
            if REAL_INVESTMENT_STORE_PATH.exists()
            else None
        )

        with patch("app.web.runtime.threading.Thread") as mocked_thread:
            response = self._post_investment_import(
                client,
                "/api/investment/transactions",
                data={
                    "broker": "zircon_hk",
                    "zircon_hk_transactions_xlsx": (
                        io.BytesIO(self._build_zircon_hk_workbook()),
                        "zircon-valid.xlsx",
                    ),
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.get_json()["success"])
        mocked_thread.assert_called_once()
        stored = load_investment_store_payload(self.investment_store_path)
        self.assertEqual(len(stored["transactions"]), 1)
        self.assertEqual(stored["transactions"][0]["broker"], "zircon_hk")
        self.assertEqual(stored["transactions"][0]["ticker"], "700.HK")
        artifact = stored["source_artifacts"][0]
        self.assertEqual(artifact["source_kind"], "manual_investment_xlsx")
        self.assertNotIn("content_base64", artifact)
        evidence_path = investment_evidence_dir_for(
            self.investment_store_path
        ) / f"{artifact['storage_key']}.bin"
        self.assertTrue(evidence_path.exists())
        self.assertEqual(
            REAL_INVESTMENT_STORE_PATH.read_bytes()
            if REAL_INVESTMENT_STORE_PATH.exists()
            else None,
            real_store_before,
        )

    def test_no_specified_broker_imports_any_standard_xlsx_workbook(self) -> None:
        client = create_app().test_client()

        with patch("app.web.runtime.threading.Thread"):
            response = self._post_investment_import(
                client,
                "/api/investment/transactions",
                data={
                    "broker": "standard_xlsx",
                    "zircon_hk_transactions_xlsx": (
                        io.BytesIO(self._build_zircon_hk_workbook()),
                        "Manual_investment_import.xlsx",
                    ),
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.get_json()["success"])
        stored = load_investment_store_payload(self.investment_store_path)
        self.assertEqual(stored["transactions"][0]["broker"], "zircon_hk")
        self.assertEqual(stored["transactions"][0]["ticker"], "700.HK")

    def test_manual_bank_institutions_use_the_standard_xlsx_import_route(self) -> None:
        client = create_app().test_client()

        with patch("app.web.runtime.threading.Thread") as mocked_thread:
            for broker in (
                "cmb_cn",
                "boc_cn",
                "boc_hk",
                "icbc_cn",
                "icbc_hk",
                "ccb_cn",
                "ccb_hk",
            ):
                with self.subTest(broker=broker):
                    response = self._post_investment_import(
                        client,
                        "/api/investment/transactions",
                        data={
                            "broker": broker,
                            "zircon_hk_transactions_xlsx": (
                                io.BytesIO(self._build_zircon_hk_workbook()),
                                f"{broker}_manual_investment.xlsx",
                            ),
                        },
                        content_type="multipart/form-data",
                    )
                    self.assertEqual(response.status_code, 200)
                    self.assertTrue(response.get_json()["success"])

        self.assertEqual(mocked_thread.call_count, 7)

    def test_zircon_hk_prevalidation_requires_session_csrf_proof(self) -> None:
        client = create_app().test_client()
        client.get("/trade/investment", base_url="http://localhost")

        with patch("app.web.runtime.parse_investment_payload") as mocked_parser:
            response = client.post(
                "/api/investment/imports/zircon-hk/validate",
                base_url="http://localhost",
                headers={
                    "Origin": "http://localhost",
                    "Sec-Fetch-Site": "same-origin",
                    INVESTMENT_CSRF_HEADER: "wrong-session-token",
                },
                data={
                    "zircon_hk_transactions_xlsx": (
                        io.BytesIO(self._build_zircon_hk_workbook()),
                        "zircon-valid.xlsx",
                    ),
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 403)
        mocked_parser.assert_not_called()
        self.assertFalse(self.investment_store_path.exists())

    def test_hsbc_statement_route_receives_one_unordered_pdf_bundle(self) -> None:
        client = create_app().test_client()

        with patch(
            "app.services.investment_import.build_investment_payload_from_hsbc_statement_bundle",
            side_effect=ValueError("bundle reached parser"),
        ) as mocked_builder:
            response = self._post_investment_import(
                client,
                "/api/investment/transactions",
                data={
                    "broker": "hsbc",
                    "hsbc_import_mode": "statement_pdf",
                    "hsbc_statement_pdfs": [
                        (io.BytesIO(b"investment"), "opaque-2.pdf"),
                        (io.BytesIO(b"composite"), "opaque-1.pdf"),
                    ],
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.get_json()["error"], "bundle reached parser")
        statement_payloads = mocked_builder.call_args.kwargs["statement_pdf_payloads"]
        self.assertEqual(
            [(payload, filename) for payload, filename in statement_payloads],
            [(b"investment", "opaque-2.pdf"), (b"composite", "opaque-1.pdf")],
        )

    def test_bochk_statement_route_receives_one_multi_file_pdf_batch(self) -> None:
        client = create_app().test_client()

        with patch(
            "app.services.investment_import.build_investment_payload_from_bochk_statement_pdfs",
            side_effect=ValueError("BOCHK batch reached parser"),
        ) as mocked_builder:
            response = self._post_investment_import(
                client,
                "/api/investment/transactions",
                data={
                    "broker": "boc_hk",
                    "boc_hk_statement_pdfs": [
                        (io.BytesIO(b"july"), "2026-07.pdf"),
                        (io.BytesIO(b"june"), "2026-06.pdf"),
                    ],
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.get_json()["error"], "BOCHK batch reached parser")
        statement_payloads = mocked_builder.call_args.kwargs["statement_pdf_payloads"]
        self.assertEqual(
            [(payload, filename) for payload, filename in statement_payloads],
            [(b"july", "2026-07.pdf"), (b"june", "2026-06.pdf")],
        )

    def test_bochk_statement_route_rejects_an_empty_pdf_batch(self) -> None:
        client = create_app().test_client()

        response = self._post_investment_import(
            client,
            "/api/investment/transactions",
            data={"broker": "boc_hk"},
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.get_json()["error"],
            "Please upload at least one BOCHK Consolidated Statement PDF.",
        )

    def test_bochk_statement_route_rejects_an_empty_file_inside_a_batch(self) -> None:
        client = create_app().test_client()

        with patch(
            "app.services.investment_import.build_investment_payload_from_bochk_statement_pdfs",
        ) as mocked_builder:
            response = self._post_investment_import(
                client,
                "/api/investment/transactions",
                data={
                    "broker": "boc_hk",
                    "boc_hk_statement_pdfs": [
                        (io.BytesIO(b"valid"), "2026-07.pdf"),
                        (io.BytesIO(b""), "2026-06.pdf"),
                    ],
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.get_json()["error"],
            "The uploaded BOCHK statement PDF '2026-06.pdf' is empty.",
        )
        mocked_builder.assert_not_called()

    def test_bochk_statement_route_rejects_a_non_pdf_filename(self) -> None:
        client = create_app().test_client()

        with patch(
            "app.services.investment_import.build_investment_payload_from_bochk_statement_pdfs",
        ) as mocked_builder:
            response = self._post_investment_import(
                client,
                "/api/investment/transactions",
                data={
                    "broker": "boc_hk",
                    "boc_hk_statement_pdfs": [
                        (io.BytesIO(b"valid"), "2026-07.txt"),
                    ],
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.get_json()["error"],
            "The uploaded BOCHK statement '2026-07.txt' must use a .pdf filename.",
        )
        mocked_builder.assert_not_called()

    def test_bochk_statement_route_rejects_an_empty_filename(self) -> None:
        client = create_app().test_client()

        with patch(
            "app.services.investment_import.build_investment_payload_from_bochk_statement_pdfs",
        ) as mocked_builder:
            response = self._post_investment_import(
                client,
                "/api/investment/transactions",
                data={
                    "broker": "boc_hk",
                    "boc_hk_statement_pdfs": [
                        (io.BytesIO(b"valid"), ""),
                    ],
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.get_json()["error"],
            "Every uploaded BOCHK statement must have a non-empty filename.",
        )
        mocked_builder.assert_not_called()

    def test_hsbc_paste_prevalidation_accepts_non_usd_cash_without_writing(self) -> None:
        client = create_app().test_client()

        response = self._post_investment_import(
            client,
            "/api/investment/imports/hsbc-paste/validate",
            json={
                "cash_account_text": self._build_hsbc_non_usd_cash_paste(),
                "portfolio_text": "",
                "order_status_text": "",
            },
        )

        self.assertEqual(response.status_code, 200)
        validation = response.get_json()
        self.assertTrue(validation["success"])
        self.assertTrue(validation["ready"])
        self.assertEqual(validation["mode"], "cash_only_non_usd")
        self.assertEqual(validation["cash_currencies"], ["CNH", "HKD"])
        self.assertEqual(validation["field_status"], {
            "cash": True,
            "portfolio": False,
            "order_status": False,
        })
        self.assertFalse(self.investment_store_path.exists())
        self.assertFalse(self.investment_cache_path.exists())

    def test_hsbc_paste_prevalidation_waits_for_usd_portfolio_and_order_status(self) -> None:
        client = create_app().test_client()

        response = self._post_investment_import(
            client,
            "/api/investment/imports/hsbc-paste/validate",
            json={
                "cash_account_text": self._build_hsbc_usd_cash_paste(),
                "portfolio_text": "",
                "order_status_text": "",
            },
        )

        self.assertEqual(response.status_code, 200)
        validation = response.get_json()
        self.assertTrue(validation["success"])
        self.assertFalse(validation["ready"])
        self.assertEqual(validation["mode"], "usd_composite")
        self.assertEqual(validation["required_fields"], ["portfolio", "order_status"])
        self.assertFalse(self.investment_store_path.exists())

    def test_hsbc_paste_prevalidation_requires_session_csrf_proof(self) -> None:
        client = create_app().test_client()
        client.get("/trade/investment", base_url="http://localhost")

        with patch("app.web.runtime.validate_hsbc_pasted_text") as mocked_validator:
            response = client.post(
                "/api/investment/imports/hsbc-paste/validate",
                base_url="http://localhost",
                headers={
                    "Origin": "http://localhost",
                    "Sec-Fetch-Site": "same-origin",
                    INVESTMENT_CSRF_HEADER: "wrong-session-token",
                },
                json={
                    "cash_account_text": self._build_hsbc_non_usd_cash_paste(),
                    "portfolio_text": "",
                    "order_status_text": "",
                },
            )

        self.assertEqual(response.status_code, 403)
        mocked_validator.assert_not_called()
        self.assertFalse(self.investment_store_path.exists())

    def test_hsbc_non_usd_cash_only_paste_import_route_persists_the_cash_ledger(self) -> None:
        client = create_app().test_client()

        with patch("app.web.runtime.threading.Thread"):
            response = self._post_investment_import(
                client,
                "/api/investment/transactions",
                data={
                    "broker": "hsbc",
                    "hsbc_import_mode": "paste",
                    "hsbc_cash_account_text": self._build_hsbc_non_usd_cash_paste(),
                    "hsbc_portfolio_text": "",
                    "hsbc_order_status_text": "",
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 200)
        response_payload = response.get_json()
        self.assertTrue(response_payload["success"])
        self.assertIn("cash-only", response_payload["message"])
        persisted = load_investment_store_payload(self.investment_store_path)
        self.assertEqual(
            persisted["ending_cash_by_currency"],
            {"HKD": "1000.00", "CNH": "12.00"},
        )
        self.assertEqual(persisted["summary"]["hsbc_paste_import_scope"], "cash_only_non_usd")

    def test_hsbc_readback_accepts_authoritative_statement_dividend_upgrade(self) -> None:
        existing_payload = {
            "broker": "hsbc",
            "account": "1234",
            "transactions": [{
                "broker": "hsbc",
                "account": "1234",
                "date": "2026-06-30",
                "type": "dividend",
                "ticker": "SGOV",
                "currency": "USD",
                "net_amount_raw": "12.34",
                "description": "Authoritative statement dividend",
                "source": {
                    "corporate_action_reference": "CA-123",
                    "file_kind": "hsbc_investment_statement_pdf",
                },
            }],
        }
        incoming_payload = {
            "broker": "hsbc",
            "account": "1234",
            "transactions": [{
                "broker": "hsbc",
                "account": "1234",
                "date": "2026-06-30",
                "type": "dividend",
                "ticker": "SGOV",
                "currency": "USD",
                "net_amount_raw": "12.34",
                "description": "CORP EVT PAYMENT",
                "source": {
                    "dividend_attribution_status": "matched",
                    "file_kind": "hsbc_usd_account_text",
                },
            }],
        }
        save_investment_store_payload(existing_payload, self.investment_store_path)
        client = create_app().test_client()

        with (
            patch(
                "app.services.investment_import.build_investment_payload_from_hsbc_pasted_text",
                return_value=incoming_payload,
            ),
            patch("app.web.runtime.ensure_latest_investment_daily_caches", return_value=[]),
        ):
            response = self._post_investment_import(
                client,
                "/api/investment/transactions",
                data={"broker": "hsbc", "hsbc_import_mode": "paste"},
            )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.get_json()["success"])
        persisted = load_investment_store_payload(self.investment_store_path)
        self.assertEqual(len(persisted["transactions"]), 1)
        self.assertEqual(
            persisted["transactions"][0]["source"]["corporate_action_reference"],
            "CA-123",
        )

    def test_more_investment_page_exposes_markdown_export_button(self) -> None:
        client = create_app().test_client()

        response = client.get("/trade/investment")
        body = response.get_data(as_text=True)

        self.assertIn('class="export-transactions-button"', body)
        self.assertIn('id="export_transactions_button"', body)
        self.assertIn('title="Export Transactions"', body)

    def test_legacy_invest_routes_redirect_to_trade_investment(self) -> None:
        client = create_app().test_client()

        response = client.get("/invest")

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], "/trade/investment")

        alias_response = client.get("/more/invest")

        self.assertEqual(alias_response.status_code, 302)
        self.assertEqual(alias_response.headers["Location"], "/trade/investment")

    def test_primary_workspace_pages_render_after_runtime_split(self) -> None:
        with (
            patch("app.web.runtime.ensure_latest_daily_caches", return_value=[]),
            patch("app.web.runtime.fetch_history", side_effect=fetch_history_stub),
            patch("app.web.runtime.fetch_quote_profile", side_effect=quote_profile_stub),
            patch("app.web.runtime.instantiate_strategy", return_value=FakeStrategy()),
            patch("app.web.runtime.run_single_ticker_backtest", return_value=backtest_result()),
            patch("app.web.runtime.record_strategy_usage"),
        ):
            client = create_app().test_client()
            responses = {
                "/workspaces/compare?ticker=QQQ&ticker=AAPL&period=3y&dividends=1": client.get(
                    "/workspaces/compare?ticker=QQQ&ticker=AAPL&period=3y&dividends=1"
                ),
                "/workspaces/portfolio?ticker=QQQ&ticker=AAPL&weight=60&weight=40&period=3y&dividends=1": client.get(
                    "/workspaces/portfolio?ticker=QQQ&ticker=AAPL&weight=60&weight=40&period=3y&dividends=1"
                ),
                "/workspaces/backtest?ticker=QQQ&strategy=buy-and-hold&period=1y&capital=10000": client.get(
                    "/workspaces/backtest?ticker=QQQ&strategy=buy-and-hold&period=1y&capital=10000"
                ),
                "/settings/network": client.get("/settings/network"),
            }

        self.assertEqual(
            {path: response.status_code for path, response in responses.items()},
            {path: 200 for path in responses},
        )
        network_body = responses["/settings/network"].get_data(as_text=True)
        self.assertIn('class="settings-service-name">Network self-check</p>', network_body)
        self.assertIn('data-network-refresh-button', network_body)
        portfolio_body = responses[
            "/workspaces/portfolio?ticker=QQQ&ticker=AAPL&weight=60&weight=40&period=3y&dividends=1"
        ].get_data(as_text=True)
        self.assertIn(
            'class="portfolio-summary-range workspace-result-date-range"',
            portfolio_body,
        )

    def test_grid_trading_uses_sentence_case_workspace_label(self) -> None:
        client = create_app().test_client()

        response = client.get("/workspaces/grid-trading")

        self.assertEqual(response.status_code, 200)
        body = response.get_data(as_text=True)
        self.assertIn('class="settings-nav-label">Grid trading</span>', body)
        self.assertIn('<p class="report-heading">Grid trading</p>', body)

    def test_exact_range_markup_exposes_shared_date_roles(self) -> None:
        client = create_app().test_client()

        responses = {
            "/workspaces/compare": client.get("/workspaces/compare?ticker=QQQ&ticker=AAPL&range=exact&from=2026-03-27&to=2026-03-28"),
            "/workspaces/portfolio": client.get(
                "/workspaces/portfolio?ticker=QQQ&ticker=AAPL&weight=60&weight=40&range=exact&from=2026-03-27&to=2026-03-28"
            ),
            "/workspaces/backtest": client.get("/workspaces/backtest?ticker=QQQ&strategy=buy-and-hold&range=exact&from=2026-03-27&to=2026-03-28"),
        }

        for response in responses.values():
            body = response.get_data(as_text=True)
            self.assertIn('data-date-role="start"', body)
            self.assertIn('data-date-role="end"', body)

    def test_refactored_runtime_apis_respond_successfully(self) -> None:
        client = create_app().test_client()

        with patch(
            "app.web.runtime.run_network_self_check",
            return_value={"rows": [], "transport_note": "Checks completed."},
        ):
            responses = {
                "/api/date-constraints?period=1y&interval=1d": client.get("/api/date-constraints?period=1y&interval=1d"),
                "/api/trade-strategy-fields?strategy=buy-and-hold": client.get(
                    "/api/trade-strategy-fields?strategy=buy-and-hold"
                ),
                "/api/settings/network-status": client.get("/api/settings/network-status"),
                "/api/settings/local-market-store/page-data?page=1": client.get(
                    "/api/settings/local-market-store/page-data?page=1"
                ),
                "/api/market-store/presence?ticker=AAPL": client.get("/api/market-store/presence?ticker=AAPL"),
            }

        self.assertEqual(
            {path: response.status_code for path, response in responses.items()},
            {path: 200 for path in responses},
        )

    def test_local_market_store_page_shows_short_history_badges_for_newly_listed_tickers(self) -> None:
        client = create_app().test_client()
        ticker = "DRAM"
        with TemporaryDirectory() as market_temp_dir:
            history_path = Path(market_temp_dir) / "historical" / f"{ticker}.parquet"
            history_path.parent.mkdir(parents=True, exist_ok=True)
            pd.DataFrame(
                {
                    "Date": pd.to_datetime(["2026-04-02"]),
                    "Close": [25.0],
                }
            ).to_parquet(history_path, index=False)

            with (
                patch("app.web.runtime.list_local_tickers", return_value=[ticker]),
                patch("app.web.runtime.history_store_path_for", return_value=history_path),
                patch("app.web.runtime.has_profile_record", return_value=True),
                patch("app.web.runtime.has_logo_asset", return_value=True),
                patch(
                    "app.web.runtime.load_profile_record",
                    return_value={"company_name": "Roundhill Memory ETF"},
                ),
                patch(
                    "app.web.runtime.resolve_stored_logo_url",
                    return_value="/market-store/logos/DRAM.png",
                ),
                patch("app.web.runtime.classify_daily_store_status", return_value="short_history"),
                patch("app.web.runtime.classify_one_minute_store_status", return_value="short_history"),
            ):
                response = client.get("/settings/local-market-store")

            self.assertEqual(response.status_code, 200)
            body = response.get_data(as_text=True)
            self.assertIn('data-local-store-ticker="DRAM"', body)
            self.assertEqual(body.count('data-local-store-status="short-history"'), 2)
            self.assertNotIn('value="refresh"', body)
            self.assertNotIn('value="refresh-1m"', body)

    def test_local_market_store_pagination_is_nested_in_the_table_shell(self) -> None:
        client = create_app().test_client()
        tickers = [f"TEST{index:02d}" for index in range(11)]
        with TemporaryDirectory() as market_temp_dir:
            history_path = Path(market_temp_dir) / "history.parquet"
            pd.DataFrame({"Date": pd.to_datetime(["2026-04-02"])}).to_parquet(history_path, index=False)

            with (
                patch("app.web.runtime.list_local_tickers", return_value=tickers),
                patch("app.web.runtime.history_store_path_for", return_value=history_path),
                patch("app.web.runtime.has_profile_record", return_value=True),
                patch("app.web.runtime.has_logo_asset", return_value=True),
                patch("app.web.runtime.load_profile_record", return_value={"company_name": "Test Company"}),
                patch("app.web.runtime.resolve_stored_logo_url", return_value="/market-store/logos/test.png"),
                patch("app.web.runtime.classify_daily_store_status", return_value="fresh"),
                patch("app.web.runtime.classify_one_minute_store_status", return_value="fresh"),
            ):
                response = client.get("/settings/local-market-store")

        self.assertEqual(response.status_code, 200)
        parser = _LocalStorePaginationStructureParser()
        parser.feed(response.get_data(as_text=True))
        self.assertEqual(parser.pagination_parent_ids, ["local_store_region"])
        self.assertEqual(
            parser.pagination_classes,
            [
                "settings-pagination local-store-pagination "
                "local-store-pagination--floating local-store-table-pagination"
            ],
        )
        self.assertIn("local-store-pagination-host", parser.pagination_parent_classes[0])
        self.assertIn("has-floating-pagination", parser.pagination_parent_classes[0])
        self.assertEqual(parser.pagination_attributes[0]["data-pagination-page-count"], "2")
        self.assertEqual(parser.pagination_attributes[0]["data-pagination-current-page"], "1")
        self.assertEqual(parser.pagination_attributes[0]["data-pagination-compact"], "1")
        self.assertEqual(parser.pagination_attributes[0]["aria-controls"], "local_store_table_body")
        self.assertEqual(
            parser.pagination_attributes[0]["data-pagination-scroll-target"],
            "local_store_table_scroll",
        )
        self.assertEqual(parser.pagination_indicator_counts, [1])
        self.assertEqual(
            [control["aria_label"] for control in parser.pagination_controls[0]],
            ["Page 1", "Page 2"],
        )

    def test_local_market_store_pagination_matches_shared_boundary_contract(self) -> None:
        client = create_app().test_client()
        tickers = [f"TEST{index:03d}" for index in range(501)]
        with TemporaryDirectory() as market_temp_dir:
            history_path = Path(market_temp_dir) / "history.parquet"
            pd.DataFrame({"Date": pd.to_datetime(["2026-04-02"])}).to_parquet(history_path, index=False)

            with (
                patch("app.web.runtime.list_local_tickers", return_value=tickers),
                patch("app.web.runtime.history_store_path_for", return_value=history_path),
                patch("app.web.runtime.has_profile_record", return_value=True),
                patch("app.web.runtime.has_logo_asset", return_value=True),
                patch("app.web.runtime.load_profile_record", return_value={"company_name": "Test Company"}),
                patch("app.web.runtime.resolve_stored_logo_url", return_value="/market-store/logos/test.png"),
                patch("app.web.runtime.classify_daily_store_status", return_value="fresh"),
                patch("app.web.runtime.classify_one_minute_store_status", return_value="fresh"),
            ):
                first_response = client.get("/settings/local-market-store")
                middle_response = client.get("/settings/local-market-store?page=6")

        self.assertEqual(first_response.status_code, 200)
        self.assertEqual(middle_response.status_code, 200)
        first_parser = _LocalStorePaginationStructureParser()
        first_parser.feed(first_response.get_data(as_text=True))
        middle_parser = _LocalStorePaginationStructureParser()
        middle_parser.feed(middle_response.get_data(as_text=True))

        def summarize_controls(parser: _LocalStorePaginationStructureParser) -> list[tuple[str | None, ...]]:
            return [
                (
                    control["aria_label"],
                    control["target"],
                    control["current"],
                    control["aria_current"],
                )
                for control in parser.pagination_controls[0]
            ]

        self.assertEqual(first_parser.pagination_attributes[0]["data-pagination-page-count"], "51")
        self.assertEqual(first_parser.pagination_attributes[0]["data-pagination-current-page"], "1")
        self.assertEqual(first_parser.pagination_attributes[0]["data-pagination-compact"], "0")
        self.assertEqual(first_parser.pagination_attributes[0]["style"], "--local-store-pagination-slots: 8;")
        self.assertEqual(first_parser.pagination_ellipsis_positions, [["trailing"]])
        first_body = first_response.get_data(as_text=True)
        self.assertIn('aria-label="Show later pages"', first_body)
        self.assertIn('role="menuitem"', first_body)
        self.assertIn('data-pagination-range-start="6"', first_body)
        self.assertIn('data-pagination-range-end="51"', first_body)
        self.assertIn('>46-51</a>', first_body)
        self.assertEqual(
            summarize_controls(first_parser),
            [
                ("Page 1", "1", "1", "page"),
                ("Page 2", "2", "0", None),
                ("Page 3", "3", "0", None),
                ("Page 4", "4", "0", None),
                ("Page 5", "5", "0", None),
                ("Page 51", "51", "0", None),
                ("Next page", "6", "0", None),
            ],
        )
        self.assertEqual(middle_parser.pagination_attributes[0]["data-pagination-current-page"], "6")
        self.assertEqual(middle_parser.pagination_attributes[0]["style"], "--local-store-pagination-slots: 11;")
        self.assertEqual(middle_parser.pagination_ellipsis_positions, [["leading", "trailing"]])
        middle_body = middle_response.get_data(as_text=True)
        self.assertIn('aria-label="Show earlier pages"', middle_body)
        self.assertIn('data-pagination-range-start="1"', middle_body)
        self.assertIn('>1-5</a>', middle_body)
        self.assertEqual(
            summarize_controls(middle_parser),
            [
                ("Previous page", "5", "0", None),
                ("Page 1", "1", "0", None),
                ("Page 6", "6", "1", "page"),
                ("Page 7", "7", "0", None),
                ("Page 8", "8", "0", None),
                ("Page 9", "9", "0", None),
                ("Page 10", "10", "0", None),
                ("Page 51", "51", "0", None),
                ("Next page", "11", "0", None),
            ],
        )

    def test_ibkr_csv_import_merges_incrementally_into_investment_store(self) -> None:
        client = create_app().test_client()
        real_store_before = (
            REAL_INVESTMENT_STORE_PATH.read_bytes()
            if REAL_INVESTMENT_STORE_PATH.exists()
            else None
        )
        original_bytes = self.investment_store_path.read_bytes() if self.investment_store_path.exists() else None
        original_payload = load_investment_store_payload(self.investment_store_path) if original_bytes else {}
        original_record_count = len(original_payload.get("transactions", []))

        try:
            with patch("app.web.runtime.ensure_latest_investment_daily_caches", return_value=[]) as mocked_refresh:
                response = self._post_investment_import(
                    client,
                    "/api/investment/transactions",
                    data={
                        "transactions_csv": (
                            io.BytesIO(self._build_sample_transactions_csv().encode("utf-8")),
                            "sample.TRANSACTIONS.1Y.csv",
                        ),
                        "positions_csv": (
                            io.BytesIO(self._build_sample_positions_csv().encode("utf-8")),
                            "sample_20260301_20260331.csv",
                        ),
                    },
                    content_type="multipart/form-data",
                )

            self.assertEqual(response.status_code, 200)
            payload = response.get_json()
            self.assertTrue(payload["success"])
            self.assertIn(
                "Exact uploaded CSV source files are retained locally as SHA-256-verified immutable evidence.",
                payload["message"],
            )
            self.assertIn("freshness_refresh_failures", payload)
            self.assertIn("summary", payload)
            self.assertTrue(str(payload["investment_store_version"]).isdigit())
            self.assertGreaterEqual(payload["transaction_count"], original_record_count)
            mocked_refresh.assert_called_once_with(["QQQ"])

            refreshed_response = client.get(
                "/api/investment/transactions",
                query_string={"store_version": payload["investment_store_version"]},
            )
            refreshed_payload = refreshed_response.get_json()
            self.assertEqual(
                refreshed_payload["investment_store_version"],
                payload["investment_store_version"],
            )

            stored = load_investment_store_payload(self.investment_store_path)
            incremental_import = stored["summary"]["incremental_import"]
            self.assertEqual(incremental_import["imported_record_count"], 1)
            self.assertIn(incremental_import["added_record_count"], {0, 1})
            self.assertIn(incremental_import["duplicate_record_count"], {0, 1})
            self.assertEqual(
                incremental_import["added_record_count"]
                + incremental_import["duplicate_record_count"],
                1,
            )
            self.assertGreaterEqual(
                stored["summary"]["total_record_count"],
                original_record_count,
            )
            self.assertTrue(
                any(
                    transaction.get("ticker") == "QQQ"
                    and transaction.get("broker") == "ibkr"
                    for transaction in stored["transactions"]
                )
            )
            real_store_after = (
                REAL_INVESTMENT_STORE_PATH.read_bytes()
                if REAL_INVESTMENT_STORE_PATH.exists()
                else None
            )
            self.assertEqual(real_store_after, real_store_before)
        finally:
            if original_bytes is None:
                if self.investment_store_path.exists():
                    self.investment_store_path.unlink()
            else:
                self.investment_store_path.parent.mkdir(parents=True, exist_ok=True)
                self.investment_store_path.write_bytes(original_bytes)

    def test_ibkr_web_paste_route_commits_provisional_trade_with_evidence(self) -> None:
        client = create_app().test_client()

        with patch(
            "app.web.runtime.ensure_latest_investment_daily_caches",
            return_value=[],
        ):
            response = self._post_investment_import(
                client,
                "/api/investment/transactions",
                data={
                    "broker": "ibkr",
                    "ibkr_import_mode": "web_paste",
                    "ibkr_trade_notifications_text": (
                        self._build_ibkr_web_trade_notifications_text()
                    ),
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 200)
        response_payload = response.get_json()
        self.assertTrue(response_payload["success"])
        self.assertIn(
            "supplemental records after the available file-snapshot cutoff",
            response_payload["message"],
        )
        stored = load_investment_store_payload(self.investment_store_path)
        self.assertEqual(len(stored["transactions"]), 1)
        transaction = stored["transactions"][0]
        self.assertEqual(transaction["datetime"], "2026-07-28 22:00:00")
        self.assertEqual(transaction["ticker"], "DRAM")
        self.assertEqual(
            transaction["source"]["file_kind"],
            "ibkr_web_trade_notification",
        )
        self.assertTrue(
            stored["summary"]["holdings_validation"]["matched"]
        )
        artifact = stored["source_artifacts"][0]
        self.assertEqual(
            artifact["source_kind"],
            "ibkr_web_trade_notifications_text",
        )
        self.assertNotIn("content_base64", artifact)
        evidence_path = investment_evidence_dir_for(
            self.investment_store_path
        ) / f"{artifact['storage_key']}.bin"
        self.assertTrue(evidence_path.exists())

    def test_ibkr_current_web_paste_route_commits_split_fills_and_calibration(
        self,
    ) -> None:
        client = create_app().test_client()

        with patch(
            "app.web.runtime.ensure_latest_investment_daily_caches",
            return_value=[],
        ):
            response = self._post_investment_import(
                client,
                "/api/investment/transactions",
                data={
                    "broker": "ibkr",
                    "ibkr_import_mode": "web_paste",
                    "ibkr_trade_notifications_text": (
                        self._build_ibkr_current_web_trade_notifications_text()
                    ),
                    "ibkr_trade_notifications_cash": "123.45",
                    "ibkr_trade_notifications_positions": "BETA 27\nALFA 9\nGAMMA 4.25",
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 200)
        response_payload = response.get_json()
        self.assertTrue(response_payload["success"])
        self.assertIn(
            "user-confirmed current IBKR position snapshot",
            response_payload["message"],
        )
        stored = load_investment_store_payload(self.investment_store_path)
        self.assertEqual(len(stored["transactions"]), 5)
        self.assertEqual(
            sum(
                1
                for transaction in stored["transactions"]
                if transaction["ticker"] == "ALFA"
                and transaction["type"] == "sell"
            ),
            2,
        )
        self.assertEqual(
            stored["position_snapshot"],
            {
                "BETA": {
                    "asset_category": "Stocks",
                    "currency": "USD",
                    "quantity": "27",
                    "as_of": "2025-08-03 08:18:00",
                    "cost_basis_status": "unknown",
                },
                "ALFA": {
                    "asset_category": "Stocks",
                    "currency": "USD",
                    "quantity": "9",
                    "as_of": "2025-08-03 08:18:00",
                    "cost_basis_status": "unknown",
                },
                "GAMMA": {
                    "asset_category": "Stocks",
                    "currency": "USD",
                    "quantity": "4.25",
                    "as_of": "2025-08-03 08:18:00",
                    "cost_basis_status": "unknown",
                },
            },
        )
        self.assertEqual(stored["ending_cash"], "123.45")
        self.assertEqual(
            stored["summary"]["position_snapshot_source"],
            "ibkr_user_verified_app_positions",
        )
        self.assertTrue(stored["summary"]["position_snapshot_authoritative"])
        self.assertEqual(
            stored["broker_snapshots"]["ibkr:U00000001"]["position_snapshot"],
            stored["position_snapshot"],
        )
        self.assertEqual(
            stored["broker_summaries"]["ibkr"]["ending_cash"],
            "123.45",
        )

        net_amounts = [
            Decimal(transaction["net_amount_raw"])
            for transaction in stored["transactions"]
        ]
        self.assertEqual(sum(net_amounts), Decimal("-21.18"))

    def test_investment_import_rejects_cross_origin_with_valid_session_token(
        self,
    ) -> None:
        client = create_app().test_client()
        client.get("/trade/investment", base_url="http://localhost")
        with client.session_transaction() as browser_session:
            csrf_token = browser_session[INVESTMENT_CSRF_SESSION_KEY]

        response = client.post(
            "/api/investment/transactions",
            base_url="http://localhost",
            headers={
                "Origin": "https://attacker.example",
                "Sec-Fetch-Site": "cross-site",
                INVESTMENT_CSRF_HEADER: csrf_token,
            },
            data={
                "broker": "ibkr",
                "ibkr_import_mode": "web_paste",
                "ibkr_trade_notifications_text": (
                    self._build_ibkr_web_trade_notifications_text()
                ),
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 403)
        self.assertFalse(response.get_json()["success"])
        self.assertIn("cross-origin", response.get_json()["error"])
        self.assertFalse(self.investment_store_path.exists())
        self.assertEqual(
            response.headers["Cache-Control"],
            "no-store, no-cache, max-age=0, must-revalidate",
        )

    def test_investment_import_rejects_invalid_same_origin_csrf_token(
        self,
    ) -> None:
        client = create_app().test_client()
        client.get("/trade/investment", base_url="http://localhost")

        response = client.post(
            "/api/investment/transactions",
            base_url="http://localhost",
            headers={
                "Origin": "http://localhost",
                "Sec-Fetch-Site": "same-origin",
                INVESTMENT_CSRF_HEADER: "invalid-session-security-token-value",
            },
            data={
                "broker": "ibkr",
                "ibkr_import_mode": "web_paste",
                "ibkr_trade_notifications_text": (
                    self._build_ibkr_web_trade_notifications_text()
                ),
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 403)
        self.assertFalse(response.get_json()["success"])
        self.assertIn("session security token", response.get_json()["error"])
        self.assertFalse(self.investment_store_path.exists())

    def test_investment_import_rejects_non_local_same_origin_hostname(
        self,
    ) -> None:
        application = create_app()
        hostile_origin = "http://rebind.attacker.example:8688"
        csrf_token = "valid-session-security-token-value-1234"
        with application.test_request_context(
            "/api/investment/transactions",
            base_url=hostile_origin,
            method="POST",
            headers={
                "Origin": hostile_origin,
                "Sec-Fetch-Site": "same-origin",
                INVESTMENT_CSRF_HEADER: csrf_token,
            },
        ):
            session[INVESTMENT_CSRF_SESSION_KEY] = csrf_token
            error = validate_investment_browser_write_request(request)

        self.assertIn("non-local", error)
        self.assertFalse(self.investment_store_path.exists())

    def test_investment_import_request_limit_returns_no_store_json_without_writing(self) -> None:
        application = create_app()
        self.assertEqual(
            application.config["MAX_CONTENT_LENGTH"],
            MAX_INVESTMENT_IMPORT_REQUEST_BYTES,
        )
        self.assertEqual(
            MAX_INVESTMENT_IMPORT_REQUEST_BYTES,
            MAX_INVESTMENT_SOURCE_ARTIFACT_BYTES + INVESTMENT_IMPORT_MULTIPART_ALLOWANCE_BYTES,
        )
        self.assertLess(
            MAX_INVESTMENT_IMPORT_REQUEST_BYTES,
            MAX_INVESTMENT_SOURCE_EVIDENCE_BYTES,
        )
        application.config["MAX_CONTENT_LENGTH"] = 1_024
        client = application.test_client()
        temporary_store_before = (
            self.investment_store_path.read_bytes()
            if self.investment_store_path.exists()
            else None
        )
        real_store_before = (
            REAL_INVESTMENT_STORE_PATH.read_bytes()
            if REAL_INVESTMENT_STORE_PATH.exists()
            else None
        )
        temporary_evidence_dir = investment_evidence_dir_for(self.investment_store_path)
        self.assertFalse(temporary_evidence_dir.exists())

        # This is the durable ledger-write boundary used by the import route.
        with (
            patch("app.web.runtime.parse_investment_payload") as mocked_parser,
            patch(
                "app.web.runtime.update_investment_store_payload"
            ) as mocked_store_write,
        ):
            response = self._post_investment_import(
                client,
                "/api/investment/transactions",
                data={
                    "broker": "ibkr",
                    "transactions_csv": (
                        io.BytesIO(b"x" * 2_048),
                        "oversized.TRANSACTIONS.csv",
                    ),
                    "positions_csv": (
                        io.BytesIO(b"x"),
                        "oversized.summary.csv",
                    ),
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 413)
        self.assertEqual(response.get_json(), {
            "success": False,
            "error": (
                "The investment import exceeds the 65 MiB total upload limit. "
                "Upload fewer or smaller files."
            ),
        })
        self.assertEqual(
            response.headers["Cache-Control"],
            "no-store, no-cache, max-age=0, must-revalidate",
        )
        mocked_parser.assert_not_called()
        mocked_store_write.assert_not_called()
        self.assertEqual(
            self.investment_store_path.read_bytes()
            if self.investment_store_path.exists()
            else None,
            temporary_store_before,
        )
        self.assertEqual(
            REAL_INVESTMENT_STORE_PATH.read_bytes()
            if REAL_INVESTMENT_STORE_PATH.exists()
            else None,
            real_store_before,
        )
        self.assertFalse(temporary_evidence_dir.exists())

    def test_ibkr_test_fixture_account_is_rejected_before_persistence(self) -> None:
        client = create_app().test_client()
        test_transactions = self._build_sample_transactions_csv().replace(
            "U***00001",
            "U***TEST",
        )

        response = self._post_investment_import(
            client,
            "/api/investment/transactions",
            data={
                "transactions_csv": (
                    io.BytesIO(test_transactions.encode("utf-8")),
                    "sample.TRANSACTIONS.1Y.csv",
                ),
                "positions_csv": (
                    io.BytesIO(self._build_sample_positions_csv().encode("utf-8")),
                    "sample_20260301_20260331.csv",
                ),
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.get_json()["error"],
            "Refusing to persist an IBKR test-fixture account into the investment store.",
        )
        self.assertFalse(self.investment_store_path.exists())

    def test_import_prewarms_only_open_investment_tickers(self) -> None:
        client = create_app().test_client()
        original_bytes = self.investment_store_path.read_bytes() if self.investment_store_path.exists() else None

        transactions_csv = "\n".join([
            "Statement,Header,Field Name,Field Value",
            "Statement,Data,Title,Transaction History",
            "Summary,Header,Field Name,Field Value",
            "Summary,Data,Starting Cash,1000.00",
            "Summary,Data,Ending Cash,1000.00",
            "Transaction History,Header,Date,Account,Description,Transaction Type,Symbol,Quantity,Price,Price Currency,Gross Amount ,Commission,Net Amount",
            "Transaction History,Data,2026-03-01,U***00001,Open Position,Buy,QQQ,1,100,USD,-100,-1,-101",
            "Transaction History,Data,2026-03-02,U***00001,Closed Position,Buy,DRAM,1,50,USD,-50,-1,-51",
            "Transaction History,Data,2026-03-03,U***00001,Closed Position Exit,Sell,DRAM,-1,55,USD,55,-1,54",
        ]) + "\n"

        try:
            with patch("app.web.runtime.ensure_latest_investment_daily_caches", return_value=[]) as mocked_refresh:
                response = self._post_investment_import(
                    client,
                    "/api/investment/transactions",
                    data={
                        "transactions_csv": (
                            io.BytesIO(transactions_csv.encode("utf-8")),
                            "sample.TRANSACTIONS.1Y.csv",
                        ),
                        "positions_csv": (
                            io.BytesIO(self._build_sample_positions_csv().encode("utf-8")),
                            "sample_20260301_20260331.csv",
                        ),
                    },
                    content_type="multipart/form-data",
                )

            self.assertEqual(response.status_code, 200)
            mocked_refresh.assert_called_once_with(["QQQ"])
        finally:
            if original_bytes is None:
                if self.investment_store_path.exists():
                    self.investment_store_path.unlink()
            else:
                self.investment_store_path.parent.mkdir(parents=True, exist_ok=True)
                self.investment_store_path.write_bytes(original_bytes)

    def test_investment_transactions_refresh_open_tickers_via_shared_freshness_helper(self) -> None:
        client = create_app().test_client()
        original_bytes = self.investment_store_path.read_bytes() if self.investment_store_path.exists() else None

        payload = {
            "starting_cash": "100.00",
            "transactions": [
                {
                    "date": "2026-03-01",
                    "type": "buy",
                    "ticker": "QQQ",
                    "quantity_raw": "1",
                    "quantity_abs": "1",
                    "price_raw": "100",
                    "net_amount_raw": "-100",
                    "normalized": {"display_quantity": "1", "net_amount": "-100"},
                }
            ],
        }

        try:
            self.investment_store_path.parent.mkdir(parents=True, exist_ok=True)
            save_investment_store_payload(payload, self.investment_store_path)
            with patch("app.web.runtime.ensure_latest_investment_daily_caches", return_value=[]) as mocked_refresh:
                response = client.get("/api/investment/transactions")

            self.assertEqual(response.status_code, 200)
            mocked_refresh.assert_called_once_with(["QQQ"])
        finally:
            if original_bytes is None:
                if self.investment_store_path.exists():
                    self.investment_store_path.unlink()
            else:
                self.investment_store_path.parent.mkdir(parents=True, exist_ok=True)
                self.investment_store_path.write_bytes(original_bytes)

    def test_investment_transactions_attempts_profile_fetch_when_logo_asset_is_missing(self) -> None:
        client = create_app().test_client()
        original_bytes = self.investment_store_path.read_bytes() if self.investment_store_path.exists() else None

        payload = {
            "starting_cash": "100.00",
            "transactions": [
                {
                    "date": "2026-03-01",
                    "type": "buy",
                    "ticker": "SNDK",
                    "quantity_raw": "1",
                    "quantity_abs": "1",
                    "price_raw": "100",
                    "net_amount_raw": "-100",
                    "normalized": {"display_quantity": "1", "net_amount": "-100"},
                }
            ],
        }

        mock_profile = type(
            "MockQuoteProfile",
            (),
            {"company_name": "Sandisk", "logo_url": "/market-store/logos/SNDK.png"},
        )()

        try:
            self.investment_store_path.parent.mkdir(parents=True, exist_ok=True)
            save_investment_store_payload(payload, self.investment_store_path)
            with patch("app.web.runtime.ensure_latest_investment_daily_caches", return_value=[]), \
                    patch("app.web.runtime.has_logo_asset", return_value=False), \
                    patch("app.web.runtime.load_profile_record", return_value={}), \
                    patch("app.web.runtime.fetch_quote_profile", return_value=mock_profile) as mocked_fetch:
                response = client.get("/api/investment/transactions")

            self.assertEqual(response.status_code, 200)
            body = response.get_json()
            self.assertEqual(body["ticker_profiles"]["SNDK"]["company_name"], "Sandisk")
            self.assertEqual(body["ticker_profiles"]["SNDK"]["logo_url"], "/market-store/logos/SNDK.png")
            mocked_fetch.assert_called_once_with("SNDK", force_refresh=True)
        finally:
            if original_bytes is None:
                if self.investment_store_path.exists():
                    self.investment_store_path.unlink()
            else:
                self.investment_store_path.parent.mkdir(parents=True, exist_ok=True)
                self.investment_store_path.write_bytes(original_bytes)

    def test_investment_transactions_keep_standard_names_when_cached_profiles_are_symbols(self) -> None:
        client = create_app().test_client()
        expected_names = {
            "AAPL": "Apple Inc.",
            "EUV": "Corgi Lithography & Semiconductor Photonics ETF",
            "GOOGL": "Alphabet Inc.",
            "IBKR": "Interactive Brokers Group, Inc.",
            "JEPQ": "JPMorgan Nasdaq Equity Premium Income ETF",
            "META": "Meta Platforms, Inc.",
            "MU": "Micron Technology, Inc.",
            "QQQ": "Invesco QQQ Trust, Series 1",
            "QCOM": "QUALCOMM Incorporated",
        }
        payload = {
            "starting_cash": "100.00",
            "transactions": [
                {
                    "date": "2026-07-21",
                    "type": "buy",
                    "ticker": ticker,
                    "quantity_raw": "1",
                    "quantity_abs": "1",
                    "price_raw": "100",
                    "net_amount_raw": "-100",
                    "normalized": {"display_quantity": "1", "net_amount": "-100"},
                }
                for ticker in expected_names
            ],
        }

        self.investment_store_path.parent.mkdir(parents=True, exist_ok=True)
        save_investment_store_payload(payload, self.investment_store_path)
        with (
            patch("app.web.runtime.ensure_latest_investment_daily_caches", return_value=[]),
            patch("app.web.runtime.fetch_longbridge_realtime_quotes", return_value=[]),
            patch("app.web.runtime.fetch_yfinance_realtime_quotes", return_value=[]),
            patch(
                "app.web.runtime.load_profile_record",
                side_effect=lambda ticker: {"company_name": ticker},
            ),
            patch(
                "app.web.runtime.resolve_stored_logo_url",
                side_effect=lambda ticker: f"/market-store/logos/{ticker}.svg",
            ),
            patch("app.web.runtime.fetch_quote_profile") as profile_refresh_mock,
        ):
            response = client.get("/api/investment/transactions")

        self.assertEqual(response.status_code, 200)
        body = response.get_json()
        self.assertEqual(
            {
                ticker: body["ticker_profiles"][ticker]["company_name"]
                for ticker in expected_names
            },
            expected_names,
        )
        self.assertEqual(
            {
                ticker: body["known_ticker_company_names"][ticker]
                for ticker in expected_names
            },
            expected_names,
        )
        profile_refresh_mock.assert_not_called()

    def test_investment_parquet_fetches_missing_history_via_shared_market_data_path(self) -> None:
        client = create_app().test_client()
        ticker = "DRAM"
        save_investment_store_payload(
            {
                "transactions": [
                    {
                        "date": "2026-03-31",
                        "type": "buy",
                        "ticker": ticker,
                        "quantity_raw": "1",
                        "normalized": {"display_quantity": "1"},
                    }
                ]
            },
            self.investment_store_path,
        )
        with TemporaryDirectory() as market_temp_dir:
            path = Path(market_temp_dir) / "historical" / f"{ticker}.parquet"

            def _mock_fetch_history(requested_ticker: str, include_dividends: bool) -> pd.DataFrame:
                self.assertEqual(requested_ticker, ticker)
                self.assertFalse(include_dividends)
                path.parent.mkdir(parents=True, exist_ok=True)
                pd.DataFrame(
                    {
                        "Date": pd.to_datetime(["2026-03-31"]),
                        "Close": [42.0],
                    }
                ).to_parquet(path, index=False)
                return pd.DataFrame({"Date": pd.to_datetime(["2026-03-31"]), "Close": [42.0]})

            with (
                patch("app.web.runtime.history_store_path_for", return_value=path),
                patch("app.web.runtime.fetch_history", side_effect=_mock_fetch_history) as mocked_fetch,
            ):
                response = client.get(f"/api/investment/parquet?ticker={ticker}")

            self.assertEqual(response.status_code, 200)
            payload = response.get_json()
            self.assertTrue(payload["success"])
            self.assertEqual(payload["ticker"], ticker)
            self.assertEqual(payload["prices"][0]["close"], 42.0)
            mocked_fetch.assert_called_once()

    def test_investment_parquet_does_not_fetch_missing_history_for_closed_ticker(self) -> None:
        client = create_app().test_client()
        ticker = "DRAM"
        save_investment_store_payload(
            {
                "transactions": [
                    {
                        "date": "2026-03-01",
                        "type": "buy",
                        "ticker": ticker,
                        "quantity_raw": "1",
                        "normalized": {"display_quantity": "1"},
                    },
                    {
                        "date": "2026-03-02",
                        "type": "sell",
                        "ticker": ticker,
                        "quantity_raw": "1",
                        "normalized": {"display_quantity": "1"},
                    },
                ]
            },
            self.investment_store_path,
        )

        with TemporaryDirectory() as market_temp_dir:
            missing_path = Path(market_temp_dir) / "historical" / f"{ticker}.parquet"
            with (
                patch("app.web.runtime.history_store_path_for", return_value=missing_path),
                patch("app.web.runtime.fetch_history") as mocked_fetch,
            ):
                response = client.get(f"/api/investment/parquet?ticker={ticker}")

        self.assertEqual(response.status_code, 404)
        mocked_fetch.assert_not_called()

    def test_investment_transactions_skip_money_market_freshness_refresh(self) -> None:
        client = create_app().test_client()
        original_bytes = self.investment_store_path.read_bytes() if self.investment_store_path.exists() else None

        payload = {
            "starting_cash": "100.00",
            "transactions": [
                {
                    "date": "2026-03-01",
                    "type": "buy",
                    "ticker": "QQQ",
                    "quantity_raw": "1",
                    "quantity_abs": "1",
                    "price_raw": "100",
                    "net_amount_raw": "-100",
                    "normalized": {"display_quantity": "1", "net_amount": "-100"},
                },
                {
                    "date": "2026-03-02",
                    "type": "buy",
                    "ticker": "005276756",
                    "quantity_raw": "1",
                    "quantity_abs": "1",
                    "price_raw": "1",
                    "net_amount_raw": "-1",
                    "normalized": {"display_quantity": "1", "net_amount": "-1"},
                },
            ],
        }

        try:
            self.investment_store_path.parent.mkdir(parents=True, exist_ok=True)
            save_investment_store_payload(payload, self.investment_store_path)
            with patch("app.web.runtime.ensure_latest_investment_daily_caches", return_value=[]) as mocked_refresh:
                response = client.get("/api/investment/transactions")

            self.assertEqual(response.status_code, 200)
            mocked_refresh.assert_called_once_with(["QQQ"])
        finally:
            if original_bytes is None:
                if self.investment_store_path.exists():
                    self.investment_store_path.unlink()
            else:
                self.investment_store_path.parent.mkdir(parents=True, exist_ok=True)
                self.investment_store_path.write_bytes(original_bytes)

    def test_investment_parquet_skips_money_market_refresh_when_local_data_exists(self) -> None:
        client = create_app().test_client()
        ticker = "005276756"
        with TemporaryDirectory() as market_temp_dir:
            path = Path(market_temp_dir) / "historical" / f"{ticker}.parquet"
            path.parent.mkdir(parents=True, exist_ok=True)
            pd.DataFrame(
                {
                    "Date": pd.to_datetime(["2026-03-31"]),
                    "Close": [1.0],
                }
            ).to_parquet(path, index=False)

            with (
                patch("app.web.runtime.history_store_path_for", return_value=path),
                patch("app.web.runtime.ensure_latest_investment_daily_caches") as mocked_refresh,
            ):
                response = client.get(f"/api/investment/parquet?ticker={ticker}")

            self.assertEqual(response.status_code, 200)
            payload = response.get_json()
            self.assertTrue(payload["success"])
            self.assertEqual(payload["ticker"], ticker)
            self.assertEqual(payload["prices"][0]["close"], 1.0)
            mocked_refresh.assert_not_called()


if __name__ == "__main__":
    unittest.main()
