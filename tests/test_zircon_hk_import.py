"""Tests for the generic typed manual-workbook import.

Code version: v0.12.0
"""

from __future__ import annotations

import base64
from datetime import datetime
from io import BytesIO
import unittest
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook

from app.services.investment_import import (
    merge_investment_payloads,
    parse_investment_payload,
)
from app.services.zircon_hk_import import (
    ZIRCON_HK_BROKER_ENTRIES,
    ZIRCON_HK_CURRENCIES,
    ZIRCON_HK_HEADERS,
    ZIRCON_HK_LISTS_SHEET,
    ZIRCON_HK_MAX_TRANSACTION_ROWS,
    ZIRCON_HK_TEMPLATE_INPUT_ROWS,
    ZIRCON_HK_TRANSACTION_SHEET,
    ZIRCON_HK_TYPE_LABELS,
    build_investment_payload_from_zircon_hk_manual_xlsx,
    build_standard_investment_xlsx,
    build_zircon_hk_template_xlsx,
)


class ZirconHkImportTests(unittest.TestCase):
    def _completed_workbook(
        self,
        *,
        overrides: dict[str, object] | None = None,
    ) -> bytes:
        workbook = load_workbook(BytesIO(build_zircon_hk_template_xlsx()))
        sheet = workbook["Transactions"]
        values: dict[str, object] = {
            "A2": "Zircon (HK)",
            "B2": "Zircon practice account",
            "C2": datetime(2026, 7, 30, 21, 15),
            "D2": "Buy",
            "E2": "HKD",
            "F2": "700.HK",
            "G2": 100,
            "H2": 20,
            "I2": None,
            "J2": 5,
            "K2": "Manual activity entry",
            "L2": "zircon-reference-1",
        }
        values.update(overrides or {})
        for coordinate, value in values.items():
            sheet[coordinate] = value
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _forex_workbook(
        self,
        *,
        acquired_amount: float = 83.22,
        acquired_datetime: datetime | None = None,
        reference_id: str = "fx-20260730-01",
    ) -> bytes:
        transaction_datetime = datetime(2026, 7, 30, 21, 15)
        return self._completed_workbook(
            overrides={
                "C2": transaction_datetime,
                "D2": "Forex trade component",
                "E2": "HKD",
                "F2": None,
                "G2": None,
                "H2": None,
                "I2": -650,
                "J2": None,
                "K2": "HKD to USD",
                "L2": reference_id,
            "A3": "Zircon (HK)",
                "B3": "Zircon practice account",
                "C3": acquired_datetime or transaction_datetime,
                "D3": "Forex trade component",
                "E3": "USD",
                "I3": acquired_amount,
                "K3": "HKD to USD",
                "L3": reference_id,
            }
        )

    def test_template_contains_dropdowns_typed_validation_and_no_sample_records(
        self,
    ) -> None:
        workbook_bytes = build_zircon_hk_template_xlsx()
        workbook = load_workbook(BytesIO(workbook_bytes))
        self.assertEqual(
            workbook.sheetnames,
            [ZIRCON_HK_TRANSACTION_SHEET, ZIRCON_HK_LISTS_SHEET],
        )
        self.assertEqual(workbook.active.title, ZIRCON_HK_TRANSACTION_SHEET)
        self.assertEqual(workbook[ZIRCON_HK_LISTS_SHEET].sheet_state, "hidden")
        sheet = workbook[ZIRCON_HK_TRANSACTION_SHEET]
        self.assertEqual(
            tuple(cell.value for cell in sheet[1]),
            ZIRCON_HK_HEADERS,
        )
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertFalse(sheet.tables)
        self.assertEqual(len(sheet.conditional_formatting), 0)
        self.assertTrue(
            all(
                sheet.cell(row=row, column=column).value is None
                for row in range(2, ZIRCON_HK_TEMPLATE_INPUT_ROWS + 2)
                for column in range(1, len(ZIRCON_HK_HEADERS) + 1)
            )
        )
        validation_types = {
            validation.type for validation in sheet.data_validations.dataValidation
        }
        self.assertEqual(validation_types, {"list", "date", "decimal"})
        validation_end_row = ZIRCON_HK_TEMPLATE_INPUT_ROWS + 1
        self.assertEqual(
            [
                workbook[ZIRCON_HK_LISTS_SHEET].cell(row=row, column=1).value
                for row in range(2, len(ZIRCON_HK_TYPE_LABELS) + 2)
            ],
            list(ZIRCON_HK_TYPE_LABELS),
        )
        self.assertEqual(
            [
                workbook[ZIRCON_HK_LISTS_SHEET].cell(row=row, column=2).value
                for row in range(2, len(ZIRCON_HK_CURRENCIES) + 2)
            ],
            list(ZIRCON_HK_CURRENCIES),
        )
        self.assertEqual(
            {
                name: workbook.defined_names[name].attr_text
                for name in ("ZirconTypes", "ZirconCurrencies", "ZirconBrokers")
            },
            {
                "ZirconTypes": (
                    f"'{ZIRCON_HK_LISTS_SHEET}'!$A$2:"
                    f"$A${len(ZIRCON_HK_TYPE_LABELS) + 1}"
                ),
                "ZirconCurrencies": (
                    f"'{ZIRCON_HK_LISTS_SHEET}'!$B$2:"
                    f"$B${len(ZIRCON_HK_CURRENCIES) + 1}"
                ),
                "ZirconBrokers": (
                    f"'{ZIRCON_HK_LISTS_SHEET}'!$C$2:"
                    f"$C${len(ZIRCON_HK_BROKER_ENTRIES) + 1}"
                ),
            },
        )
        self.assertEqual(
            {
                (
                    validation.type,
                    validation.formula1,
                    validation.formula2,
                    str(validation.sqref),
                )
                for validation in sheet.data_validations.dataValidation
            },
            {
                ("list", "ZirconBrokers", None, f"A2:A{validation_end_row}"),
                ("list", "ZirconTypes", None, f"D2:D{validation_end_row}"),
                ("list", "ZirconCurrencies", None, f"E2:E{validation_end_row}"),
                (
                    "date",
                    "DATE(2000,1,1)",
                    "DATE(2100,12,31)",
                    f"C2:C{validation_end_row}",
                ),
                ("decimal", "0", "1000000000000", f"G2:G{validation_end_row}"),
                ("decimal", "0", "1000000000000", f"H2:H{validation_end_row}"),
                (
                    "decimal",
                    "-1000000000000",
                    "1000000000000",
                    f"I2:I{validation_end_row}",
                ),
                ("decimal", "0", "1000000000000", f"J2:J{validation_end_row}"),
            },
        )
        with ZipFile(BytesIO(workbook_bytes)) as archive:
            transactions_xml = archive.read("xl/worksheets/sheet1.xml")
            archive_members = set(archive.namelist())
        self.assertNotIn(b"<conditionalFormatting", transactions_xml)
        self.assertNotIn(b"<tableParts", transactions_xml)
        self.assertFalse(any(name.startswith("xl/tables/") for name in archive_members))
        for name in (b"ZirconBrokers", b"ZirconTypes", b"ZirconCurrencies"):
            self.assertIn(b"<formula1>" + name + b"</formula1>", transactions_xml)
            self.assertNotIn(b"<formula1>=" + name + b"</formula1>", transactions_xml)
        self.assertEqual(
            [
                workbook["Lists"].cell(row=row, column=3).value
                for row in range(2, len(ZIRCON_HK_BROKER_ENTRIES) + 2)
            ],
            [entry.label for entry in ZIRCON_HK_BROKER_ENTRIES],
        )

        listed_brokers = {
            workbook["Lists"].cell(row=row, column=3).value
            for row in range(2, len(ZIRCON_HK_BROKER_ENTRIES) + 2)
        }
        self.assertTrue({
            "China Merchants Bank",
            "Bank of China",
            "Bank of China (Hong Kong)",
            "Industrial and Commercial Bank of China",
            "Industrial and Commercial Bank of China (Asia)",
            "China Construction Bank",
            "China Construction Bank (Asia)",
        }.issubset(listed_brokers))
        with ZipFile(BytesIO(workbook_bytes)) as archive:
            self.assertIsNone(archive.testzip())
            archive_members = set(archive.namelist())
        self.assertIn("[Content_Types].xml", archive_members)
        self.assertIn("xl/workbook.xml", archive_members)
        self.assertFalse(
            any(
                member == "xl/vbaProject.bin"
                or member == "xl/connections.xml"
                or member.startswith(("xl/embeddings/", "xl/externalLinks/", "xl/oleObjects/"))
                for member in archive_members
            )
        )

    def test_standard_export_round_trips_every_supported_transaction_type(self) -> None:
        def transaction(
            ledger_no: int,
            transaction_type: str,
            *,
            currency: str = "USD",
            ticker: str = "",
            source: dict[str, object] | None = None,
            **fields: object,
        ) -> dict[str, object]:
            return {
                "ledger_no": ledger_no,
                "broker": "ibkr",
                "account": "U-XLSX-HEALTH",
                "datetime": "2026-07-30 09:15:00",
                "type": transaction_type,
                "currency": currency,
                "ticker": ticker,
                "source": source or {},
                **fields,
            }

        transactions = [
            transaction(1, "buy", ticker="AAPL", quantity_raw="2", price_raw="100"),
            transaction(2, "sell", ticker="AAPL", quantity_raw="1", price_raw="120"),
            transaction(3, "deposit", net_amount_raw="1000"),
            transaction(4, "withdrawal", net_amount_raw="-50"),
            transaction(5, "virtual_balance_reset", currency="CNY", net_amount_raw="-21511.9"),
            transaction(6, "dividend", ticker="AAPL", net_amount_raw="5"),
            transaction(
                7,
                "dividend_reinvestment",
                ticker="AAPL",
                quantity_raw="1",
                price_raw="5",
            ),
            transaction(8, "fee", net_amount_raw="-2"),
            transaction(9, "credit_interest", net_amount_raw="1"),
            transaction(10, "debit_interest", net_amount_raw="-1"),
            transaction(
                11,
                "foreign_tax_withholding",
                ticker="AAPL",
                net_amount_raw="-0.5",
            ),
            transaction(12, "payment_in_lieu", ticker="AAPL", net_amount_raw="0.5"),
            transaction(13, "adjustment", net_amount_raw="2"),
            transaction(14, "grant", ticker="AAPL", quantity_raw="3", price_raw="0"),
            transaction(15, "kol_reward", net_amount_raw="10"),
            transaction(16, "fx_translation_pnl", net_amount_raw="-3"),
            transaction(17, "transfer_in", ticker="AAPL", quantity_raw="4", price_raw="100"),
            transaction(18, "transfer_out", ticker="AAPL", quantity_raw="2", price_raw="100"),
            transaction(
                19,
                "forex_trade_component",
                currency="HKD",
                net_amount_raw="-100",
                source={"reference_id": "xlsx-health-fx"},
            ),
            transaction(
                20,
                "forex_trade_component",
                currency="USD",
                net_amount_raw="12",
                source={"reference_id": "xlsx-health-fx"},
            ),
            transaction(21, "virtual_deposit", net_amount_raw="25"),
            transaction(22, "virtual_withdrawal", net_amount_raw="-25"),
        ]

        workbook_bytes = build_standard_investment_xlsx(transactions)
        payload = build_investment_payload_from_zircon_hk_manual_xlsx(
            xlsx_bytes=workbook_bytes,
            filename="Standard_investment_export.xlsx",
        )

        self.assertEqual(
            {transaction["type"] for transaction in payload["transactions"]},
            set(ZIRCON_HK_TYPE_LABELS.values()),
        )
        self.assertEqual(payload["summary"]["transaction_count"], len(transactions))
        with ZipFile(BytesIO(workbook_bytes)) as archive:
            self.assertIsNone(archive.testzip())

    def test_standard_export_round_trips_every_importable_broker_choice(self) -> None:
        transactions = [
            {
                "ledger_no": index,
                "broker": entry.code,
                "account": f"account-{entry.code}",
                "datetime": "2026-07-30 09:15:00",
                "type": "buy",
                "currency": "USD",
                "ticker": "AAPL",
                "quantity_raw": "1",
                "price_raw": "100",
                "commission_raw": "0",
                "source": {},
            }
            for index, entry in enumerate(ZIRCON_HK_BROKER_ENTRIES, start=1)
        ]

        payload = build_investment_payload_from_zircon_hk_manual_xlsx(
            xlsx_bytes=build_standard_investment_xlsx(transactions),
            filename="Standard_investment_export.xlsx",
        )

        self.assertEqual(
            {transaction["broker"] for transaction in payload["transactions"]},
            {entry.code for entry in ZIRCON_HK_BROKER_ENTRIES},
        )
        self.assertEqual(
            payload["summary"]["transaction_count"],
            len(ZIRCON_HK_BROKER_ENTRIES),
        )

    def test_parser_accepts_legacy_zircon_hk_label_without_parentheses(self) -> None:
        payload = build_investment_payload_from_zircon_hk_manual_xlsx(
            xlsx_bytes=self._completed_workbook(overrides={"A2": "Zircon HK"}),
            filename="Zircon_manual_investment_import.xlsx",
        )

        self.assertEqual(payload["broker"], "zircon_hk")
        self.assertEqual(payload["transactions"][0]["broker"], "zircon_hk")

    def test_standard_export_shares_fx_execution_identity_across_currency_legs(self) -> None:
        transactions = [
            {
                "ledger_no": 101,
                "broker": "ibkr",
                "account": "U-FX-IDENTITY",
                "datetime": "2026-07-30 09:15:00",
                "type": "forex_trade_component",
                "currency": "HKD",
                "net_amount_raw": "-100",
                "source": {
                    "execution_key": "provider-fx-001:HKD",
                    "row_number": 42,
                },
            },
            {
                "ledger_no": 102,
                "broker": "ibkr",
                "account": "U-FX-IDENTITY",
                "datetime": "2026-07-30 09:15:00",
                "type": "forex_trade_component",
                "currency": "USD",
                "net_amount_raw": "12",
                "source": {
                    "execution_key": "provider-fx-001:USD",
                    "row_number": 43,
                },
            },
        ]

        workbook_bytes = build_standard_investment_xlsx(transactions)
        workbook = load_workbook(BytesIO(workbook_bytes))
        sheet = workbook[ZIRCON_HK_TRANSACTION_SHEET]
        self.assertEqual(sheet["L2"].value, sheet["L3"].value)
        self.assertRegex(sheet["L2"].value, r"^antigravity-[0-9a-f]{40}$")

        payload = build_investment_payload_from_zircon_hk_manual_xlsx(
            xlsx_bytes=workbook_bytes,
            filename="Standard_investment_export.xlsx",
        )
        self.assertEqual(payload["summary"]["transaction_count"], 2)

    def test_standard_export_suffixes_repeated_references_deterministically(self) -> None:
        transactions = [
            {
                "ledger_no": 201,
                "broker": "hsbc",
                "account": "HSBC-REFERENCE-COLLISION",
                "datetime": "2026-07-30 09:15:00",
                "type": "deposit",
                "currency": "USD",
                "net_amount_raw": "100",
                "source": {"reference_id": "provider-ref-001"},
            },
            {
                "ledger_no": 202,
                "broker": "hsbc",
                "account": "HSBC-REFERENCE-COLLISION",
                "datetime": "2026-07-30 09:16:00",
                "type": "withdrawal",
                "currency": "USD",
                "net_amount_raw": "-20",
                "source": {"reference_id": "provider-ref-001"},
            },
        ]

        workbook_bytes = build_standard_investment_xlsx(transactions)
        first_workbook = load_workbook(BytesIO(workbook_bytes))
        second_workbook = load_workbook(BytesIO(workbook_bytes))
        first_references = [
            first_workbook[ZIRCON_HK_TRANSACTION_SHEET].cell(row=row, column=12).value
            for row in (2, 3)
        ]
        second_references = [
            second_workbook[ZIRCON_HK_TRANSACTION_SHEET].cell(row=row, column=12).value
            for row in (2, 3)
        ]
        self.assertEqual(first_references[0], "provider-ref-001")
        self.assertEqual(first_references, second_references)
        self.assertRegex(
            first_references[1],
            r"^provider-ref-001::antigravity-[0-9a-f]{16}$",
        )
        self.assertEqual(len(set(first_references)), 2)

        payload = build_investment_payload_from_zircon_hk_manual_xlsx(
            xlsx_bytes=workbook_bytes,
            filename="Standard_investment_export.xlsx",
        )
        self.assertEqual(payload["summary"]["transaction_count"], 2)

    def test_parser_reports_the_current_maximum_transaction_row_contract(self) -> None:
        workbook = load_workbook(BytesIO(build_zircon_hk_template_xlsx()))
        sheet = workbook[ZIRCON_HK_TRANSACTION_SHEET]
        sheet.cell(row=ZIRCON_HK_MAX_TRANSACTION_ROWS + 2, column=1, value="Zircon (HK)")
        buffer = BytesIO()
        workbook.save(buffer)

        with self.assertRaisesRegex(ValueError, r"10,000 transaction rows"):
            build_investment_payload_from_zircon_hk_manual_xlsx(
                xlsx_bytes=buffer.getvalue(),
                filename="too-many-rows.xlsx",
            )

    def test_manual_workbook_accepts_new_bank_institutions_and_requested_currencies(self) -> None:
        cases = (
            ("China Merchants Bank", "CNY", "cmb-cn-cny"),
            ("Bank of China", "CNY", "boc-cn-cny"),
            ("Bank of China", "HKD", "boc-cn-hkd"),
            ("Bank of China", "USD", "boc-cn-usd"),
            ("Bank of China (Hong Kong)", "CNY", "boc-hk-cny"),
            ("Bank of China (Hong Kong)", "CNH", "boc-hk-cnh"),
            ("Bank of China (Hong Kong)", "HKD", "boc-hk-hkd"),
            ("Bank of China (Hong Kong)", "USD", "boc-hk-usd"),
            ("Industrial and Commercial Bank of China", "CNY", "icbc-cn-cny"),
            ("Industrial and Commercial Bank of China (Asia)", "HKD", "icbc-hk-hkd"),
            ("China Construction Bank", "CNY", "ccb-cn-cny"),
            ("China Construction Bank (Asia)", "HKD", "ccb-hk-hkd"),
        )

        for label, currency, reference_id in cases:
            with self.subTest(label=label, currency=currency):
                payload = build_investment_payload_from_zircon_hk_manual_xlsx(
                    xlsx_bytes=self._completed_workbook(
                        overrides={
                            "A2": label,
                            "E2": currency,
                            "L2": reference_id,
                        }
                    ),
                    filename="manual-bank-account.xlsx",
                )

                transaction = payload["transactions"][0]
                self.assertEqual(transaction["currency"], currency)
                self.assertEqual(transaction["broker"], {
                    "China Merchants Bank": "cmb_cn",
                    "Bank of China": "boc_cn",
                    "Bank of China (Hong Kong)": "boc_hk",
                    "Industrial and Commercial Bank of China": "icbc_cn",
                    "Industrial and Commercial Bank of China (Asia)": "icbc_hk",
                    "China Construction Bank": "ccb_cn",
                    "China Construction Bank (Asia)": "ccb_hk",
                }[label])

    def test_parser_normalizes_typed_trade_and_retains_exact_workbook_evidence(
        self,
    ) -> None:
        workbook_bytes = self._completed_workbook()

        payload = parse_investment_payload(
            "zircon_hk",
            "manual_xlsx",
            xlsx_bytes=workbook_bytes,
            filename="zircon-activity.xlsx",
        )

        self.assertEqual(payload["broker"], "zircon_hk")
        self.assertEqual(payload["summary"]["transaction_count"], 1)
        transaction = payload["transactions"][0]
        self.assertEqual(transaction["datetime"], "2026-07-30 09:15:00")
        self.assertEqual(transaction["ticker"], "700.HK")
        self.assertEqual(transaction["quantity_raw"], "100")
        self.assertEqual(transaction["price_raw"], "20")
        self.assertEqual(transaction["gross_amount_raw"], "-2000")
        self.assertEqual(transaction["commission_raw"], "-5")
        self.assertEqual(transaction["net_amount_raw"], "-2005")
        self.assertEqual(transaction["source"]["source_row"], 2)
        self.assertEqual(transaction["source"]["source_timezone"], "Asia/Hong_Kong")
        artifact = payload["source_artifacts"][0]
        self.assertEqual(
            base64.b64decode(artifact["content_base64"]),
            workbook_bytes,
        )

    def test_standard_export_round_trips_selected_ledger_rows(self) -> None:
        workbook_bytes = build_standard_investment_xlsx([
            {
                "ledger_no": 42,
                "broker": "ibkr",
                "account": "U1234567",
                "datetime": "2026-07-30 09:15:00",
                "type": "buy",
                "currency": "USD",
                "ticker": "AAPL",
                "quantity_raw": "2",
                "price_raw": "210.25",
                "commission_raw": "-1.25",
                "description": "Authoritative AAPL purchase",
                "source": {},
            },
        ])

        workbook = load_workbook(BytesIO(workbook_bytes))
        values = tuple(
            workbook["Transactions"].cell(row=2, column=column).value
            for column in range(1, len(ZIRCON_HK_HEADERS) + 1)
        )
        self.assertEqual(values[0], "IBKR")
        self.assertEqual(values[3], "Buy")
        self.assertEqual(values[5], "AAPL")
        self.assertRegex(values[11], r"^antigravity-[0-9a-f]{40}$")

        payload = build_investment_payload_from_zircon_hk_manual_xlsx(
            xlsx_bytes=workbook_bytes,
            filename="AAPL_standard_investment_export.xlsx",
        )
        transaction = payload["transactions"][0]
        self.assertEqual(transaction["broker"], "ibkr")
        self.assertEqual(transaction["account"], "U1234567")
        self.assertEqual(transaction["datetime"], "2026-07-30 09:15:00")
        self.assertEqual(transaction["quantity_raw"], "2")
        self.assertEqual(transaction["price_raw"], "210.25")
        self.assertEqual(transaction["commission_raw"], "-1.25")

    def test_standard_export_preserves_incomplete_and_reversed_cash_events(self) -> None:
        transactions = [
            {
                "ledger_no": 1,
                "broker": "usmart_hk",
                "account": "uSMART-1",
                "datetime": "2026-07-30 09:15:00",
                "type": "buy",
                "currency": "USD",
                "net_amount_raw": "-100",
                "description": "Fractional Shares Purchase (symbol unavailable in statement)",
                "source": {"file_kind": "usmart_hk_statement_pdf", "row_number": 1},
            },
            {
                "ledger_no": 2,
                "broker": "hsbc",
                "account": "HSBC-1",
                "datetime": "2026-07-30 09:16:00",
                "type": "withdrawal",
                "currency": "USD",
                "net_amount_raw": "40",
                "description": "Cancel Withdrawal",
                "source": {"file_kind": "hsbc_usd_account_text", "row_number": 2},
            },
            {
                "ledger_no": 3,
                "broker": "hsbc",
                "account": "HSBC-1",
                "datetime": "2026-07-30 09:17:00",
                "type": "deposit",
                "currency": "USD",
                "net_amount_raw": "-43.87",
                "description": "Returned cheque interest",
                "source": {"file_kind": "hsbc_usd_account_text", "row_number": 3},
            },
            {
                "ledger_no": 4,
                "broker": "futuhk",
                "account": "Futu-1",
                "datetime": "2026-07-30 09:18:00",
                "type": "dividend",
                "currency": "USD",
                "net_amount_raw": "1.65",
                "description": "SPLG 9.00000000 SHARES DIVIDENDS 0.18284148 USD PER SHARE",
                "source": {"file_kind": "futuhk_statement_pdf", "row_number": 4},
            },
            {
                "ledger_no": 5,
                "broker": "futuhk",
                "account": "Futu-1",
                "datetime": "2026-07-30 09:19:00",
                "type": "foreign_tax_withholding",
                "currency": "USD",
                "net_amount_raw": "-0.16",
                "description": "SPLG 9.00000000 SHARES WITHHOLDING TAX -0.01828836 USD PER SHARE",
                "source": {"file_kind": "futuhk_statement_pdf", "row_number": 5},
            },
            {
                "ledger_no": 6,
                "broker": "longbridge_hk",
                "account": "Longbridge-1",
                "datetime": "2026-07-30 09:20:00",
                "type": "adjustment",
                "currency": "HKD",
                "net_amount_raw": "0",
                "description": "Non-cash money-market valuation event",
                "source": {"file_kind": "longbridge_cash_flow", "row_number": 6},
            },
            {
                "ledger_no": 7,
                "broker": "longbridge_hk",
                "account": "Longbridge-1",
                "datetime": "2026-07-30 09:21:00",
                "type": "forex_trade_component",
                "currency": "HKD",
                "net_amount_raw": "-100",
                "description": "Currency Conversion (Debit)",
                "source": {"file_kind": "longbridge_cash_flow", "row_number": 7},
            },
        ]

        workbook_bytes = build_standard_investment_xlsx(transactions)
        workbook = load_workbook(BytesIO(workbook_bytes))
        rows = [
            tuple(
                workbook["Transactions"].cell(row=row, column=column).value
                for column in range(1, 13)
            )
            for row in range(2, len(transactions) + 2)
        ]
        self.assertEqual(rows[0][3], "Adjustment")
        self.assertIn("original Type 'Buy'", rows[0][10])
        self.assertEqual(rows[1][3], "Adjustment")
        self.assertEqual(rows[2][3], "Adjustment")
        self.assertEqual(rows[3][3], "Dividend")
        self.assertEqual(rows[3][5], "SPLG")
        self.assertEqual(rows[4][3], "Foreign tax withholding")
        self.assertEqual(rows[4][5], "SPLG")
        self.assertEqual(rows[5][8], 0)
        self.assertEqual(rows[6][3], "Adjustment")

        payload = build_investment_payload_from_zircon_hk_manual_xlsx(
            xlsx_bytes=workbook_bytes,
            filename="Standard_investment_export.xlsx",
        )
        self.assertEqual(payload["summary"]["transaction_count"], len(transactions))
        self.assertTrue(any(
            transaction["description"].startswith("[Standard XLSX fallback:")
            for transaction in payload["transactions"]
        ))

    def test_standard_export_uses_ibkr_statement_base_currency_for_blank_cash_rows(
        self,
    ) -> None:
        workbook_bytes = build_standard_investment_xlsx([
            {
                "ledger_no": 4601,
                "broker": "ibkr",
                "account": "U11131870",
                "datetime": "2026-01-01 20:00:00",
                "type": "deposit",
                "currency": "",
                "gross_amount_raw": "1284.5987154",
                "net_amount_raw": "1284.5987154",
                "description": "Electronic Fund Transfer",
                "source": {"broker": "ibkr", "file_kind": "transactions"},
            },
        ])

        workbook = load_workbook(BytesIO(workbook_bytes))
        self.assertEqual(workbook["Transactions"]["E2"].value, "USD")

        payload = build_investment_payload_from_zircon_hk_manual_xlsx(
            xlsx_bytes=workbook_bytes,
            filename="ibkr_cash_export.xlsx",
        )
        transaction = payload["transactions"][0]
        self.assertEqual(transaction["currency"], "USD")
        self.assertEqual(transaction["net_amount_raw"], "1284.5987154")

    def test_standard_export_rejects_blank_currency_without_ibkr_statement_evidence(
        self,
    ) -> None:
        with self.assertRaisesRegex(
            ValueError,
            r"Investment transaction 4601 uses unsupported standard XLSX currency ''",
        ):
            build_standard_investment_xlsx([
                {
                    "ledger_no": 4601,
                    "broker": "ibkr",
                    "account": "U11131870",
                    "datetime": "2026-01-01 20:00:00",
                    "type": "deposit",
                    "currency": "",
                    "net_amount_raw": "1284.5987154",
                    "description": "Electronic Fund Transfer",
                    "source": {"broker": "ibkr", "file_kind": "manual_xlsx"},
                },
            ])

    def test_standard_export_round_trips_in_kind_security_transfers_without_cash(
        self,
    ) -> None:
        workbook_bytes = build_standard_investment_xlsx([
            {
                "ledger_no": 5096,
                "broker": "ibkr",
                "account": "U11131870",
                "datetime": "2026-07-31 23:00:00",
                "type": "transfer_out",
                "currency": "USD",
                "ticker": "QQQI",
                "quantity_raw": "5",
                "price_raw": "52.68",
                "commission_raw": "0",
                "gross_amount_raw": "0",
                "net_amount_raw": "0",
                "description": "FOP security transfer out: QQQI",
                "source": {"reference_id": "ibkr-fop-20260731-qqqi"},
            },
            {
                "ledger_no": 5097,
                "broker": "schwab",
                "account": "Individual ...342",
                "datetime": "2026-07-31 23:00:00",
                "type": "transfer_in",
                "currency": "USD",
                "ticker": "QQQI",
                "quantity_raw": "5",
                "commission_raw": "0",
                "gross_amount_raw": "0",
                "net_amount_raw": "0",
                "description": "Security transfer in: QQQI",
                "source": {"reference_id": "schwab-transfer-20260731-qqqi"},
            },
        ])

        workbook = load_workbook(BytesIO(workbook_bytes))
        sheet = workbook["Transactions"]
        self.assertEqual(sheet["D2"].value, "Transfer Out")
        self.assertEqual(sheet["D3"].value, "Transfer In")
        self.assertEqual(sheet["G2"].value, 5)
        self.assertIsNone(sheet["I2"].value)
        self.assertIsNone(sheet["I3"].value)

        payload = build_investment_payload_from_zircon_hk_manual_xlsx(
            xlsx_bytes=workbook_bytes,
            filename="QQQI_standard_investment_export.xlsx",
        )
        exported_outbound, exported_inbound = payload["transactions"]
        self.assertEqual(exported_outbound["type"], "transfer_out")
        self.assertEqual(exported_outbound["broker"], "ibkr")
        self.assertEqual(exported_outbound["quantity_raw"], "5")
        self.assertEqual(exported_outbound["price_raw"], "52.68")
        self.assertEqual(exported_outbound["gross_amount_raw"], "0")
        self.assertEqual(exported_outbound["net_amount_raw"], "0")
        self.assertFalse(exported_outbound["normalized"]["is_cash_flow"])
        self.assertEqual(exported_outbound["normalized"]["side"], "sell")
        self.assertEqual(exported_inbound["type"], "transfer_in")
        self.assertEqual(exported_inbound["broker"], "schwab")
        self.assertEqual(exported_inbound["quantity_raw"], "5")
        self.assertEqual(exported_inbound["gross_amount_raw"], "0")
        self.assertEqual(exported_inbound["net_amount_raw"], "0")
        self.assertFalse(exported_inbound["normalized"]["is_cash_flow"])
        self.assertEqual(exported_inbound["normalized"]["side"], "buy")

    def test_parser_rejects_cash_consideration_for_in_kind_security_transfers(
        self,
    ) -> None:
        with self.assertRaisesRegex(
            ValueError,
            r"Transactions!I2.*blank or zero for an in-kind security transfer",
        ):
            build_investment_payload_from_zircon_hk_manual_xlsx(
                xlsx_bytes=self._completed_workbook(
                    overrides={
                        "D2": "Transfer Out",
                        "E2": "USD",
                        "F2": "QQQI",
                        "G2": 5,
                        "H2": 52.68,
                        "I2": 1,
                        "J2": None,
                    }
                ),
                filename="invalid-security-transfer.xlsx",
            )

    def test_parser_rejects_text_dates_numbers_formulas_and_wrong_cash_signs(
        self,
    ) -> None:
        invalid_cases = (
            ({"C2": "30 Jul 2026 21:15"}, r"Transactions!C2.*typed Excel date or date-time"),
            ({"G2": "one hundred"}, r"Transactions!G2.*numeric Excel cell"),
            ({"K2": "=WEBSERVICE(\"https://example.invalid\")"}, r"Transactions!K2.*formula"),
            ({"D2": "Sell", "I2": -1995}, r"Transactions!I2.*blank for trades"),
        )
        for overrides, message in invalid_cases:
            with self.subTest(overrides=overrides):
                with self.assertRaisesRegex(ValueError, message):
                    build_investment_payload_from_zircon_hk_manual_xlsx(
                        xlsx_bytes=self._completed_workbook(overrides=overrides),
                        filename="zircon-activity.xlsx",
                    )

    def test_parser_reports_each_invalid_cell_before_any_payload_is_returned(
        self,
    ) -> None:
        workbook_bytes = self._completed_workbook(
            overrides={
                "C2": "not a date",
                "A3": "Wrong broker",
                "C3": datetime(2026, 7, 30, 22, 0),
                "D3": "Deposit",
                "E3": "USD",
                "I3": 100,
            }
        )

        with self.assertRaises(ValueError) as raised:
            build_investment_payload_from_zircon_hk_manual_xlsx(
                xlsx_bytes=workbook_bytes,
                filename="zircon-activity.xlsx",
            )

        message = str(raised.exception)
        self.assertIn("Transactions!C2", message)
        self.assertIn("Transactions!A3", message)
        self.assertIn("Fix these 2 cells", message)

    def test_parser_ignores_rows_with_only_prefilled_broker_and_account(self) -> None:
        payload = build_investment_payload_from_zircon_hk_manual_xlsx(
            xlsx_bytes=self._completed_workbook(
                overrides={
                    "A3": "Zircon (HK)",
                    "B3": "Zircon practice account",
                    "B4": "Zircon practice account",
                }
            ),
            filename="manual-prefilled-rows.xlsx",
        )

        self.assertEqual(len(payload["transactions"]), 1)

    def test_parser_rejects_active_content_before_openpyxl_parsing(self) -> None:
        archive_buffer = BytesIO(build_zircon_hk_template_xlsx())
        with ZipFile(archive_buffer, mode="a", compression=ZIP_DEFLATED) as archive:
            archive.writestr("xl/vbaProject.bin", b"unsafe active content")

        with self.assertRaisesRegex(
            ValueError,
            "must not contain macros, embedded objects",
        ):
            build_investment_payload_from_zircon_hk_manual_xlsx(
                xlsx_bytes=archive_buffer.getvalue(),
                filename="zircon-active-content.xlsx",
            )

    def test_stable_reference_updates_a_corrected_manual_entry(self) -> None:
        initial = build_investment_payload_from_zircon_hk_manual_xlsx(
            xlsx_bytes=self._completed_workbook(),
            filename="zircon-initial.xlsx",
        )
        corrected = build_investment_payload_from_zircon_hk_manual_xlsx(
            xlsx_bytes=self._completed_workbook(
                overrides={"H2": 21}
            ),
            filename="zircon-corrected.xlsx",
        )

        merged = merge_investment_payloads(initial, corrected)

        self.assertEqual(len(merged["transactions"]), 1)
        self.assertEqual(merged["transactions"][0]["price_raw"], "21")
        self.assertEqual(merged["transactions"][0]["net_amount_raw"], "-2105")

    def test_date_only_defaults_to_2300_hong_kong_and_allows_other_brokers(
        self,
    ) -> None:
        payload = build_investment_payload_from_zircon_hk_manual_xlsx(
            xlsx_bytes=self._completed_workbook(
                overrides={
                    "A2": "HSBC",
                    "C2": datetime(2026, 7, 30),
                }
            ),
            filename="manual-investment.xlsx",
        )

        transaction = payload["transactions"][0]
        self.assertEqual(transaction["broker"], "hsbc")
        self.assertEqual(transaction["datetime"], "2026-07-30 11:00:00")
        self.assertTrue(
            transaction["source"]["source_time_defaulted_to_2300"]
        )
        self.assertEqual(payload["broker"], "hsbc")
        self.assertEqual(
            payload["source_artifacts"][0]["source_kind"],
            "manual_investment_xlsx",
        )

    def test_forex_conversion_requires_and_enriches_two_signed_currency_legs(
        self,
    ) -> None:
        payload = build_investment_payload_from_zircon_hk_manual_xlsx(
            xlsx_bytes=self._forex_workbook(),
            filename="manual-forex.xlsx",
        )

        self.assertEqual(len(payload["transactions"]), 2)
        transactions = {
            transaction["currency"]: transaction
            for transaction in payload["transactions"]
        }
        self.assertEqual(transactions["HKD"]["net_amount_raw"], "-650")
        self.assertEqual(transactions["USD"]["net_amount_raw"], "83.22")
        self.assertEqual(
            transactions["HKD"]["source"]["forex_pair_reference_id"],
            "fx-20260730-01",
        )
        self.assertEqual(transactions["HKD"]["source"]["forex_leg"], "sold")
        self.assertEqual(transactions["USD"]["source"]["forex_leg"], "acquired")
        self.assertEqual(transactions["USD"]["source"]["forex_pair"], "USD.HKD")
        self.assertEqual(
            transactions["USD"]["source"]["exchange_rate_convention"],
            "HKD per USD",
        )

    def test_forex_pair_uses_currency_scoped_stable_correction_identity(
        self,
    ) -> None:
        initial = build_investment_payload_from_zircon_hk_manual_xlsx(
            xlsx_bytes=self._forex_workbook(),
            filename="manual-forex-initial.xlsx",
        )
        corrected = build_investment_payload_from_zircon_hk_manual_xlsx(
            xlsx_bytes=self._forex_workbook(acquired_amount=83.25),
            filename="manual-forex-corrected.xlsx",
        )

        merged = merge_investment_payloads(initial, corrected)

        self.assertEqual(len(merged["transactions"]), 2)
        amounts = {
            transaction["currency"]: transaction["net_amount_raw"]
            for transaction in merged["transactions"]
        }
        self.assertEqual(amounts, {"HKD": "-650", "USD": "83.25"})

    def test_forex_pair_rejects_missing_or_mismatched_second_leg(self) -> None:
        invalid_workbooks = (
            (
                self._completed_workbook(
                    overrides={
                        "D2": "Forex trade component",
                        "E2": "HKD",
                        "F2": None,
                        "G2": None,
                        "H2": None,
                        "I2": -650,
                        "J2": None,
                        "L2": "fx-20260730-01",
                    }
                ),
                r"Transactions!L2.*requires exactly two",
            ),
            (
                self._forex_workbook(
                    acquired_datetime=datetime(2026, 7, 30, 21, 16)
                ),
                r"Transactions!L3.*same transaction date and time",
            ),
            (
                self._forex_workbook(acquired_amount=-83.22),
                r"Transactions!L3.*one negative sold-currency Amount",
            ),
        )
        for workbook_bytes, message in invalid_workbooks:
            with self.subTest(message=message):
                with self.assertRaisesRegex(ValueError, message):
                    build_investment_payload_from_zircon_hk_manual_xlsx(
                        xlsx_bytes=workbook_bytes,
                        filename="manual-forex-invalid.xlsx",
                    )

    def test_duplicate_reference_id_in_one_account_is_rejected(self) -> None:
        workbook_bytes = self._completed_workbook(
            overrides={
                "A3": "Zircon (HK)",
                "B3": "Zircon practice account",
                "C3": datetime(2026, 7, 30, 22, 0),
                "D3": "Deposit",
                "E3": "HKD",
                "I3": 100,
                "L3": "zircon-reference-1",
            }
        )

        with self.assertRaisesRegex(
            ValueError,
            r"Transactions!L3.*Reference ID must be unique",
        ):
            build_investment_payload_from_zircon_hk_manual_xlsx(
                xlsx_bytes=workbook_bytes,
                filename="zircon-duplicate-reference.xlsx",
            )


if __name__ == "__main__":
    unittest.main()
