"""Regression tests for Longbridge SG file imports.

Code version: v0.5.0
"""

from __future__ import annotations

import unittest
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from unittest.mock import patch
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook

import app.services.investment_import as investment_import_service
from app.services.investment_import import (
    LONGBRIDGE_SG_FLAT_HOLDING_TICKERS,
    LONGBRIDGE_SG_IMPORTER_VERSION,
    _build_longbridge_sg_order_records,
    _build_broker_reported_performance_calibrations,
    _longbridge_sg_fund_entry_to_cash_flow_row,
    _parse_longbridge_sg_eastern_datetime,
    _parse_longbridge_sg_fund_details_entries,
    _parse_longbridge_sg_history_orders_xlsx,
    _replay_holdings,
    build_investment_payload_from_longbridge_sg_files,
)


SOURCE_DIR = Path("/Users/example/Desktop/IBKR/Longbridge SG")
FUND_DETAILS_PATH = SOURCE_DIR / "SG99999999 Fund Details.txt"
HISTORY_ORDERS_PATH = SOURCE_DIR / "SG99999999 Historyorders.xlsx"
LONGBRIDGE_SG_EXPECTED_REALIZED_PNL = {
    "NVDA": "2.34",
    "TQQQ": "-1.11",
}
SYNTHETIC_PRIVATE_INVESTMENT_EVIDENCE = {
    "longbridge_sg_performance_calibrations": {
        "SG99999999": LONGBRIDGE_SG_EXPECTED_REALIZED_PNL,
    },
}


class LongbridgeSgImportTests(unittest.TestCase):
    def setUp(self) -> None:
        self.private_evidence_patcher = patch.object(
            investment_import_service,
            "_load_local_private_investment_evidence",
            return_value=SYNTHETIC_PRIVATE_INVESTMENT_EVIDENCE,
        )
        self.private_evidence_patcher.start()
        self.addCleanup(self.private_evidence_patcher.stop)

    def test_account_performance_calibrations_are_scoped_to_longbridge_sg(self) -> None:
        calibrations = _build_broker_reported_performance_calibrations(
            "longbridge_sg",
            "SG99999999",
        )

        self.assertEqual(
            {ticker: entry["realized_total"] for ticker, entry in calibrations.items()},
            LONGBRIDGE_SG_EXPECTED_REALIZED_PNL,
        )
        self.assertTrue(all(
            entry["realized_total_includes_nonperformance"] is True
            for entry in calibrations.values()
        ))
        self.assertEqual(
            sum(Decimal(entry["realized_total"]) for entry in calibrations.values()),
            Decimal("1.23"),
        )

    def test_eastern_filled_time_converts_to_utc_and_keeps_local_trade_date(self) -> None:
        parsed = _parse_longbridge_sg_eastern_datetime("2024-11-15 01:12:24 ET", market="US")
        self.assertIsNotNone(parsed)
        assert parsed is not None
        self.assertEqual(
            parsed.astimezone(__import__("datetime").timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
            "2024-11-15 06:12:24",
        )

        workbook_rows = _parse_longbridge_sg_history_orders_xlsx(_build_history_orders_workbook_bytes())
        order_row = next(
            row for row in workbook_rows if row.get("order_id") == "TEST-ORDER-SG-001"
        )
        warnings: list[str] = []
        order_records = _build_longbridge_sg_order_records([order_row], warnings)
        self.assertEqual(warnings, [])
        self.assertEqual(len(order_records), 1)
        self.assertEqual(order_records[0]["date"], "2024-11-15")
        self.assertEqual(order_records[0]["datetime"], "2024-11-15 06:12:24")

    def test_history_orders_rejects_missing_required_columns(self) -> None:
        workbook = Workbook()
        orders_sheet = workbook.active
        orders_sheet.title = "Orders"
        orders_sheet.append(["Order Status", "Market", "Symbol"])
        orders_sheet.append(["Filled", "US", "AAPL"])
        buffer = BytesIO()
        workbook.save(buffer)

        with self.assertRaisesRegex(
            ValueError,
            r'History Orders sheet "Orders" is missing required columns:.*avg price',
        ):
            _parse_longbridge_sg_history_orders_xlsx(buffer.getvalue())

    def test_history_orders_rejects_macros_before_pandas_reads_the_workbook(self) -> None:
        source_bytes = _build_history_orders_workbook_bytes()
        output = BytesIO()
        with ZipFile(BytesIO(source_bytes), "r") as source_archive:
            with ZipFile(output, "w", ZIP_DEFLATED) as output_archive:
                for member in source_archive.infolist():
                    output_archive.writestr(member, source_archive.read(member.filename))
                output_archive.writestr("xl/vbaProject.bin", b"not executable")

        with self.assertRaisesRegex(ValueError, r"must not contain macros"):
            _parse_longbridge_sg_history_orders_xlsx(output.getvalue())

    def test_source_files_replay_flat_tqqq_and_nvda_positions(self) -> None:
        if not FUND_DETAILS_PATH.is_file() or not HISTORY_ORDERS_PATH.is_file():
            self.skipTest("Longbridge (SG) source fixtures are unavailable on this machine.")

        payload = build_investment_payload_from_longbridge_sg_files(
            fund_details_text=FUND_DETAILS_PATH.read_text(encoding="utf-8"),
            history_orders_xlsx_bytes=HISTORY_ORDERS_PATH.read_bytes(),
            fund_details_filename=FUND_DETAILS_PATH.name,
            history_orders_filename=HISTORY_ORDERS_PATH.name,
        )
        replayed = _replay_holdings(payload["transactions"])
        for ticker in LONGBRIDGE_SG_FLAT_HOLDING_TICKERS:
            self.assertEqual(
                replayed.get(ticker, Decimal("0")),
                Decimal("0"),
                msg=f"Expected flat holdings for {ticker}, got {replayed.get(ticker)}",
            )

        order_ids = {
            txn.get("source", {}).get("order_id")
            for txn in payload["transactions"]
            if txn.get("type") in {"buy", "sell"}
        }
        self.assertNotIn("TEST-ORDER-SG-002", order_ids)

        fund_entry_count = len(
            _parse_longbridge_sg_fund_details_entries(
                FUND_DETAILS_PATH.read_text(encoding="utf-8")
            )
        )
        self.assertEqual(fund_entry_count, 56)
        self.assertEqual(payload["generator"]["fund_details_entry_count"], 56)
        self.assertEqual(payload["generator"]["version"], LONGBRIDGE_SG_IMPORTER_VERSION)
        self.assertEqual(payload["generator"]["accounted_fund_details_entry_count"], 56)
        self.assertEqual(payload["generator"]["unaccounted_fund_details_entry_count"], 0)
        self.assertEqual(payload["generator"]["matched_cash_flow_row_count"], 32)
        self.assertEqual(payload["generator"]["booking_date_applied_order_count"], 16)
        self.assertEqual(
            {
                ticker: entry["realized_total"]
                for ticker, entry in payload["performance_snapshot"].items()
            },
            LONGBRIDGE_SG_EXPECTED_REALIZED_PNL,
        )
        self.assertTrue(payload["summary"]["performance_snapshot_authoritative"])
        self.assertEqual(
            payload["summary"]["performance_snapshot_source"],
            "user_confirmed_broker_performance_calibration",
        )
        self.assertEqual(
            {artifact["bundle_role"] for artifact in payload["source_artifacts"]},
            {"fund_details", "history_orders"},
        )
        self.assertEqual(len(payload["transactions"]), 40)
        self.assertTrue(payload["summary"]["cash_reconciliation_matched"])
        self.assertTrue(all(
            Decimal(summary["difference"]) == Decimal("0")
            for summary in payload["summary"]["cash_reconciliation"].values()
        ))

        kol_rewards = [
            txn
            for txn in payload["transactions"]
            if txn.get("type") == "kol_reward"
        ]
        self.assertEqual(len(kol_rewards), 3)
        self.assertTrue(all("KOL Rewards" in str(txn.get("description", "")) for txn in kol_rewards))

        trade_cash_flow_count = sum(
            1
            for txn in payload["transactions"]
            if txn.get("source", {}).get("file_kind") == "longbridge_cash_flow"
            and txn.get("type") in {"buy", "sell"}
        )
        self.assertEqual(trade_cash_flow_count, 0)

        order = next(
            txn
            for txn in payload["transactions"]
            if txn.get("source", {}).get("order_id") == "TEST-ORDER-SG-001"
        )
        self.assertEqual(order["date"], "2024-11-15")
        self.assertEqual(order["datetime"], "2024-11-15 20:00:00")
        self.assertEqual(order["source"]["history_order_datetime"], "2024-11-15 06:12:24")
        self.assertEqual(order["source"]["fund_details_booking_date"], "2024-11-15")

        tqqq_dividends = [
            txn
            for txn in payload["transactions"]
            if txn.get("ticker") == "TQQQ" and txn.get("type") == "dividend"
        ]
        nvda_dividends = [
            txn
            for txn in payload["transactions"]
            if txn.get("ticker") == "NVDA" and txn.get("type") == "dividend"
        ]
        self.assertEqual(len(tqqq_dividends), 2)
        self.assertEqual(len(nvda_dividends), 2)
        self.assertEqual(
            sum(
                1
                for txn in payload["transactions"]
                if (txn.get("source") or {}).get("excluded_from_broker_pnl")
            ),
            4,
        )

    def test_fund_details_booking_date_uses_new_york_convention_without_splitting_fx_legs(self) -> None:
        booking_date = "2024-09-04"
        for currency, amount in (("HKD", Decimal("-600.00")), ("SGD", Decimal("-103.90")), ("USD", Decimal("76.62"))):
            row = _longbridge_sg_fund_entry_to_cash_flow_row({
                "flow_name": "Currency Conversion (Debit)" if amount < 0 else "Currency Conversion (Credit)",
                "description": f"FX FROM {currency} TO USD @ 0.1277",
                "balance": amount,
                "currency": currency,
                "time": booking_date,
                "symbol": "",
            })
            self.assertEqual(row["booking_date"], booking_date)
            self.assertTrue(str(row["time"]))

        if not FUND_DETAILS_PATH.is_file() or not HISTORY_ORDERS_PATH.is_file():
            self.skipTest("Longbridge (SG) source fixtures are unavailable on this machine.")

        payload = build_investment_payload_from_longbridge_sg_files(
            fund_details_text=FUND_DETAILS_PATH.read_text(encoding="utf-8"),
            history_orders_xlsx_bytes=HISTORY_ORDERS_PATH.read_bytes(),
            fund_details_filename=FUND_DETAILS_PATH.name,
            history_orders_filename=HISTORY_ORDERS_PATH.name,
        )
        target_amounts = {"-600.00", "76.62", "79.24", "-103.90"}
        fx_rows = [
            txn
            for txn in payload["transactions"]
            if txn.get("type") == "forex_trade_component"
            and str(txn.get("gross_amount_raw") or "") in target_amounts
        ]
        self.assertEqual(len(fx_rows), 4)
        for txn in fx_rows:
            self.assertEqual(
                txn["date"],
                "2024-09-04",
                msg=f"FX row {txn.get('currency')} should keep booking date 2024-09-04, got {txn.get('date')}",
            )


def _build_history_orders_workbook_bytes() -> bytes:
    workbook = Workbook()
    orders_sheet = workbook.active
    orders_sheet.title = "Orders"
    orders_sheet.append([
        "Order Status",
        "Market",
        "Symbol",
        "Stock Name",
        "Direction",
        "Order Type",
        "Quantity",
        "Price",
        "Trigger Price",
        "Currency",
        "Order Time",
        "Avg Price",
        "Executed Qty",
        "Turnover",
        "Remaining Orders",
        "Canceled/Invalid Orders",
        "Validity",
        "Session",
        "Trigger Status",
        "Order No.",
        "Reason",
    ])
    orders_sheet.append([
        "Filled",
        "US",
        "TQQQ",
        "Proshares UltraPro QQQ",
        "Buy",
        "Limit",
        1,
        78.5,
        "--",
        "USD",
        "2024-11-15 00:57:36 ET",
        78.5,
        1,
        78.5,
        0,
        0,
        "Valid for the day",
        "Overnight Trading",
        "Untriggered",
        "TEST-ORDER-SG-001",
        "Execution of 1",
    ])
    orders_sheet.append([
        "Expired",
        "US",
        "NVDA",
        "NVIDIA",
        "Sell",
        "Limit",
        1,
        142.0,
        "--",
        "USD",
        "2025-02-19 09:00:38 ET",
        0,
        0,
        0,
        0,
        1,
        "Valid for the day",
        "RTH+Pre/Post-Mkt",
        "Untriggered",
        "TEST-ORDER-SG-002",
        None,
    ])

    trade_log_sheet = workbook.create_sheet("Trade Log")
    trade_log_sheet.append([
        "Order No.",
        "Market",
        "Symbol",
        "Stock Name",
        "Direction",
        "Time",
        "Order Status",
        "Price",
        "Quantity",
        "Message",
    ])
    trade_log_sheet.append([
        "TEST-ORDER-SG-001",
        "US",
        "TQQQ",
        "Proshares UltraPro QQQ",
        "Buy",
        "2024-11-15 01:12:24 ET",
        "Filled",
        78.5,
        1,
        "Execution of 1",
    ])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


if __name__ == "__main__":
    unittest.main()
